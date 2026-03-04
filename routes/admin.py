"""
Admin Routes Blueprint
All admin panel functionality
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, Response
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
from models import db, User, Match, Tip, Round, Team, Achievement, UserAchievement, ExtraAnswer, System
from services.email_service import send_welcome_email_for_imported_user
from services.scoring import calculate_points, check_and_award_achievements, update_user_scores
from parsers.smart_parser import smart_parse_matches
from datetime import datetime, timedelta
from sqlalchemy import func, desc, or_
import json
import csv
import io
import base64

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


# Helper function - admin_required
def admin_required():
    """Check if current user is admin, redirect if not"""
    if not current_user.is_authenticated:
        flash("Musíš být přihlášený!", "error")
        return redirect(url_for("auth.login"))
    if not current_user.is_admin:
        flash("Nemáš admin oprávnění!", "error")
        return redirect(url_for("index"))


    @admin_bp.route("/users", methods=["GET", "POST"])
    @login_required
    def admin_users():
        admin_required()

        if request.method == "POST":
            action = request.form.get("bulk_action")
            user_ids = request.form.getlist("user_ids")

            if not user_ids:
                flash("Nevybral jsi žádné uživatele.", "error")
                return redirect(url_for("admin_users"))

            user_ids = [int(uid) for uid in user_ids]
            affected_users = User.query.filter(User.id.in_(user_ids)).all()

            # Ochrana proti smazání/úpravě ownera a sebe sama
            for u in affected_users:
                if u.is_owner:
                    flash(f"Nelze hromadně upravit ownera ({u.username}).", "error")
                    return redirect(url_for("admin_users"))
                if u.id == current_user.id:
                    flash("Nelze hromadně upravit sám sebe.", "error")
                    return redirect(url_for("admin_users"))

            if action == "delete":
                count = len(affected_users)
                for u in affected_users:
                    db.session.delete(u)
                db.session.commit()
                audit("users.bulk_delete", "User", None, count=count)
                flash(f"Smazáno {count} uživatelů.", "ok")

            elif action == "set_role":
                new_role = request.form.get("new_role")
                if new_role not in ["user", "viewer", "moderator", "admin"]:
                    flash("Neplatná role.", "error")
                    return redirect(url_for("admin_users"))

                for u in affected_users:
                    u.role = new_role
                db.session.commit()
                audit("users.bulk_role", "User", None, role=new_role, count=len(affected_users))
                flash(f"Změněna role pro {len(affected_users)} uživatelů na '{new_role}'.", "ok")

            elif action == "reset_password":
                new_password = request.form.get("new_password", "").strip()
                if not new_password:
                    flash("Zadej nové heslo.", "error")
                    return redirect(url_for("admin_users"))

                # Validace síly hesla
                is_valid, error_msg = validate_password(new_password)
                if not is_valid:
                    flash(error_msg, "error")
                    return redirect(url_for("admin_users"))

                for u in affected_users:
                    u.set_password(new_password)
                db.session.commit()
                audit("users.bulk_password", "User", None, count=len(affected_users))
                flash(f"Resetováno heslo pro {len(affected_users)} uživatelů.", "ok")

            elif action == "send_welcome_reset":
                # Pošli welcome email s reset linkem
                base_url = request.url_root.rstrip('/')
                sent_count = 0

                for u in affected_users:
                    try:
                        if send_welcome_with_reset_link(u, base_url):
                            sent_count += 1
                    except Exception as e:
                        continue

                audit("users.bulk_welcome", "User", None, count=sent_count)
                flash(f"📧 Welcome emaily odeslány: {sent_count} uživatelům. Mají 24h na nastavení hesla.", "ok")

            return redirect(url_for("admin_users"))

        users = User.query.order_by(User.username.asc()).all()
        users = [u for u in users if can_see_user_in_admin(u)]
        return render_page(r"""
<div class="card">
  <div class="row" style="justify-content:space-between;">
    <div>
      <h2 style="margin:0 0 8px 0;">Uživatelé</h2>
      <div class="muted">Owner = <b>{{ owner }}</b>. Tajný user je skrytý pro jiné adminy.</div>
    </div>
    <div class="row">
      {% if current_user.is_owner %}<span class="tag pill-ok">Owner admin</span>{% endif %}
      <a class="btn btn-primary" href="{{ url_for('admin_user_new') }}">➕ Nový uživatel</a>
      <a class="btn" href="{{ url_for('admin_users_import') }}" style="background:#667eea; color:white;">📤 Import uživatelů</a>
    </div>
  </div>

  <hr class="sep">

  {% if current_user.is_owner %}
  {# Hromadná správa #}
  <form method="post" id="bulkForm">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="card" style="background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); margin-bottom:16px; padding:16px;">
      <h3 style="margin:0 0 12px 0;">Hromadná správa</h3>

      <div style="display:grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap:10px;">
        {# Změna role #}
        <div>
          <select name="new_role" id="bulk_role" style="width:100%;">
            <option value="">-- Vybrat roli --</option>
            <option value="user">User</option>
            <option value="viewer">Viewer</option>
            <option value="moderator">Moderátor</option>
            <option value="admin">Admin</option>
          </select>
        </div>
        <button type="button" onclick="submitBulkAction('set_role')" class="btn btn-sm">
          👥 Změnit roli
        </button>

        {# Reset hesla #}
        <input type="password" name="new_password" id="bulk_password" placeholder="Nové heslo" style="width:100%;">
        <button type="button" onclick="submitBulkAction('reset_password')" class="btn btn-sm">
          🔑 Resetovat heslo
        </button>

        {# Poslat welcome email #}
        <div></div>
        <button type="button" onclick="submitBulkAction('send_welcome_reset')" class="btn btn-sm" style="background:rgba(110,168,254,0.3); color:#0b1020;">
          📧 Poslat welcome email
        </button>

        {# Smazat #}
        <div></div>
        <button type="button" onclick="submitBulkDelete()" class="btn btn-sm"
                style="background:rgba(255,77,109,0.2); color:#ff4d6d; border:1px solid rgba(255,77,109,0.4);">
          🗑️ Smazat vybrané
        </button>
      </div>

      <input type="hidden" name="bulk_action" id="bulk_action">
      <div class="muted" style="margin-top:10px; font-size:12px;">
        💡 Označ uživatele checkboxem, vyber akci a klikni na tlačítko<br>
        📧 Welcome email = pošle reset link pro nastavení hesla (24h platnost)
      </div>
    </div>

  <table class="datatable">
    <thead>
      <tr>
        <th style="width:40px;">
          <input type="checkbox" id="selectAll" onclick="toggleAll(this)"
                 title="Vybrat všechny">
        </th>
        <th>Username</th>
        <th>Email</th>
        <th style="text-align:center;">Role</th>
        <th style="text-align:center;">Registrace</th>
        <th style="text-align:center; min-width:400px;">Akce</th>
      </tr>
    </thead>
    <tbody>
      {% for u in users %}
      <tr>
        <td>
          {% if u.id != current_user.id and not u.is_owner %}
            <input type="checkbox" name="user_ids" value="{{ u.id }}" class="user-checkbox">
          {% endif %}
        </td>
        <td>
          <strong>{{ u.display_name }}</strong>
          {% if u.is_owner %} <span class="tag pill-ok">Owner</span>{% endif %}
          {% if u.nickname %}<br><span class="muted" style="font-size:11px;">Nick: {{ u.nickname }}</span>{% endif %}
          <div class="muted" style="font-size:11px;">Login: {{ u.username }}</div>
        </td>
        <td class="muted">{{ u.email }}</td>
        <td style="text-align:center;">
          {% if u.effective_role == 'admin' %}
            <span class="tag pill-bad">Admin</span>
          {% elif u.effective_role == 'moderator' %}
            <span class="tag pill-ok">Moderátor</span>
          {% elif u.effective_role == 'viewer' %}
            <span class="tag">Viewer</span>
          {% else %}
            <span class="tag pill-primary">User</span>
          {% endif %}
        </td>
        <td style="text-align:center;" class="muted">{{ u.created_at.strftime('%d.%m.%Y') }}</td>
        <td style="text-align:center;">
          <div style="display:flex; justify-content:center; gap:6px; flex-wrap:nowrap;">
              <a class="btn btn-sm" href="{{ url_for('admin_user_edit', user_id=u.id) }}">Upravit</a>
              <a class="btn btn-sm" href="{{ url_for('admin_user_reset_password', user_id=u.id) }}">Reset hesla</a>
              {% if u.id != current_user.id and not u.is_owner %}
                <a class="btn btn-sm" href="{{ url_for('admin_user_change_role', user_id=u.id) }}">Změnit roli</a>
                <form method="post" action="{{ url_for('admin_user_delete', user_id=u.id) }}" style="display:inline;"
                      onsubmit="return confirm('Opravdu smazat uživatele {{ u.username }}?')">
                  <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                  <button type="submit" class="btn btn-sm btn-danger"
                          style="background:rgba(255,77,109,0.2); color:#ff4d6d; border:none; cursor:pointer;">
                    Smazat
                  </button>
                </form>
              {% endif %}
          </div>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  </form>
  {% else %}
  {# Non-owner zobrazí jen tabulku bez hromadné správy #}
  <table class="datatable">
    <thead>
      <tr>
        <th>Username</th>
        <th>Email</th>
        <th style="text-align:center;">Role</th>
        <th style="text-align:center;">Registrace</th>
      </tr>
    </thead>
    <tbody>
      {% for u in users %}
      <tr>
        <td>
          <strong>{{ u.display_name }}</strong>
          {% if u.nickname %}<br><span class="muted" style="font-size:11px;">Nick: {{ u.nickname }}</span>{% endif %}
          <div class="muted" style="font-size:11px;">Login: {{ u.username }}</div>
        </td>
        <td class="muted">{{ u.email }}</td>
        <td style="text-align:center;">
          {% if u.effective_role == 'admin' %}
            <span class="tag pill-bad">Admin</span>
          {% elif u.effective_role == 'moderator' %}
            <span class="tag pill-ok">Moderátor</span>
          {% elif u.effective_role == 'viewer' %}
            <span class="tag">Viewer</span>
          {% else %}
            <span class="tag pill-primary">User</span>
          {% endif %}
        </td>
        <td style="text-align:center;" class="muted">{{ u.created_at.strftime('%d.%m.%Y') }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  {% endif %}
</div>

<script>
function toggleAll(checkbox) {
  const checkboxes = document.querySelectorAll('.user-checkbox');
  checkboxes.forEach(cb => cb.checked = checkbox.checked);
}

function getSelectedCount() {
  const checkboxes = document.querySelectorAll('.user-checkbox:checked');
  return checkboxes.length;
}

function submitBulkAction(action) {
  const count = getSelectedCount();
  if (count === 0) {
    alert('Nevybral jsi žádné uživatele!');
    return;
  }

  let confirmMsg = '';
  if (action === 'set_role') {
    const role = document.getElementById('bulk_role').value;
    if (!role) {
      alert('Vyber roli!');
      return;
    }
    confirmMsg = `Opravdu změnit roli pro ${count} uživatelů na '${role}'?`;
  } else if (action === 'reset_password') {
    const password = document.getElementById('bulk_password').value;
    if (!password) {
      alert('Zadej nové heslo!');
      return;
    }
    confirmMsg = `Opravdu resetovat heslo pro ${count} uživatelů?`;
  } else if (action === 'send_welcome_reset') {
    confirmMsg = `Poslat welcome email ${count} uživatelům?\n\nKaždý dostane email s odkazem pro nastavení hesla (platnost 24h).`;
  }

  if (confirmMsg && confirm(confirmMsg)) {
    document.getElementById('bulk_action').value = action;
    document.getElementById('bulkForm').submit();
  }
}

function submitBulkDelete() {
  const count = getSelectedCount();
  if (count === 0) {
    alert('Nevybral jsi žádné uživatele!');
    return;
  }

  if (confirm(`POZOR: Opravdu smazat ${count} uživatelů?\n\nTato akce je NEVRATNÁ!`)) {
    document.getElementById('bulk_action').value = 'delete';
    document.getElementById('bulkForm').submit();
  }
}
</script>

<style>
  .btn-danger {
    background: rgba(255, 77, 109, 0.15);
    color: #ff4d6d;
  }
  .btn-danger:hover {
    background: rgba(255, 77, 109, 0.25);
  }
</style>
""", users=users, owner=OWNER_ADMIN_EMAIL)



    @admin_bp.route("/user/<int:user_id>/toggle-admin")
    @login_required
    def admin_toggle_admin(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)
        if u.id == current_user.id:
            flash("Sebe si nepřepínej.", "error")
            return redirect(url_for("admin_users"))
        u.is_admin = not u.is_admin
        db.session.commit()
        audit("user.toggle_admin", "User", u.id, is_admin=u.is_admin)
        return redirect(url_for("admin_users"))



    @admin_bp.route("/user/new", methods=["GET", "POST"])
    @login_required
    def admin_user_new():
        admin_required()
        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            email = (request.form.get("email") or "").strip().lower()
            password = (request.form.get("password") or "").strip()
            role = (request.form.get("role") or "user").strip()
            first_name = (request.form.get("first_name") or "").strip()
            last_name = (request.form.get("last_name") or "").strip()
            nickname = (request.form.get("nickname") or "").strip()

            if not username or not email or not password:
                flash("Vyplň všechna povinná pole.", "error")
                return redirect(url_for("admin_user_new"))

            # Validace síly hesla
            is_valid, error_msg = validate_password(password)
            if not is_valid:
                flash(error_msg, "error")
                return redirect(url_for("admin_user_new"))

            # Kontrola, jestli už username nebo email neexistuje
            if User.query.filter_by(username=username).first():
                flash(f"Username '{username}' už existuje.", "error")
                return redirect(url_for("admin_user_new"))

            if User.query.filter_by(email=email).first():
                flash(f"Email '{email}' už existuje.", "error")
                return redirect(url_for("admin_user_new"))

            # Vytvoření nového uživatele
            new_user = User(
                username=username,
                email=email,
                role=role,
                first_name=first_name or None,
                last_name=last_name or None,
                nickname=nickname or None,
                email_verified=True  # Admin vytvořil uživatele, email je ověřený
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            audit("user.create", "User", new_user.id, username=username, email=email, role=role)
            flash(f"Uživatel '{username}' byl vytvořen.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Nový uživatel</h2>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Username (login) *</label>
      <input name="username" placeholder="username" required>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Email *</label>
      <input name="email" type="email" placeholder="email@example.com" required>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Heslo *</label>
      <input name="password" type="password" placeholder="Min. 8 znaků" required minlength="8">
      <div class="muted" style="font-size:12px; margin-top:4px;">
        Požadavky: min. 8 znaků, velké/malé písmeno, číslo
      </div>
    </div>
    <div class="grid2">
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Jméno</label>
        <input name="first_name" placeholder="Jan">
      </div>
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Příjmení</label>
        <input name="last_name" placeholder="Novák">
      </div>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Přezdívka (Nick)</label>
      <input name="nickname" placeholder="JN23">
      <div class="muted" style="font-size:12px; margin-top:4px;">Zobrazí se v žebříčku</div>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Role</label>
      <select name="role">
        <option value="user">User (může tipovat)</option>
        <option value="viewer">Viewer (jen prohlížet)</option>
        <option value="moderator">Moderátor (může tipovat + částečná správa)</option>
        <option value="admin">Admin (plná správa)</option>
      </select>
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Vytvořit</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
  </form>
</div>
""")



    @admin_bp.route("/users/import", methods=["GET", "POST"])
    @login_required
    def admin_users_import():
        """Import uživatelů z Excel souboru - krok 1: Upload"""
        admin_required()

        if request.method == "POST":
            if 'file' not in request.files:
                flash("Žádný soubor nebyl vybrán.", "error")
                return redirect(url_for("admin_users_import"))

            file = request.files['file']
            if file.filename == '':
                flash("Žádný soubor nebyl vybrán.", "error")
                return redirect(url_for("admin_users_import"))

            if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
                flash("Podporovány jsou pouze .xlsx, .xls a .csv soubory.", "error")
                return redirect(url_for("admin_users_import"))

            try:
                import openpyxl
                import pandas as pd
                import pickle

                # Načti soubor
                if file.filename.endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                # Očekávané sloupce
                required_cols = ['email', 'username', 'password']
                missing_cols = [col for col in required_cols if col not in df.columns]

                if missing_cols:
                    flash(f"Chybí povinné sloupce: {', '.join(missing_cols)}", "error")
                    return redirect(url_for("admin_users_import"))

                # Parsuj a validuj všechny uživatele
                users_preview = []

                for index, row in df.iterrows():
                    email = str(row['email']).strip().lower() if pd.notna(row['email']) else ''
                    username = str(row['username']).strip() if pd.notna(row['username']) else ''
                    password = str(row['password']).strip() if pd.notna(row['password']) else ''

                    # Validace
                    errors = []
                    status = 'ok'

                    if not email or not username or not password:
                        errors.append("Prázdné povinné pole")
                        status = 'error'

                    # Kontrola duplicit
                    if email and User.query.filter_by(email=email).first():
                        errors.append(f"Email '{email}' už existuje")
                        status = 'duplicate'

                    if username and User.query.filter_by(username=username).first():
                        errors.append(f"Username '{username}' už existuje")
                        status = 'duplicate'

                    # Validace hesla
                    if password and status == 'ok':
                        is_valid, error_msg = validate_password(password)
                        if not is_valid:
                            errors.append(error_msg)
                            status = 'error'

                    # Volitelné sloupce
                    first_name = str(row.get('first_name', '')).strip() if pd.notna(row.get('first_name')) else None
                    last_name = str(row.get('last_name', '')).strip() if pd.notna(row.get('last_name')) else None
                    nickname = str(row.get('nickname', '')).strip() if pd.notna(row.get('nickname')) else None
                    role = str(row.get('role', 'user')).strip().lower() if pd.notna(row.get('role')) else 'user'

                    if role not in ['user', 'viewer', 'moderator', 'admin']:
                        role = 'user'

                    users_preview.append({
                        'row_num': index + 2,  # +2 protože 1=header, Excel je 1-indexed
                        'email': email,
                        'username': username,
                        'password': password,
                        'first_name': first_name,
                        'last_name': last_name,
                        'nickname': nickname,
                        'role': role,
                        'status': status,
                        'errors': errors
                    })

                # Počet podle statusů
                ok_count = sum(1 for u in users_preview if u['status'] == 'ok')
                error_count = sum(1 for u in users_preview if u['status'] in ['error', 'duplicate'])

                # Ulož do temp file
                preview_fd, preview_path = tempfile.mkstemp(suffix='.pkl', prefix='users_import_')
                os.close(preview_fd)

                with open(preview_path, 'wb') as f:
                    pickle.dump(users_preview, f)

                session['users_import_preview_file'] = preview_path

                return redirect(url_for("admin_users_import_preview"))

            except Exception as e:
                flash(f"Chyba při zpracování souboru: {str(e)}", "error")
                return redirect(url_for("admin_users_import"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Import uživatelů - Krok 1</h2>
  <div class="muted">Nahraj Excel nebo CSV soubor s uživateli</div>
  <hr class="sep">

  <div style="background:rgba(110,168,254,0.08); padding:16px; border-radius:8px; margin-bottom:16px;">
    <h3 style="margin:0 0 8px 0;">📋 Formát souboru</h3>
    <p><strong>Povinné sloupce:</strong></p>
    <ul>
      <li><code>email</code> - Email uživatele</li>
      <li><code>username</code> - Uživatelské jméno (login)</li>
      <li><code>password</code> - Heslo (min. 8 znaků, velké/malé/číslo)</li>
    </ul>
    <p><strong>Volitelné sloupce:</strong></p>
    <ul>
      <li><code>first_name</code> - Jméno</li>
      <li><code>last_name</code> - Příjmení</li>
      <li><code>nickname</code> - Přezdívka (zobrazí se v žebříčku)</li>
      <li><code>role</code> - Role (user/viewer/moderator/admin, default: user)</li>
    </ul>
  </div>

  <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; gap:16px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    <div style="background:rgba(110,168,254,0.1); padding:12px; border-left:4px solid #6ea8fe; margin-bottom:8px;">
      <strong>💡 Nemáš Excel?</strong> Stáhni si šablonu s hlavičkami a příkladem:
      <a href="{{ url_for('admin_users_import_template') }}" class="btn btn-sm" style="margin-left:8px;">
        📥 Stáhnout šablonu Excel
      </a>
    </div>

    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Soubor (.xlsx, .xls, .csv)</label>
      <input type="file" name="file" accept=".xlsx,.xls,.csv" required>
    </div>

    <div class="row" style="gap:10px;">
      <button class="btn btn-primary" type="submit">➡️ Pokračovat na Preview</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
  </form>

  <hr class="sep">

  <div class="muted" style="font-size:12px;">
    <p><strong>💡 Postup:</strong></p>
    <ol>
      <li><strong>Upload</strong> - Nahraj Excel soubor</li>
      <li><strong>Preview</strong> - Zkontroluj kdo se bude importovat a vyber komu poslat email</li>
      <li><strong>Potvrzení</strong> - Importuj vybrané uživatele</li>
    </ol>
  </div>
</div>
""")



    @admin_bp.route("/users/import/template")
    @login_required
    def admin_users_import_template():
        """Stáhne Excel šablonu pro import uživatelů"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            # Vytvoř workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Uživatelé"

            # Hlavička
            headers = ['email', 'username', 'password', 'first_name', 'last_name', 'nickname', 'role']
            ws.append(headers)

            # Stylování hlavičky
            header_fill = PatternFill(start_color="6EA8FE", end_color="6EA8FE", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Příklad řádku (volitelně, pro ukázku)
            example_row = [
                'jan.novak@email.cz',
                'jan123',
                'Test1234',
                'Jan',
                'Novák',
                'JN',
                'user'
            ]
            ws.append(example_row)

            # Nastavení šířky sloupců
            ws.column_dimensions['A'].width = 25  # email
            ws.column_dimensions['B'].width = 15  # username
            ws.column_dimensions['C'].width = 15  # password
            ws.column_dimensions['D'].width = 15  # first_name
            ws.column_dimensions['E'].width = 15  # last_name
            ws.column_dimensions['F'].width = 12  # nickname
            ws.column_dimensions['G'].width = 12  # role

            # Přidej poznámku do druhého sheetu
            notes_ws = wb.create_sheet("Poznámky")
            notes_ws['A1'] = "NÁVOD K POUŽITÍ"
            notes_ws['A1'].font = Font(bold=True, size=14)

            notes = [
                "",
                "POVINNÉ SLOUPCE:",
                "• email - Email uživatele (musí být unikátní)",
                "• username - Uživatelské jméno pro login (musí být unikátní)",
                "• password - Heslo (min. 8 znaků, velké/malé písmeno, číslice)",
                "",
                "VOLITELNÉ SLOUPCE:",
                "• first_name - Křestní jméno",
                "• last_name - Příjmení",
                "• nickname - Přezdívka (zobrazí se v žebříčku)",
                "• role - Role (user/viewer/moderator/admin, výchozí: user)",
                "",
                "POSTUP:",
                "1. Vyplň uživatele do 1. sheetu (Uživatelé)",
                "2. První řádek s ukázkou můžeš smazat nebo upravit",
                "3. Ulož soubor",
                "4. Nahraj na stránce Import uživatelů",
                "",
                "PŘÍKLAD HESEL:",
                "✅ Test1234 - OK (8+ znaků, velké, malé, číslice)",
                "✅ Password123 - OK",
                "❌ test - Špatně (moc krátké, bez velkého, bez čísla)",
                "❌ TestTest - Špatně (bez čísla)",
            ]

            for i, note in enumerate(notes, 2):
                notes_ws[f'A{i}'] = note

            notes_ws.column_dimensions['A'].width = 60

            # Ulož do BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='import_uzivatelu_sablona.xlsx'
            )

        except Exception as e:
            flash(f"Chyba při vytváření šablony: {str(e)}", "error")
            return redirect(url_for("admin_users_import"))



    @admin_bp.route("/users/import/preview")
    @login_required
    def admin_users_import_preview():
        """Import uživatelů - krok 2: Preview a výběr komu poslat email"""
        admin_required()

        preview_file = session.get('users_import_preview_file')
        if not preview_file or not os.path.exists(preview_file):
            flash("Preview data nenalezena. Nahraj soubor znovu.", "error")
            return redirect(url_for("admin_users_import"))

        try:
            import pickle
            with open(preview_file, 'rb') as f:
                users_preview = pickle.load(f)

            ok_users = [u for u in users_preview if u['status'] == 'ok']
            error_users = [u for u in users_preview if u['status'] in ['error', 'duplicate']]

            return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Import uživatelů - Preview</h2>
  <div class="muted">Zkontroluj kdo se bude importovat</div>
  <hr class="sep">

  <div class="row" style="gap:12px; margin-bottom:16px;">
    <div class="tag pill-ok">✅ K importu: {{ ok_count }}</div>
    {% if error_count > 0 %}<div class="tag pill-error">❌ Chyby: {{ error_count }}</div>{% endif %}
  </div>

  <form method="post" action="{{ url_for('admin_users_import_confirm') }}">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    {% if ok_users %}
    <h3 style="margin:16px 0 8px 0;">✅ Připraveno k importu ({{ ok_count }})</h3>
    <div style="background:#f8f9fa; padding:8px; border-radius:8px; margin-bottom:16px; max-height:400px; overflow-y:auto;">
      <table style="width:100%; font-size:13px;">
        <thead style="position:sticky; top:0; background:#e9ecef;">
          <tr>
            <th style="padding:8px; text-align:left;">Email</th>
            <th style="padding:8px; text-align:left;">Username</th>
            <th style="padding:8px; text-align:left;">Jméno</th>
            <th style="padding:8px; text-align:left;">Role</th>
          </tr>
        </thead>
        <tbody>
          {% for user in ok_users %}
          <tr style="border-bottom:1px solid #dee2e6;">
            <td style="padding:8px;">{{ user.email }}</td>
            <td style="padding:8px;">{{ user.username }}</td>
            <td style="padding:8px;">{{ user.first_name or '' }} {{ user.last_name or '' }}</td>
            <td style="padding:8px;"><span class="tag">{{ user.role }}</span></td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <div style="background:rgba(110,168,254,0.1); padding:12px; border-left:4px solid #6ea8fe; margin-bottom:16px;">
      <strong>💡 Info:</strong> Uživatelé budou importováni s ověřeným emailem a budou se moci přihlásit.
      Welcome emaily můžeš poslat později z Admin → Uživatelé → Hromadná správa.
    </div>
    {% endif %}

    {% if error_users %}
    <h3 style="margin:16px 0 8px 0;">❌ Chyby - nebudou importováni ({{ error_count }})</h3>
    <div style="background:#fff3cd; padding:12px; border-radius:8px; margin-bottom:16px; max-height:300px; overflow-y:auto;">
      {% for user in error_users %}
      <div style="padding:8px; border-bottom:1px solid #ffc107;">
        <strong>Řádek {{ user.row_num }}:</strong> {{ user.email }} / {{ user.username }}
        <div class="muted" style="font-size:12px;">
          {% for error in user.errors %}• {{ error }}<br>{% endfor %}
        </div>
      </div>
      {% endfor %}
    </div>
    {% endif %}

    <hr class="sep">

    <div class="row" style="gap:10px;">
      <button class="btn btn-primary" type="submit" {% if ok_count == 0 %}disabled{% endif %}>
        ✅ Potvrdit import ({{ ok_count }} uživatelů)
      </button>
      <a class="btn" href="{{ url_for('admin_users_import') }}">⬅️ Zpět na upload</a>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
  </form>
</div>
""", ok_users=ok_users, error_users=error_users, ok_count=len(ok_users), error_count=len(error_users))

        except Exception as e:
            flash(f"Chyba při načítání preview: {str(e)}", "error")
            return redirect(url_for("admin_users_import"))



    @admin_bp.route("/users/import/confirm", methods=["POST"])
    @login_required
    def admin_users_import_confirm():
        """Import uživatelů - krok 3: Potvrzení a provedení importu"""
        admin_required()

        preview_file = session.get('users_import_preview_file')
        if not preview_file or not os.path.exists(preview_file):
            flash("Preview data nenalezena. Nahraj soubor znovu.", "error")
            return redirect(url_for("admin_users_import"))

        try:
            import pickle
            with open(preview_file, 'rb') as f:
                users_preview = pickle.load(f)

            created = 0
            ok_users = [u for u in users_preview if u['status'] == 'ok']

            for user_data in ok_users:
                try:
                    # Vytvoř uživatele
                    new_user = User(
                        email=user_data['email'],
                        username=user_data['username'],
                        first_name=user_data['first_name'],
                        last_name=user_data['last_name'],
                        nickname=user_data['nickname'],
                        role=user_data['role'],
                        email_verified=True  # Importovaní users mají ověřený email
                    )
                    new_user.set_password(user_data['password'])
                    db.session.add(new_user)
                    db.session.flush()

                    audit("user.imported", "User", new_user.id, username=new_user.username, email=new_user.email)
                    created += 1

                except Exception as e:
                    # Rollback tohoto usera, pokračuj dál
                    db.session.rollback()
                    continue

            db.session.commit()

            # Cleanup temp file
            try:
                os.remove(preview_file)
                session.pop('users_import_preview_file', None)
            except:
                pass

            # Flash výsledek
            flash(f"✅ Importováno: {created} uživatelů. Pro poslání welcome emailů jdi do Uživatelé → Hromadná správa.", "ok")
            return redirect(url_for("admin_users"))

        except Exception as e:
            flash(f"Chyba při importu: {str(e)}", "error")
            return redirect(url_for("admin_users_import"))



    @admin_bp.route("/user/<int:user_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_user_edit(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if request.method == "POST":
            new_username = (request.form.get("username") or "").strip()
            new_email = (request.form.get("email") or "").strip().lower()
            new_first_name = (request.form.get("first_name") or "").strip()
            new_last_name = (request.form.get("last_name") or "").strip()
            new_nickname = (request.form.get("nickname") or "").strip()

            if not new_username or not new_email:
                flash("Vyplň username a email.", "error")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            # Kontrola duplicit
            existing_user = User.query.filter_by(username=new_username).first()
            if existing_user and existing_user.id != u.id:
                flash(f"Username '{new_username}' už existuje.", "error")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            existing_email = User.query.filter_by(email=new_email).first()
            if existing_email and existing_email.id != u.id:
                flash(f"Email '{new_email}' už existuje.", "error")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            old_username = u.username
            old_email = u.email
            u.username = new_username
            u.email = new_email
            u.first_name = new_first_name or None
            u.last_name = new_last_name or None
            u.nickname = new_nickname or None
            db.session.commit()
            audit("user.update", "User", u.id, old_username=old_username, new_username=new_username, old_email=old_email, new_email=new_email)
            flash(f"Uživatel aktualizován.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Upravit uživatele</h2>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Username (login) *</label>
      <input name="username" value="{{ u.username }}" required>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Email *</label>
      <input name="email" type="email" value="{{ u.email }}" required>
    </div>
    <div class="grid2">
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Jméno</label>
        <input name="first_name" value="{{ u.first_name or '' }}">
      </div>
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Příjmení</label>
        <input name="last_name" value="{{ u.last_name or '' }}">
      </div>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Přezdívka (Nick)</label>
      <input name="nickname" value="{{ u.nickname or '' }}">
      <div class="muted" style="font-size:12px; margin-top:4px;">
        Zobrazí se v žebříčku. Současný nick: <strong>{{ u.display_name }}</strong>
      </div>
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Uložit</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
  </form>
</div>
""", u=u)



    @admin_bp.route("/user/<int:user_id>/reset-password", methods=["GET", "POST"])
    @login_required
    def admin_user_reset_password(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if request.method == "POST":
            new_password = (request.form.get("password") or "").strip()
            if not new_password:
                flash("Vyplň nové heslo.", "error")
                return redirect(url_for("admin_user_reset_password", user_id=user_id))

            # Validace síly hesla
            is_valid, error_msg = validate_password(new_password)
            if not is_valid:
                flash(error_msg, "error")
                return redirect(url_for("admin_user_reset_password", user_id=user_id))

            u.set_password(new_password)
            db.session.commit()
            audit("user.reset_password", "User", u.id)
            flash(f"Heslo pro '{u.username}' bylo resetováno.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Resetovat heslo</h2>
  <div class="muted">Uživatel: <b>{{ u.username }}</b> ({{ u.email }})</div>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Nové heslo *</label>
      <input name="password" type="password" placeholder="Min. 8 znaků" required minlength="8">
      <div class="muted" style="font-size:12px; margin-top:4px;">
        Požadavky: min. 8 znaků, velké/malé písmeno, číslo
      </div>
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Resetovat heslo</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
  </form>
</div>
""", u=u)



    @admin_bp.route("/user/<int:user_id>/change-role", methods=["GET", "POST"])
    @login_required
    def admin_user_change_role(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if u.id == current_user.id:
            flash("Nemůžeš měnit roli sám sobě.", "error")
            return redirect(url_for("admin_users"))

        if request.method == "POST":
            new_role = (request.form.get("role") or "user").strip()
            old_role = u.effective_role

            # Nastavení role
            if new_role == "admin":
                u.is_admin = True
                u.role = "admin"
            else:
                u.is_admin = False
                u.role = new_role

            db.session.commit()
            audit("user.change_role", "User", u.id, old_role=old_role, new_role=new_role)
            flash(f"Role uživatele '{u.username}' změněna na '{new_role}'.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Změnit roli</h2>
  <div class="muted">Uživatel: <b>{{ u.username }}</b> ({{ u.email }})</div>
  <div class="muted">Současná role: <b>{{ u.effective_role }}</b></div>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Nová role</label>
      <select name="role">
        <option value="user" {% if u.effective_role == 'user' %}selected{% endif %}>User (může tipovat)</option>
        <option value="viewer" {% if u.effective_role == 'viewer' %}selected{% endif %}>Viewer (jen prohlížet)</option>
        <option value="moderator" {% if u.effective_role == 'moderator' %}selected{% endif %}>Moderátor (může tipovat + částečná správa)</option>
        <option value="admin" {% if u.effective_role == 'admin' %}selected{% endif %}>Admin (plná správa)</option>
      </select>
    </div>
    <div class="muted" style="margin-top:10px; padding:10px; background:rgba(249,199,79,0.1); border-radius:8px;">
      <b>Vysvětlení rolí:</b><br>
      • <b>User</b> - může tipovat zápasy a odpovídat na extra otázky<br>
      • <b>Viewer</b> - může jen prohlížet žebříček a výsledky, nemůže tipovat<br>
      • <b>Moderátor</b> - může tipovat + má přístup k některým admin funkcím<br>
      • <b>Admin</b> - plný přístup ke všem funkcím
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Změnit roli</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
  </form>
</div>
""", u=u)



    @admin_bp.route("/user/<int:user_id>/delete", methods=["POST"])
    @login_required
    def admin_user_delete(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if u.id == current_user.id:
            flash("Nemůžeš smazat sám sebe.", "error")
            return redirect(url_for("admin_users"))

        username = u.username
        # Smazat všechny tipy uživatele
        Tip.query.filter_by(user_id=u.id).delete()
        # Smazat všechny odpovědi na extra otázky
        ExtraAnswer.query.filter_by(user_id=u.id).delete()
        # Smazat uživatele
        db.session.delete(u)
        db.session.commit()
        audit("user.delete", "User", user_id, username=username)
        flash(f"Uživatel '{username}' byl smazán.", "ok")
        return redirect(url_for("admin_users"))

    # --- ADMIN ROUNDS ---
    # --- ADMIN SMART IMPORT ---
    # =========================================================
    # ULTRA SMART IMPORT - AI-POWERED MATCH PARSING
    # =========================================================

    def extract_text_from_screenshot(image_data: bytes) -> Optional[str]:
        """
        Extract text from screenshot using Tesseract OCR (FREE)

        Args:
            image_data: Image file bytes (PNG, JPEG, etc)

        Returns:
            Extracted text or None if failed
        """

        if not TESSERACT_AVAILABLE:
            print("❌ Tesseract not available")
            return None

        try:
            # Open image
            image = Image.open(io.BytesIO(image_data))

            # Preprocessing for better OCR
            # Convert to grayscale
            image = image.convert('L')

            # Optional: increase contrast
            from PIL import ImageEnhance
            enhancer = ImageEnhance.Contrast(image)
            image = enhancer.enhance(2.0)

            # Extract text
            # Use custom config for better table recognition
            custom_config = r'--oem 3 --psm 6'
            text = pytesseract.image_to_string(image, lang='ces+eng', config=custom_config)

            print(f"✅ OCR extracted {len(text)} characters")
            print(f"📝 First 200 chars: {text[:200]}")

            return text.strip()

        except Exception as e:
            print(f"❌ OCR error: {e}")
            return None


    def _split_joined_lines(text: str) -> str:
        """
        CRITICAL FIX: Table copy/paste joins all lines

        New strategy: Extract complete match patterns
        Pattern: DD. MM. YYYY + text + time/score (N:N or NN:NN)
        """

        # Skip if already properly formatted
        lines = text.split('\n')
        if len(lines) > 3:
            return text

        # Rejoin text
        text_joined = ' '.join(lines)

        # Pattern for complete match:
        # Date + any text + final number:number
        # Use non-greedy match and lookahead
        pattern = r'(\d{1,2}\.\s*\d{1,2}\.\s*\d{4}.*?\d{1,2}:\d{1,2})(?=\d{1,2}\.\s*\d{1,2}\.\s*\d{4}|$)'

        matches = re.findall(pattern, text_joined)

        if len(matches) >= 2:
            # Add any header before first match
            parts = []
            first_match_pos = text_joined.find(matches[0])
            if first_match_pos > 0:
                header = text_joined[:first_match_pos].strip()
                if header:
                    parts.append(header)

            # Add all matches
            for match in matches:
                parts.append(match.strip())

            if len(parts) > 1:
                result = '\n'.join(parts)
                print(f"🔧 Extracted {len(parts)} lines from joined input")
                return result

        return text




    def _split_joined_lines(text: str) -> str:
        """
        🔧 FIX: Multiple matches joined on one line (from table copy/paste)

        Example:
        Input:  "27. 2. 2026DuklaSlavia0:228. 2. 2026Liberec..."
        Output: Each match on separate line

        Finds all dates and splits before each (except first)
        """

        result_lines = []

        for line in text.strip().split('\n'):
            line = line.strip()
            if not line:
                continue

            # Find all date patterns in this line
            date_pattern = r'\d{1,2}\.\s*\d{1,2}\.\s*\d{4}'
            matches = list(re.finditer(date_pattern, line))

            if len(matches) <= 1:
                # 0 or 1 date - keep as is
                result_lines.append(line)
                continue

            # Multiple dates found - split before each (except first)
            print(f"🔧 Found {len(matches)} dates in one line - splitting...")

            parts = []
            last_end = 0

            for i, match in enumerate(matches):
                if i == 0:
                    # First date - take from start until next date
                    if len(matches) > 1:
                        next_start = matches[1].start()
                        parts.append(line[last_end:next_start])
                        last_end = next_start
                    else:
                        parts.append(line[last_end:])
                else:
                    # Subsequent dates
                    if i < len(matches) - 1:
                        next_start = matches[i+1].start()
                        parts.append(line[last_end:next_start])
                        last_end = next_start
                    else:
                        # Last date - take rest of line
                        parts.append(line[last_end:])

            # Add all parts
            for part in parts:
                part = part.strip()
                if part:
                    result_lines.append(part)

        return '\n'.join(result_lines)


    def _clean_ocr_artifacts(text: str) -> str:
        """
        Clean up OCR artifacts from screenshot text

        OCR often adds extra characters like:
        - "vice >" or "více>" (from web UI buttons)
        - "|" pipes
        - Extra whitespace
        - Weird unicode characters

        Returns cleaned text ready for parsing
        """
        lines = text.strip().split('\n')
        cleaned_lines = []

        for line in lines:
            # Remove common OCR artifacts
            # Remove "vice >", "více>", "vic>", "víc>" and similar
            line = re.sub(r'\s*v[ií]ce?\s*>?\s*', ' ', line, flags=re.IGNORECASE)

            # Remove pipe symbols (often from table borders)
            line = re.sub(r'\s*\|\s*', ' ', line)

            # Remove extra whitespace between words (but keep spaces)
            line = re.sub(r'\s+', ' ', line)

            # Remove leading/trailing whitespace
            line = line.strip()

            if line:
                cleaned_lines.append(line)

        result = '\n'.join(cleaned_lines)
        print(f"🧹 OCR cleanup: {len(lines)} lines → {len(cleaned_lines)} cleaned lines")
        return result

    def _parse_multiline_app_format(text: str) -> List[Dict]:
        """
        Parse multi-line format from app/website:

        07.03. 15:00
        Bohemians
        Slovan Liberec
        --
        """
        matches = []
        blocks = text.split('--')
        current_year = 2026

        for block in blocks:
            lines = [line.strip() for line in block.strip().split('\n') if line.strip()]

            if len(lines) < 3:
                continue

            # Skip "25. kolo" headers
            if lines[0].lower().endswith('kolo'):
                lines = lines[1:]

            if len(lines) < 3:
                continue

            datetime_line = lines[0]
            home_team = lines[1]
            away_team = lines[2]

            # Parse "07.03. 15:00"
            match = re.match(r'(\d{2})\.(\d{2})\.\s+(\d{1,2}):(\d{2})', datetime_line)
            if not match:
                continue

            day = int(match.group(1))
            month = int(match.group(2))
            hour = int(match.group(3))
            minute = int(match.group(4))

            try:
                dt = datetime(current_year, month, day, hour, minute)
                matches.append({
                    'home_team': home_team.strip(),
                    'away_team': away_team.strip(),
                    'start_time': dt,
                })
            except ValueError:
                continue

        return matches

    def smart_parse_matches(text: str, round_id: int = None) -> List[Dict]:
        """
        🤖 ULTRA SMART PARSER V2 - Better whitespace handling + OCR cleanup

        Podporované formáty:
        - Table format: "27. 2. 2026DuklaSlavia18:00"
        - With spaces: "27. 2. 2026 Sparta - Slavia 18:00"
        - Date + time first: "27. 2. 20:00 Sparta vs Slavia"
        - CSV, Fortuna, UEFA, etc.
        - OCR from screenshots (with cleanup)



    @admin_bp.route("/smart-import", methods=["GET", "POST"])
    @login_required
    def admin_smart_import():
        """
        🤖 ULTRA SMART IMPORT

        Nakopíruj zápasy odkudkoliv → AI je naparsuje → Preview → Import
        """
        admin_required()

        rounds = Round.query.order_by(Round.name).all()

        if request.method == "POST":
            action = request.form.get("action", "parse")

            if action == "parse":
                # KROK 1: Parsování
                round_id = request.form.get("round_id")
                import_mode = request.form.get("import_mode", "fixtures")

                # Check if screenshot pasted from clipboard
                screenshot_data = request.form.get('screenshot_data', '')

                if screenshot_data and screenshot_data.startswith('data:image'):
                    # Clipboard screenshot mode
                    print("📸 Clipboard screenshot detected")

                    try:
                        # Extract base64 data
                        import base64
                        # Format: data:image/png;base64,iVBORw0KG...
                        base64_data = screenshot_data.split(',')[1]
                        image_data = base64.b64decode(base64_data)

                        # Extract text using OCR
                        text = extract_text_from_screenshot(image_data)

                        if not text:
                            flash("❌ Nepodařilo se přečíst screenshot. Zkus paste text ručně.", "error")
                            return redirect(url_for("admin_smart_import"))

                        print(f"✅ OCR extracted text ({len(text)} chars)")

                    except Exception as e:
                        print(f"❌ Screenshot processing error: {e}")
                        flash(f"❌ Chyba při zpracování screenshotu: {e}", "error")
                        return redirect(url_for("admin_smart_import"))
                else:
                    # Text mode (default)
                    text = request.form.get("raw_text", "")

                if not text.strip():
                    flash("❌ Zadej nějaký text k parsování!", "error")
                    return redirect(url_for("admin_smart_import"))

                # Parse!
                matches = smart_parse_matches(text, round_id=int(round_id) if round_id else None)

                if not matches:
                    flash("❌ Nepodařilo se naparsovat žádné zápasy. Zkus jiný formát.", "error")
                    return redirect(url_for("admin_smart_import"))

                # Normalize team names
                if round_id:
                    for m in matches:
                        m['home_team'] = normalize_team_name(m['home_team'], int(round_id))
                        m['away_team'] = normalize_team_name(m['away_team'], int(round_id))

                # Store preview server-side (DB) – v cookie jen ID (spolehlivé i na multi-worker hostingu)
                try:
                    payload = json.dumps({"round_id": round_id, "import_mode": import_mode, "matches": matches}, ensure_ascii=False)
                    imp = ImportSession(user_id=current_user.id, kind="smart_import_matches", payload_json=payload)
                    db.session.add(imp)
                    db.session.commit()
                    session['smart_import_session_id'] = imp.id
                    session['import_round_id'] = round_id
                except Exception as e:
                    db.session.rollback()
                    flash(f"❌ Nepodařilo se uložit preview importu: {e}", "error")
                    return redirect(url_for("admin_smart_import"))

                flash(f"✅ Naparsováno {len(matches)} zápasů! Zkontroluj a uprav pokud potřeba.", "ok")
                return redirect(url_for("admin_smart_import") + "?preview=1")

            elif action == "import":
                # KROK 2: Import do DB
                matches_json = request.form.get("matches_data", "[]")
                round_id = request.form.get("round_id")
                import_mode = request.form.get("import_mode", "fixtures")

                if not round_id:
                    flash("❌ Vyber soutěž/kolo!", "error")
                    return redirect(url_for("admin_smart_import"))

                try:
                    matches = json.loads(matches_json)
                except:
                    flash("❌ Chyba při parsování dat!", "error")
                    return redirect(url_for("admin_smart_import"))

                # Get round first
                r = db.session.get(Round, int(round_id))
                if not r:
                    flash("❌ Kolo nenalezeno!", "error")
                    return redirect(url_for("admin_smart_import"))

                # Store round data before expunging
                round_id_int = r.id
                round_name = r.name

                # CRITICAL: Expunge all objects from session to avoid pollution
                db.session.expunge_all()

                print(f"🔍 Using round_id: {round_id_int}, round_name: {round_name}")

                # Import!
                imported = 0
                errors = []
                skipped = 0

                for match_data in matches:
                    print(f"\n🔍 Processing match_data: {match_data}")
                    print(f"🔍 match_data type: {type(match_data)}")
                    print(f"🔍 match_data keys: {match_data.keys() if isinstance(match_data, dict) else 'NOT A DICT!'}")

                    try:
                        home_team = match_data.get('home_team', '').strip()
                        away_team = match_data.get('away_team', '').strip()

                        print(f"🔍 Extracted teams: home='{home_team}' ({type(home_team)}), away='{away_team}' ({type(away_team)})")

                        if not home_team or not away_team:
                            print(f"⚠️ Přeskakuji zápas - chybí tým: home='{home_team}', away='{away_team}'")
                            skipped += 1
                            continue

                        # Start time
                        start_time = None
                        if match_data.get('start_time'):
                            try:
                                start_time_str = match_data['start_time']
                                if isinstance(start_time_str, str):
                                    start_time = datetime.fromisoformat(start_time_str)
                                elif isinstance(start_time_str, datetime):
                                    start_time = start_time_str
                                else:
                                    print(f"⚠️ Neznámý typ start_time: {type(start_time_str)}")
                                    start_time = None
                            except Exception as e:
                                print(f"⚠️ Chyba parsování času '{match_data.get('start_time')}': {e}")
                                start_time = None

                        # Scores
                        home_score = match_data.get('home_score')
                        away_score = match_data.get('away_score')

                        # Import mode handling
                        if import_mode == 'fixtures':
                            home_score = None
                            away_score = None

                        # Validate scores are integers or None
                        if home_score is not None:
                            try:
                                home_score = int(home_score)
                            except:
                                home_score = None

                        if away_score is not None:
                            try:
                                away_score = int(away_score)
                            except:
                                away_score = None

                        # Debug log before creating Match
                        print(f"🔍 Match data:")
                        print(f"   round_id={round_id_int} (type={type(round_id_int)})")
                        print(f"   home_team={home_team} (type={type(home_team)})")
                        print(f"   away_team={away_team} (type={type(away_team)})")
                        print(f"   start_time={start_time} (type={type(start_time)})")
                        print(f"   home_score={home_score} (type={type(home_score)})")
                        print(f"   away_score={away_score} (type={type(away_score)})")

                        # Ensure all values are correct types
                        if not isinstance(round_id_int, int):
                            print(f"❌ round_id není int: {type(round_id_int)}")
                            continue

                        if not isinstance(home_team, str) or not isinstance(away_team, str):
                            print(f"❌ Týmy nejsou string")
                            continue

                        if start_time is not None and not isinstance(start_time, datetime):
                            print(f"❌ start_time není datetime: {type(start_time)}")
                            start_time = None

                        if home_score is not None and not isinstance(home_score, int):
                            print(f"❌ home_score není int: {type(home_score)}")
                            home_score = None

                        if away_score is not None and not isinstance(away_score, int):
                            print(f"❌ away_score není int: {type(away_score)}")
                            away_score = None

                                                # Results mode: must have scores
                        if import_mode == 'results':
                            if home_score is None or away_score is None:
                                skipped += 1
                                continue

                        # Create match
                        # Resolve Team objects (Match expects FK IDs, not strings)
                        home_team_obj = Team.query.filter_by(
                            round_id=round_id_int,
                            name=home_team,
                            is_deleted=False
                        ).first()
                        if not home_team_obj:
                            home_team_obj = Team(round_id=round_id_int, name=home_team)
                            db.session.add(home_team_obj)
                            db.session.flush()

                        away_team_obj = Team.query.filter_by(
                            round_id=round_id_int,
                            name=away_team,
                            is_deleted=False
                        ).first()
                        if not away_team_obj:
                            away_team_obj = Team(round_id=round_id_int, name=away_team)
                            db.session.add(away_team_obj)
                            db.session.flush()

                        # Results mode: try to find existing match and update scores
                        if import_mode == 'results':
                            q = Match.query.filter_by(round_id=round_id_int, home_team_id=home_team_obj.id, away_team_id=away_team_obj.id)
                            if start_time is not None:
                                exact = q.filter(Match.start_time == start_time).first()
                                if exact:
                                    exact.home_score = home_score
                                    exact.away_score = away_score
                                    imported += 1
                                    print(f"✅ Aktualizován výsledek (exact): {home_team} - {away_team} {home_score}:{away_score}")
                                    continue
                            existing = q.order_by(Match.start_time.asc()).first()
                            if existing:
                                existing.home_score = home_score
                                existing.away_score = away_score
                                imported += 1
                                print(f"✅ Aktualizován výsledek: {home_team} - {away_team} {home_score}:{away_score}")
                                continue
                            print(f"⚠️ Nenalezen existující zápas pro výsledek: {home_team} - {away_team}")
                            skipped += 1
                            continue

                        m = Match(
                            round_id=round_id_int,
                            home_team_id=home_team_obj.id,
                            away_team_id=away_team_obj.id,
                            start_time=start_time,
                            home_score=home_score,
                            away_score=away_score,
                        )

                        print(f"🔍 Match object created: {m}")
                        print(f"🔍 Match type: {type(m)}")
                        print(f"🔍 Match.__dict__: {m.__dict__}")

                        db.session.add(m)

                        # Flush immediately to catch errors
                        try:
                            db.session.flush()
                            imported += 1
                            print(f"✅ Přidán zápas: {home_team} - {away_team}")
                        except Exception as flush_error:
                            db.session.rollback()
                            error_msg = f"{home_team} - {away_team}: FLUSH ERROR: {str(flush_error)}"
                            errors.append(error_msg)
                            print(f"❌ Flush error: {flush_error}")
                            print(f"❌ Match object: {m.__dict__}")
                            continue

                    except Exception as e:
                        error_msg = f"{home_team} - {away_team}: {str(e)}"
                        errors.append(error_msg)
                        print(f"❌ Chyba při importu: {error_msg}")

                try:
                    db.session.commit()
                    print(f"✅ DB commit úspěšný: {imported} zápasů")
                except Exception as e:
                    db.session.rollback()
                    print(f"❌ DB commit selhal: {e}")
                    flash(f"❌ Chyba při ukládání do databáze: {str(e)}", "error")
                    return redirect(url_for("admin_smart_import"))

                # Clear session
                session.pop('parsed_matches', None)
                session.pop('import_round_id', None)

                if errors:
                    flash(f"⚠️ Importováno {imported} zápasů, přeskočeno {skipped}, {len(errors)} chyb: {', '.join(errors[:3])}", "warning")
                elif skipped > 0:
                    flash(f"✅ Importováno {imported} zápasů, přeskočeno {skipped} (chybějící týmy)", "ok")
                else:
                    flash(f"✅ Úspěšně importováno {imported} zápasů!", "ok")

                audit("smart_import", "Match", None, count=imported, round=round_name)

                return redirect(url_for("admin_rounds"))

        # GET request
        preview_mode = request.args.get("preview") == "1"
        parsed_matches: List[Dict[str, Any]] = []
        import_round_id = session.get('import_round_id')
        import_mode = 'fixtures'

        if preview_mode:
            sid = session.get('smart_import_session_id')
            if sid:
                imp = db.session.get(ImportSession, int(sid))
                if imp and imp.user_id == current_user.id and imp.kind == "smart_import_matches":
                    try:
                        payload = json.loads(imp.payload_json or "{}")
                        import_round_id = payload.get("round_id", import_round_id)
                        import_mode = payload.get('import_mode', import_mode) or import_mode
                        parsed_matches = payload.get("matches", []) or []
                    except Exception:
                        parsed_matches = []

        return render_template_string(SMART_IMPORT_TEMPLATE,
                                      rounds=rounds,
                                      preview_mode=preview_mode,
                                      parsed_matches=parsed_matches,
                                      import_round_id=import_round_id,
                                      import_mode=import_mode)


    # Template pro Smart Import
    SMART_IMPORT_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>🤖 Smart Import | Tipovačka</title>
  <style>
    * { margin: 0; padding: 0; box-sizing: border-box; }

    body {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
      min-height: 100vh;
      padding: 20px;
      color: #e0e0e0;
    }

    .container {
      max-width: 1200px;
      margin: 0 auto;
    }

    .card {
      background: rgba(30, 30, 46, 0.95);
      border-radius: 16px;
      padding: 32px;
      box-shadow: 0 20px 60px rgba(0,0,0,0.5);
      margin-bottom: 24px;
      border: 1px solid rgba(255,255,255,0.1);
      backdrop-filter: blur(10px);
    }

    h1 {
      font-size: 32px;
      margin-bottom: 8px;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
    }

    .subtitle {
      color: #a0a0a0;
      margin-bottom: 24px;
      font-size: 16px;
    }

    .form-group {
      margin-bottom: 20px;
    }

    label {
      display: block;
      margin-bottom: 8px;
      font-weight: 600;
      color: #e0e0e0;
    }

    .muted {
      color: #888;
      font-size: 14px;
      margin-top: 4px;
    }

    select, textarea, input[type="text"], input[type="number"], input[type="datetime-local"] {
      width: 100%;
      padding: 12px;
      border: 2px solid rgba(255,255,255,0.1);
      border-radius: 8px;
      font-size: 16px;
      font-family: inherit;
      transition: all 0.2s;
      background: rgba(20, 20, 30, 0.8);
      color: #e0e0e0;
    }

    select:focus, textarea:focus, input:focus {
      outline: none;
      border-color: #667eea;
      background: rgba(20, 20, 30, 0.95);
    }

    textarea {
      min-height: 300px;
      font-family: 'Monaco', 'Courier New', monospace;
      resize: vertical;
      line-height: 1.6;
    }

    .btn {
      padding: 14px 28px;
      border: none;
      border-radius: 8px;
      font-size: 16px;
      font-weight: 600;
      cursor: pointer;
      transition: all 0.2s;
      display: inline-block;
    }

    .btn-primary {
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      color: white;
    }

    .btn-primary:hover {
      transform: translateY(-2px);
      box-shadow: 0 8px 24px rgba(102, 126, 234, 0.4);
    }

    .btn-success {
      background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
      color: white;
    }

    .btn-success:hover {
      transform: translateY(-2px);
      box-shadow: 0 8px 24px rgba(46, 204, 113, 0.4);
    }

    .btn-secondary {
      background: rgba(255,255,255,0.1);
      color: #e0e0e0;
      border: 1px solid rgba(255,255,255,0.2);
    }

    .btn-secondary:hover {
      background: rgba(255,255,255,0.15);
    }

    .examples {
      background: rgba(20, 20, 30, 0.6);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 20px;
      border: 1px solid rgba(255,255,255,0.05);
    }

    .examples h3 {
      margin-bottom: 12px;
      font-size: 18px;
      color: #667eea;
    }

    .example-item {
      background: rgba(10, 10, 15, 0.8);
      padding: 8px 12px;
      border-radius: 4px;
      margin-bottom: 8px;
      font-family: 'Monaco', 'Courier New', monospace;
      font-size: 14px;
      border-left: 4px solid #667eea;
      color: #a0a0a0;
    }

    .preview-table {
      width: 100%;
      border-collapse: collapse;
      margin-top: 20px;
    }

    .preview-table th {
      background: rgba(102, 126, 234, 0.2);
      padding: 12px;
      text-align: left;
      font-weight: 600;
      border-bottom: 2px solid #667eea;
      color: #e0e0e0;
    }

    .preview-table td {
      padding: 10px 12px;
      border-bottom: 1px solid rgba(255,255,255,0.05);
    }

    .preview-table input[type="text"],
    .preview-table input[type="number"],
    .preview-table input[type="datetime-local"],
    .preview-table input[type="checkbox"] {
      padding: 8px;
      font-size: 14px;
      background: rgba(20, 20, 30, 0.8);
      border: 1px solid rgba(255,255,255,0.1);
      color: #e0e0e0;
    }

    .preview-table input:focus {
      border-color: #667eea;
      background: rgba(20, 20, 30, 0.95);
    }

    .radio-group {
      display: flex;
      gap: 20px;
      margin-bottom: 20px;
      flex-wrap: wrap;
    }

    .radio-group label {
      display: flex;
      align-items: center;
      gap: 8px;
      cursor: pointer;
      padding: 10px 16px;
      background: rgba(20, 20, 30, 0.6);
      border-radius: 8px;
      border: 2px solid rgba(255,255,255,0.1);
      transition: all 0.2s;
      margin-bottom: 0;
      font-weight: normal;
    }

    .radio-group label:hover {
      border-color: #667eea;
      background: rgba(20, 20, 30, 0.8);
    }

    .radio-group input[type="radio"] {
      width: auto;
      margin: 0;
    }

    .radio-group input[type="radio"]:checked + span {
      color: #667eea;
      font-weight: 600;
    }

    .flash {
      padding: 12px 20px;
      border-radius: 8px;
      margin-bottom: 20px;
      border-left: 4px solid;
    }

    .flash.ok {
      background: rgba(46, 204, 113, 0.1);
      border-color: #2ecc71;
      color: #2ecc71;
    }

    .flash.error {
      background: rgba(231, 76, 60, 0.1);
      border-color: #e74c3c;
      color: #e74c3c;
    }

    .flash.warning {
      background: rgba(241, 196, 15, 0.1);
      border-color: #f1c40f;
      color: #f1c40f;
    }

    .back-link {
      color: #667eea;
      text-decoration: none;
      display: inline-block;
      margin-bottom: 20px;
      font-weight: 600;
    }

    .back-link:hover {
      text-decoration: underline;
    }

    @media (max-width: 768px) {
      body { padding: 10px; }
      .card { padding: 20px; }
      h1 { font-size: 24px; }
      .radio-group { flex-direction: column; }
    }
      .paste-zone {
      position: relative;
    }

    .paste-instruction {
      text-align: center;
      padding: 20px;
      background: rgba(102, 126, 234, 0.1);
      border-radius: 8px;
      margin-bottom: 16px;
      border: 2px dashed rgba(102, 126, 234, 0.3);
    }

    .paste-zone.has-image .paste-instruction {
      display: none;
    }
  </style>
</head>
<body>
  <div class="container">
    <a href="{{ url_for('admin_dashboard') }}" class="back-link">← Zpět na Admin</a>

    {% with messages = get_flashed_messages(with_categories=True) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="flash {{ category }}">{{ message }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    {% if not preview_mode %}
    <div class="card">
      <h1>🤖 Smart Import V2</h1>
      <div class="subtitle">Nakopíruj zápasy odkudkoliv → AI je naparsuje → Preview → Import</div>

      <form method="post" >
        <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
        <input type="hidden" name="action" value="parse"/>

        <div class="form-group">
          <label>Soutěž / Kolo *</label>
          <select name="round_id" required>
            <option value="">-- Vyber soutěž --</option>
            {% for r in rounds %}
            <option value="{{ r.id }}" {% if import_round_id and import_round_id|string == r.id|string %}selected{% endif %}>
              {{ r.name }}
            </option>
            {% endfor %}
          </select>
        </div>

        <div class="form-group">
          <label>Import Mode</label>
          <div class="radio-group">
            <label>
              <input type="radio" name="import_mode" value="fixtures" {% if import_mode != 'results' %}checked{% endif %}>
              <span>📅 Fixtures (rozpisy)</span>
            </label>
            <label>
              <input type="radio" name="import_mode" value="results" {% if import_mode == 'results' %}checked{% endif %}>
              <span>🎯 Results (výsledky)</span>
            </label>
          </div>
          <div class="muted">
            Fixtures = nové zápasy bez skóre | Results = aktualizuje skóre existujících
          </div>
        </div>

        <div class="form-group">
          <label>Zápasy - screenshot NEBO text</label>

          <!-- Screenshot Upload -->


          <div class="examples">
            <h3>💡 Podporované formáty:</h3>
            <div class="example-item">27. 2. 2026DuklaSlavia18:00</div>
            <div class="example-item">27. 2. 2026 Sparta - Slavia 18:00</div>
            <div class="example-item">27. 2. 20:00 Sparta vs Slavia</div>
            <div class="example-item">1. 3. SpartaOstrava20:00 (auto rok)</div>
            <div class="example-item">Sparta - Slavia 2:1</div>
          </div>
          <div class="paste-zone" id="pasteZone">
            <div id="pasteInstruction" class="paste-instruction">
              📸 <strong>Napiš nebo vlož screenshot</strong>
              <div class="muted" style="margin-top: 8px;">
                Win+Shift+S → Ctrl+V zde → Auto OCR ✨
              </div>
            </div>

            <textarea name="raw_text"
                      id="rawText"
                      placeholder="Paste text NEBO screenshot (Ctrl+V)..."
                      style="min-height: 300px;"></textarea>

            <input type="hidden" name="screenshot_data" id="screenshotData">

            <div id="imagePreview" style="display: none; margin-top: 12px;">
              <div style="color: #2ecc71; margin-bottom: 8px;">
                ✅ Screenshot načten - OCR se spustí při parsování
              </div>
              <img id="previewImg" style="max-width: 100%; max-height: 200px; border-radius: 8px; border: 2px solid #667eea;">
              <button type="button" onclick="clearPastedImage()" class="btn btn-secondary" style="margin-top: 8px;">
                ❌ Smazat screenshot
              </button>
            </div>
          </div>
          <div class="muted">💡 Tip: Headers jako "24. kolo" se automaticky přeskočí</div>
        </div>

        <button type="submit" class="btn btn-primary">🔍 Parsovat & Preview</button>
      </form>
    </div>
    {% else %}
    <div class="card">
      <h1>🔍 Preview - zkontroluj a uprav</h1>
      <div class="subtitle">Naparsováno {{ parsed_matches|length }} zápasů</div>

      <form method="post" id="importForm">
        <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
        <input type="hidden" name="action" value="import"/>
        <input type="hidden" name="round_id" value="{{ import_round_id }}"/>
        <input type="hidden" name="import_mode" value="{{ import_mode }}"/>
        <input type="hidden" name="matches_data" id="matchesData" value=""/>

        <table class="preview-table">
          <thead>
            <tr>
              <th style="width: 40px;">
                <input type="checkbox" id="selectAll" checked>
              </th>
              <th>Domácí</th>
              <th>Hosté</th>
              <th style="width: 80px;">Skóre D</th>
              <th style="width: 80px;">Skóre H</th>
              <th style="width: 200px;">Datum & Čas</th>
            </tr>
          </thead>
          <tbody>
            {% for match in parsed_matches %}
            <tr>
              <td>
                <input type="checkbox" class="match-select" checked>
              </td>
              <td>
                <input type="text"
                       class="home-team"
                       value="{{ match.home_team }}"
                       data-original="{{ match.home_team }}">
              </td>
              <td>
                <input type="text"
                       class="away-team"
                       value="{{ match.away_team }}"
                       data-original="{{ match.away_team }}">
              </td>
              <td>
                <input type="number"
                       class="home-score"
                       value="{{ match.home_score if match.home_score is not none else '' }}"
                       min="0"
                       placeholder="-">
              </td>
              <td>
                <input type="number"
                       class="away-score"
                       value="{{ match.away_score if match.away_score is not none else '' }}"
                       min="0"
                       placeholder="-">
              </td>
              <td>
                <input type="datetime-local"
                       class="start-time"
                       value="{{ match.start_time[:16] if match.start_time else '' }}">
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>

        <div style="margin-top: 24px; display: flex; gap: 12px;">
          <button type="submit" class="btn btn-success">✅ Importovat vybrané zápasy</button>
          <a href="{{ url_for('admin_smart_import') }}" class="btn btn-secondary">← Zpět</a>
        </div>
      </form>
    </div>
    {% endif %}
  </div>

  <script>
    // Select all checkbox
    document.getElementById('selectAll')?.addEventListener('change', function() {
      document.querySelectorAll('.match-select').forEach(cb => cb.checked = this.checked);
    });

    // Collect data before submit
    document.getElementById('importForm')?.addEventListener('submit', function(e) {
      const rows = document.querySelectorAll('.preview-table tbody tr');
      const matches = [];

      rows.forEach(row => {
        if (row.querySelector('.match-select').checked) {
          const homeTeam = row.querySelector('.home-team').value;
          const awayTeam = row.querySelector('.away-team').value;
          const homeScore = row.querySelector('.home-score').value;
          const awayScore = row.querySelector('.away-score').value;
          const startTime = row.querySelector('.start-time').value;

          matches.push({
            home_team: homeTeam,
            away_team: awayTeam,
            home_score: homeScore ? parseInt(homeScore) : null,
            away_score: awayScore ? parseInt(awayScore) : null,
            start_time: startTime || null
          });
        }
      });

      document.getElementById('matchesData').value = JSON.stringify(matches);
    });
  </script>

  <script>
    // Image preview
    document.getElementById('screenshot')?.addEventListener('change', function(e) {
      const file = e.target.files[0];
      if (file) {
        const reader = new FileReader();
        reader.onload = function(e) {
          document.getElementById('previewImg').src = e.target.result;
          document.getElementById('preview').style.display = 'block';
          document.getElementById('uploadZone').querySelector('.upload-label').style.display = 'none';
        };
        reader.readAsDataURL(file);
      }
    });

    function clearImage() {
      document.getElementById('screenshot').value = '';
      document.getElementById('preview').style.display = 'none';
      document.getElementById('uploadZone').querySelector('.upload-label').style.display = 'block';
    }

    // Drag & drop
    const uploadZone = document.getElementById('uploadZone');
    if (uploadZone) {
      ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
        uploadZone.addEventListener(eventName, preventDefaults, false);
      });

      function preventDefaults(e) {
        e.preventDefault();
        e.stopPropagation();
      }

      ['dragenter', 'dragover'].forEach(eventName => {
        uploadZone.addEventListener(eventName, () => {
          uploadZone.classList.add('drag-over');
        });
      });

      ['dragleave', 'drop'].forEach(eventName => {
        uploadZone.addEventListener(eventName, () => {
          uploadZone.classList.remove('drag-over');
        });
      });

      uploadZone.addEventListener('drop', function(e) {
        const dt = e.dataTransfer;
        const files = dt.files;

        if (files.length > 0) {
          document.getElementById('screenshot').files = files;
          const event = new Event('change');
          document.getElementById('screenshot').dispatchEvent(event);
        }
      });
    }
  </script>

  <script>
    let pastedImageData = null;

    // Listen for paste events on textarea
    document.getElementById('rawText')?.addEventListener('paste', function(e) {
      const items = (e.clipboardData || e.originalEvent.clipboardData).items;

      for (let item of items) {
        if (item.type.indexOf('image') !== -1) {
          // Image pasted!
          e.preventDefault();

          const blob = item.getAsFile();
          const reader = new FileReader();

          reader.onload = function(event) {
            const base64 = event.target.result;

            // Store in hidden field
            document.getElementById('screenshotData').value = base64;

            // Show preview
            document.getElementById('previewImg').src = base64;
            document.getElementById('imagePreview').style.display = 'block';
            document.getElementById('pasteZone').classList.add('has-image');

            // Clear textarea (user pasted image, not text)
            document.getElementById('rawText').value = '';

            console.log('✅ Screenshot pasted, will OCR on submit');
          };

          reader.readAsDataURL(blob);
          break;
        }
      }
    });

    function clearPastedImage() {
      document.getElementById('screenshotData').value = '';
      document.getElementById('imagePreview').style.display = 'none';
      document.getElementById('pasteZone').classList.remove('has-image');
      document.getElementById('rawText').focus();
    }
  </script>
</body>
</html>
"""




    @admin_bp.route("/rounds")
    @login_required
    def admin_rounds():
        admin_required()
        rounds = Round.query.order_by(Round.is_active.desc(), Round.id.desc()).all()
        return render_page(r"""
<div class="card">
  <div class="row" style="justify-content:space-between;">
    <div>
      <h2 style="margin:0;">Soutěže</h2>
      <div class="muted">Tipy/extra uzávěrky jsou v UTC (pro jednoduchost). Pokud chceš, doplníme lokální čas.</div>
    </div>
    <a class="btn btn-primary" href="{{ url_for('admin_round_new') }}">Nová soutěž</a>
  </div>

  <hr class="sep">

  {% for rr in rounds %}
    <div class="card" style="background:rgba(255,255,255,.03); margin-bottom:10px;">
      <div class="row" style="justify-content:space-between;">
        <div>
          <div style="font-weight:900;">
            {% if rr.is_active %}★ {% endif %}
            {% if rr.is_archived %}📦 {% endif %}
            {{ rr.name }}
          </div>
          <div class="muted">Sport: {{ rr.sport.name }}</div>
          <div class="muted">Tipy: {{ rr.tips_close_time.strftime("%Y-%m-%d %H:%M") if rr.tips_close_time else "—" }} |
            Extra: {{ rr.extra_close_time.strftime("%Y-%m-%d %H:%M") if rr.extra_close_time else "—" }}</div>
        </div>
        <div class="row" style="gap: 8px; flex-wrap: wrap;">
          <a class="btn" href="{{ url_for('admin_round_edit', round_id=rr.id) }}">Edit</a>
          <a class="btn" href="{{ url_for('admin_round_toggle', round_id=rr.id) }}">
            {% if rr.is_active %}Deaktivovat{% else %}Aktivovat{% endif %}
          </a>
          <a class="btn" href="{{ url_for('admin_round_toggle_archive', round_id=rr.id) }}"
             style="{% if rr.is_archived %}background:rgba(51,209,122,.15); color:#33d17a; border:1px solid rgba(51,209,122,.3);{% else %}background:rgba(139,92,246,.15); color:#8b5cf6; border:1px solid rgba(139,92,246,.3);{% endif %}">
            {% if rr.is_archived %}📤 Odarchivovat{% else %}📦 Archivovat{% endif %}
          </a>
          <a class="btn" href="{{ url_for('admin_round_delete_confirm', round_id=rr.id) }}"
             style="background:rgba(255,77,109,0.2); color:#ff4d6d; border:1px solid rgba(255,77,109,0.4);">
            🗑️ Smazat
          </a>
        </div>
      </div>
    </div>
  {% endfor %}
</div>
""", rounds=rounds)



    @admin_bp.route("/round/new", methods=["GET", "POST"])
    @login_required
    def admin_round_new():
        admin_required()
        sports = Sport.query.order_by(Sport.name.asc()).all()
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            sport_id = int(request.form.get("sport_id") or "0")
            tips_close = parse_naive_datetime(request.form.get("tips_close") or "")
            extra_close = parse_naive_datetime(request.form.get("extra_close") or "")

            if not name or not sport_id:
                flash("Vyplň název a sport.", "error")
                return redirect(url_for("admin_round_new"))

            rr = Round(name=name, sport_id=sport_id, tips_close_time=tips_close, extra_close_time=extra_close, is_active=True)
            db.session.add(rr)
            for other in Round.query.all():
                other.is_active = False
            rr.is_active = True
            db.session.commit()
            set_selected_round_id(rr.id)
            audit("round.create", "Round", rr.id, name=rr.name)
            flash("Soutěž vytvořena.", "ok")
            return redirect(url_for("admin_rounds"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Nová soutěž</h2>
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input name="name" placeholder="Název soutěže" required>
    <select name="sport_id" required>
      {% for sp in sports %}
        <option value="{{ sp.id }}">{{ sp.name }}</option>
      {% endfor %}
    </select>
    <div class="grid2">
      <div>
        <div class="muted" style="margin-bottom:6px;">Uzávěrka tipů (volitelné)</div>
        <input name="tips_close" type="datetime-local">
      </div>
      <div>
        <div class="muted" style="margin-bottom:6px;">Uzávěrka extra (volitelné)</div>
        <input name="extra_close" type="datetime-local">
      </div>
    </div>
    <button class="btn btn-primary" type="submit">Vytvořit</button>
    <a class="btn" href="{{ url_for('admin_rounds') }}">Zpět</a>
  </form>
</div>
""", sports=sports)



    @admin_bp.route("/round/<int:round_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_round_edit(round_id: int):
        admin_required()
        rr = db.session.get(Round, round_id)
        if not rr:
            abort(404)

        sports = Sport.query.order_by(Sport.name.asc()).all()

        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            sport_id = int(request.form.get("sport_id") or rr.sport_id)
            tips_close = parse_naive_datetime(request.form.get("tips_close") or "")
            extra_close = parse_naive_datetime(request.form.get("extra_close") or "")

            if not name:
                flash("Vyplň název.", "error")
                return redirect(url_for("admin_round_edit", round_id=rr.id))

            rr.name = name
            rr.sport_id = sport_id
            rr.tips_close_time = tips_close
            rr.extra_close_time = extra_close
            db.session.commit()
            audit("round.edit", "Round", rr.id, name=rr.name)
            flash("Soutěž upravena.", "ok")
            return redirect(url_for("admin_rounds"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Upravit soutěž</h2>
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <div class="muted" style="margin-bottom:6px;">Název soutěže</div>
      <input name="name" value="{{ rr.name }}" required>
    </div>
    <div>
      <div class="muted" style="margin-bottom:6px;">Sport</div>
      <select name="sport_id" required>
        {% for sp in sports %}
          <option value="{{ sp.id }}" {% if sp.id == rr.sport_id %}selected{% endif %}>{{ sp.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="grid2">
      <div>
        <div class="muted" style="margin-bottom:6px;">Uzávěrka tipů (volitelné)</div>
        <input name="tips_close" type="datetime-local" value="{{ dt_tips }}">
      </div>
      <div>
        <div class="muted" style="margin-bottom:6px;">Uzávěrka extra (volitelné)</div>
        <input name="extra_close" type="datetime-local" value="{{ dt_extra }}">
      </div>
    </div>
    <button class="btn btn-primary" type="submit">Uložit</button>
    <a class="btn" href="{{ url_for('admin_rounds') }}">Zpět</a>
  </form>
</div>
""", rr=rr, sports=sports, dt_tips=dt_to_input_value(rr.tips_close_time), dt_extra=dt_to_input_value(rr.extra_close_time))



    @admin_bp.route("/round/<int:round_id>/toggle")
    @login_required
    def admin_round_toggle(round_id: int):
        admin_required()
        r = db.session.get(Round, round_id)
        if not r:
            abort(404)
        if not r.is_active:
            for other in Round.query.all():
                other.is_active = False
            r.is_active = True
            set_selected_round_id(r.id)
        else:
            r.is_active = False
        db.session.commit()
        audit("round.toggle_active", "Round", r.id, is_active=r.is_active)
        return redirect(url_for("admin_rounds"))



    @admin_bp.route("/round/<int:round_id>/toggle-archive")
    @login_required
    def admin_round_toggle_archive(round_id: int):
        admin_required()
        r = db.session.get(Round, round_id)
        if not r:
            abort(404)

        # Toggle archived flag
        r.is_archived = not r.is_archived

        # Pokud archivujeme, deaktivujeme
        if r.is_archived and r.is_active:
            r.is_active = False

        db.session.commit()
        audit("round.toggle_archive", "Round", r.id, is_archived=r.is_archived)

        msg = "Soutěž archivována." if r.is_archived else "Soutěž vrácena z archivu."
        flash(msg, "ok")
        return redirect(url_for("admin_rounds"))



    @admin_bp.route("/round/<int:round_id>/delete/confirm")
    @login_required
    def admin_round_delete_confirm(round_id: int):
        admin_required()
        r = db.session.get(Round, round_id)
        if not r:
            abort(404)

        # Spočítat co se smaže
        matches_count = Match.query.filter_by(round_id=r.id).count()
        teams_count = Team.query.filter_by(round_id=r.id).count()
        tips_count = db.session.query(Tip).join(Match).filter(Match.round_id == r.id).count()
        extra_questions_count = ExtraQuestion.query.filter_by(round_id=r.id).count()
        extra_answers_count = db.session.query(ExtraAnswer).join(ExtraQuestion).filter(ExtraQuestion.round_id == r.id).count()

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 16px 0; color:#ff4d6d;">⚠️ Smazat soutěž?</h2>

  <div class="card" style="background:rgba(255,77,109,0.1); border:2px solid #ff4d6d; padding:20px; margin-bottom:20px;">
    <h3 style="margin:0 0 12px 0; color:#ff4d6d;">POZOR: Tato akce je NEVRATNÁ!</h3>
    <div style="font-size:16px; line-height:1.8;">
      Chystáš se smazat soutěž: <strong style="font-size:18px;">{{ round.name }}</strong>
    </div>
  </div>

  <div class="card" style="background:rgba(255,255,255,0.03); padding:20px; margin-bottom:20px;">
    <h3 style="margin:0 0 16px 0;">Co se trvale odstraní:</h3>
    <div style="display:grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap:16px;">
      <div class="card" style="background:rgba(255,77,109,0.05); padding:16px; text-align:center;">
        <div style="font-size:32px; font-weight:900; color:#ff4d6d;">{{ matches_count }}</div>
        <div class="muted">Zápasů</div>
      </div>
      <div class="card" style="background:rgba(255,77,109,0.05); padding:16px; text-align:center;">
        <div style="font-size:32px; font-weight:900; color:#ff4d6d;">{{ tips_count }}</div>
        <div class="muted">Tipů</div>
      </div>
      <div class="card" style="background:rgba(255,77,109,0.05); padding:16px; text-align:center;">
        <div style="font-size:32px; font-weight:900; color:#ff4d6d;">{{ teams_count }}</div>
        <div class="muted">Týmů</div>
      </div>
      <div class="card" style="background:rgba(255,77,109,0.05); padding:16px; text-align:center;">
        <div style="font-size:32px; font-weight:900; color:#ff4d6d;">{{ extra_questions_count }}</div>
        <div class="muted">Extra otázek</div>
      </div>
      <div class="card" style="background:rgba(255,77,109,0.05); padding:16px; text-align:center;">
        <div style="font-size:32px; font-weight:900; color:#ff4d6d;">{{ extra_answers_count }}</div>
        <div class="muted">Extra odpovědí</div>
      </div>
    </div>
  </div>

  <div class="card" style="background:rgba(255,199,79,0.08); border:1px solid rgba(255,199,79,0.4); padding:20px; margin-bottom:20px;">
    <h3 style="margin:0 0 12px 0;">⚠️ Potvrzení smazání</h3>
    <div class="muted" style="margin-bottom:12px;">
      Pro potvrzení napiš přesný název soutěže:
    </div>
    <div style="font-weight:900; font-size:18px; margin-bottom:16px; padding:12px; background:rgba(0,0,0,0.3); border-radius:8px; text-align:center;">
      {{ round.name }}
    </div>

    <form method="post" action="{{ url_for('admin_round_delete', round_id=round.id) }}" onsubmit="return validateDelete()">
      <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
      <input type="text" id="confirm_name" name="confirm_name"
             placeholder="Napiš název soutěže pro potvrzení"
             style="width:100%; margin-bottom:16px; padding:12px; font-size:16px;"
             autocomplete="off" required>

      <div style="display:flex; gap:12px;">
        <button type="submit" class="btn"
                style="flex:1; padding:14px; font-size:16px; font-weight:900; background:rgba(255,77,109,0.3); color:#ff4d6d; border:2px solid #ff4d6d;">
          🗑️ ANO, SMAZAT TRVALE
        </button>
        <a href="{{ url_for('admin_rounds') }}" class="btn btn-primary" style="flex:1; padding:14px; font-size:16px; font-weight:900; text-align:center;">
          ✖️ Zrušit
        </a>
      </div>
    </form>
  </div>
</div>

<script>
const expectedName = {{ round.name|tojson }};

function validateDelete() {
  const input = document.getElementById('confirm_name').value.trim();
  if (input !== expectedName) {
    alert('Název nesouhlasí! Zkontroluj překlepy.\n\nOčekáváno: ' + expectedName + '\nZadáno: ' + input);
    return false;
  }
  return confirm('POSLEDNÍ VAROVÁNÍ!\n\nOpravdu TRVALE smazat soutěž "' + expectedName + '" a všechna související data?\n\nTato akce JE NEVRATNÁ!');
}
</script>
""", round=r, matches_count=matches_count, tips_count=tips_count, teams_count=teams_count,
     extra_questions_count=extra_questions_count, extra_answers_count=extra_answers_count)



    @admin_bp.route("/round/<int:round_id>/delete", methods=["POST"])
    @login_required
    def admin_round_delete(round_id: int):
        admin_required()
        r = db.session.get(Round, round_id)
        if not r:
            abort(404)

        # Ověřit potvrzovací text
        confirm_name = request.form.get('confirm_name', '').strip()
        if confirm_name != r.name:
            flash("Název nesouhlasí! Soutěž nebyla smazána.", "error")
            return redirect(url_for("admin_round_delete_confirm", round_id=round_id))

        round_name = r.name

        # Smazat všechny extra odpovědi
        extra_questions = ExtraQuestion.query.filter_by(round_id=r.id).all()
        for eq in extra_questions:
            ExtraAnswer.query.filter_by(question_id=eq.id).delete()
            db.session.delete(eq)

        # Smazat všechny tipy
        matches = Match.query.filter_by(round_id=r.id).all()
        for m in matches:
            Tip.query.filter_by(match_id=m.id).delete()
            db.session.delete(m)

        # Smazat všechny týmy
        teams = Team.query.filter_by(round_id=r.id).all()
        for t in teams:
            db.session.delete(t)

        # Smazat soutěž
        db.session.delete(r)
        db.session.commit()

        audit("round.delete", "Round", round_id, name=round_name)
        flash(f"✅ Soutěž '{round_name}' byla trvale smazána včetně všech zápasů, tipů a týmů.", "ok")

        # Pokud byla aktivní, aktivovat jinou
        if r.is_active:
            other = Round.query.filter_by(is_active=False).first()
            if other:
                other.is_active = True
                db.session.commit()
                set_selected_round_id(other.id)

        return redirect(url_for("admin_rounds"))

    # --- ADMIN TEAM NEW ---


    @admin_bp.route("/team/new", methods=["GET", "POST"])
    @login_required
    def admin_team_new():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            return redirect(url_for("admin_rounds"))

        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            if not name:
                flash("Vyplň název týmu.", "error")
                return redirect(url_for("admin_team_new"))
            if Team.query.filter_by(round_id=r.id, name=name, is_deleted=False).first():
                flash("Tým už existuje.", "error")
                return redirect(url_for("admin_team_new"))
            t = Team(round_id=r.id, name=name)
            db.session.add(t)
            db.session.commit()
            audit("team.create", "Team", t.id, round_id=r.id, name=t.name)
            flash("Tým přidán.", "ok")
            return redirect(url_for("teams"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Přidat tým</h2>
  <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input name="name" placeholder="Název týmu" required>
    <button class="btn btn-primary" type="submit">Vytvořit</button>
    <a class="btn" href="{{ url_for('teams') }}">Zpět</a>
  </form>
</div>
""", r=r)


    # --- ADMIN TEAM ALIASES ---


    @admin_bp.route("/team-aliases", methods=["GET", "POST"])
    @login_required
    def admin_team_aliases():
        """
        Správa aliasů týmů (pro Smart Import a další importéry)
        - alias (krátký název / zkratka) -> canonical_name (plný název v DB)
        """
        admin_required()

        # Vybranou soutěž ber z query paramu, fallback na selected round
        rid = request.values.get("round_id")
        if rid is None or str(rid).strip() == "":
            rid = ensure_selected_round()
        try:
            rid = int(rid) if rid else None
        except:
            rid = None

        rounds = Round.query.order_by(Round.name.asc()).all()
        r = db.session.get(Round, rid) if rid else None

        if request.method == "POST":
            action = request.form.get("action", "").strip()

            # Switch selected round
            if action == "select_round":
                rid2 = request.form.get("round_id")
                try:
                    rid2 = int(rid2) if rid2 else None
                except:
                    rid2 = None
                if rid2:
                    set_selected_round_id(rid2)
                return redirect(url_for("admin_team_aliases", round_id=rid2))

            if not r:
                flash("Nejdřív vyber soutěž/kolo.", "error")
                return redirect(url_for("admin_team_aliases"))

            if action in ("add", "edit"):
                alias = (request.form.get("alias") or "").strip()
                canonical = (request.form.get("canonical_name") or "").strip()

                if not alias or not canonical:
                    flash("Vyplň alias i cílový (kanonický) název.", "error")
                    return redirect(url_for("admin_team_aliases", round_id=r.id))

                # normalizace (bez lower na bool apod.)
                alias_n = normalize_team_name(alias, round_id=r.id) if alias else ""
                # alias necháváme tak jak ho user zadal (trim), ale odstraníme nadbytečné mezery
                alias_n = re.sub(r"\s+", " ", alias).strip()
                canonical_n = re.sub(r"\s+", " ", canonical).strip()

                if action == "add":
                    exists = TeamAlias.query.filter(
                        TeamAlias.round_id == r.id,
                        db.func.lower(TeamAlias.alias) == alias_n.lower()
                    ).first()
                    if exists:
                        flash("Alias už existuje (v této soutěži).", "error")
                        return redirect(url_for("admin_team_aliases", round_id=r.id))
                    ta = TeamAlias(round_id=r.id, alias=alias_n, canonical_name=canonical_n)
                    db.session.add(ta)
                    db.session.commit()
                    audit("team_alias.create", "TeamAlias", ta.id, round_id=r.id, alias=ta.alias, canonical=ta.canonical_name)
                    flash("Alias uložen.", "ok")
                    return redirect(url_for("admin_team_aliases", round_id=r.id))

                # edit
                alias_id = request.form.get("alias_id")
                try:
                    alias_id = int(alias_id)
                except:
                    alias_id = None
                ta = db.session.get(TeamAlias, alias_id) if alias_id else None
                if not ta or ta.round_id != r.id:
                    flash("Alias nenalezen.", "error")
                    return redirect(url_for("admin_team_aliases", round_id=r.id))

                ta.alias = alias_n
                ta.canonical_name = canonical_n
                db.session.commit()
                audit("team_alias.edit", "TeamAlias", ta.id, round_id=r.id, alias=ta.alias, canonical=ta.canonical_name)
                flash("Alias upraven.", "ok")
                return redirect(url_for("admin_team_aliases", round_id=r.id))

            if action == "delete":
                if not r:
                    flash("Nejdřív vyber soutěž/kolo.", "error")
                    return redirect(url_for("admin_team_aliases"))

                alias_id = request.form.get("alias_id")
                try:
                    alias_id = int(alias_id)
                except:
                    alias_id = None
                ta = db.session.get(TeamAlias, alias_id) if alias_id else None
                if not ta or ta.round_id != r.id:
                    flash("Alias nenalezen.", "error")
                    return redirect(url_for("admin_team_aliases", round_id=r.id))

                db.session.delete(ta)
                db.session.commit()
                audit("team_alias.delete", "TeamAlias", alias_id, round_id=r.id)
                flash("Alias smazán.", "ok")
                return redirect(url_for("admin_team_aliases", round_id=r.id))

            flash("Neznámá akce.", "error")
            return redirect(url_for("admin_team_aliases", round_id=(r.id if r else None)))

        # GET
        aliases = []
        if r:
            aliases = TeamAlias.query.filter_by(round_id=r.id).order_by(TeamAlias.alias.asc()).all()

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">🔁 Aliasy týmů</h2>
  <div class="muted">Alias slouží pro importéry (Smart Import, API import), aby se krátké názvy mapovaly na plné názvy ve tvé DB.</div>
</div>

<div class="card">
  <form method="post" class="row" style="gap:10px; align-items:flex-end;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="action" value="select_round"/>
    <div style="flex:1; min-width:240px;">
      <div class="muted" style="margin-bottom:6px;">Soutěž / kolo</div>
      <select name="round_id" onchange="this.form.submit()">
        <option value="">— vyber —</option>
        {% for rr in rounds %}
          <option value="{{ rr.id }}" {% if r and rr.id==r.id %}selected{% endif %}>{{ rr.name }}</option>
        {% endfor %}
      </select>
    </div>
    {% if r %}
      <a class="btn" href="{{ url_for('teams') }}">➡️ Týmy</a>
      <a class="btn" href="{{ url_for('admin_smart_import') }}">🤖 Smart Import</a>
    {% endif %}
  </form>
</div>

{% if r %}
<div class="card">
  <h3 style="margin:0 0 10px 0;">➕ Přidat alias</h3>
  <form method="post" class="row" style="gap:10px; align-items:flex-end;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="action" value="add"/>
    <input type="hidden" name="round_id" value="{{ r.id }}"/>
    <div style="flex:1; min-width:220px;">
      <div class="muted" style="margin-bottom:6px;">Alias (co importér najde)</div>
      <input name="alias" placeholder="např. Sparta / Hradec Kr. / Ml. Boleslav" required>
    </div>
    <div style="flex:1; min-width:240px;">
      <div class="muted" style="margin-bottom:6px;">Kanonický název (jak chceš mít v DB)</div>
      <input name="canonical_name" placeholder="např. AC Sparta Praha" required>
    </div>
    <button class="btn btn-primary" type="submit">Uložit</button>
  </form>
</div>

<div class="card">
  <h3 style="margin:0 0 10px 0;">📋 Seznam aliasů</h3>
  {% if not aliases %}
    <div class="muted">Zatím žádné aliasy.</div>
  {% else %}
  <div style="overflow:auto;">
    <table class="table">
      <thead>
        <tr>
          <th>Alias</th>
          <th>Kanonický název</th>
          <th style="width:240px;">Akce</th>
        </tr>
      </thead>
      <tbody>
        {% for a in aliases %}
          <tr>
            <td><code>{{ a.alias }}</code></td>
            <td>{{ a.canonical_name }}</td>
            <td>
              <details>
                <summary class="btn" style="display:inline-block;">Upravit</summary>
                <div style="margin-top:10px;">
                  <form method="post" class="row" style="gap:8px; align-items:flex-end;">
                    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                    <input type="hidden" name="action" value="edit"/>
                    <input type="hidden" name="round_id" value="{{ r.id }}"/>
                    <input type="hidden" name="alias_id" value="{{ a.id }}"/>
                    <div style="flex:1; min-width:180px;">
                      <div class="muted" style="margin-bottom:6px;">Alias</div>
                      <input name="alias" value="{{ a.alias }}" required>
                    </div>
                    <div style="flex:1; min-width:220px;">
                      <div class="muted" style="margin-bottom:6px;">Kanonický název</div>
                      <input name="canonical_name" value="{{ a.canonical_name }}" required>
                    </div>
                    <button class="btn btn-primary" type="submit">Uložit</button>
                  </form>

                  <form method="post" style="margin-top:8px;">
                    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                    <input type="hidden" name="action" value="delete"/>
                    <input type="hidden" name="round_id" value="{{ r.id }}"/>
                    <input type="hidden" name="alias_id" value="{{ a.id }}"/>
                    <button class="btn btn-danger" type="submit" onclick="return confirm('Smazat alias {{ a.alias }}?')">Smazat</button>
                  </form>
                </div>
              </details>
            </td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endif %}
</div>
{% else %}
<div class="card"><div class="muted">Vyber soutěž/kolo pro správu aliasů.</div></div>
{% endif %}
""", rounds=rounds, r=r, aliases=aliases)


    # --- ADMIN MATCH NEW/EDIT ---


    @admin_bp.route("/match/new", methods=["GET", "POST"])
    @login_required
    def admin_match_new():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            return redirect(url_for("admin_rounds"))
        teams_q = Team.query.filter_by(round_id=r.id, is_deleted=False).order_by(Team.name.asc()).all()
        if not teams_q:
            flash("Nejdřív přidej týmy pro tuhle soutěž.", "error")
            return redirect(url_for("admin_team_new"))

        if request.method == "POST":
            home_id = int(request.form.get("home_team_id") or "0")
            away_id = int(request.form.get("away_team_id") or "0")
            start = parse_naive_datetime(request.form.get("start_time") or "")
            if not home_id or not away_id or home_id == away_id:
                flash("Vyber domácí a hosty (různé týmy).", "error")
                return redirect(url_for("admin_match_new"))

            ht = db.session.get(Team, home_id)
            at = db.session.get(Team, away_id)
            if not ht or not at or ht.round_id != r.id or at.round_id != r.id:
                flash("Týmy nepatří do vybrané soutěže.", "error")
                return redirect(url_for("admin_match_new"))

            m = Match(round_id=r.id, home_team_id=ht.id, away_team_id=at.id, start_time=start)
            db.session.add(m)
            db.session.commit()
            audit("match.create", "Match", m.id, round_id=r.id)
            flash("Zápas vytvořen.", "ok")
            return redirect(url_for("matches"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Přidat zápas</h2>
  <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <select name="home_team_id" required>
      <option value="">Domácí tým…</option>
      {% for t in teams %}<option value="{{ t.id }}">{{ t.name }}</option>{% endfor %}
    </select>
    <select name="away_team_id" required>
      <option value="">Hosté…</option>
      {% for t in teams %}<option value="{{ t.id }}">{{ t.name }}</option>{% endfor %}
    </select>
    <div>
      <div class="muted" style="margin-bottom:6px;">Začátek (volitelné)</div>
      <input name="start_time" type="datetime-local">
    </div>
    <button class="btn btn-primary" type="submit">Vytvořit</button>
    <a class="btn" href="{{ url_for('matches') }}">Zpět</a>
  </form>
</div>
""", r=r, teams=teams_q)



    @admin_bp.route("/match/<int:match_id>/quick-score", methods=["POST"])
    @login_required
    def admin_quick_score(match_id: int):
        admin_required()
        m = db.session.get(Match, match_id)
        if not m:
            abort(404)

        home_score_val = request.form.get("home_score", "").strip()
        away_score_val = request.form.get("away_score", "").strip()

        # Prázdné = None (smazat výsledek)
        m.home_score = int(home_score_val) if home_score_val else None
        m.away_score = int(away_score_val) if away_score_val else None

        db.session.commit()
        audit("match.quick_score", "Match", m.id, home=m.home_score, away=m.away_score)
        flash(f"Výsledek uložen: {m.home_team.name} {m.home_score or '-'}:{m.away_score or '-'} {m.away_team.name}", "ok")
        return redirect(url_for("leaderboard"))



    @admin_bp.route("/match/<int:match_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_match_edit(match_id: int):
        admin_required()
        m = db.session.get(Match, match_id)
        if not m:
            abort(404)
        r = db.session.get(Round, m.round_id)
        teams_q = Team.query.filter_by(round_id=r.id, is_deleted=False).order_by(Team.name.asc()).all()

        def parse_int_or_none(x):
            x = (x or "").strip()
            if x == "":
                return None
            return int(x)

        if request.method == "POST":
            home_id = int(request.form.get("home_team_id") or m.home_team_id)
            away_id = int(request.form.get("away_team_id") or m.away_team_id)
            start = parse_naive_datetime(request.form.get("start_time") or "")
            hs = parse_int_or_none(request.form.get("home_score"))
            aas = parse_int_or_none(request.form.get("away_score"))
            if home_id == away_id:
                flash("Domácí a hosté musí být různé týmy.", "error")
                return redirect(url_for("admin_match_edit", match_id=m.id))

            ht = db.session.get(Team, home_id)
            at = db.session.get(Team, away_id)
            if not ht or not at or ht.round_id != r.id or at.round_id != r.id:
                flash("Týmy nepatří do téhle soutěže.", "error")
                return redirect(url_for("admin_match_edit", match_id=m.id))

            m.home_team_id = ht.id
            m.away_team_id = at.id
            m.start_time = start
            m.home_score = hs
            m.away_score = aas
            db.session.commit()
            audit("match.edit", "Match", m.id)
            flash("Zápas upraven.", "ok")
            return redirect(url_for("matches"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Edit zápasu</h2>
  <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <select name="home_team_id" required>
      {% for t in teams %}<option value="{{ t.id }}" {% if t.id == m.home_team_id %}selected{% endif %}>{{ t.name }}</option>{% endfor %}
    </select>
    <select name="away_team_id" required>
      {% for t in teams %}<option value="{{ t.id }}" {% if t.id == m.away_team_id %}selected{% endif %}>{{ t.name }}</option>{% endfor %}
    </select>
    <div>
      <div class="muted" style="margin-bottom:6px;">Začátek</div>
      <input name="start_time" type="datetime-local" value="{{ dt }}">
    </div>
    <div class="row">
      <input name="home_score" type="number" min="0" style="width:140px;" value="{{ m.home_score if m.home_score is not none else '' }}" placeholder="Domácí skóre">
      <div class="muted">:</div>
      <input name="away_score" type="number" min="0" style="width:140px;" value="{{ m.away_score if m.away_score is not none else '' }}" placeholder="Hosté skóre">
    </div>
    <button class="btn btn-primary" type="submit">Uložit</button>
    <a class="btn" href="{{ url_for('matches') }}">Zpět</a>
  </form>
</div>
""", r=r, m=m, teams=teams_q, dt=dt_to_input_value(m.start_time))



    @admin_bp.route("/match/<int:match_id>/delete")
    @login_required
    def admin_match_delete(match_id: int):
        admin_required()
        m = db.session.get(Match, match_id)
        if not m:
            abort(404)

        # Smazat všechny tipy na tento zápas
        Tip.query.filter_by(match_id=m.id).delete()

        # Soft delete zápasu
        m.is_deleted = True
        db.session.commit()
        audit("match.delete", "Match", m.id)
        flash(f"Zápas smazán (včetně všech tipů).", "ok")
        return redirect(url_for("matches"))

    # --- ADMIN EXTRA NEW ---


    @admin_bp.route("/extra/new", methods=["GET", "POST"])
    @login_required
    def admin_extra_new():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            return redirect(url_for("admin_rounds"))
        if request.method == "POST":
            q = (request.form.get("question") or "").strip()
            deadline_str = (request.form.get("deadline") or "").strip()

            if not q:
                flash("Vyplň otázku.", "error")
                return redirect(url_for("admin_extra_new"))

            # Parse deadline pokud je vyplněno
            deadline = None
            if deadline_str:
                try:
                    # Očekáváme formát: YYYY-MM-DD HH:MM nebo YYYY-MM-DDTHH:MM
                    deadline = datetime.strptime(deadline_str.replace('T', ' '), '%Y-%m-%d %H:%M')
                except ValueError:
                    flash("Nesprávný formát data. Použij formát: YYYY-MM-DD HH:MM", "error")
                    return redirect(url_for("admin_extra_new"))

            eq = ExtraQuestion(round_id=r.id, question=q, deadline=deadline)
            db.session.add(eq)
            db.session.commit()
            audit("extra.question.create", "ExtraQuestion", eq.id)
            flash("Extra otázka přidána.", "ok")
            return redirect(url_for("extras"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Přidat extra otázku</h2>
  <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
  <hr class="sep">
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="form-group">
      <label>Text otázky</label>
      <input name="question" placeholder="Text otázky" required>
    </div>
    <div class="form-group">
      <label>Uzávěrka odpovědí (volitelné)</label>
      <input type="datetime-local" name="deadline" placeholder="YYYY-MM-DD HH:MM">
      <small class="muted">Po tomto datu budou odpovědi viditelné v žebříčku. Pokud nevyplníte, odpovědi budou viditelné ihned.</small>
    </div>
    <button class="btn btn-primary" type="submit">Vytvořit</button>
    <a class="btn" href="{{ url_for('extras') }}">Zpět</a>
  </form>
</div>
""", r=r)

    # --- ADMIN EXTRA MANAGE (správa odpovědí) ---


    @admin_bp.route("/extra/manage")
    @login_required
    def admin_extra_manage():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            return redirect(url_for("admin_rounds"))

        # Načti všechny Extra otázky
        questions = ExtraQuestion.query.filter_by(
            round_id=r.id,
            is_deleted=False
        ).order_by(ExtraQuestion.id.asc()).all()

        # Načti všechny uživatele
        users = User.query.order_by(User.username.asc()).all()

        # Načti všechny odpovědi pro tuto soutěž
        all_answers = ExtraAnswer.query.join(ExtraQuestion).filter(
            ExtraQuestion.round_id == r.id,
            ExtraQuestion.is_deleted == False
        ).all()

        # Vytvoř mapu odpovědí: {user_id: {question_id: answer}}
        answer_map = {}
        for ans in all_answers:
            if ans.user_id not in answer_map:
                answer_map[ans.user_id] = {}
            answer_map[ans.user_id][ans.question_id] = ans

        return render_page(r"""
<style>
  .extra-manage-table{ width:100%; border-collapse:collapse; }
  .extra-manage-table th, .extra-manage-table td{
    border:1px solid var(--line);
    padding:8px;
    text-align:left;
  }
  .extra-manage-table th{ background:rgba(255,255,255,.05); font-weight:800; }
  .extra-manage-table td{ font-size:13px; }
  .answer-cell{ position:relative; }
  .answer-text{ display:block; margin-bottom:4px; }
  .answer-actions{ display:flex; gap:4px; }
  .answer-status{
    display:inline-block;
    padding:2px 6px;
    border-radius:4px;
    font-size:11px;
    font-weight:800;
  }
  .status-correct{ background:#28a745; color:white; }
  .status-wrong{ background:#dc3545; color:white; }
  .status-none{ background:#6c757d; color:white; }
  .q-actions { display:flex; gap:4px; margin-top:6px; justify-content:center; }
</style>

<div class="card">
  <div class="row" style="justify-content:space-between;">
    <div>
      <h2 style="margin:0;">Správa Extra odpovědí</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <div class="row" style="gap:8px;">
      <a class="btn" href="{{ url_for('admin_extra_new') }}">➕ Přidat otázku</a>
      <a class="btn" href="{{ url_for('extras') }}">Zpět</a>
    </div>
  </div>

  <hr class="sep">

  {% if questions|length == 0 %}
    <div class="muted">Zatím nejsou žádné extra otázky.</div>
  {% else %}
    <div style="overflow:auto;">
      <table class="extra-manage-table">
        <thead>
          <tr>
            <th style="min-width:120px;">Uživatel</th>
            {% for q in questions %}
              <th style="min-width:200px; text-align:center;">
                <div>{{ q.question }}</div>
                {% if q.deadline %}
                  <div><small class="muted">Uzávěrka: {{ q.deadline.strftime('%d.%m. %H:%M') }}</small></div>
                {% endif %}
                <div class="q-actions">
                  <a href="{{ url_for('admin_extra_edit_question', question_id=q.id) }}"
                     class="btn" style="font-size:11px; padding:3px 8px;">✏️ Edit</a>
                  <a href="{{ url_for('admin_extra_delete_question', question_id=q.id) }}"
                     class="btn" style="font-size:11px; padding:3px 8px; background:rgba(255,77,109,.2); color:#ff4d6d; border:1px solid rgba(255,77,109,.4);"
                     onclick="return confirm('Smazat otázku „{{ q.question }}" a VŠECHNY odpovědi na ni?')">🗑️</a>
                </div>
              </th>
            {% endfor %}
          </tr>
        </thead>
        <tbody>
          {% for u in users %}
            <tr>
              <td><b>{{ u.display_name }}</b><br><small class="muted">{{ u.email }}</small></td>
              {% for q in questions %}
                {% set ans = answer_map.get(u.id, {}).get(q.id) %}
                <td class="answer-cell">
                  {% if ans %}
                    <div class="answer-text">
                      <span class="answer-status {% if ans.is_correct %}status-correct{% else %}status-wrong{% endif %}">
                        {% if ans.is_correct %}✓ Správně{% else %}✗ Špatně{% endif %}
                      </span>
                      <br>
                      <span style="margin-top:4px; display:block;">{{ ans.answer_text }}</span>
                    </div>
                    <div class="answer-actions">
                      <a href="{{ url_for('admin_extra_edit_answer', question_id=q.id, user_id=u.id) }}"
                         class="btn btn-sm" style="font-size:11px; padding:4px 8px;">Editovat</a>
                      <a href="{{ url_for('admin_extra_delete_answer', answer_id=ans.id) }}"
                         class="btn btn-sm" style="font-size:11px; padding:4px 8px; background:#dc3545;"
                         onclick="return confirm('Opravdu smazat odpověď?')">Smazat</a>
                    </div>
                  {% else %}
                    <span class="answer-status status-none">Bez odpovědi</span>
                    <div style="margin-top:4px;">
                      <a href="{{ url_for('admin_extra_edit_answer', question_id=q.id, user_id=u.id) }}"
                         class="btn btn-sm btn-primary" style="font-size:11px; padding:4px 8px;">Přidat</a>
                    </div>
                  {% endif %}
                </td>
              {% endfor %}
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  {% endif %}
</div>
""", r=r, questions=questions, users=users, answer_map=answer_map)

    # --- ADMIN EXTRA EDIT QUESTION (editace otázky) ---


    @admin_bp.route("/extra/question/<int:question_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_extra_edit_question(question_id: int):
        admin_required()
        q = db.session.get(ExtraQuestion, question_id)
        if not q or q.is_deleted:
            abort(404)

        if request.method == "POST":
            text = (request.form.get("question") or "").strip()
            deadline_str = (request.form.get("deadline") or "").strip()

            if not text:
                flash("Otázka nesmí být prázdná.", "error")
                return redirect(url_for("admin_extra_edit_question", question_id=question_id))

            deadline = None
            if deadline_str:
                try:
                    deadline = datetime.strptime(deadline_str.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
                except ValueError:
                    flash("Nesprávný formát data.", "error")
                    return redirect(url_for("admin_extra_edit_question", question_id=question_id))

            q.question = text
            q.deadline = deadline
            db.session.commit()
            audit("admin.extra.question.edit", "ExtraQuestion", q.id)
            flash("Otázka upravena.", "ok")
            return redirect(url_for("admin_extra_manage"))

        # Formát deadlinu pro input
        deadline_val = q.deadline.strftime("%Y-%m-%dT%H:%M") if q.deadline else ""

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">✏️ Editovat extra otázku</h2>
  <div class="muted">Soutěž: <b>{{ q.round.name }}</b></div>
  <hr class="sep">
  <form method="post" style="display:flex; flex-direction:column; gap:14px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="form-group">
      <label>Text otázky</label>
      <input name="question" value="{{ q.question }}" required autofocus>
    </div>
    <div class="form-group">
      <label>Uzávěrka odpovědí (volitelné)</label>
      <input type="datetime-local" name="deadline" value="{{ deadline_val }}">
      <small class="muted">Po tomto datu budou odpovědi viditelné v žebříčku.</small>
    </div>
    <div class="row" style="gap:10px;">
      <button class="btn btn-primary" type="submit">💾 Uložit</button>
      <a class="btn" href="{{ url_for('admin_extra_manage') }}">Zrušit</a>
    </div>
  </form>
</div>
""", q=q, deadline_val=deadline_val)

    # --- ADMIN EXTRA DELETE QUESTION (smazání otázky + odpovědí) ---


    @admin_bp.route("/extra/question/<int:question_id>/delete")
    @login_required
    def admin_extra_delete_question(question_id: int):
        admin_required()
        q = db.session.get(ExtraQuestion, question_id)
        if not q or q.is_deleted:
            abort(404)

        # Soft-delete otázky
        q.is_deleted = True
        # Smaž všechny odpovědi (hard delete, nemají is_deleted)
        ExtraAnswer.query.filter_by(question_id=q.id).delete()
        db.session.commit()
        audit("admin.extra.question.delete", "ExtraQuestion", q.id)
        flash("Otazka smazana vcetne vsech odpovedi.", "ok")
        return redirect(url_for("admin_extra_manage"))

    # --- ADMIN EXTRA EDIT ANSWER (editace/přidání odpovědi) ---


    @admin_bp.route("/extra/answer/<int:question_id>/<int:user_id>", methods=["GET", "POST"])
    @login_required
    def admin_extra_edit_answer(question_id: int, user_id: int):
        admin_required()

        q = db.session.get(ExtraQuestion, question_id)
        if not q:
            abort(404)

        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        # Načti existující odpověď (pokud existuje)
        ans = ExtraAnswer.query.filter_by(question_id=question_id, user_id=user_id).first()

        if request.method == "POST":
            answer_text = (request.form.get("answer_text") or "").strip()
            is_correct = request.form.get("is_correct") == "1"

            if not answer_text:
                flash("Vyplň odpověď.", "error")
                return redirect(url_for("admin_extra_edit_answer", question_id=question_id, user_id=user_id))

            if ans:
                # Editace existující
                ans.answer_text = answer_text
                ans.is_correct = is_correct
                audit("admin.extra.answer.edit", "ExtraAnswer", ans.id)
                flash("Odpověď upravena.", "ok")
            else:
                # Vytvoření nové
                ans = ExtraAnswer(
                    question_id=question_id,
                    user_id=user_id,
                    answer_text=answer_text,
                    is_correct=is_correct
                )
                db.session.add(ans)
                audit("admin.extra.answer.create", "ExtraAnswer", None, question_id=question_id, user_id=user_id)
                flash("Odpověď přidána.", "ok")

            db.session.commit()
            return redirect(url_for("admin_extra_manage"))

        # GET - zobraz formulář
        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">{% if ans %}Editovat{% else %}Přidat{% endif %} odpověď</h2>
  <div class="muted">
    <b>Uživatel:</b> {{ u.display_name }} ({{ u.email }})<br>
    <b>Otázka:</b> {{ q.question }}
  </div>
  <hr class="sep">

  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="form-group">
      <label>Odpověď</label>
      <input name="answer_text" value="{{ ans.answer_text if ans else '' }}" required autofocus>
    </div>

    <div class="form-group">
      <label>Hodnocení</label>
      <div class="row" style="gap:10px;">
        <label style="display:flex; align-items:center; gap:6px;">
          <input type="radio" name="is_correct" value="1"
                 {% if ans and ans.is_correct %}checked{% endif %}
                 {% if not ans %}checked{% endif %}>
          <span style="color:#28a745; font-weight:800;">✓ Správně</span>
        </label>
        <label style="display:flex; align-items:center; gap:6px;">
          <input type="radio" name="is_correct" value="0"
                 {% if ans and not ans.is_correct %}checked{% endif %}>
          <span style="color:#dc3545; font-weight:800;">✗ Špatně</span>
        </label>
      </div>
    </div>

    <div class="form-actions">
      <button type="submit" class="btn btn-primary">Uložit</button>
      <a href="{{ url_for('admin_extra_manage') }}" class="btn">Zrušit</a>
    </div>
  </form>
</div>
""", q=q, u=u, ans=ans)

    # --- ADMIN EXTRA DELETE ANSWER (smazání odpovědi) ---


    @admin_bp.route("/extra/answer/<int:answer_id>/delete")
    @login_required
    def admin_extra_delete_answer(answer_id: int):
        admin_required()

        ans = db.session.get(ExtraAnswer, answer_id)
        if not ans:
            abort(404)

        db.session.delete(ans)
        db.session.commit()
        audit("admin.extra.answer.delete", "ExtraAnswer", answer_id)
        flash("Odpověď smazána.", "ok")
        return redirect(url_for("admin_extra_manage"))

    # --- ADMIN DASHBOARD ---


    @admin_bp.route("/dashboard")
    @login_required
    def admin_dashboard():
        admin_required()

        # Stats
        total_users = User.query.count()
        total_rounds = Round.query.count()
        active_rounds = Round.query.filter_by(is_active=True).count()
        total_matches = Match.query.filter_by(is_deleted=False).count()
        total_tips = Tip.query.count()

        # Recent users
        recent_users = User.query.order_by(User.id.desc()).limit(5).all()

        # Active round stats
        active_round = Round.query.filter_by(is_active=True).first()
        active_round_stats = None
        if active_round:
            matches_total = Match.query.filter_by(round_id=active_round.id, is_deleted=False).count()
            matches_with_results = Match.query.filter(
                Match.round_id == active_round.id,
                Match.is_deleted == False,
                Match.home_score != None,
                Match.away_score != None
            ).count()
            tips_total = Tip.query.join(Match).filter(Match.round_id == active_round.id).count()

            active_round_stats = {
                'round': active_round,
                'matches_total': matches_total,
                'matches_with_results': matches_with_results,
                'tips_total': tips_total,
                'completion': int((matches_with_results / matches_total * 100) if matches_total > 0 else 0)
            }

        # Pending tasks
        matches_no_results = Match.query.filter(
            Match.is_deleted == False,
            db.or_(Match.home_score == None, Match.away_score == None)
        ).count()

        return render_page(r"""
<style>
  .admin-dashboard {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
  }

  .stat-card {
    background: rgba(255,255,255,.03);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 20px;
    transition: all 0.3s ease;
  }

  .stat-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(0,0,0,.2);
  }

  .stat-value {
    font-size: 36px;
    font-weight: 900;
    margin-bottom: 8px;
    line-height: 1;
  }

  .progress-bar-container {
    background: rgba(255,255,255,.05);
    border-radius: 8px;
    height: 8px;
    margin-top: 12px;
    overflow: hidden;
  }

  .progress-bar-fill {
    height: 100%;
    background: linear-gradient(90deg, var(--accent), var(--ok));
    border-radius: 8px;
    transition: width 0.5s ease;
  }

  /* Admin cards */
  .admin-cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 12px;
  }

  .admin-card {
    background: rgba(110,168,254,0.08);
    border: 1px solid rgba(110,168,254,0.2);
    border-radius: 16px;
    padding: 16px;
    cursor: pointer;
    transition: all 0.2s;
    text-decoration: none;
    color: #e9eefc;
    display: flex;
    flex-direction: column;
    gap: 8px;
  }

  .admin-card:hover {
    background: rgba(110,168,254,0.15);
    border-color: rgba(110,168,254,0.3);
    transform: translateY(-2px);
    box-shadow: 0 4px 16px rgba(0,0,0,0.2);
  }

  .admin-card-header {
    display: flex;
    align-items: center;
    gap: 12px;
  }

  .admin-card-icon {
    font-size: 24px;
  }

  .admin-card-title {
    font-size: 15px;
    font-weight: 700;
  }

  .admin-card-desc {
    font-size: 12px;
    color: #94a3b8;
    line-height: 1.4;
  }
</style>

<div class="card">
  <h1 style="margin: 0 0 8px 0;">👨‍💼 Admin Dashboard</h1>
  <div class="muted">Přehled a rychlé akce</div>
</div>

<div class="admin-dashboard">
  <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">👥 UŽIVATELÉ</h3>
    <div class="stat-value" style="color: var(--accent);">{{ total_users }}</div>
    <div class="muted" style="font-size: 13px;">Celkem registrovaných</div>
  </div>

  <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">🏆 SOUTĚŽE</h3>
    <div class="stat-value" style="color: var(--ok);">{{ total_rounds }}</div>
    <div class="muted" style="font-size: 13px;">{{ active_rounds }} aktivních</div>
  </div>

  <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">⚽ ZÁPASY</h3>
    <div class="stat-value" style="color: var(--warn);">{{ total_matches }}</div>
    <div class="muted" style="font-size: 13px;">V databázi</div>
  </div>

  <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">🎯 TIPY</h3>
    <div class="stat-value" style="color: var(--danger);">{{ total_tips }}</div>
    <div class="muted" style="font-size: 13px;">Celkem odeslaných</div>
  </div>
</div>

{% if active_round_stats %}
<div class="card">
  <h3 style="margin: 0 0 16px 0;">⭐ Aktivní soutěž: {{ active_round_stats.round.name }}</h3>

  <div class="row" style="justify-content: space-between; margin-bottom: 16px;">
    <div>
      <div class="muted" style="font-size: 13px;">Zápasy s výsledky</div>
      <div style="font-size: 24px; font-weight: 900;">
        {{ active_round_stats.matches_with_results }} / {{ active_round_stats.matches_total }}
      </div>
    </div>

    <div>
      <div class="muted" style="font-size: 13px;">Odesláno tipů</div>
      <div style="font-size: 24px; font-weight: 900;">
        {{ active_round_stats.tips_total }}
      </div>
    </div>

    <div>
      <div class="muted" style="font-size: 13px;">Dokončení</div>
      <div style="font-size: 24px; font-weight: 900; color: var(--ok);">
        {{ active_round_stats.completion }}%
      </div>
    </div>
  </div>

  <div class="progress-bar-container">
    <div class="progress-bar-fill" style="width: {{ active_round_stats.completion }}%;"></div>
  </div>
</div>
{% endif %}

{% if matches_no_results > 0 %}
<div class="card" style="background: rgba(249,199,79,.08); border-color: rgba(249,199,79,.3);">
  <h3 style="margin: 0 0 12px 0; color: #f9c74f;">⚠️ Čeká na vyřízení</h3>
  <div>
    <strong>{{ matches_no_results }}</strong> zápasů bez výsledků
    <a href="{{ url_for('admin_bulk_edit') }}" class="btn btn-sm" style="margin-left: 12px;">Zadat výsledky</a>
  </div>
</div>
{% endif %}

<div class="card">
  <h3 style="margin: 0 0 16px 0;">👥 Noví uživatelé</h3>
  {% for user in recent_users %}
    <div style="padding: 8px 0; border-bottom: 1px solid var(--line);">
      <strong>{{ user.display_name }}</strong>
      <span class="muted" style="font-size: 12px; margin-left: 8px;">@{{ user.username }}</span>
      {% if user.is_admin %}<span class="tag pill-ok" style="margin-left: 8px;">Admin</span>{% endif %}
    </div>
  {% endfor %}
</div>

<div class="card">
  <h3 style="margin: 0 0 16px 0;">🔧 Admin nástroje</h3>
  <div class="admin-cards">

    <a href="{{ url_for('admin_bulk_edit') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">✏️</div>
        <div class="admin-card-title">Bulk Edit</div>
      </div>
      <div class="admin-card-desc">Zadávání výsledků</div>
    </a>

    <a href="{{ url_for('admin_import') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">📥</div>
        <div class="admin-card-title">Import</div>
      </div>
      <div class="admin-card-desc">Importovat data</div>
    </a>

    <a href="{{ url_for('admin_export_hub') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">📤</div>
        <div class="admin-card-title">Export</div>
      </div>
      <div class="admin-card-desc">Exportovat data</div>
    </a>

    <a href="{{ url_for('admin_undo') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">⏪</div>
        <div class="admin-card-title">Undo</div>
      </div>
      <div class="admin-card-desc">Vrátit změny</div>
    </a>

    <a href="{{ url_for('admin_rounds') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">🎮</div>
        <div class="admin-card-title">Soutěže</div>
      </div>
      <div class="admin-card-desc">Správa soutěží</div>
    </a>

    <a href="{{ url_for('admin_users') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">👥</div>
        <div class="admin-card-title">Uživatelé</div>
      </div>
      <div class="admin-card-desc">Správa userů</div>
    </a>

    <a href="{{ url_for('admin_api_sources') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">🔌</div>
        <div class="admin-card-title">API Zdroje</div>
      </div>
      <div class="admin-card-desc">Fotbal & hokej API</div>
    </a>

    <a href="{{ url_for('admin_smart_import') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">🤖</div>
        <div class="admin-card-title">Smart Import</div>
      </div>
      <div class="admin-card-desc">AI parsování zápasů</div>
    </a>

  </div>
</div>

""", total_users=total_users, total_rounds=total_rounds, active_rounds=active_rounds,
     total_matches=total_matches, total_tips=total_tips, recent_users=recent_users,
     active_round_stats=active_round_stats, matches_no_results=matches_no_results)

    # --- ADMIN BULK EDIT ---


    @admin_bp.route("/bulk-edit")
    @login_required
    def admin_bulk_edit():
        admin_required()

        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None

        if not r:
            flash("Vyber soutěž pro bulk editaci.", "error")
            return redirect(url_for("admin_rounds"))

        # Načti všechny zápasy
        matches = Match.query.filter_by(
            round_id=r.id,
            is_deleted=False
        ).order_by(Match.start_time.asc(), Match.id.asc()).all()

        # Stats
        total = len(matches)
        with_results = sum(1 for m in matches if m.home_score is not None and m.away_score is not None)
        without_results = total - with_results

        return render_page(r"""
<style>
  .bulk-table {
    width: 100%;
    border-collapse: collapse;
  }

  .bulk-table th,
  .bulk-table td {
    padding: 12px 8px;
    text-align: left;
    border-bottom: 1px solid var(--line);
  }

  .bulk-table th {
    background: rgba(255,255,255,.03);
    font-weight: 900;
    position: sticky;
    top: 0;
  }

  .bulk-table input[type="number"] {
    width: 60px;
    text-align: center;
  }

  .match-row:hover {
    background: rgba(255,255,255,.03);
  }

  .match-row.has-result {
    background: rgba(51,209,122,.05);
  }
</style>

<div class="card">
  <div class="row" style="justify-content: space-between; align-items: center;">
    <div>
      <h2 style="margin: 0 0 8px 0;">✏️ Bulk Edit - Hromadné úpravy</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <div class="row" style="gap: 8px;">
      <div class="tag pill-ok">✅ {{ with_results }}</div>
      <div class="tag pill-warn">⏳ {{ without_results }}</div>
    </div>
  </div>
</div>

<form method="post" action="{{ url_for('admin_bulk_edit_save') }}">
  <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
  <input type="hidden" name="round_id" value="{{ r.id }}">

  <div class="card">
    <div class="row" style="justify-content: space-between; align-items: center; margin-bottom: 16px;">
      <h3 style="margin: 0;">📋 Zápasy ({{ total }})</h3>
      <button type="submit" class="btn btn-primary">💾 Uložit všechny změny</button>
    </div>

    <div class="bulk-table-wrapper" style="overflow-x: auto;">
      <table class="bulk-table">
        <thead>
          <tr>
            <th style="width: 40px;">#</th>
            <th>Domácí</th>
            <th style="width: 60px; text-align: center;">Skóre</th>
            <th>Hosté</th>
            <th style="width: 60px; text-align: center;">Skóre</th>
          </tr>
        </thead>
        <tbody>
          {% for m in matches %}
            <tr class="match-row {% if m.home_score is not none and m.away_score is not none %}has-result{% endif %}">
              <td>{{ loop.index }}</td>
              <td><strong>{{ m.home_team.name if m.home_team else '?' }}</strong></td>
              <td style="text-align: center;">
                <input type="number"
                       name="match_{{ m.id }}_home"
                       value="{{ m.home_score if m.home_score is not none else '' }}"
                       min="0" max="20">
              </td>
              <td><strong>{{ m.away_team.name if m.away_team else '?' }}</strong></td>
              <td style="text-align: center;">
                <input type="number"
                       name="match_{{ m.id }}_away"
                       value="{{ m.away_score if m.away_score is not none else '' }}"
                       min="0" max="20">
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <div style="margin-top: 16px; text-align: right;">
      <button type="submit" class="btn btn-primary">💾 Uložit všechny změny</button>
    </div>
  </div>
</form>

""", r=r, matches=matches, total=total, with_results=with_results, without_results=without_results)



    @admin_bp.route("/bulk-edit/save", methods=["POST"])
    @login_required
    def admin_bulk_edit_save():
        admin_required()

        round_id = int(request.form.get("round_id"))
        r = db.session.get(Round, round_id)

        if not r:
            flash("Soutěž nenalezena.", "error")
            return redirect(url_for("admin_bulk_edit"))

        matches = Match.query.filter_by(round_id=r.id, is_deleted=False).all()

        updated_count = 0
        for match in matches:
            match_id = match.id

            home_score_str = request.form.get(f"match_{match_id}_home")
            away_score_str = request.form.get(f"match_{match_id}_away")

            old_home = match.home_score
            old_away = match.away_score

            if home_score_str and away_score_str:
                new_home = int(home_score_str)
                new_away = int(away_score_str)

                if old_home != new_home or old_away != new_away:
                    # Vytvoř undo point PŘED změnou
                    home_team_name = match.home_team.name if match.home_team else "?"
                    away_team_name = match.away_team.name if match.away_team else "?"
                    old_score_str = f"{old_home}:{old_away}" if old_home is not None and old_away is not None else "—"
                    new_score_str = f"{new_home}:{new_away}"

                    create_undo_point(
                        action_type='update_score',
                        entity_type='Match',
                        entity_id=match.id,
                        before_state={
                            'home_score': old_home,
                            'away_score': old_away
                        },
                        description=f"{home_team_name} {old_score_str} → {new_score_str} {away_team_name}"
                    )

                    # Teď proveď změnu
                    match.home_score = new_home
                    match.away_score = new_away
                    updated_count += 1
            elif home_score_str == '' and away_score_str == '':
                if match.home_score is not None or match.away_score is not None:
                    # Vytvoř undo point PŘED smazáním
                    home_team_name = match.home_team.name if match.home_team else "?"
                    away_team_name = match.away_team.name if match.away_team else "?"
                    old_score_str = f"{old_home}:{old_away}" if old_home is not None and old_away is not None else "—"

                    create_undo_point(
                        action_type='clear_score',
                        entity_type='Match',
                        entity_id=match.id,
                        before_state={
                            'home_score': old_home,
                            'away_score': old_away
                        },
                        description=f"{home_team_name} {old_score_str} → smazáno {away_team_name}"
                    )

                    match.home_score = None
                    match.away_score = None
                    updated_count += 1

        db.session.commit()
        audit("bulk_edit.save", "Match", None, details=f"Updated {updated_count} matches")

        # Pošli push notifikace o zadaných výsledcích
        if updated_count > 0:
            try:
                send_results_notification(round_id)
            except Exception as e:
                print(f"Error sending push notifications: {e}")

        flash(f"✅ Aktualizováno {updated_count} zápasů!", "ok")
        return redirect(url_for("admin_bulk_edit"))

    # --- ADMIN BULK IMPORT (CSV) WITH PREVIEW ---


    @admin_bp.route("/bulk-import/template")
    @login_required
    def admin_bulk_import_template():
        """Stáhne CSV šablonu pro bulk import"""
        admin_required()

        try:
            from io import StringIO

            # Create CSV template
            output = StringIO()
            output.write("Domácí,Hosté,Datum,Čas\n")
            output.write("Sparta Praha,Slavia Praha,2024-03-15,18:00\n")
            output.write("Plzeň,Brno,2024-03-16,16:30\n")
            output.write("Baník,Bohemians,2024-03-17,15:00\n")

            # Create response
            from flask import Response
            response = Response(output.getvalue(), mimetype='text/csv')
            response.headers['Content-Disposition'] = 'attachment; filename=bulk_import_sablona.csv'
            return response

        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_bulk_import"))



    @admin_bp.route("/bulk-import", methods=["GET", "POST"])
    @login_required
    def admin_bulk_import():
        """Hromadný import týmů a zápasů z CSV - STEP 1: Upload"""
        admin_required()

        if request.method == "POST":
            import_type = request.form.get("import_type")
            round_id = request.form.get("round_id")

            if not round_id:
                flash("Vyber soutěž!", "error")
                return redirect(url_for("admin_bulk_import"))

            round_id = int(round_id)
            r = db.session.get(Round, round_id)

            if not r:
                flash("Soutěž nenalezena!", "error")
                return redirect(url_for("admin_bulk_import"))

            if 'csv_file' not in request.files:
                flash("Žádný soubor nevybrán!", "error")
                return redirect(url_for("admin_bulk_import"))

            file = request.files['csv_file']

            if file.filename == '':
                flash("Žádný soubor nevybrán!", "error")
                return redirect(url_for("admin_bulk_import"))

            if not file.filename.endswith('.csv'):
                flash("Musí být CSV soubor!", "error")
                return redirect(url_for("admin_bulk_import"))

            try:
                # Read CSV and store in temporary file (session cookies are limited to 4KB)
                csv_content = file.stream.read().decode("utf-8")

                # Create temporary file
                temp_fd, temp_path = tempfile.mkstemp(suffix='.csv', prefix='bulk_import_')
                with os.fdopen(temp_fd, 'w', encoding='utf-8') as f:
                    f.write(csv_content)

                # Store only the file path in session (small!)
                session['bulk_import_file'] = temp_path
                session['bulk_import_type'] = import_type
                session['bulk_import_round_id'] = round_id

                # Redirect to preview
                return redirect(url_for("admin_bulk_import_preview"))

            except Exception as e:
                flash(f"❌ Chyba při čtení CSV: {str(e)}", "error")
                return redirect(url_for("admin_bulk_import"))

        # GET request - show form
        rounds = Round.query.order_by(Round.id.desc()).all()

        return render_page(r"""
<div class="card">
  <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <div>
      <h2>📥 Hromadný import z CSV</h2>
      <div class="muted">Import týmů a zápasů z CSV souborů s preview</div>
    </div>
    <a href="{{ url_for('admin_bulk_import_template') }}" class="btn" style="background:#6ea8fe; color:white;">
      📥 Stáhnout šablonu CSV
    </a>
  </div>
</div>

<div class="card">
  <h3>📋 Import týmů</h3>
  <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="import_type" value="teams">

    <div style="margin-bottom: 16px;">
      <label class="muted">Soutěž *</label>
      <select name="round_id" required>
        <option value="">-- Vyber soutěž --</option>
        {% for r in rounds %}
          <option value="{{ r.id }}">{{ r.name }}</option>
        {% endfor %}
      </select>
    </div>

    <div style="margin-bottom: 16px;">
      <label class="muted">CSV soubor *</label>
      <input type="file" name="csv_file" accept=".csv" required>
      <div class="muted" style="font-size: 13px; margin-top: 4px;">
        Formát: <code>name</code> (1 sloupec)
      </div>
    </div>

    <button type="submit" class="btn btn-primary">👁️ Zobrazit preview</button>
  </form>
</div>

<div class="card">
  <h3>⚽ Import zápasů</h3>
  <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="import_type" value="matches">

    <div style="margin-bottom: 16px;">
      <label class="muted">Soutěž *</label>
      <select name="round_id" required>
        <option value="">-- Vyber soutěž --</option>
        {% for r in rounds %}
          <option value="{{ r.id }}">{{ r.name }}</option>
        {% endfor %}
      </select>
    </div>

    <div style="margin-bottom: 16px;">
      <label class="muted">CSV soubor *</label>
      <input type="file" name="csv_file" accept=".csv" required>
      <div class="muted" style="font-size: 13px; margin-top: 4px;">
        Formát: <code>home_team,away_team,start_time,home_score,away_score</code><br>
        Start time: <code>YYYY-MM-DD HH:MM</code> nebo <code>YYYY-MM-DD</code><br>
        Scores: Nechej prázdné pokud zápas ještě nebyl
      </div>
    </div>

    <button type="submit" class="btn btn-primary">👁️ Zobrazit preview</button>
  </form>
</div>

<div class="card">
  <h3>👤 Import tipů</h3>
  <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="import_type" value="tips">

    <div style="margin-bottom: 16px;">
      <label class="muted">Soutěž *</label>
      <select name="round_id" required>
        <option value="">-- Vyber soutěž --</option>
        {% for r in rounds %}
          <option value="{{ r.id }}">{{ r.name }}</option>
        {% endfor %}
      </select>
    </div>

    <div style="margin-bottom: 16px;">
      <label class="muted">CSV soubor *</label>
      <input type="file" name="csv_file" accept=".csv" required>
      <div class="muted" style="font-size: 13px; margin-top: 4px;">
        Formát: <code>user_email,home_team,away_team,home_score,away_score</code><br>
        Příklad: <code>mejlacz@gmail.com,Česko,Kanada,3,2</code><br>
        <strong>Možnost přepsání:</strong> Existující tipy budou označeny jako "Update"
      </div>
    </div>

    <button type="submit" class="btn btn-primary">👁️ Zobrazit preview</button>
  </form>
</div>

<div class="card" style="background: rgba(255,255,255,.02);">
  <h3>💡 Jak na to</h3>
  <ol style="margin: 0; padding-left: 20px;">
    <li>Vyber CSV soubor → zobrazí se <strong>preview</strong></li>
    <li>Zkontroluj data v tabulce</li>
    <li>Klikni "Potvrdit" nebo "Zrušit"</li>
    <li>Pořadí: <strong>týmy</strong> → <strong>zápasy</strong> → <strong>tipy</strong></li>
    <li><strong>Update tipů:</strong> Existující tipy lze přepsat (status "Update")</li>
  </ol>
</div>
""", rounds=rounds)



    @admin_bp.route("/bulk-import/preview")
    @login_required
    def admin_bulk_import_preview():
        """STEP 2: Preview dat před importem"""
        admin_required()

        # Get data from session
        temp_file = session.get('bulk_import_file')
        import_type = session.get('bulk_import_type')
        round_id = session.get('bulk_import_round_id')

        if not temp_file or not import_type or not round_id:
            flash("Session expirovala, nahraj CSV znovu", "error")
            return redirect(url_for("admin_bulk_import"))

        # Check if temp file still exists
        if not os.path.exists(temp_file):
            flash("CSV soubor expiroval, nahraj znovu", "error")
            session.pop('bulk_import_file', None)
            return redirect(url_for("admin_bulk_import"))

        r = db.session.get(Round, round_id)
        if not r:
            flash("Soutěž nenalezena", "error")
            return redirect(url_for("admin_bulk_import"))

        try:
            # Read CSV from temp file
            with open(temp_file, 'r', encoding='utf-8') as f:
                csv_content = f.read()

            stream = io.StringIO(csv_content, newline=None)
            csv_reader = csv.DictReader(stream)
            rows = list(csv_reader)

            preview_data = []
            new_count = 0
            overwrite_count = 0
            error_count = 0
            errors = []

            if import_type == 'teams':
                for idx, row in enumerate(rows):
                    team_name = row.get('name', '').strip()

                    if not team_name:
                        error_count += 1
                        errors.append("Prázdný název týmu")
                        continue

                    existing = Team.query.filter_by(round_id=round_id, name=team_name).first()

                    status = "overwrite" if existing else "new"
                    if status == "new":
                        new_count += 1
                    else:
                        overwrite_count += 1

                    preview_data.append({
                        'index': idx,
                        'name': team_name,
                        'status': status,
                        'selected': status == 'new',  # Auto-select only new
                        'existing_id': existing.id if existing else None
                    })

            elif import_type == 'matches':
                for idx, row in enumerate(rows):
                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    start_time_str = row.get('start_time', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    if not home_name or not away_name:
                        error_count += 1
                        errors.append(f"Chybí název týmu")
                        preview_data.append({
                            'index': idx,
                            'home_team': home_name or '?',
                            'away_team': away_name or '?',
                            'start_time': start_time_str,
                            'home_score': home_score_str,
                            'away_score': away_score_str,
                            'status': 'error',
                            'error': 'Chybí název týmu',
                            'selected': False
                        })
                        continue

                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()

                    if not home_team or not away_team:
                        error_count += 1
                        missing = []
                        if not home_team:
                            missing.append(home_name)
                        if not away_team:
                            missing.append(away_name)
                        error_msg = f"Tým nenalezen: {', '.join(missing)}"
                        errors.append(error_msg)
                        preview_data.append({
                            'index': idx,
                            'home_team': home_name,
                            'away_team': away_name,
                            'start_time': start_time_str,
                            'home_score': home_score_str,
                            'away_score': away_score_str,
                            'status': 'error',
                            'error': error_msg,
                            'selected': False
                        })
                        continue

                    # Check existing
                    existing = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    status = "new"
                    if existing:
                        # Zápas již existuje - můžeš ho přepsat
                        status = "overwrite"
                        overwrite_count += 1  # Count as overwrite
                    else:
                        new_count += 1

                    preview_data.append({
                        'index': idx,
                        'home_team': home_name,
                        'away_team': away_name,
                        'start_time': start_time_str,
                        'home_score': home_score_str if home_score_str else '—',
                        'away_score': away_score_str if away_score_str else '—',
                        'status': status,
                        'selected': status == 'new',  # Auto-select only NEW (not overwrite by default)
                        'existing_id': existing.id if existing else None  # Store ID for update
                    })

            elif import_type == 'tips':
                for idx, row in enumerate(rows):
                    user_email = row.get('user_email', '').strip()
                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    # Validate basic data
                    if not user_email or not home_name or not away_name or not home_score_str or not away_score_str:
                        error_count += 1
                        errors.append(f"Neúplná data na řádku {idx+1}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email or '?',
                            'home_team': home_name or '?',
                            'away_team': away_name or '?',
                            'tip': f"{home_score_str or '?'}:{away_score_str or '?'}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Neúplná data',
                            'selected': False
                        })
                        continue

                    # Find user
                    user = User.query.filter_by(email=user_email).first()
                    if not user:
                        error_count += 1
                        errors.append(f"Uživatel nenalezen: {user_email}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Uživatel nenalezen',
                            'selected': False
                        })
                        continue

                    # Find teams
                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()

                    if not home_team or not away_team:
                        error_count += 1
                        missing = []
                        if not home_team:
                            missing.append(home_name)
                        if not away_team:
                            missing.append(away_name)
                        error_msg = f"Tým nenalezen: {', '.join(missing)}"
                        errors.append(error_msg)
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': error_msg,
                            'selected': False
                        })
                        continue

                    # Find match
                    match = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if not match:
                        error_count += 1
                        errors.append(f"Zápas nenalezen: {home_name} vs {away_name}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Zápas nenalezen',
                            'selected': False
                        })
                        continue

                    # Parse tip scores
                    try:
                        new_home = int(home_score_str)
                        new_away = int(away_score_str)
                    except:
                        error_count += 1
                        errors.append(f"Neplatné skóre: {home_score_str}:{away_score_str}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Neplatné skóre',
                            'selected': False
                        })
                        continue

                    # Find existing tip
                    existing_tip = Tip.query.filter_by(
                        user_id=user.id,
                        match_id=match.id
                    ).first()

                    # Determine status
                    status = "new"
                    current_tip = "—"

                    if existing_tip:
                        current_tip = f"{existing_tip.home_score}:{existing_tip.away_score}"
                        if existing_tip.home_score != new_home or existing_tip.away_score != new_away:
                            status = "update"
                        else:
                            status = "skip"

                    if status == "new":
                        new_count += 1
                    elif status == "skip":
                        skip_count += 1
                    elif status == "update":
                        update_count += 1

                    preview_data.append({
                        'index': idx,
                        'user_email': user_email,
                        'user_name': user.username,
                        'home_team': home_name,
                        'away_team': away_name,
                        'tip': f"{new_home}:{new_away}",
                        'current_tip': current_tip,
                        'status': status,
                        'selected': status in ['new', 'update']  # Auto-select new and updates
                    })

            # Render preview
            if import_type == 'teams':
                table_html = """
<table style="width: 100%; border-collapse: collapse;">
  <tr style="border-bottom: 2px solid var(--line);">
    <th style="padding: 12px 8px; text-align: center; width: 50px;">
      <input type="checkbox" id="select-all" checked>
    </th>
    <th style="padding: 12px 8px; text-align: left;">Název týmu</th>
    <th style="padding: 12px 8px; text-align: center; width: 120px;">Status</th>
  </tr>
  {% for item in preview_data %}
  <tr style="border-bottom: 1px solid var(--line);">
    <td style="padding: 10px 8px; text-align: center;">
      <input type="checkbox" name="selected_rows" value="{{ item.index }}" {% if item.selected %}checked{% endif %} class="row-checkbox">
    </td>
    <td style="padding: 10px 8px;">{{ item.name }}</td>
    <td style="padding: 10px 8px; text-align: center;">
      {% if item.status == 'new' %}
        <span class="tag" style="background: rgba(51,209,122,.15); color: #33d17a;">Nový</span>
      {% elif item.status == 'overwrite' %}
        <span class="tag" style="background: rgba(99,179,237,.15); color: #63b3ed;">Přepsat</span>
      {% endif %}
    </td>
  </tr>
  {% endfor %}
</table>
"""
            elif import_type == 'matches':
                table_html = """
<table style="width: 100%; border-collapse: collapse; font-size: 14px;">
  <tr style="border-bottom: 2px solid var(--line);">
    <th style="padding: 10px 8px; text-align: center; width: 50px;">
      <input type="checkbox" id="select-all" checked>
    </th>
    <th style="padding: 10px 8px; text-align: left;">Domácí</th>
    <th style="padding: 10px 8px; text-align: left;">Hosté</th>
    <th style="padding: 10px 8px; text-align: center; width: 100px;">Skóre</th>
    <th style="padding: 10px 8px; text-align: center; width: 120px;">Status</th>
  </tr>
  {% for item in preview_data %}
  <tr style="border-bottom: 1px solid var(--line);">
    <td style="padding: 8px; text-align: center;">
      <input type="checkbox" name="selected_rows" value="{{ item.index }}" {% if item.selected %}checked{% endif %} class="row-checkbox">
    </td>
    <td style="padding: 8px;">{{ item.home_team }}</td>
    <td style="padding: 8px;">{{ item.away_team }}</td>
    <td style="padding: 8px; text-align: center;">
      {% if item.home_score != '—' and item.away_score != '—' %}
        {{ item.home_score }}:{{ item.away_score }}
      {% else %}
        <span class="muted">—</span>
      {% endif %}
    </td>
    <td style="padding: 8px; text-align: center;">
      {% if item.status == 'new' %}
        <span class="tag" style="background: rgba(51,209,122,.15); color: #33d17a;">Nový</span>
      {% elif item.status == 'skip' %}
        <span class="tag" style="background: rgba(255,255,255,.05); color: var(--muted);">Přeskočit</span>
      {% elif item.status == 'overwrite' %}
        <span class="tag" style="background: rgba(99,179,237,.15); color: #63b3ed;">Přepsat</span>
      {% elif item.status == 'update' %}
        <span class="tag" style="background: rgba(249,199,79,.15); color: #f9c74f;">Update</span>
      {% elif item.status == 'error' %}
        <span class="tag" style="background: rgba(255,77,109,.15); color: #ff4d6d;">Chyba</span>
      {% endif %}
    </td>
  </tr>
  {% endfor %}
</table>
"""
            else:  # tips
                table_html = """
<table style="width: 100%; border-collapse: collapse; font-size: 14px;">
  <tr style="border-bottom: 2px solid var(--line);">
    <th style="padding: 10px 8px; text-align: center; width: 50px;">
      <input type="checkbox" id="select-all" checked>
    </th>
    <th style="padding: 10px 8px; text-align: left;">Uživatel</th>
    <th style="padding: 10px 8px; text-align: left;">Domácí</th>
    <th style="padding: 10px 8px; text-align: left;">Hosté</th>
    <th style="padding: 10px 8px; text-align: center; width: 80px;">Nový tip</th>
    <th style="padding: 10px 8px; text-align: center; width: 80px;">Současný</th>
    <th style="padding: 10px 8px; text-align: center; width: 100px;">Status</th>
  </tr>
  {% for item in preview_data %}
  <tr style="border-bottom: 1px solid var(--line);">
    <td style="padding: 8px; text-align: center;">
      <input type="checkbox" name="selected_rows" value="{{ item.index }}" {% if item.selected %}checked{% endif %} class="row-checkbox">
    </td>
    <td style="padding: 8px;">
      {% if item.user_name %}
        <strong>{{ item.user_name }}</strong><br>
        <span class="muted" style="font-size: 11px;">{{ item.user_email }}</span>
      {% else %}
        {{ item.user_email }}
      {% endif %}
    </td>
    <td style="padding: 8px;">{{ item.home_team }}</td>
    <td style="padding: 8px;">{{ item.away_team }}</td>
    <td style="padding: 8px; text-align: center;">
      <strong>{{ item.tip }}</strong>
    </td>
    <td style="padding: 8px; text-align: center;">
      {% if item.current_tip != '—' %}
        {{ item.current_tip }}
      {% else %}
        <span class="muted">—</span>
      {% endif %}
    </td>
    <td style="padding: 8px; text-align: center;">
      {% if item.status == 'new' %}
        <span class="tag" style="background: rgba(51,209,122,.15); color: #33d17a;">Nový</span>
      {% elif item.status == 'skip' %}
        <span class="tag" style="background: rgba(255,255,255,.05); color: var(--muted);">Stejný</span>
      {% elif item.status == 'update' %}
        <span class="tag" style="background: rgba(249,199,79,.15); color: #f9c74f;">Přepsat</span>
      {% elif item.status == 'error' %}
        <span class="tag" style="background: rgba(255,77,109,.15); color: #ff4d6d;">Chyba</span>
      {% endif %}
    </td>
  </tr>
  {% endfor %}
</table>
"""

            return render_page(r"""
<div class="card">
  <h2>👁️ Preview importu</h2>
  <div class="muted">Zkontroluj data před potvrzením</div>
</div>

<div class="card">
  <div class="row" style="gap: 32px;">
    <div>
      <div class="muted">Celkem</div>
      <div style="font-size: 28px; font-weight: 900;">{{ total }}</div>
    </div>
    <div>
      <div class="muted">Nové</div>
      <div style="font-size: 28px; font-weight: 900; color: #33d17a;">{{ new_count }}</div>
    </div>
    {% if overwrite_count > 0 %}
    <div>
      <div class="muted">Přepsat</div>
      <div style="font-size: 28px; font-weight: 900; color: #63b3ed;">{{ overwrite_count }}</div>
    </div>
    {% endif %}
    {% if error_count > 0 %}
    <div>
      <div class="muted">Chyby</div>
      <div style="font-size: 28px; font-weight: 900; color: #ff4d6d;">{{ error_count }}</div>
    </div>
    {% endif %}
  </div>
</div>

{% if errors|length > 0 %}
<div class="card" style="background: rgba(255,77,109,.08);">
  <h3>⚠️ Chyby ({{ errors|length }})</h3>
  {% for error in errors[:10] %}
    <div class="muted" style="font-size: 13px; margin-bottom: 4px;">• {{ error }}</div>
  {% endfor %}
  {% if errors|length > 10 %}
    <div class="muted" style="font-size: 13px; margin-top: 8px;">... a {{ errors|length - 10 }} dalších</div>
  {% endif %}
</div>
{% endif %}

<div class="card">
  <h3>📊 Data ({{ preview_data|length }})</h3>

  <div style="margin-bottom: 12px; display: flex; gap: 12px; flex-wrap: wrap;">
    <button type="button" onclick="selectAll()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      ✅ Vybrat vše
    </button>
    <button type="button" onclick="deselectAll()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      ❌ Zrušit vše
    </button>
    <button type="button" onclick="selectNew()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      🟢 Jen nové
    </button>
    <button type="button" onclick="selectOverwrite()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      🔵 Přepsat existující
    </button>
    <div style="flex: 1; text-align: right; line-height: 32px;">
      <span class="muted" style="font-size: 13px;">
        Vybráno: <strong id="selected-count">{{ new_count }}</strong> / {{ preview_data|length }}
      </span>
    </div>
  </div>

  <div style="overflow-x: auto; margin-top: 12px;">
    """ + table_html + """
  </div>
</div>

<script>
function updateCount() {
  const checked = document.querySelectorAll('.row-checkbox:checked').length;
  document.getElementById('selected-count').textContent = checked;
  document.getElementById('confirm-count').textContent = checked;
}

function selectAll() {
  document.querySelectorAll('.row-checkbox').forEach(cb => cb.checked = true);
  document.getElementById('select-all').checked = true;
  updateCount();
}

function deselectAll() {
  document.querySelectorAll('.row-checkbox').forEach(cb => cb.checked = false);
  document.getElementById('select-all').checked = false;
  updateCount();
}

function selectNew() {
  deselectAll();
  document.querySelectorAll('tr').forEach(row => {
    const tag = row.querySelector('.tag');
    const checkbox = row.querySelector('.row-checkbox');
    if (tag && tag.textContent.trim() === 'Nový' && checkbox) {
      checkbox.checked = true;
    }
  });
  updateCount();
}

function selectOverwrite() {
  deselectAll();
  document.querySelectorAll('tr').forEach(row => {
    const tag = row.querySelector('.tag');
    const checkbox = row.querySelector('.row-checkbox');
    if (tag && tag.textContent.trim() === 'Přepsat' && checkbox) {
      checkbox.checked = true;
    }
  });
  updateCount();
}

function selectUpdates() {
  deselectAll();
  document.querySelectorAll('tr').forEach(row => {
    const tag = row.querySelector('.tag');
    const checkbox = row.querySelector('.row-checkbox');
    if (tag && tag.textContent.trim() === 'Update' && checkbox) {
      checkbox.checked = true;
    }
  });
  updateCount();
}

// Select all checkbox handler
document.getElementById('select-all').addEventListener('change', function() {
  if (this.checked) {
    selectAll();
  } else {
    deselectAll();
  }
});

// Individual checkbox handler
document.querySelectorAll('.row-checkbox').forEach(cb => {
  cb.addEventListener('change', updateCount);
});

// Initial count
updateCount();
</script>

<form method="post" action="{{ url_for('admin_bulk_import_confirm') }}">
  <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
  <div class="card">
    <div class="row" style="gap: 12px;">
      <button type="submit" name="action" value="confirm" class="btn btn-primary" style="flex: 1;">
        ✅ Potvrdit import (<span id="confirm-count">{{ new_count + update_count }}</span>)
      </button>
      <a href="{{ url_for('admin_bulk_import') }}" class="btn" style="flex: 1; text-align: center; line-height: 40px;">
        ❌ Zrušit
      </a>
    </div>
  </div>
</form>
""",
            preview_data=preview_data,
            total=len(preview_data),
            new_count=new_count,
            overwrite_count=overwrite_count,
            error_count=error_count,
            errors=errors,
            import_type=import_type,
            round_name=r.name
        )

        except Exception as e:
            flash(f"❌ Chyba při preview: {str(e)}", "error")
            import traceback
            traceback.print_exc()
            return redirect(url_for("admin_bulk_import"))



    @admin_bp.route("/bulk-import/confirm", methods=["POST"])
    @login_required
    def admin_bulk_import_confirm():
        """STEP 3: Potvrzení a provedení importu"""
        admin_required()

        action = request.form.get("action")

        # Get temp file path
        temp_file = session.get('bulk_import_file')

        if action != "confirm":
            flash("Import zrušen", "ok")
            # Cleanup temp file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
            session.pop('bulk_import_file', None)
            session.pop('bulk_import_type', None)
            session.pop('bulk_import_round_id', None)
            return redirect(url_for("admin_bulk_import"))

        # Get data from session
        import_type = session.get('bulk_import_type')
        round_id = session.get('bulk_import_round_id')

        if not temp_file or not import_type or not round_id:
            flash("Session expirovala, nahraj CSV znovu", "error")
            return redirect(url_for("admin_bulk_import"))

        # Check if temp file still exists
        if not os.path.exists(temp_file):
            flash("CSV soubor expiroval, nahraj znovu", "error")
            session.pop('bulk_import_file', None)
            return redirect(url_for("admin_bulk_import"))

        r = db.session.get(Round, round_id)
        if not r:
            flash("Soutěž nenalezena", "error")
            return redirect(url_for("admin_bulk_import"))

        # Get selected rows from form
        selected_rows = request.form.getlist('selected_rows')
        if not selected_rows:
            flash("Nevybrané žádné záznamy k importu", "error")
            # Cleanup temp file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
            return redirect(url_for("admin_bulk_import_preview"))

        # Convert to set of integers for fast lookup
        selected_indices = set(int(idx) for idx in selected_rows)

        try:
            # Read CSV from temp file
            with open(temp_file, 'r', encoding='utf-8') as f:
                csv_content = f.read()

            stream = io.StringIO(csv_content, newline=None)
            csv_reader = csv.DictReader(stream)

            if import_type == 'teams':
                imported = 0
                skipped = 0

                for idx, row in enumerate(csv_reader):
                    # Skip if not selected
                    if idx not in selected_indices:
                        continue

                    team_name = row.get('name', '').strip()

                    if not team_name:
                        continue

                    existing = Team.query.filter_by(round_id=round_id, name=team_name).first()

                    if existing:
                        skipped += 1
                    else:
                        team = Team(round_id=round_id, name=team_name)
                        db.session.add(team)
                        imported += 1

                db.session.commit()
                audit("bulk_import.teams", "Team", None, details=f"Imported {imported}, skipped {skipped}")
                flash(f"✅ Importováno {imported} týmů, přeskočeno {skipped}", "ok")

            elif import_type == 'matches':
                imported = 0
                skipped = 0
                updated = 0
                errors = []

                for idx, row in enumerate(csv_reader):
                    # Skip if not selected
                    if idx not in selected_indices:
                        continue

                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    start_time_str = row.get('start_time', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    if not home_name or not away_name:
                        continue

                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()

                    if not home_team or not away_team:
                        continue

                    # Parse start time
                    start_time = None
                    if start_time_str:
                        try:
                            start_time = datetime.strptime(start_time_str, '%Y-%m-%d %H:%M')
                        except:
                            try:
                                start_time = datetime.strptime(start_time_str, '%Y-%m-%d')
                            except:
                                continue

                    # Parse scores
                    home_score = None
                    away_score = None

                    if home_score_str and away_score_str:
                        try:
                            home_score = int(home_score_str)
                            away_score = int(away_score_str)
                        except:
                            pass

                    # Check if match exists
                    existing = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if existing:
                        # OVERWRITE - přepsat existující zápas VŠEMI daty z CSV
                        if start_time is not None:
                            existing.start_time = start_time
                        existing.home_score = home_score
                        existing.away_score = away_score
                        updated += 1
                    else:
                        # Create new match
                        match = Match(
                            round_id=round_id,
                            home_team_id=home_team.id,
                            away_team_id=away_team.id,
                            start_time=start_time,
                            home_score=home_score,
                            away_score=away_score
                        )
                        db.session.add(match)
                        imported += 1

                db.session.commit()
                audit("bulk_import.matches", "Match", None, details=f"Imported {imported}, overwritten {updated}")

                if updated > 0:
                    flash(f"✅ Importováno {imported} nových zápasů, přepsáno {updated} existujících", "ok")
                else:
                    flash(f"✅ Importováno {imported} nových zápasů", "ok")

            elif import_type == 'tips':
                imported = 0
                updated = 0
                skipped = 0
                errors = []

                for idx, row in enumerate(csv_reader):
                    # Skip if not selected
                    if idx not in selected_indices:
                        continue

                    user_email = row.get('user_email', '').strip()
                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    # Validate
                    if not user_email or not home_name or not away_name or not home_score_str or not away_score_str:
                        continue

                    # Find user
                    user = User.query.filter_by(email=user_email).first()
                    if not user:
                        continue

                    # Find teams
                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()
                    if not home_team or not away_team:
                        continue

                    # Find match
                    match = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()
                    if not match:
                        continue

                    # Parse scores
                    try:
                        new_home = int(home_score_str)
                        new_away = int(away_score_str)
                    except:
                        continue

                    # Find existing tip
                    existing_tip = Tip.query.filter_by(
                        user_id=user.id,
                        match_id=match.id
                    ).first()

                    if existing_tip:
                        # Update if different
                        if existing_tip.home_score != new_home or existing_tip.away_score != new_away:
                            existing_tip.home_score = new_home
                            existing_tip.away_score = new_away
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        # Create new tip
                        tip = Tip(
                            user_id=user.id,
                            match_id=match.id,
                            home_score=new_home,
                            away_score=new_away
                        )
                        db.session.add(tip)
                        imported += 1

                db.session.commit()
                audit("bulk_import.tips", "Tip", None, details=f"Imported {imported}, updated {updated}, skipped {skipped}")

                if updated > 0:
                    flash(f"✅ Importováno {imported} tipů, přepsáno {updated}, přeskočeno {skipped}", "ok")
                else:
                    flash(f"✅ Importováno {imported} tipů, přeskočeno {skipped}", "ok")

            # Clear session and cleanup temp file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
            session.pop('bulk_import_file', None)
            session.pop('bulk_import_type', None)
            session.pop('bulk_import_round_id', None)

        except Exception as e:
            flash(f"❌ Chyba při importu: {str(e)}", "error")
            import traceback
            traceback.print_exc()
            # Cleanup temp file on error too
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass

        return redirect(url_for("admin_bulk_import"))

    # --- ADMIN EXPORT ---


    @admin_bp.route("/export/<what>")
    @login_required
    def admin_export(what):
        admin_required()

        if what == "users":
            users = User.query.all()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['ID', 'Username', 'Email', 'Jméno', 'Příjmení', 'Admin'])

            for u in users:
                writer.writerow([
                    u.id,
                    u.username,
                    u.email,
                    u.first_name or '',
                    u.last_name or '',
                    'Ano' if u.is_admin else 'Ne'
                ])

            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment; filename=users_export.csv'}
            )

        elif what == "matches":
            rid = ensure_selected_round()
            r = db.session.get(Round, rid) if rid else None

            if not r:
                flash("Vyber soutěž.", "error")
                return redirect(url_for("admin_dashboard"))

            matches = Match.query.filter_by(round_id=r.id, is_deleted=False).order_by(Match.start_time).all()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['ID', 'Domácí', 'Hosté', 'Skóre domácí', 'Skóre hosté'])

            for m in matches:
                writer.writerow([
                    m.id,
                    m.home_team.name if m.home_team else '',
                    m.away_team.name if m.away_team else '',
                    m.home_score if m.home_score is not None else '',
                    m.away_score if m.away_score is not None else ''
                ])

            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=matches_{r.name.replace(" ", "_")}.csv'}
            )

        else:
            flash("Neznámý export typ.", "error")
            return redirect(url_for("admin_dashboard"))

    # --- ADMIN UNDO ---


    @admin_bp.route("/undo")
    @login_required
    def admin_undo():
        admin_required()

        # Načti poslední undo pointy (maximálně 20)
        recent_undos = UndoStack.query.filter_by(
            user_id=current_user.id,
            is_undone=False
        ).order_by(UndoStack.created_at.desc()).limit(20).all()

        # Stats
        total_undos = UndoStack.query.filter_by(user_id=current_user.id).count()
        available_undos = UndoStack.query.filter_by(user_id=current_user.id, is_undone=False).count()

        return render_page(r"""
<div class="card">
  <h2 style="margin: 0 0 8px 0;">🔄 Undo - Vrátit změny</h2>
  <div class="muted">Možnost vrátit poslední akce zpět</div>
  <hr class="sep">

  <div class="row" style="gap: 12px; margin-bottom: 16px;">
    <div class="tag pill-ok">✅ {{ available_undos }} dostupných</div>
    <div class="tag">📊 {{ total_undos }} celkem</div>
  </div>

  {% if not recent_undos %}
    <div class="card" style="background: rgba(255,255,255,.03); text-align: center; padding: 32px;">
      <div style="font-size: 48px; margin-bottom: 16px;">📝</div>
      <div class="muted">Žádné akce k vrácení.</div>
      <div class="muted" style="font-size: 13px; margin-top: 8px;">
        Změny se zaznamenávají automaticky při Bulk Edit.
      </div>
    </div>
  {% else %}
    <table class="datatable">
      <thead>
        <tr>
          <th style="width: 140px;">Čas</th>
          <th style="width: 120px;">Typ</th>
          <th>Popis</th>
          <th style="width: 120px; text-align: center;">Akce</th>
        </tr>
      </thead>
      <tbody>
        {% for undo in recent_undos %}
          <tr>
            <td>{{ undo.created_at.strftime("%d.%m. %H:%M") }}</td>
            <td>
              <span class="tag pill-ok">{{ undo.entity_type }}</span>
            </td>
            <td>{{ undo.description or '-' }}</td>
            <td style="text-align: center;">
              <form method="post" action="{{ url_for('admin_undo_perform', undo_id=undo.id) }}" style="margin: 0;">
                <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                <button type="submit" class="btn btn-sm btn-primary"
                        onclick="return confirm('Opravdu vrátit tuto změnu?')">
                  ↶ Vrátit
                </button>
              </form>
            </td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
  {% endif %}
</div>

<div class="card" style="background: rgba(110,168,254,.08); border-color: rgba(110,168,254,.3);">
  <h3 style="margin: 0 0 12px 0;">💡 Jak to funguje?</h3>
  <ul style="margin: 0; padding-left: 20px;">
    <li>Při každé důležité změně (Bulk Edit) se vytvoří undo point</li>
    <li>Můžeš kdykoli vrátit změnu pomocí tlačítka "Vrátit"</li>
    <li>Každý admin vidí pouze své vlastní undo pointy</li>
    <li>Akce lze vrátit pouze jednou</li>
    <li>Zobrazuje se max 20 posledních změn</li>
  </ul>
</div>

""", recent_undos=recent_undos, total_undos=total_undos, available_undos=available_undos)



    @admin_bp.route("/undo/<int:undo_id>/perform", methods=["POST"])
    @login_required
    def admin_undo_perform(undo_id):
        admin_required()

        result = perform_undo(undo_id)

        if result['success']:
            flash(result['message'], "ok")
        else:
            flash(f"❌ {result['message']}", "error")

        return redirect(url_for("admin_undo"))

    # === BACKUP DATABÁZE ===



    @admin_bp.route("/backup")
    @login_required
    def admin_backup():
        """Stránka pro správu záloh databáze"""
        admin_required()

        # Zjisti velikost aktuální databáze
        db_path = os.path.join(app.instance_path, 'tipovacka.db')
        try:
            db_size = os.path.getsize(db_path)
            db_size_mb = db_size / (1024 * 1024)
        except:
            db_size_mb = 0

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">💾 Záloha databáze</h2>
  <div class="muted">Pravidelně zálohuj data tipovačky</div>
  <hr class="sep">

  <div style="background:rgba(110,168,254,0.08); padding:16px; border-radius:8px; margin-bottom:16px;">
    <div class="row" style="justify-content:space-between;">
      <div>
        <strong>Aktuální databáze:</strong>
      </div>
      <div>
        <span class="tag">{{ "%.2f"|format(db_size_mb) }} MB</span>
      </div>
    </div>
  </div>

  <h3 style="margin:16px 0 8px 0;">Manual záloha</h3>
  <div class="muted" style="margin-bottom:12px;">
    Vytvoř zálohu kdykoliv kliknutím na tlačítko
  </div>

  <div class="row" style="gap:10px; margin-bottom:24px;">
    <a href="{{ url_for('admin_backup_download') }}" class="btn btn-primary">
      📥 Stáhnout backup
    </a>
    <a href="{{ url_for('admin_backup_email') }}" class="btn" style="background:#667eea; color:white;">
      📧 Poslat na email
    </a>
  </div>

  <div style="background:rgba(255,193,7,0.1); padding:12px; border-left:4px solid #ffc107; font-size:13px; margin-bottom:24px;">
    <strong>📧 Email backup:</strong> Odešle se na <strong>{{ current_user.email }}</strong>
  </div>

  <hr class="sep">

  <h3 style="margin:16px 0 8px 0;">⏰ Automatický backup</h3>
  <div class="muted" style="margin-bottom:12px;">
    Záloha se pošle automaticky každý den v 3:00
  </div>

  <div style="background:rgba(110,168,254,0.1); padding:16px; border-radius:8px; margin-bottom:16px;">
    <div class="row" style="gap:16px; align-items:flex-start;">
      <div style="flex:1;">
        <strong style="display:block; margin-bottom:8px;">FREE účet (PythonAnywhere)</strong>
        <div class="muted" style="font-size:13px; line-height:1.6;">
          ❌ Scheduled tasks nejsou dostupné<br>
          ✅ Použij manual backup tlačítka výše
        </div>
      </div>
      <div style="flex:1;">
        <strong style="display:block; margin-bottom:8px;">PAID účet / Wedos</strong>
        <div class="muted" style="font-size:13px; line-height:1.6;">
          ✅ Cron job dostupný<br>
          ✅ Denní automatický backup<br>
          ✅ Email všem adminům
        </div>
      </div>
    </div>
  </div>

  <details style="margin-top:16px;">
    <summary style="cursor:pointer; padding:12px; background:rgba(0,0,0,0.05); border-radius:4px; font-weight:600;">
      📖 Návod: Nastavení automatického backupu (pro budoucnost)
    </summary>
    <div style="padding:16px; background:rgba(0,0,0,0.02); border-radius:4px; margin-top:8px; font-size:13px;">
      <p><strong>Soubor: backup_daily.py</strong></p>
      <pre style="background:#0b1020; color:#6ea8fe; padding:12px; border-radius:4px; overflow-x:auto; font-size:11px; line-height:1.4;">#!/usr/bin/env python3
import os, sys
sys.path.insert(0, '/path/to/tipovacka')

os.environ['SECRET_KEY'] = 'tvuj-secret-key'
os.environ['SEND_REAL_EMAILS'] = 'true'
os.environ['FROM_EMAIL'] = 'noreply@tvoje-domena.cz'
os.environ['FROM_NAME'] = 'Tipovačka Backup'
os.environ['SMTP_SERVER'] = 'smtp.wedos.com'
os.environ['SMTP_PORT'] = '587'
os.environ['SMTP_USERNAME'] = 'noreply@tvoje-domena.cz'
os.environ['SMTP_PASSWORD'] = 'heslo'

from app2 import app, db, User, send_email_with_attachment
from datetime import datetime
import zipfile, os
from io import BytesIO

with app.app_context():
    admins = User.query.filter(User.role.in_(['admin', 'owner'])).all()
    db_path = os.path.join(app.instance_path, 'tipovacka.db')

    memory_file = BytesIO()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(db_path, f'tipovacka_{timestamp}.db')

    memory_file.seek(0)
    zip_data = memory_file.read()
    size_mb = len(zip_data) / (1024 * 1024)

    for admin in admins:
        html = f"&lt;h1&gt;Backup {datetime.now().strftime('%d.%m.%Y')}&lt;/h1&gt;&lt;p&gt;Velikost: {size_mb:.2f} MB&lt;/p&gt;"
        send_email_with_attachment(admin.email, f"Záloha {datetime.now().strftime('%d.%m.%Y')}", html, "Backup", zip_data, f'backup_{timestamp}.zip')

    print(f"{datetime.now()} - Backup odeslán {len(admins)} adminům")</pre>

      <p style="margin-top:16px;"><strong>Cron job (Wedos/VPS):</strong></p>
      <pre style="background:#0b1020; color:#6ea8fe; padding:12px; border-radius:4px; font-size:12px;">0 3 * * * /usr/bin/python3 /path/to/backup_daily.py >> /var/log/backup.log 2>&1</pre>

      <p class="muted" style="margin-top:12px;">
        = Každý den ve 3:00 se spustí backup a pošle email všem adminům
      </p>
    </div>
  </details>

  <hr class="sep">
  <a href="{{ url_for('admin_users') }}" class="btn">← Zpět do admin</a>
</div>
""", db_size_mb=db_size_mb)



    @admin_bp.route("/backup/download")
    @login_required
    def admin_backup_download():
        """Stáhne aktuální databázi jako .zip"""
        admin_required()

        try:
            import zipfile
            from io import BytesIO
            from datetime import datetime

            db_path = os.path.join(app.instance_path, 'tipovacka.db')

            if not os.path.exists(db_path):
                flash("Databáze nenalezena!", "error")
                return redirect(url_for("admin_backup"))

            # Vytvoř zip v paměti
            memory_file = BytesIO()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(db_path, f'tipovacka_{timestamp}.db')

            memory_file.seek(0)

            audit("backup.download", "Database", None, timestamp=timestamp)

            return send_file(
                memory_file,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'tipovacka_backup_{timestamp}.zip'
            )

        except Exception as e:
            flash(f"Chyba při vytváření backupu: {str(e)}", "error")
            return redirect(url_for("admin_backup"))



    @admin_bp.route("/backup/email")
    @login_required
    def admin_backup_email():
        """Pošle backup databáze na email aktuálního admina"""
        admin_required()

        try:
            import zipfile
            from io import BytesIO
            from datetime import datetime

            db_path = os.path.join(app.instance_path, 'tipovacka.db')

            if not os.path.exists(db_path):
                flash("Databáze nenalezena!", "error")
                return redirect(url_for("admin_backup"))

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
            timestamp_file = datetime.now().strftime('%Y%m%d_%H%M%S')

            # Vytvoř zip
            memory_file = BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(db_path, f'tipovacka_{timestamp_file}.db')

            memory_file.seek(0)
            zip_data = memory_file.read()

            # Velikost backupu
            size_mb = len(zip_data) / (1024 * 1024)

            # Pošli email s přílohou
            html = f"""
            <!DOCTYPE html>
            <html>
            <head><meta charset="utf-8"></head>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <div style="background: #0b1020; color: white; padding: 20px; text-align: center;">
                        <h1>💾 Záloha databáze</h1>
                    </div>
                    <div style="background: #f4f4f4; padding: 30px;">
                        <h2>Záloha tipovačky</h2>
                        <p><strong>Datum:</strong> {timestamp}</p>
                        <p><strong>Velikost:</strong> {size_mb:.2f} MB</p>
                        <p><strong>Soubor:</strong> tipovacka_backup_{timestamp_file}.zip</p>

                        <div style="background: #e3f2fd; padding: 15px; border-left: 4px solid #2196f3; margin: 20px 0;">
                            <strong>💡 Doporučení:</strong><br>
                            • Ulož si backup na bezpečné místo<br>
                            • Nezveřejňuj ho (obsahuje všechna data)<br>
                            • Gmail: 15 GB free = ~5000 backupů
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """

            text = f"""Záloha tipovačky

Datum: {timestamp}
Velikost: {size_mb:.2f} MB
Soubor: tipovacka_backup_{timestamp_file}.zip
            """

            # Odešli s přílohou
            success = send_email_with_attachment(
                to_email=current_user.email,
                subject=f"Záloha tipovačky - {timestamp}",
                html_body=html,
                text_body=text,
                attachment_data=zip_data,
                attachment_name=f'tipovacka_backup_{timestamp_file}.zip'
            )

            if success:
                flash(f"✅ Backup odeslán na: {current_user.email} ({size_mb:.2f} MB)", "ok")
                audit("backup.email", "Database", None, size_mb=f"{size_mb:.2f}")
            else:
                flash(f"⚠️ Chyba při odesílání. Zkus stáhnout manuálně.", "warning")

            return redirect(url_for("admin_backup"))

        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_backup"))

    # --- ADMIN IMPORT ---


    @admin_bp.route("/import")
    @login_required
    def admin_import():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Import (CSV)</h2>
  <div class="muted">Vybraná soutěž: <b>{{ r.name if r else "—" }}</b>. Pokud CSV obsahuje sloupec <b>round_id</b>, použije se.</div>
  <hr class="sep">
  <div class="row">
    <a class="btn btn-primary" href="{{ url_for('admin_import_teams') }}">Import týmů</a>
    <a class="btn btn-primary" href="{{ url_for('admin_import_matches') }}">Import zápasů</a>
    <a class="btn btn-primary" href="{{ url_for('admin_import_extras') }}">Import extra</a>
  </div>
  <hr class="sep">
  <div style="margin-top:20px;">
    <h3 style="margin:0 0 8px 0;">Import dat z jiné tipovačky</h3>
    <div class="muted" style="margin-bottom:10px;">Importuj kompletní data (uživatele, zápasy a tipy) z Excel žebříčku</div>
    <div class="row" style="gap: 12px;">
      <a class="btn btn-primary" href="{{ url_for('admin_import_leaderboard_smart') }}" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
        ✨ Smart Import (s kontrolou duplicit)
      </a>
      <a class="btn" href="{{ url_for('admin_import_leaderboard') }}">
        📥 Klasický Import (bez kontroly)
      </a>
    </div>
    <div class="muted" style="margin-top: 8px; font-size: 12px;">
      💡 <strong>Smart Import</strong> ti ukáže preview a označí možné duplicity před importem!
    </div>
  </div>
</div>
""", r=r)

    # --- ADMIN SMART IMPORT (s kontrolou duplicit) ---


    @admin_bp.route("/import/leaderboard-smart/template")
    @login_required
    def admin_import_leaderboard_smart_template():
        """Stáhne Excel šablonu pro smart import žebříčku"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Smart Import"

            # Hlavička - Jméno | Zápasy
            headers = ['Jméno', 'Sparta-Slavia', 'Plzeň-Brno', 'Baník-Bohemians']
            ws.append(headers)

            # Stylování
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Příklady
            ws.append(['Jan Novák', '2:1', '1:1', '3:0'])
            ws.append(['Petr Svoboda', '1:2', '2:0', '1:1'])

            # Šířka
            ws.column_dimensions['A'].width = 20
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 15

            # Poznámky
            notes = wb.create_sheet("Poznámky")
            notes['A1'] = "SMART IMPORT - S KONTROLOU DUPLICIT"
            notes['A1'].font = Font(bold=True, size=14)
            notes['A3'] = "Formát stejný jako běžný import žebříčku."
            notes['A4'] = "Výhoda: Automaticky detekuje duplicitní zápasy!"
            notes['A6'] = "HLAVIČKA:"
            notes['A7'] = "   Jméno | Domácí-Hosté | Domácí-Hosté ..."
            notes['A9'] = "TIPY:"
            notes['A10'] = "   Jan | 2:1 | 1:0 | 3:2"
            notes.column_dimensions['A'].width = 50

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name='smart_import_sablona.xlsx')
        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_import"))



    @admin_bp.route("/import/leaderboard-smart", methods=["GET", "POST"])
    @login_required
    def admin_import_leaderboard_smart():
        admin_required()

        if request.method == "POST":
            file = request.files.get('excel_file')
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_leaderboard_smart"))

            # Určit cílovou soutěž
            import_target = request.form.get('import_target', 'existing')

            if import_target == 'new':
                new_round_name = request.form.get('new_round_name', '').strip()
                if not new_round_name:
                    flash("Zadej název nové soutěže.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                sport = Sport.query.filter_by(name="Fotbal").first()
                if not sport:
                    sport = Sport(name="Fotbal")
                    db.session.add(sport)
                    db.session.flush()

                r = Round(name=new_round_name, sport_id=sport.id, is_active=False)
                db.session.add(r)
                db.session.flush()
                round_id = r.id
                audit("round.create.import", "Round", r.id, name=r.name)
            else:
                round_id = int(request.form.get('round_id', 0))
                r = db.session.get(Round, round_id)
                if not r:
                    flash("Vybraná soutěž neexistuje.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

            try:
                import openpyxl
                from io import BytesIO

                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active

                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    flash("Soubor musí mít alespoň hlavičku a jeden řádek.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                # Parsovat zápasy z hlavičky
                matches_data = []
                for i, match_name in enumerate(header[1:], start=1):
                    if not match_name or str(match_name).strip() == "":
                        continue

                    match_str = str(match_name).strip()
                    for sep in ['-', ' vs ', ' x ', ':', ' – ']:
                        if sep in match_str:
                            parts = match_str.split(sep, 1)
                            if len(parts) == 2:
                                home = parts[0].strip()
                                away = parts[1].strip()

                                # Detekovat možnou duplicitu
                                home_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == home.lower()
                                ).first()

                                away_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == away.lower()
                                ).first()

                                is_duplicate = False
                                existing_match_id = None
                                status = 'new'  # NOVÉ!

                                if home_team_check and away_team_check:
                                    existing_match = Match.query.filter_by(
                                        round_id=round_id,
                                        home_team_id=home_team_check.id,
                                        away_team_id=away_team_check.id,
                                        is_deleted=False
                                    ).first()

                                    if existing_match:
                                        is_duplicate = True
                                        existing_match_id = existing_match.id
                                        status = 'overwrite'  # Může být přepsán!

                                matches_data.append({
                                    'col': i,
                                    'home': home,
                                    'away': away,
                                    'is_duplicate': is_duplicate,  # Zpětná kompatibilita
                                    'status': status,  # NOVÉ!
                                    'existing_match_id': existing_match_id,
                                    'match_str': match_str
                                })
                                break

                if not matches_data:
                    flash("Nenalezeny žádné platné zápasy v hlavičce.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                # Uložit preview data do TEMP FILE (ne session - cookie overflow!)
                import tempfile
                import pickle

                preview_data = {
                    'round_id': round_id,
                    'round_name': r.name,
                    'matches': matches_data,  # Velké data!
                    'total_matches': len(matches_data),
                    'overwrite_count': sum(1 for m in matches_data if m['status'] == 'overwrite'),
                    'new_count': sum(1 for m in matches_data if m['status'] == 'new')
                }

                # Uložit preview data do pickle file
                preview_fd, preview_path = tempfile.mkstemp(suffix='.pkl', prefix='smart_preview_')
                with os.fdopen(preview_fd, 'wb') as f:
                    pickle.dump(preview_data, f)

                # Do session jen cesta + základní info (malé!)
                session['smart_import_preview'] = {
                    'preview_file': preview_path,
                    'round_id': round_id,
                    'round_name': r.name,
                    'total_matches': len(matches_data),
                    'overwrite_count': sum(1 for m in matches_data if m['status'] == 'overwrite'),
                    'new_count': sum(1 for m in matches_data if m['status'] == 'new')
                }

                # Uložit Excel file do TEMP FILE
                file.seek(0)  # Reset file pointer
                excel_fd, excel_path = tempfile.mkstemp(suffix='.xlsx', prefix='smart_excel_')
                with os.fdopen(excel_fd, 'wb') as f:
                    f.write(file.read())
                session['smart_import_file'] = excel_path

                return redirect(url_for("admin_import_leaderboard_smart_preview"))

            except Exception as e:
                flash(f"Chyba při načítání souboru: {str(e)}", "error")
                db.session.rollback()
                return redirect(url_for("admin_import_leaderboard_smart"))

        # GET
        all_rounds = Round.query.order_by(Round.id.desc()).all()

        return render_page(r"""
<div class="card">
  <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <div>
      <h2 style="margin: 0 0 8px 0;">✨ Smart Import Žebříčku</h2>
      <div class="muted">Import s kontrolou duplicit a možností výběru</div>
    </div>
    <a href="{{ url_for('admin_import_leaderboard_smart_template') }}" class="btn" style="background:#667eea; color:white;">
      📥 Stáhnout šablonu Excel
    </a>
  </div>
  <hr class="sep">

  <div class="card" style="background: rgba(110,168,254,.08); border-color: rgba(110,168,254,.3); margin-bottom: 16px;">
    <h3 style="margin: 0 0 12px 0;">💡 Jak to funguje?</h3>
    <ul style="margin: 0; padding-left: 20px;">
      <li><strong>Preview před importem</strong> - Vidíš CO se chystá importovat</li>
      <li><strong>Detekce existujících</strong> - Modře označí zápasy které již existují (mohou být použity)</li>
      <li><strong>Checkboxy pro výběr</strong> - Ručně vyber co chceš/nechceš importovat</li>
      <li><strong>Quick select</strong> - Tlačítka "Jen nové" / "Přepsat existující"</li>
    </ul>
  </div>

  <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; align-items:stretch; gap:16px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">1️⃣ Kam importovat data?</h3>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="existing" {% if all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_existing">
        <span>Do existující soutěže</span>
      </label>

      <div id="existing_round_select" style="margin-left:28px; margin-bottom:16px;">
        <select name="round_id" style="width:100%; max-width:400px;">
          {% for rnd in all_rounds %}
            <option value="{{ rnd.id }}">
              {% if rnd.is_active %}★ {% endif %}{{ rnd.name }}
            </option>
          {% endfor %}
          {% if not all_rounds %}
            <option value="">-- Žádné soutěže --</option>
          {% endif %}
        </select>
      </div>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="new" {% if not all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_new">
        <span>Vytvořit novou soutěž</span>
      </label>

      <div id="new_round_input" style="margin-left:28px; display:none;">
        <input type="text" name="new_round_name" placeholder="Název nové soutěže"
               style="width:100%; max-width:400px;">
      </div>
    </div>

    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">2️⃣ Vyber Excel soubor</h3>
      <input type="file" name="excel_file" accept=".xlsx,.xls" required>
    </div>

    <button class="btn btn-primary" type="submit" style="padding:14px; font-size:16px; font-weight:900;">
      📥 Načíst preview
    </button>
    <a class="btn" href="{{ url_for('admin_import') }}">Zpět</a>
  </form>
</div>

<script>
function toggleImportTarget() {
  const existingChecked = document.getElementById('radio_existing').checked;
  const existingDiv = document.getElementById('existing_round_select');
  const newDiv = document.getElementById('new_round_input');

  if (existingChecked) {
    existingDiv.style.display = 'block';
    newDiv.style.display = 'none';
  } else {
    existingDiv.style.display = 'none';
    newDiv.style.display = 'block';
  }
}

toggleImportTarget();
</script>
""", all_rounds=all_rounds)



    @admin_bp.route("/import/leaderboard-smart/preview")
    @login_required
    def admin_import_leaderboard_smart_preview():
        admin_required()

        preview_meta = session.get('smart_import_preview')
        if not preview_meta:
            flash("Nejprve nahraj soubor.", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

        # Načti matches data z pickle file
        preview_file = preview_meta.get('preview_file')
        if not preview_file or not os.path.exists(preview_file):
            flash("Preview data expirovala, nahraj soubor znovu.", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

        import pickle
        with open(preview_file, 'rb') as f:
            preview = pickle.load(f)

        return render_page(r"""
<style>
  .match-preview {
    display: flex;
    align-items: center;
    padding: 12px;
    border: 1px solid var(--line);
    border-radius: 8px;
    margin-bottom: 8px;
    transition: all 0.2s ease;
  }

  .match-preview:hover {
    background: rgba(255,255,255,.03);
  }

  .match-preview.overwrite {
    background: rgba(99,179,237,.08);
    border-color: rgba(99,179,237,.3);
  }

  .match-preview.new {
    background: rgba(51,209,122,.05);
    border-color: rgba(51,209,122,.2);
  }

  .match-checkbox {
    margin-right: 12px;
    width: 20px;
    height: 20px;
    cursor: pointer;
  }
</style>

<div class="card">
  <h2 style="margin: 0 0 8px 0;">✨ Preview importu</h2>
  <div class="muted">Soutěž: <b>{{ preview.round_name }}</b></div>
  <hr class="sep">

  <div class="row" style="justify-content: space-between; margin-bottom: 16px;">
    <div>
      <div class="tag pill-ok" style="font-size: 14px;">✅ {{ preview.new_count }} nových</div>
      <div class="tag pill-info" style="font-size: 14px; background: rgba(99,179,237,.15); color: #63b3ed;">🔵 {{ preview.overwrite_count }} přepsat</div>
    </div>

    <div class="row" style="gap: 8px;">
      <button type="button" onclick="selectAll()" class="btn btn-sm">☑️ Vybrat vše</button>
      <button type="button" onclick="deselectAll()" class="btn btn-sm">☐ Zrušit výběr</button>
      <button type="button" onclick="selectNew()" class="btn btn-sm">🟢 Jen nové</button>
      <button type="button" onclick="selectOverwrite()" class="btn btn-sm btn-primary">🔵 Přepsat existující</button>
    </div>
  </div>
</div>

<form method="post" action="{{ url_for('admin_import_leaderboard_smart_confirm') }}">
  <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
  <div class="card">
    <h3 style="margin: 0 0 16px 0;">📋 Zápasy k importu ({{ preview.total_matches }})</h3>

    {% for match in preview.matches %}
      <div class="match-preview {{ match.status }}" data-status="{{ match.status }}">
        <input type="checkbox"
               name="selected_matches"
               value="{{ loop.index0 }}"
               class="match-checkbox"
               {% if match.status == 'new' %}checked{% endif %}
               id="match_{{ loop.index0 }}">

        <label for="match_{{ loop.index0 }}" style="flex: 1; cursor: pointer; margin: 0;">
          <div class="row" style="justify-content: space-between; align-items: center;">
            <div>
              <strong>{{ match.home }}</strong>
              <span class="muted">vs</span>
              <strong>{{ match.away }}</strong>
            </div>

            <div>
              {% if match.status == 'overwrite' %}
                <span class="tag pill-info" style="background: rgba(99,179,237,.15); color: #63b3ed;">🔵 PŘEPSAT</span>
              {% else %}
                <span class="tag pill-ok">✅ NOVÝ</span>
              {% endif %}
            </div>
          </div>
        </label>
      </div>
    {% endfor %}
  </div>

  <div class="card">
    <div class="row" style="justify-content: space-between; align-items: center;">
      <div>
        <div style="font-size: 14px; margin-bottom: 4px;">
          Importuje se: <strong><span id="selected-count">{{ preview.new_count }}</span></strong> zápasů
        </div>
        <div class="muted" style="font-size: 12px;">
          Duplicity budou přeskočeny
        </div>
      </div>

      <div class="row" style="gap: 12px;">
        <a href="{{ url_for('admin_import_leaderboard_smart') }}" class="btn">❌ Zrušit</a>
        <button type="submit" class="btn btn-primary" style="font-size: 16px; padding: 12px 24px;">
          ✅ Potvrdit a importovat
        </button>
      </div>
    </div>
  </div>
</form>

<script>
function updateCount() {
  const checkboxes = document.querySelectorAll('.match-checkbox:checked');
  document.getElementById('selected-count').textContent = checkboxes.length;
}

function selectAll() {
  document.querySelectorAll('.match-checkbox').forEach(cb => {
    cb.checked = true;
  });
  updateCount();
}

function deselectAll() {
  document.querySelectorAll('.match-checkbox').forEach(cb => {
    cb.checked = false;
  });
  updateCount();
}

function selectNew() {
  deselectAll();
  document.querySelectorAll('.match-preview').forEach(preview => {
    const status = preview.getAttribute('data-status');
    const checkbox = preview.querySelector('.match-checkbox');
    if (status === 'new' && checkbox) {
      checkbox.checked = true;
    }
  });
  updateCount();
}

function selectOverwrite() {
  deselectAll();
  document.querySelectorAll('.match-preview').forEach(preview => {
    const status = preview.getAttribute('data-status');
    const checkbox = preview.querySelector('.match-checkbox');
    if (status === 'overwrite' && checkbox) {
      checkbox.checked = true;
    }
  });
  updateCount();
}

// Update count on checkbox change
document.querySelectorAll('.match-checkbox').forEach(cb => {
  cb.addEventListener('change', updateCount);
});
</script>
""", preview=preview)

    def _resolve_round_id(row: dict[str, str], fallback_round_id: int) -> int:
        val = (row.get("round_id") or row.get("round") or "").strip()
        if val:
            try:
                return int(val)
            except Exception:
                pass
        return fallback_round_id



    @admin_bp.route("/import/leaderboard-smart/confirm", methods=["POST"])
    @login_required
    def admin_import_leaderboard_smart_confirm():
        admin_required()

        preview_meta = session.get('smart_import_preview')
        temp_file = session.get('smart_import_file')

        if not preview_meta or not temp_file:
            flash("Session vypršela, nahraj soubor znovu.", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

        # Zkontroluj jestli temp files existují
        preview_file = preview_meta.get('preview_file')
        if not os.path.exists(temp_file) or not os.path.exists(preview_file):
            flash("Soubor expiroval, nahraj znovu.", "error")
            session.pop('smart_import_file', None)
            session.pop('smart_import_preview', None)
            return redirect(url_for("admin_import_leaderboard_smart"))

        try:
            # Načti preview data z pickle file
            import pickle
            with open(preview_file, 'rb') as f:
                preview = pickle.load(f)

            # Načti vybrané indexy
            selected_indices = request.form.getlist('selected_matches')
            selected_indices = [int(i) for i in selected_indices]

            if not selected_indices:
                flash("Nevybral jsi žádné zápasy k importu.", "error")
                return redirect(url_for("admin_import_leaderboard_smart_preview"))

            # Načti Excel file z TEMP FILE
            import openpyxl

            wb = openpyxl.load_workbook(temp_file)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]

            round_id = preview['round_id']
            r = db.session.get(Round, round_id)

            if not r:
                flash("Soutěž nenalezena.", "error")
                return redirect(url_for("admin_import_leaderboard_smart"))

            # Importuj pouze vybrané zápasy (vytvoř nové nebo použij existující)
            created_matches = 0
            skipped_matches = 0
            match_map = {}

            # KROK 1: Vytvoř/použij vybrané zápasy
            for idx in selected_indices:
                if idx >= len(preview['matches']):
                    continue

                match_data = preview['matches'][idx]

                home_team = _get_or_create_team(r.id, match_data['home'])
                away_team = _get_or_create_team(r.id, match_data['away'])

                # Zkusit najít existující zápas
                m = Match.query.filter_by(
                    round_id=r.id,
                    home_team_id=home_team.id,
                    away_team_id=away_team.id,
                    is_deleted=False
                ).first()

                if not m:
                    m = Match(round_id=r.id, home_team_id=home_team.id, away_team_id=away_team.id)
                    db.session.add(m)
                    db.session.flush()
                    created_matches += 1
                else:
                    skipped_matches += 1

                match_map[match_data['col']] = m

            # KROK 2: Pro NEVybrané zápasy - najdi existující (pro import tipů)
            for idx, match_data in enumerate(preview['matches']):
                if idx in selected_indices:
                    continue  # Už je v match_map

                # Zkus najít existující zápas (bez vytváření nového)
                # CASE-INSENSITIVE vyhledávání!
                home_team = Team.query.filter(
                    Team.round_id == r.id,
                    db.func.lower(Team.name) == match_data['home'].lower(),
                    Team.is_deleted == False
                ).first()

                away_team = Team.query.filter(
                    Team.round_id == r.id,
                    db.func.lower(Team.name) == match_data['away'].lower(),
                    Team.is_deleted == False
                ).first()

                if home_team and away_team:
                    existing_match = Match.query.filter_by(
                        round_id=r.id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if existing_match:
                        # Přidej do match_map pro import tipů
                        match_map[match_data['col']] = existing_match

            # Import tipů (stejně jako v klasickém importu, ale jen pro vybrané zápasy)
            results_imported = 0
            data_rows_start = 1
            data_rows_end = len(rows)

            # Kontrola výsledků v prvním řádku
            if len(rows) > 1:
                first_data_row = rows[1]
                first_cell = str(first_data_row[0]).strip() if first_data_row[0] else ""

                if first_cell.lower() in ['', 'výsledek', 'result', 'skóre', 'score', 'vysledek']:
                    for col_idx, match in match_map.items():
                        if col_idx >= len(first_data_row):
                            continue

                        result_value = first_data_row[col_idx]
                        if not result_value:
                            continue

                        home_score = None
                        away_score = None

                        if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                            home_score = result_value.hour if result_value.hour < 20 else None
                            away_score = result_value.minute if result_value.minute < 60 else None
                        elif isinstance(result_value, str) and ':' in result_value:
                            parts = result_value.split(':')
                            if len(parts) == 2:
                                try:
                                    home_score = int(parts[0].strip())
                                    away_score = int(parts[1].strip())
                                except:
                                    pass

                        if home_score is not None and away_score is not None:
                            match.home_score = home_score
                            match.away_score = away_score
                            results_imported += 1

                    data_rows_start = 2

            # Kontrola výsledků v posledním řádku
            if len(rows) > data_rows_start:
                last_row = rows[-1]
                last_cell = str(last_row[0]).strip() if last_row[0] else ""

                if last_cell.lower() in ['výsledek', 'result', 'skóre', 'score', 'vysledek']:
                    for col_idx, match in match_map.items():
                        if col_idx >= len(last_row):
                            continue

                        result_value = last_row[col_idx]
                        if not result_value:
                            continue

                        home_score = None
                        away_score = None

                        if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                            home_score = result_value.hour if result_value.hour < 20 else None
                            away_score = result_value.minute if result_value.minute < 60 else None
                        elif isinstance(result_value, str) and ':' in result_value:
                            parts = result_value.split(':')
                            if len(parts) == 2:
                                try:
                                    home_score = int(parts[0].strip())
                                    away_score = int(parts[1].strip())
                                except:
                                    pass

                        if home_score is not None and away_score is not None:
                            match.home_score = home_score
                            match.away_score = away_score
                            results_imported += 1

                    data_rows_end = len(rows) - 1

            # Import tipů
            users_created = 0
            tips_imported = 0
            tips_overwritten = 0

            for row_idx in range(data_rows_start, data_rows_end):
                if row_idx >= len(rows):
                    break

                row = rows[row_idx]
                if not row or len(row) < 2:
                    continue

                username = str(row[0]).strip() if row[0] else ""
                if not username or username.lower() in ['výsledek', 'result', 'skóre', 'score', 'vysledek']:
                    continue

                # Normalize Unicode
                import unicodedata
                username = unicodedata.normalize('NFC', username)

                # Flush aby jsme viděli nově vytvořené uživatele
                db.session.flush()

                # PYTHON-based case-insensitive search (SPOLEHLIVĚJŠÍ!)
                all_users = User.query.all()
                u = None
                username_lower = username.lower()
                for user in all_users:
                    user_norm = unicodedata.normalize('NFC', user.username)
                    if user_norm.lower() == username_lower:
                        u = user
                        break

                if not u:
                    u = User(
                        username=username,
                        email=f"{username.lower()}@imported.local",
                        password_hash=generate_password_hash("changeme123"),
                        is_admin=False
                    )
                    db.session.add(u)
                    db.session.flush()
                    users_created += 1

                # Import tipů pro vybrané zápasy
                for col_idx, match in match_map.items():
                    if col_idx >= len(row):
                        continue

                    tip_value = row[col_idx]
                    if not tip_value:
                        continue

                    home_guess = None
                    away_guess = None

                    if hasattr(tip_value, 'hour') and hasattr(tip_value, 'minute'):
                        home_guess = tip_value.hour if tip_value.hour < 20 else None
                        away_guess = tip_value.minute if tip_value.minute < 60 else None
                    elif isinstance(tip_value, str) and ':' in tip_value:
                        parts = tip_value.split(':')
                        if len(parts) == 2:
                            try:
                                home_guess = int(parts[0].strip())
                                away_guess = int(parts[1].strip())
                            except:
                                pass

                    if home_guess is not None and away_guess is not None:
                        existing_tip = Tip.query.filter_by(user_id=u.id, match_id=match.id).first()
                        if existing_tip:
                            # PŘEPSAT existující tip novými daty
                            existing_tip.tip_home = home_guess
                            existing_tip.tip_away = away_guess
                            tips_overwritten += 1
                        else:
                            # Vytvořit nový tip
                            tip = Tip(
                                user_id=u.id,
                                match_id=match.id,
                                tip_home=home_guess,
                                tip_away=away_guess
                            )
                            db.session.add(tip)
                            tips_imported += 1

            db.session.commit()

            # Smaž temp files
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass

            if preview_file and os.path.exists(preview_file):
                try:
                    os.unlink(preview_file)
                except:
                    pass

            # Vyčisti session
            session.pop('smart_import_preview', None)
            session.pop('smart_import_file', None)

            audit("leaderboard.smart_import", "Round", r.id,
                  created_matches=created_matches,
                  skipped_matches=skipped_matches,
                  users_created=users_created,
                  tips_imported=tips_imported,
                  tips_overwritten=tips_overwritten,
                  results_imported=results_imported)

            tips_message = f"Importováno {tips_imported} nových tipů"
            if tips_overwritten > 0:
                tips_message += f", přepsáno {tips_overwritten} existujících"

            flash(f"""✅ Smart Import dokončen!
                  Vytvořeno {created_matches} nových zápasů
                  Použito {skipped_matches} existujících zápasů
                  Vytvořeno {users_created} nových uživatelů
                  {tips_message}
                  Importováno {results_imported} výsledků""", "ok")

            return redirect(url_for("admin_import"))

        except Exception as e:
            db.session.rollback()

            # Smaž temp files i při chybě
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass

            if preview_file and os.path.exists(preview_file):
                try:
                    os.unlink(preview_file)
                except:
                    pass

            flash(f"Chyba při importu: {str(e)}", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

    def _get_or_create_team(round_id: int, name: str) -> Team:
        import unicodedata

        name = (name or "").strip()
        if not name:
            raise ValueError("empty team")

        # Normalize Unicode (pro české znaky Š, Č, Ř, ...)
        name = unicodedata.normalize('NFC', name)

        # Flush změny aby jsme viděli všechny týmy včetně nově vytvořených
        db.session.flush()

        # PYTHON-based case-insensitive search (NEJSPOLEHLIVĚJŠÍ!)
        # Načti všechny týmy pro tento round
        all_teams = Team.query.filter_by(
            round_id=round_id,
            is_deleted=False
        ).all()

        # Najdi case-insensitive match s Unicode normalizací
        name_lower = name.lower()
        for t in all_teams:
            t_normalized = unicodedata.normalize('NFC', t.name)
            if t_normalized.lower() == name_lower:
                return t  # Našli jsme!

        # Tým neexistuje - vytvoř nový
        t = Team(round_id=round_id, name=name)
        db.session.add(t)
        db.session.flush()  # Flush aby byl dostupný pro další volání

        audit("team.create.auto", "Team", t.id, round_id=round_id, name=name)
        return t



    @admin_bp.route("/import/teams/template")
    @login_required
    def admin_import_teams_template():
        """Stáhni Excel šablonu pro import týmů"""
        admin_required()
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Týmy"

        # Header
        headers = ["name", "group", "country_code"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Příklady
        examples = [
            ("Sparta Praha", "A", "CZ"),
            ("Slavia Praha", "A", "CZ"),
            ("Barcelona", "B", "ES"),
            ("Real Madrid", "B", "ES"),
        ]

        gray_fill = PatternFill("solid", fgColor="F2F2F2")
        for row_i, (name, grp, code) in enumerate(examples, 2):
            fill = gray_fill if row_i % 2 == 0 else PatternFill()
            for col, val in enumerate([name, grp, code], 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.fill = fill

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15

        # Popis
        ws2 = wb.create_sheet("Popis")
        ws2["A1"], ws2["B1"] = "Sloupec", "Popis"
        ws2["A1"].font = ws2["B1"].font = Font(bold=True)

        popis = [
            ("name", "POVINNÉ - název týmu"),
            ("group", "VOLITELNÉ - skupina (A, B, C...)"),
            ("country_code", "VOLITELNÉ - kód země (CZ, SK, DE...)"),
        ]
        for row_i, (col, desc) in enumerate(popis, 2):
            ws2[f"A{row_i}"] = col
            ws2[f"B{row_i}"] = desc
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 40

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        from flask import send_file
        return send_file(out,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="sablona_tymy.xlsx",
                         as_attachment=True)



    @admin_bp.route("/import/teams", methods=["GET", "POST"])
    @login_required
    def admin_import_teams():
        admin_required()
        import os, pickle
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            abort(400)

        if request.method == "POST":
            action = request.form.get("action", "upload")

            # ── KROK 2: POTVRZENÍ ──────────────────────────────────────────
            if action == "confirm":
                preview_file = session.get("teams_import_preview_file")
                round_id = session.get("teams_import_round_id")
                overwrite = session.get("teams_import_overwrite", False)

                if not preview_file or not os.path.exists(preview_file):
                    flash("Session vypršela, nahraj soubor znovu.", "error")
                    return redirect(url_for("admin_import_teams"))

                r2 = db.session.get(Round, round_id) if round_id else r

                with open(preview_file, "rb") as f:
                    rows_data = pickle.load(f)
                os.unlink(preview_file)
                session.pop("teams_import_preview_file", None)
                session.pop("teams_import_round_id", None)
                session.pop("teams_import_overwrite", None)

                created = skipped = updated = 0
                for row in rows_data:
                    if row["status"] == "skip":
                        skipped += 1
                        continue

                    name = row["name"]
                    existing = Team.query.filter_by(
                        round_id=r2.id, is_deleted=False
                    ).filter(db.func.lower(Team.name) == name.lower()).first()

                    if existing:
                        if overwrite:
                            # Aktualizuj group/country_code pokud je v importu
                            if row.get("group"):
                                existing.group = row["group"]
                            if row.get("country_code"):
                                existing.country_code = row["country_code"]
                            db.session.flush()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        team = Team(
                            round_id=r2.id,
                            name=name,
                            group=row.get("group"),
                            country_code=row.get("country_code")
                        )
                        db.session.add(team)
                        db.session.flush()
                        created += 1

                db.session.commit()
                audit("import.teams.excel", "Round", r2.id, created=created, updated=updated, skipped=skipped)
                flash(f"Import týmů hotov: vytvořeno {created}, aktualizováno {updated}, přeskočeno {skipped}.", "ok")
                return redirect(url_for("teams"))

            # ── KROK 1: UPLOAD + PREVIEW ───────────────────────────────────
            import openpyxl, tempfile
            from io import BytesIO

            file = request.files.get("excel_file")
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_teams"))

            overwrite = request.form.get("overwrite") == "1"

            try:
                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))

                if len(rows) < 2:
                    flash("Soubor musí mít hlavičku a alespoň jeden řádek.", "error")
                    return redirect(url_for("admin_import_teams"))

                header = [str(h).strip().lower() if h else "" for h in rows[0]]

                # Najdi sloupce
                def col(name):
                    return header.index(name) if name in header else None

                name_col = col("name")
                if name_col is None:
                    flash("Chybí sloupec 'name'. Zkontroluj šablonu.", "error")
                    return redirect(url_for("admin_import_teams"))

                grp_col = col("group")
                code_col = col("country_code")

                # Existující týmy
                existing_teams = Team.query.filter_by(
                    round_id=r.id, is_deleted=False
                ).all()
                existing_lower = {t.name.lower(): t for t in existing_teams}

                preview_rows = []
                for data_row in rows[1:]:
                    name = str(data_row[name_col]).strip() if data_row[name_col] else ""
                    if not name or name.lower() == "none":
                        continue

                    group = str(data_row[grp_col]).strip() if grp_col is not None and data_row[grp_col] else None
                    country_code = str(data_row[code_col]).strip() if code_col is not None and data_row[code_col] else None

                    # Status
                    if name.lower() in existing_lower:
                        status = "overwrite" if overwrite else "skip"
                    else:
                        status = "new"

                    preview_rows.append({
                        "name": name,
                        "group": group,
                        "country_code": country_code,
                        "status": status,
                    })

                if not preview_rows:
                    flash("Soubor neobsahuje žádné platné týmy.", "error")
                    return redirect(url_for("admin_import_teams"))

                # Ulož do temp file
                fd, preview_path = tempfile.mkstemp(suffix=".pkl", prefix="teams_import_")
                with os.fdopen(fd, "wb") as f:
                    pickle.dump(preview_rows, f)

                session["teams_import_preview_file"] = preview_path
                session["teams_import_round_id"] = r.id
                session["teams_import_overwrite"] = overwrite

                new_count = sum(1 for x in preview_rows if x["status"] == "new")
                overwrite_count = sum(1 for x in preview_rows if x["status"] == "overwrite")
                skip_count = sum(1 for x in preview_rows if x["status"] == "skip")

                return render_page(r"""
<style>
.preview-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.preview-table th { background: rgba(255,255,255,.07); padding: 10px 12px; text-align: left; }
.preview-table td { padding: 10px 12px; border-bottom: 1px solid var(--line); }
.badge-new    { background: rgba(51,209,122,.2);  color: #33d17a; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }
.badge-skip   { background: rgba(167,178,214,.15); color: #a7b2d6; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }
.badge-over   { background: rgba(110,168,254,.2);  color: #6ea8fe; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }
</style>

<div class="card">
  <h2 style="margin:0 0 4px 0;">📋 Preview – Týmy</h2>
  <div class="muted">Soutěž: <b>{{ r.name }}</b></div>

  <div class="row" style="gap:16px; margin: 16px 0; flex-wrap:wrap;">
    <div style="padding:12px 20px; background:rgba(51,209,122,.1); border:1px solid rgba(51,209,122,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#33d17a;">{{ new_count }}</div>
      <div class="muted" style="font-size:12px;">Nových</div>
    </div>
    <div style="padding:12px 20px; background:rgba(110,168,254,.1); border:1px solid rgba(110,168,254,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#6ea8fe;">{{ overwrite_count }}</div>
      <div class="muted" style="font-size:12px;">Aktualizovaných</div>
    </div>
    <div style="padding:12px 20px; background:rgba(167,178,214,.1); border:1px solid rgba(167,178,214,.2); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#a7b2d6;">{{ skip_count }}</div>
      <div class="muted" style="font-size:12px;">Přeskočených</div>
    </div>
  </div>

  <div style="overflow-x:auto; margin-bottom:20px;">
    <table class="preview-table">
      <thead>
        <tr>
          <th>#</th>
          <th>Název týmu</th>
          <th>Skupina</th>
          <th>Země</th>
          <th>Status</th>
        </tr>
      </thead>
      <tbody>
        {% for row in preview_rows %}
        <tr>
          <td class="muted">{{ loop.index }}</td>
          <td>{{ row.name }}</td>
          <td class="muted">{{ row.group or '—' }}</td>
          <td class="muted">{{ row.country_code or '—' }}</td>
          <td>
            {% if row.status == 'new' %}
              <span class="badge-new">✨ Nový</span>
            {% elif row.status == 'overwrite' %}
              <span class="badge-over">✏️ Aktualizovat</span>
            {% else %}
              <span class="badge-skip">⏭ Přeskočit</span>
            {% endif %}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <form method="post" style="display:flex; gap:12px; flex-wrap:wrap;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="action" value="confirm">
    <button type="submit" class="btn btn-primary">✅ Potvrdit import</button>
    <a href="{{ url_for('admin_import_teams') }}" class="btn">✕ Zrušit</a>
  </form>
</div>
""", r=r, preview_rows=preview_rows,
     new_count=new_count, overwrite_count=overwrite_count, skip_count=skip_count)

            except Exception as e:
                flash(f"Chyba při čtení souboru: {e}", "error")
                return redirect(url_for("admin_import_teams"))

        # === GET: Zobraz formulář ===
        return render_page(r"""
<style>
.upload-zone {
  border: 2px dashed rgba(110,168,254,0.4);
  border-radius: 12px;
  padding: 40px;
  text-align: center;
  background: rgba(110,168,254,0.05);
  cursor: pointer;
  transition: all 0.2s;
}
.upload-zone:hover {
  border-color: rgba(110,168,254,0.7);
  background: rgba(110,168,254,0.1);
}
</style>

<div class="card">
  <div class="row" style="justify-content:space-between; margin-bottom:20px;">
    <div>
      <h2 style="margin:0 0 4px 0;">📥 Import týmů (Excel)</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <a href="{{ url_for('admin_import_teams_template') }}" class="btn"
       style="background:rgba(51,209,122,.15); color:#33d17a; border:1px solid rgba(51,209,122,.3);">
      📄 Stáhnout šablonu
    </a>
  </div>

  <div class="muted" style="margin-bottom:20px; padding:12px 16px; background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); border-radius:8px;">
    💡 <strong>Formát Excel (sloupce):</strong><br>
    <code>name</code> – povinný, název týmu<br>
    <code>group</code> – volitelný, skupina (A, B, C...)<br>
    <code>country_code</code> – volitelný, kód země (CZ, SK, DE...)
  </div>

  <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="upload-zone" onclick="document.getElementById('xl').click()">
      <div style="font-size:48px; margin-bottom:12px;">📊</div>
      <div style="font-weight:700; margin-bottom:8px;">Klikni nebo přetáhni Excel soubor</div>
      <div class="muted">.xlsx • max 10 MB</div>
      <input type="file" id="xl" name="excel_file" accept=".xlsx" style="display:none"
             onchange="document.getElementById('fname').textContent = this.files[0]?.name || ''">
    </div>
    <div id="fname" class="muted" style="margin-bottom:16px; text-align:center;"></div>

    <div style="padding:14px; background:rgba(110,168,254,.08); border:1px solid rgba(110,168,254,.2); border-radius:8px; margin-bottom:20px;">
      <label style="display:flex; align-items:center; gap:10px; cursor:pointer;">
        <input type="checkbox" name="overwrite" value="1" style="width:18px; height:18px;">
        <div>
          <strong>Přepsat existující týmy</strong>
          <div class="muted" style="font-size:12px;">Aktualizuje skupinu a kód země. Bez zaškrtnutí se přeskočí.</div>
        </div>
      </label>
    </div>

    <div class="row" style="gap:12px;">
      <button type="submit" class="btn btn-primary">📋 Zobrazit preview</button>
      <a href="{{ url_for('admin_import') }}" class="btn">← Zpět</a>
    </div>
  </form>
</div>
""", r=r)



    @admin_bp.route("/import/matches/template")
    @login_required
    def admin_import_matches_template():
        """Stáhne Excel šablonu pro import zápasů"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Zápasy"

            # Hlavička
            headers = ['Datum', 'Domácí', 'Hosté', 'Čas']
            ws.append(headers)

            # Stylování
            for col in range(1, 5):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="6EA8FE", end_color="6EA8FE", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Příklad
            ws.append(['2024-03-15', 'Sparta Praha', 'Slavia Praha', '18:00'])
            ws.append(['2024-03-16', 'Plzeň', 'Brno', '16:30'])

            # Šířka sloupců
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10

            # Poznámky
            notes = wb.create_sheet("Poznámky")
            notes['A1'] = "FORMÁT ZÁPASŮ"
            notes['A1'].font = Font(bold=True, size=14)
            notes['A3'] = "Datum: YYYY-MM-DD (např. 2024-03-15)"
            notes['A4'] = "Domácí: Název týmu"
            notes['A5'] = "Hosté: Název týmu"
            notes['A6'] = "Čas: HH:MM (např. 18:00)"
            notes.column_dimensions['A'].width = 50

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name='import_zapasu_sablona.xlsx')
        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_import"))



    @admin_bp.route("/import/matches", methods=["GET", "POST"])
    @login_required
    def admin_import_matches():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            abort(400)

        if request.method == "POST":
            # Kontrola, jestli je CSV nebo Excel
            if 'excel_file' in request.files and request.files['excel_file'].filename:
                # Import z Excelu
                file = request.files['excel_file']
                try:
                    import openpyxl
                    from io import BytesIO

                    wb = openpyxl.load_workbook(BytesIO(file.read()))
                    ws = wb.active

                    created = 0
                    skipped = 0

                    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                        if not row or not any(row):
                            continue

                        # Formát: Datum, Domácí, Hosté, Čas, Liga
                        date_val = row[0] if len(row) > 0 else None
                        home = (str(row[1]) if len(row) > 1 and row[1] else "").strip()
                        away = (str(row[2]) if len(row) > 2 and row[2] else "").strip()
                        time_val = row[3] if len(row) > 3 else None

                        if not home or not away or home == away:
                            skipped += 1
                            continue

                        # Parsování datumu a času
                        start = None
                        if date_val and time_val:
                            try:
                                from datetime import datetime as dt_mod
                                if isinstance(date_val, str):
                                    date_str = date_val
                                else:
                                    date_str = date_val.strftime("%Y-%m-%d")

                                if isinstance(time_val, str):
                                    time_str = time_val
                                else:
                                    time_str = time_val.strftime("%H:%M")

                                start = dt_mod.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                            except:
                                pass

                        ht = _get_or_create_team(r.id, home)
                        at = _get_or_create_team(r.id, away)

                        existing = Match.query.filter_by(
                            round_id=r.id,
                            home_team_id=ht.id,
                            away_team_id=at.id,
                            is_deleted=False
                        ).first()

                        if existing:
                            skipped += 1
                            continue

                        m = Match(round_id=r.id, home_team_id=ht.id, away_team_id=at.id, start_time=start)
                        db.session.add(m)
                        created += 1

                    db.session.commit()
                    audit("import.matches.excel", "Round", r.id, created=created, skipped=skipped)
                    flash(f"Import z Excelu hotov: vytvořeno {created}, přeskočeno {skipped}.", "ok")
                    return redirect(url_for("matches"))

                except Exception as e:
                    flash(f"Chyba při importu z Excelu: {str(e)}", "error")
                    return redirect(url_for("admin_import_matches"))
            else:
                # CSV import (původní kód)
                csv_text = (request.form.get("csv_text") or "").strip()
                if not csv_text:
                    flash("Vlož CSV nebo nahraj Excel soubor.", "error")
                    return redirect(url_for("admin_import_matches"))
                reader = csv.DictReader(io.StringIO(csv_text))
                created = 0
                skipped = 0
                for row in reader:
                    rrid = _resolve_round_id(row, r.id)
                    home = (row.get("home_team") or row.get("home") or "").strip()
                    away = (row.get("away_team") or row.get("away") or "").strip()
                    start = parse_naive_datetime((row.get("start_time") or row.get("start") or "").strip())
                    if not home or not away or home == away:
                        skipped += 1
                        continue
                    ht = _get_or_create_team(rrid, home)
                    at = _get_or_create_team(rrid, away)
                    existing = Match.query.filter_by(round_id=rrid, home_team_id=ht.id, away_team_id=at.id, start_time=start, is_deleted=False).first()
                    if existing:
                        skipped += 1
                        continue
                    m = Match(round_id=rrid, home_team_id=ht.id, away_team_id=at.id, start_time=start)
                    hs = (row.get("home_score") or "").strip()
                    aas = (row.get("away_score") or "").strip()
                    if hs != "":
                        m.home_score = int(hs)
                    if aas != "":
                        m.away_score = int(aas)
                    db.session.add(m)
                    created += 1
                db.session.commit()
                audit("import.matches", "Round", r.id, created=created, skipped=skipped)
                flash(f"Import zápasů hotov: vytvořeno {created}, přeskočeno {skipped}.", "ok")
                return redirect(url_for("matches"))

        return render_page(r"""
<div class="card">
  <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <div>
      <h2 style="margin:0 0 8px 0;">Import zápasů</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <a href="{{ url_for('admin_import_matches_template') }}" class="btn" style="background:#6ea8fe; color:white;">
      📥 Stáhnout šablonu Excel
    </a>
  </div>
  <hr class="sep">

  <h3 style="margin:20px 0 10px 0;">Import z Excel souboru</h3>
  <div class="muted" style="margin-bottom:10px;">
    Formát: 1. sloupec = Datum, 2. = Domácí, 3. = Hosté, 4. = Čas, 5. = Liga (volitelné)<br>
    První řádek je hlavička (přeskočí se)
  </div>
  <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="file" name="excel_file" accept=".xlsx,.xls" required>
    <button class="btn btn-primary" type="submit">Importovat z Excelu</button>
  </form>

  <hr class="sep" style="margin:20px 0;">

  <h3 style="margin:20px 0 10px 0;">Import z CSV (klasický)</h3>
  <div class="muted">Sloupce: <b>round_id</b> (volitelné), <b>home_team</b>, <b>away_team</b>, <b>start_time</b> (YYYY-MM-DD HH:MM), volitelně <b>home_score</b>, <b>away_score</b></div>
  <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <textarea name="csv_text" placeholder="round_id,home_team,away_team,start_time,home_score,away_score&#10;{{ r.id }},Sparta,Slavia,2026-02-10 18:00,,"></textarea>
    <button class="btn btn-primary" type="submit">Importovat CSV</button>
  </form>

  <hr class="sep">
  <a class="btn" href="{{ url_for('admin_import') }}">Zpět</a>
</div>
""", r=r)



    @admin_bp.route("/import/extras/template")
    @login_required
    def admin_import_extras_template():
        """Stáhni Excel šablonu pro import extra otázek"""
        admin_required()
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extra otázky"

        # Header - user + příklady otázek
        headers = ["user", "Vítězný tým", "Kanadaské bodování", "Střelec turnaje"]
        header_fill_user = PatternFill("solid", fgColor="2F4F8F")
        header_fill_q    = PatternFill("solid", fgColor="4472C4")
        header_font      = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill_user if col == 1 else header_fill_q
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Příklady
        examples = [
            ("Václav",   "Kanada",  "MacKinnon", "McDavid"),
            ("Mejla",    "Kanada",  "McDavid",   "McDavid"),
            ("Ondejsek", "Švédsko", "McDavid",   "MacKinnon"),
        ]
        gray_fill = PatternFill("solid", fgColor="F2F2F2")
        for row_i, row_data in enumerate(examples, 2):
            fill = gray_fill if row_i % 2 == 0 else PatternFill()
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.fill = fill

        ws.column_dimensions["A"].width = 18
        for col_letter in ["B", "C", "D"]:
            ws.column_dimensions[col_letter].width = 28

        # Sheet 2 - popis
        ws2 = wb.create_sheet("Popis")
        ws2["A1"], ws2["B1"] = "Sloupec", "Popis"
        ws2["A1"].font = ws2["B1"].font = Font(bold=True)
        ws2["A2"] = "user"
        ws2["B2"] = "POVINNÉ - přesný username uživatele (case-insensitive)"
        ws2["A3"] = "Název otázky (2. sloupec a dál)"
        ws2["B3"] = "Každý sloupec = jedna extra otázka. Buňka = odpověď uživatele."
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 55

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        from flask import send_file
        return send_file(out,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="sablona_extra_otazky.xlsx",
                         as_attachment=True)



    @admin_bp.route("/import/extras", methods=["GET", "POST"])
    @login_required
    def admin_import_extras():
        admin_required()
        import os, pickle
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            abort(400)

        if request.method == "POST":
            action = request.form.get("action", "upload")

            # ── KROK 2: POTVRZENÍ ──────────────────────────────────────────
            if action == "confirm":

                preview_file = session.get("extra_import_preview_file")
                round_id     = session.get("extra_import_round_id")
                overwrite    = session.get("extra_import_overwrite", False)

                if not preview_file or not os.path.exists(preview_file):
                    flash("Session vypršela, nahraj soubor znovu.", "error")
                    return redirect(url_for("admin_import_extras"))

                r2 = db.session.get(Round, round_id) if round_id else r

                with open(preview_file, "rb") as f:
                    preview_data = pickle.load(f)
                os.unlink(preview_file)
                session.pop("extra_import_preview_file", None)
                session.pop("extra_import_round_id", None)
                session.pop("extra_import_overwrite", None)

                questions_data = preview_data["questions"]   # [{text, q_obj_or_None}]
                rows_data      = preview_data["rows"]        # [{user, answers:[{q_idx, answer, status}]}]

                # 1) Vytvoř/najdi otázky
                q_objects = []
                for qd in questions_data:
                    q_text = qd["text"]
                    existing_q = ExtraQuestion.query.filter_by(
                        round_id=r2.id, is_deleted=False
                    ).filter(db.func.lower(ExtraQuestion.question) == q_text.lower()).first()
                    if not existing_q:
                        existing_q = ExtraQuestion(round_id=r2.id, question=q_text)
                        db.session.add(existing_q)
                        db.session.flush()
                    q_objects.append(existing_q)

                # 2) Zpracuj odpovědi
                created = skipped = updated = 0
                for row in rows_data:
                    user = row["user"]
                    if user is None:
                        skipped += len(row["answers"])
                        continue
                    for ans_data in row["answers"]:
                        if ans_data["status"] == "skip":
                            skipped += 1
                            continue
                        q_obj   = q_objects[ans_data["q_idx"]]
                        ans_txt = ans_data["answer"]
                        existing_ans = ExtraAnswer.query.filter_by(
                            question_id=q_obj.id, user_id=user.id
                        ).first()
                        if existing_ans:
                            if overwrite:
                                existing_ans.answer_text = ans_txt
                                updated += 1
                            else:
                                skipped += 1
                        else:
                            db.session.add(ExtraAnswer(
                                question_id=q_obj.id,
                                user_id=user.id,
                                answer_text=ans_txt
                            ))
                            created += 1

                db.session.commit()
                audit("import.extras.excel", "Round", r2.id,
                      created=created, updated=updated, skipped=skipped)
                flash(f"Import extra otázek hotov: vytvořeno {created}, "
                      f"aktualizováno {updated}, přeskočeno {skipped}.", "ok")
                return redirect(url_for("extras"))

            # ── KROK 1: UPLOAD + PREVIEW ───────────────────────────────────
            import openpyxl, tempfile
            from io import BytesIO

            file = request.files.get("excel_file")
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_extras"))

            overwrite = request.form.get("overwrite") == "1"

            try:
                wb   = openpyxl.load_workbook(BytesIO(file.read()))
                ws   = wb.active
                rows = list(ws.iter_rows(values_only=True))

                if len(rows) < 2:
                    flash("Soubor musí mít hlavičku a alespoň jeden řádek.", "error")
                    return redirect(url_for("admin_import_extras"))

                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce (user + otázka).", "error")
                    return redirect(url_for("admin_import_extras"))

                # Otázky = sloupce 2..N
                question_texts = []
                for h in header[1:]:
                    if h and str(h).strip():
                        question_texts.append(str(h).strip())

                if not question_texts:
                    flash("Nenašel jsem žádné otázky v hlavičce.", "error")
                    return redirect(url_for("admin_import_extras"))

                # Existující otázky a odpovědi
                existing_qs = {
                    eq.question.lower(): eq
                    for eq in ExtraQuestion.query.filter_by(round_id=r.id, is_deleted=False).all()
                }
                existing_answers = {}  # (question_id, user_id) → ExtraAnswer
                for ans in ExtraAnswer.query.join(ExtraQuestion).filter(
                    ExtraQuestion.round_id == r.id,
                    ExtraQuestion.is_deleted == False
                ).all():
                    existing_answers[(ans.question_id, ans.user_id)] = ans

                # Všichni uživatelé (case-insensitive lookup)
                import unicodedata
                all_users = User.query.all()
                user_map  = {}
                for u in all_users:
                    norm = unicodedata.normalize("NFC", u.username).lower()
                    user_map[norm] = u

                # Info o otázkách pro preview
                questions_preview = []
                for qt in question_texts:
                    is_new = qt.lower() not in existing_qs
                    questions_preview.append({"text": qt, "is_new": is_new})

                # Zpracuj řádky
                preview_rows = []
                for data_row in rows[1:]:
                    username_raw = str(data_row[0]).strip() if data_row[0] else ""
                    if not username_raw or username_raw.lower() == "none":
                        continue

                    norm_name = unicodedata.normalize("NFC", username_raw).lower()
                    user_obj  = user_map.get(norm_name)

                    answers = []
                    for q_idx, qt in enumerate(question_texts):
                        col_idx = q_idx + 1
                        ans_val = str(data_row[col_idx]).strip() if (col_idx < len(data_row) and data_row[col_idx]) else ""
                        if not ans_val or ans_val.lower() == "none":
                            status = "skip"
                        elif user_obj is None:
                            status = "unknown_user"
                        else:
                            existing_q = existing_qs.get(qt.lower())
                            if existing_q and (existing_q.id, user_obj.id) in existing_answers:
                                status = "overwrite" if overwrite else "skip"
                            else:
                                status = "new"

                        answers.append({
                            "q_idx":  q_idx,
                            "answer": ans_val,
                            "status": status,
                        })

                    preview_rows.append({
                        "username_raw": username_raw,
                        "user":         user_obj,
                        "answers":      answers,
                    })

                if not preview_rows:
                    flash("Soubor neobsahuje žádné platné řádky.", "error")
                    return redirect(url_for("admin_import_extras"))

                # Ulož do temp souboru
                fd, preview_path = tempfile.mkstemp(suffix=".pkl", prefix="extra_import_")
                with os.fdopen(fd, "wb") as f:
                    pickle.dump({
                        "questions": [{"text": qt} for qt in question_texts],
                        "rows":      preview_rows,
                    }, f)

                session["extra_import_preview_file"] = preview_path
                session["extra_import_round_id"]     = r.id
                session["extra_import_overwrite"]    = overwrite

                # Počty pro preview
                new_cnt  = sum(1 for row in preview_rows for a in row["answers"] if a["status"] == "new")
                over_cnt = sum(1 for row in preview_rows for a in row["answers"] if a["status"] == "overwrite")
                skip_cnt = sum(1 for row in preview_rows for a in row["answers"] if a["status"] in ("skip", "unknown_user"))
                unk_users = [row["username_raw"] for row in preview_rows if row["user"] is None]

                return render_page(r"""
<style>
.prev-table { width:100%; border-collapse:collapse; font-size:13px; }
.prev-table th { background:rgba(255,255,255,.07); padding:10px 12px; text-align:left; white-space:nowrap; }
.prev-table td { padding:10px 12px; border-bottom:1px solid var(--line); vertical-align:middle; }
.badge-new  { background:rgba(51,209,122,.2);  color:#33d17a; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
.badge-skip { background:rgba(167,178,214,.15);color:#a7b2d6; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
.badge-over { background:rgba(110,168,254,.2); color:#6ea8fe; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
.badge-unk  { background:rgba(255,77,109,.2);  color:#ff4d6d; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
.q-header   { writing-mode:vertical-rl; transform:rotate(180deg); white-space:nowrap; padding:8px 4px; max-height:120px; font-size:12px; }
</style>

<div class="card">
  <h2 style="margin:0 0 4px 0;">📋 Preview – Extra otázky</h2>
  <div class="muted">Soutěž: <b>{{ r.name }}</b></div>

  {% if unk_users %}
  <div style="margin:16px 0; padding:12px 16px; background:rgba(255,77,109,.1); border:1px solid rgba(255,77,109,.3); border-radius:8px;">
    ⚠️ <strong>Nenalezení uživatelé</strong> (budou přeskočeni):
    {% for u in unk_users %}<code style="margin-left:8px;">{{ u }}</code>{% endfor %}
  </div>
  {% endif %}

  <div class="row" style="gap:16px; margin:16px 0; flex-wrap:wrap;">
    <div style="padding:12px 20px; background:rgba(51,209,122,.1); border:1px solid rgba(51,209,122,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#33d17a;">{{ new_cnt }}</div>
      <div class="muted" style="font-size:12px;">Nových odpovědí</div>
    </div>
    <div style="padding:12px 20px; background:rgba(110,168,254,.1); border:1px solid rgba(110,168,254,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#6ea8fe;">{{ over_cnt }}</div>
      <div class="muted" style="font-size:12px;">Aktualizovaných</div>
    </div>
    <div style="padding:12px 20px; background:rgba(167,178,214,.1); border:1px solid rgba(167,178,214,.2); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#a7b2d6;">{{ skip_cnt }}</div>
      <div class="muted" style="font-size:12px;">Přeskočených</div>
    </div>
  </div>

  <div style="overflow-x:auto; margin-bottom:20px;">
    <table class="prev-table">
      <thead>
        <tr>
          <th>User</th>
          {% for q in questions_preview %}
            <th style="text-align:center;">
              <div class="q-header">
                {{ q.text }}
                {% if q.is_new %}<span style="color:#33d17a; font-size:10px;"> ✨</span>{% endif %}
              </div>
            </th>
          {% endfor %}
        </tr>
      </thead>
      <tbody>
        {% for row in preview_rows %}
        <tr>
          <td>
            {% if row.user %}
              <strong>{{ row.user.display_name }}</strong>
            {% else %}
              <span class="badge-unk">❓ {{ row.username_raw }}</span>
            {% endif %}
          </td>
          {% for ans in row.answers %}
          <td style="text-align:center;">
            {% if ans.status == 'new' %}
              <div style="font-size:12px; margin-bottom:3px;">{{ ans.answer }}</div>
              <span class="badge-new">✨</span>
            {% elif ans.status == 'overwrite' %}
              <div style="font-size:12px; margin-bottom:3px;">{{ ans.answer }}</div>
              <span class="badge-over">✏️</span>
            {% elif ans.status == 'unknown_user' %}
              <span class="badge-unk">—</span>
            {% else %}
              <span class="badge-skip">—</span>
            {% endif %}
          </td>
          {% endfor %}
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <form method="post" style="display:flex; gap:12px; flex-wrap:wrap;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="action" value="confirm">
    <button type="submit" class="btn btn-primary">✅ Potvrdit import</button>
    <a href="{{ url_for('admin_import_extras') }}" class="btn">✕ Zrušit</a>
  </form>
</div>
""", r=r,
     preview_rows=preview_rows,
     questions_preview=questions_preview,
     new_cnt=new_cnt, over_cnt=over_cnt, skip_cnt=skip_cnt,
     unk_users=unk_users)

            except Exception as e:
                import traceback
                flash(f"Chyba při čtení souboru: {e}", "error")
                return redirect(url_for("admin_import_extras"))

        # ── GET: formulář ──────────────────────────────────────────────────
        return render_page(r"""
<style>
.upload-zone {
  border:2px dashed rgba(110,168,254,0.4); border-radius:12px; padding:40px;
  text-align:center; background:rgba(110,168,254,0.05); cursor:pointer; transition:all .2s;
}
.upload-zone:hover { border-color:rgba(110,168,254,0.7); background:rgba(110,168,254,0.1); }
</style>

<div class="card">
  <div class="row" style="justify-content:space-between; margin-bottom:20px;">
    <div>
      <h2 style="margin:0 0 4px 0;">📥 Import extra otázek (Excel)</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <a href="{{ url_for('admin_import_extras_template') }}" class="btn"
       style="background:rgba(51,209,122,.15); color:#33d17a; border:1px solid rgba(51,209,122,.3);">
      📄 Stáhnout šablonu
    </a>
  </div>

  <div style="margin-bottom:20px; padding:14px 16px; background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); border-radius:8px; font-size:13px;">
    💡 <strong>Formát Excel:</strong><br><br>
    <table style="border-collapse:collapse; font-size:12px;">
      <thead>
        <tr style="background:rgba(255,255,255,.07);">
          <th style="padding:6px 12px; border:1px solid var(--line);">user</th>
          <th style="padding:6px 12px; border:1px solid var(--line);">Vítězný tým</th>
          <th style="padding:6px 12px; border:1px solid var(--line);">Kanadaské bodování</th>
          <th style="padding:6px 12px; border:1px solid var(--line);">Střelec</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:6px 12px; border:1px solid var(--line);">Václav</td><td style="padding:6px 12px; border:1px solid var(--line);">Kanada</td><td style="padding:6px 12px; border:1px solid var(--line);">MacKinnon</td><td style="padding:6px 12px; border:1px solid var(--line);">McDavid</td></tr>
        <tr><td style="padding:6px 12px; border:1px solid var(--line);">Mejla</td><td style="padding:6px 12px; border:1px solid var(--line);">Kanada</td><td style="padding:6px 12px; border:1px solid var(--line);">McDavid</td><td style="padding:6px 12px; border:1px solid var(--line);">McDavid</td></tr>
      </tbody>
    </table>
    <div class="muted" style="margin-top:10px;">
      • 1. sloupec = <strong>user</strong> (username uživatele)<br>
      • Další sloupce = názvy extra otázek<br>
      • Buňky = odpovědi uživatelů<br>
      • Nové otázky se vytvoří automaticky ✨
    </div>
  </div>

  <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="upload-zone" onclick="document.getElementById('xl').click()">
      <div style="font-size:48px; margin-bottom:12px;">📊</div>
      <div style="font-weight:700; margin-bottom:8px;">Klikni nebo přetáhni Excel soubor</div>
      <div class="muted">.xlsx • max 10 MB</div>
      <input type="file" id="xl" name="excel_file" accept=".xlsx" style="display:none"
             onchange="document.getElementById('fname').textContent = this.files[0]?.name || ''">
    </div>
    <div id="fname" class="muted" style="margin:8px 0 16px 0; text-align:center;"></div>

    <div style="padding:14px; background:rgba(110,168,254,.08); border:1px solid rgba(110,168,254,.2); border-radius:8px; margin-bottom:20px;">
      <label style="display:flex; align-items:center; gap:10px; cursor:pointer;">
        <input type="checkbox" name="overwrite" value="1" style="width:18px; height:18px;">
        <div>
          <strong>Přepsat existující odpovědi</strong>
          <div class="muted" style="font-size:12px;">Bez zaškrtnutí se existující odpovědi přeskočí.</div>
        </div>
      </label>
    </div>

    <div class="row" style="gap:12px;">
      <button type="submit" class="btn btn-primary">📋 Zobrazit preview</button>
      <a href="{{ url_for('admin_import') }}" class="btn">← Zpět</a>
    </div>
  </form>
</div>
""", r=r)



    @admin_bp.route("/import/leaderboard/template")
    @login_required
    def admin_import_leaderboard_template():
        """Stáhne Excel šablonu pro import žebříčku"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Žebříček"

            # Hlavička - Jméno | Zápasy (formát: "Domácí-Hosté")
            headers = ['Jméno', 'Sparta-Slavia', 'Plzeň-Brno', 'Baník-Bohemians']
            ws.append(headers)

            # Stylování
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="6EA8FE", end_color="6EA8FE", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Příklady
            ws.append(['Jan Novák', '2:1', '1:1', '3:0'])
            ws.append(['Petr Svoboda', '1:2', '2:0', '1:1'])

            # Šířka sloupců
            ws.column_dimensions['A'].width = 20
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 15

            # Poznámky
            notes = wb.create_sheet("Poznámky")
            notes['A1'] = "FORMÁT ŽEBŘÍČKU"
            notes['A1'].font = Font(bold=True, size=14)
            notes['A3'] = "1. řádek = Hlavička"
            notes['A4'] = "   - První sloupec: 'Jméno'"
            notes['A5'] = "   - Další sloupce: názvy zápasů (formát: 'Domácí-Hosté')"
            notes['A7'] = "2+ řádky = Tipy uživatelů"
            notes['A8'] = "   - První sloupec: Jméno uživatele"
            notes['A9'] = "   - Další sloupce: tipy (formát: '2:1' nebo '1:0')"
            notes.column_dimensions['A'].width = 50

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name='import_zebricku_sablona.xlsx')
        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_import"))



    @admin_bp.route("/import/leaderboard", methods=["GET", "POST"])
    @login_required
    def admin_import_leaderboard():
        admin_required()

        if request.method == "POST":
            file = request.files.get('excel_file')
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_leaderboard"))

            # Určit cílovou soutěž
            import_target = request.form.get('import_target', 'existing')

            if import_target == 'new':
                new_round_name = request.form.get('new_round_name', '').strip()
                if not new_round_name:
                    flash("Zadej název nové soutěže.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                sport = Sport.query.filter_by(name="Fotbal").first()
                if not sport:
                    sport = Sport(name="Fotbal")
                    db.session.add(sport)
                    db.session.flush()

                r = Round(name=new_round_name, sport_id=sport.id, is_active=False)
                db.session.add(r)
                db.session.flush()
                round_id = r.id
            else:
                round_id = int(request.form.get('round_id', 0))
                r = db.session.get(Round, round_id)
                if not r:
                    flash("Vybraná soutěž neexistuje.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

            try:
                import openpyxl
                from io import BytesIO
                from datetime import datetime as dt_parse

                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active

                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    flash("Soubor musí mít alespoň hlavičku a jeden řádek.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Parsovat zápasy z hlavičky
                matches_data = []
                for i, match_name in enumerate(header[1:], start=1):
                    if not match_name or str(match_name).strip() == "":
                        continue

                    match_str = str(match_name).strip()
                    for sep in ['-', ' vs ', ' x ', ':', ' – ']:
                        if sep in match_str:
                            parts = match_str.split(sep, 1)
                            if len(parts) == 2:
                                home = parts[0].strip()
                                away = parts[1].strip()

                                # Detekovat možnou duplicitu
                                home_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == home.lower()
                                ).first()

                                away_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == away.lower()
                                ).first()

                                is_duplicate = False
                                if home_team_check and away_team_check:
                                    existing_match = Match.query.filter_by(
                                        round_id=round_id,
                                        home_team_id=home_team_check.id,
                                        away_team_id=away_team_check.id,
                                        is_deleted=False
                                    ).first()

                                    if existing_match:
                                        is_duplicate = True

                                matches_data.append({
                                    'col': i,
                                    'home': home,
                                    'away': away,
                                    'is_duplicate': is_duplicate,
                                    'match_str': match_str
                                })
                                break

                if not matches_data:
                    flash("Nenalezeny žádné platné zápasy v hlavičce.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Uložit preview data do session
                session['import_preview'] = {
                    'round_id': round_id,
                    'round_name': r.name,
                    'matches': matches_data,
                    'total_matches': len(matches_data),
                    'duplicates_count': sum(1 for m in matches_data if m['is_duplicate'])
                }

                return redirect(url_for("admin_import_leaderboard_preview"))

            except Exception as e:
                flash(f"Chyba při načítání souboru: {str(e)}", "error")
                db.session.rollback()
                return redirect(url_for("admin_import_leaderboard"))

            try:
                import openpyxl
                from io import BytesIO

                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active

                # Načíst všechny řádky
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    flash("Soubor musí mít alespoň hlavičku a jeden řádek.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # První řádek = hlavička
                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce (Tipér + zápasy).", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Parsovat hlavičku zápasů (sloupce od 2. dál)
                matches_data = []
                for i, match_name in enumerate(header[1:], start=1):
                    if not match_name or str(match_name).strip() == "":
                        continue

                    match_str = str(match_name).strip()
                    # Očekáváme formát jako "Slavia-Sparta" nebo "Slavia vs Sparta"
                    for sep in ['-', ' vs ', ' x ', ':']:
                        if sep in match_str:
                            parts = match_str.split(sep, 1)
                            if len(parts) == 2:
                                home = parts[0].strip()
                                away = parts[1].strip()
                                matches_data.append({'col': i, 'home': home, 'away': away})
                                break

                if not matches_data:
                    flash("Nenalezeny žádné platné zápasy v hlavičce (očekáván formát 'Domácí-Hosté').", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Vytvořit/najít zápasy
                created_matches = 0
                match_map = {}  # col_index -> Match
                for md in matches_data:
                    home_team = _get_or_create_team(r.id, md['home'])
                    away_team = _get_or_create_team(r.id, md['away'])

                    # Zkusit najít existující zápas
                    m = Match.query.filter_by(
                        round_id=r.id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if not m:
                        m = Match(round_id=r.id, home_team_id=home_team.id, away_team_id=away_team.id)
                        db.session.add(m)
                        db.session.flush()
                        created_matches += 1

                    match_map[md['col']] = m

                # Zkontrolovat první řádek po hlavičce - může obsahovat výsledky
                # NEBO poslední řádek - výsledky mohou být i dole
                results_imported = 0
                data_rows_start = 1  # Od kterého řádku začínají tipy
                data_rows_end = len(rows)  # Kde končí tipy (může být zkráceno o poslední řádek)

                # KONTROLA 1: První řádek po hlavičce (rows[1])
                if len(rows) > 1:
                    first_data_row = rows[1]
                    first_cell = str(first_data_row[0]).strip() if first_data_row[0] else ""

                    # Pokud první buňka je prázdná, "Výsledek", "Result" nebo "Skóre", jedná se o řádek s výsledky
                    if first_cell.lower() in ['', 'výsledek', 'result', 'skóre', 'score', 'vysledek']:
                        # Importovat výsledky z druhého řádku
                        for col_idx, match in match_map.items():
                            if col_idx >= len(first_data_row):
                                continue

                            result_value = first_data_row[col_idx]
                            if not result_value:
                                continue

                            # Parsovat výsledek
                            home_score = None
                            away_score = None

                            # Excel time format
                            if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                                home_score = result_value.hour
                                away_score = result_value.minute
                            else:
                                # String format
                                result_str = str(result_value).strip()
                                if result_str and result_str not in ['-', '—', '']:
                                    for sep in [':', '-']:
                                        if sep in result_str:
                                            parts = result_str.split(sep, 1)
                                            if len(parts) == 2:
                                                try:
                                                    home_score = int(parts[0].strip())
                                                    away_score = int(parts[1].strip())
                                                    break
                                                except:
                                                    pass

                            if home_score is not None and away_score is not None:
                                match.home_score = home_score
                                match.away_score = away_score
                                results_imported += 1

                        # Přeskočit tento řádek při zpracování tipů
                        data_rows_start = 2
                        db.session.flush()

                # KONTROLA 2: Poslední řádek (rows[-1]) - NOVĚ!
                # Pouze pokud jsme NENAŠLI výsledky v prvním řádku
                if results_imported == 0 and len(rows) > 2:
                    last_row = rows[-1]
                    last_cell = str(last_row[0]).strip() if last_row[0] else ""

                    # Pokud poslední řádek má "Výsledek" nebo je prázdný
                    if last_cell.lower() in ['', 'výsledek', 'result', 'skóre', 'score', 'vysledek']:
                        # Importovat výsledky z posledního řádku
                        for col_idx, match in match_map.items():
                            if col_idx >= len(last_row):
                                continue

                            result_value = last_row[col_idx]
                            if not result_value:
                                continue

                            # Parsovat výsledek
                            home_score = None
                            away_score = None

                            # Excel time format
                            if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                                home_score = result_value.hour
                                away_score = result_value.minute
                            else:
                                # String format
                                result_str = str(result_value).strip()
                                if result_str and result_str not in ['-', '—', '']:
                                    for sep in [':', '-']:
                                        if sep in result_str:
                                            parts = result_str.split(sep, 1)
                                            if len(parts) == 2:
                                                try:
                                                    home_score = int(parts[0].strip())
                                                    away_score = int(parts[1].strip())
                                                    break
                                                except:
                                                    pass

                            if home_score is not None and away_score is not None:
                                match.home_score = home_score
                                match.away_score = away_score
                                results_imported += 1

                        # Vynechat poslední řádek při zpracování tipů
                        data_rows_end = len(rows) - 1
                        db.session.flush()

                # Procházet řádky s tipéry
                created_users = 0
                created_tips = 0
                skipped_tips = 0
                skipped_tips_details = []  # Seznam přeskočených tipů pro Excel report

                for row in rows[data_rows_start:data_rows_end]:  # Skip header, případně výsledky nahoře a případně výsledky dole
                    if not row or not any(row):
                        continue

                    username = str(row[0]).strip() if row[0] else ""
                    if not username:
                        continue

                    # Najít nebo vytvořit uživatele
                    email = f"{username.lower().replace(' ', '_')}@import.tipovacka"
                    user = User.query.filter_by(username=username).first()

                    if not user:
                        user = User(
                            email=email,
                            username=username,
                            role='user'
                        )
                        user.set_password('tipovacka123')  # Výchozí heslo
                        db.session.add(user)
                        db.session.flush()
                        created_users += 1

                    # Importovat tipy
                    for col_idx, match in match_map.items():
                        if col_idx >= len(row):
                            continue

                        cell_value = row[col_idx]
                        if not cell_value:
                            continue

                        # Parsovat tip
                        tip_home = None
                        tip_away = None

                        # OPRAVA: Excel často převádí "2:1" na časový formát "2:01:00"
                        # Zkontrolovat jestli je to datetime.time objekt
                        if hasattr(cell_value, 'hour') and hasattr(cell_value, 'minute'):
                            # Je to time objekt - použít hour jako domácí, minute jako hosté
                            tip_home = cell_value.hour
                            tip_away = cell_value.minute
                        else:
                            # Je to string - parsovat běžně
                            tip_str = str(cell_value).strip()
                            if not tip_str or tip_str in ['-', '—', '']:
                                continue

                            # Parsovat formát "2:1" nebo "2-1"
                            for sep in [':', '-']:
                                if sep in tip_str:
                                    parts = tip_str.split(sep, 1)
                                    if len(parts) == 2:
                                        try:
                                            tip_home = int(parts[0].strip())
                                            tip_away = int(parts[1].strip())
                                            break
                                        except:
                                            pass

                        if tip_home is None or tip_away is None:
                            skipped_tips += 1
                            # Zaznamenat detail přeskočeného tipu
                            match_name = f"{match.home_team.name if match.home_team else '?'}-{match.away_team.name if match.away_team else '?'}"
                            skipped_tips_details.append({
                                'user': username,
                                'match': match_name,
                                'value': str(cell_value) if cell_value else '(prázdné)',
                                'reason': 'Nepodařilo se parsovat tip (očekáván formát "2:1" nebo "2-1")'
                            })
                            continue

                        # Vytvořit nebo aktualizovat tip
                        existing_tip = Tip.query.filter_by(
                            user_id=user.id,
                            match_id=match.id
                        ).first()

                        if existing_tip:
                            existing_tip.tip_home = tip_home
                            existing_tip.tip_away = tip_away
                        else:
                            db.session.add(Tip(
                                user_id=user.id,
                                match_id=match.id,
                                tip_home=tip_home,
                                tip_away=tip_away
                            ))
                            created_tips += 1

                db.session.commit()
                audit("import.leaderboard", "Round", r.id,
                      users=created_users, matches=created_matches, tips=created_tips, skipped=skipped_tips, results=results_imported)

                # Vytvořit Excel report pro přeskočené tipy
                skipped_report_path = None
                if skipped_tips > 0 and skipped_tips_details:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment

                    wb_report = openpyxl.Workbook()
                    ws_report = wb_report.active
                    ws_report.title = "Přeskočené tipy"

                    # Hlavička
                    headers = ['Tipér', 'Zápas', 'Hodnota v buňce', 'Důvod přeskočení']
                    ws_report.append(headers)

                    # Styling hlavičky
                    for cell in ws_report[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="FF4D6D", end_color="FF4D6D", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Data
                    for detail in skipped_tips_details:
                        ws_report.append([
                            detail['user'],
                            detail['match'],
                            detail['value'],
                            detail['reason']
                        ])

                    # Automatická šířka sloupců
                    for column in ws_report.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_report.column_dimensions[column_letter].width = adjusted_width

                    # Uložit report do outputs
                    import os
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_dir = "/mnt/user-data/outputs"
                    os.makedirs(output_dir, exist_ok=True)
                    skipped_report_filename = f"preskocene_tipy_{timestamp}.xlsx"
                    skipped_report_path = os.path.join(output_dir, skipped_report_filename)
                    wb_report.save(skipped_report_path)

                msg = f"✅ Import dokončen do soutěže '{r.name}'! Vytvořeno: {created_users} uživatelů, {created_matches} zápasů, {created_tips} tipů."
                if results_imported > 0:
                    msg += f" Importováno {results_imported} výsledků."
                if skipped_tips > 0:
                    msg += f" Přeskočeno: {skipped_tips} tipů."
                flash(msg, "ok")

                # Přepnout na nově importovanou soutěž
                set_selected_round_id(r.id)

                # Pokud byly přeskočené tipy, nabídnout stažení reportu
                if skipped_report_path:
                    # Uložit cestu do session pro download
                    from flask import session as flask_session
                    flask_session['skipped_report_path'] = skipped_report_path
                    flask_session['skipped_report_filename'] = skipped_report_filename
                    return redirect(url_for("admin_import_leaderboard_report"))

                return redirect(url_for("leaderboard"))

            except Exception as e:
                db.session.rollback()
                flash(f"Chyba při importu: {str(e)}", "error")
                return redirect(url_for("admin_import_leaderboard"))

        # GET request - zobrazit formulář
        all_rounds = Round.query.order_by(Round.is_active.desc(), Round.id.desc()).all()

        return render_page(r"""
<div class="card">
  <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <h2 style="margin:0 0 8px 0;">📊 Import žebříčku z jiné tipovačky</h2>
    <a href="{{ url_for('admin_import_leaderboard_template') }}" class="btn" style="background:#6ea8fe; color:white;">
      📥 Stáhnout šablonu Excel
    </a>
  </div>

  <div class="muted" style="margin-bottom:16px;">
    Importuj kompletní data (uživatele, zápasy a tipy) z Excel souboru s žebříčkem.
  </div>

  <div class="card" style="background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); margin-bottom:16px;">
    <h3 style="margin:0 0 8px 0;">📋 Formát souboru</h3>
    <div class="muted" style="font-size:13px; line-height:1.6;">
      <strong>První řádek (hlavička):</strong><br>
      • Sloupec A: "Tipér" nebo "Jméno" (ignoruje se)<br>
      • Sloupec B+: Názvy zápasů ve formátu <strong>"Domácí-Hosté"</strong><br>
      &nbsp;&nbsp;Příklad: "Slavia-Sparta", "Plzeň-Liberec"<br><br>

      <strong>Druhý řádek NEBO poslední řádek (VOLITELNĚ - výsledky):</strong><br>
      • Sloupec A: prázdné NEBO "Výsledek" NEBO "Result"<br>
      • Sloupec B+: Výsledky zápasů <strong>"2:1"</strong><br>
      &nbsp;&nbsp;💡 Výsledky můžou být na <strong>druhém řádku</strong> (hned po hlavičce)<br>
      &nbsp;&nbsp;💡 NEBO na <strong>posledním řádku</strong> (dole pod tipéry) - NOVĚ!<br>
      &nbsp;&nbsp;💡 Pokud první sloupec je prázdný nebo obsahuje "Výsledek", importují se výsledky<br>
      &nbsp;&nbsp;💡 Pokud první sloupec obsahuje jméno, považuje se za tipéra<br><br>

      <strong>Další řádky (tipéři):</strong><br>
      • Sloupec A: Jméno tipéra<br>
      • Sloupec B+: Tipy ve formátu <strong>"2:1"</strong> nebo <strong>"2-1"</strong><br>
      &nbsp;&nbsp;💡 Funguje i časový formát "2:01" (Excel často automaticky převádí)<br><br>

      <strong>Co se stane:</strong><br>
      ✅ Vytvoří se uživatelé (heslo: "tipovacka123")<br>
      ✅ Vytvoří se zápasy a týmy<br>
      ✅ Naimportují se tipy<br>
      ✅ Naimportují se výsledky (pokud jsou)<br>
      ℹ️ Email uživatelů: <code>jmeno@import.tipovacka</code>
    </div>
  </div>

  <div class="card" style="background:rgba(255,199,79,0.08); border:1px solid rgba(255,199,79,0.2); margin-bottom:16px;">
    <strong>⚠️ Příklad Excel souboru S VÝSLEDKY:</strong>
    <table style="margin-top:8px; border-collapse:collapse; font-size:12px;">
      <tr style="background:rgba(255,255,255,0.05);">
        <th style="border:1px solid var(--line); padding:4px;">Tipér</th>
        <th style="border:1px solid var(--line); padding:4px;">Slavia-Sparta</th>
        <th style="border:1px solid var(--line); padding:4px;">Plzeň-Liberec</th>
        <th style="border:1px solid var(--line); padding:4px;">Baník-Sigma</th>
      </tr>
      <tr style="background:rgba(76,175,80,0.1);">
        <td style="border:1px solid var(--line); padding:4px; font-style:italic; color:#4CAF50;">Výsledek</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">3:1</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">2:2</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">1:0</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Petr</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">1:0</td>
        <td style="border:1px solid var(--line); padding:4px;">3:3</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Jana</td>
        <td style="border:1px solid var(--line); padding:4px;">1:1</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">0:2</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Martin</td>
        <td style="border:1px solid var(--line); padding:4px;">2:01</td>
        <td style="border:1px solid var(--line); padding:4px; font-size:10px;" class="muted">časový formát OK ✓</td>
        <td style="border:1px solid var(--line); padding:4px;">1:3</td>
      </tr>
    </table>
    <div class="muted" style="font-size:11px; margin-top:6px;">
      💡 Řádek "Výsledek" je volitelný. Pokud první sloupec je prázdný nebo obsahuje "Výsledek"/"Result"/"Skóre", importují se výsledky zápasů.
    </div>
  </div>

  <div class="card" style="background:rgba(255,199,79,0.08); border:1px solid rgba(255,199,79,0.2); margin-bottom:16px;">
    <strong>⚠️ Příklad Excel souboru s výsledky DOLE (NOVĚ):</strong>
    <table style="margin-top:8px; border-collapse:collapse; font-size:12px;">
      <tr style="background:rgba(255,255,255,0.05);">
        <th style="border:1px solid var(--line); padding:4px;">Tipér</th>
        <th style="border:1px solid var(--line); padding:4px;">Slavia-Sparta</th>
        <th style="border:1px solid var(--line); padding:4px;">Plzeň-Liberec</th>
        <th style="border:1px solid var(--line); padding:4px;">Baník-Sigma</th>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Petr</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">1:0</td>
        <td style="border:1px solid var(--line); padding:4px;">3:3</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Jana</td>
        <td style="border:1px solid var(--line); padding:4px;">1:1</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">0:2</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Martin</td>
        <td style="border:1px solid var(--line); padding:4px;">2:01</td>
        <td style="border:1px solid var(--line); padding:4px;">1:1</td>
        <td style="border:1px solid var(--line); padding:4px;">1:3</td>
      </tr>
      <tr style="background:rgba(76,175,80,0.1);">
        <td style="border:1px solid var(--line); padding:4px; font-style:italic; color:#4CAF50;">Výsledek</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">3:1</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">2:2</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">1:0</td>
      </tr>
    </table>
    <div class="muted" style="font-size:11px; margin-top:6px;">
      💡 <strong>NOVĚ:</strong> Výsledky můžou být i v posledním řádku! Import automaticky detekuje řádek "Výsledek" ať je nahoře nebo dole.
    </div>
  </div>

  <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; align-items:stretch; gap:16px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    {# Krok 1: Kam importovat #}
    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">1️⃣ Kam importovat data?</h3>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="existing" {% if all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_existing">
        <span>Do existující soutěže</span>
      </label>

      <div id="existing_round_select" style="margin-left:28px; margin-bottom:16px;">
        <select name="round_id" style="width:100%; max-width:400px;">
          {% for rnd in all_rounds %}
            <option value="{{ rnd.id }}">
              {% if rnd.is_active %}★ {% endif %}{{ rnd.name }}
            </option>
          {% endfor %}
          {% if not all_rounds %}
            <option value="">-- Žádné soutěže --</option>
          {% endif %}
        </select>
      </div>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="new" {% if not all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_new">
        <span>Vytvořit novou soutěž</span>
      </label>

      <div id="new_round_input" style="margin-left:28px; display:none;">
        <input type="text" name="new_round_name" placeholder="Název nové soutěže (např. CZ Liga fotbal 2026)"
               style="width:100%; max-width:400px;">
        <div class="muted" style="margin-top:6px; font-size:12px;">
          Vytvoří se nová soutěž se sportem "Fotbal"
        </div>
      </div>
    </div>

    {# Krok 2: Soubor #}
    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">2️⃣ Vyber Excel soubor</h3>
      <input type="file" name="excel_file" accept=".xlsx,.xls" required>
    </div>

    <button class="btn btn-primary" type="submit" style="padding:14px; font-size:16px; font-weight:900;">
      📥 Importovat žebříček
    </button>
    <a class="btn" href="{{ url_for('admin_import') }}">Zpět</a>
  </form>
</div>

<script>
function toggleImportTarget() {
  const existingChecked = document.getElementById('radio_existing').checked;
  const existingDiv = document.getElementById('existing_round_select');
  const newDiv = document.getElementById('new_round_input');

  if (existingChecked) {
    existingDiv.style.display = 'block';
    newDiv.style.display = 'none';
  } else {
    existingDiv.style.display = 'none';
    newDiv.style.display = 'block';
  }
}

// Inicializace při načtení
toggleImportTarget();
</script>
""", all_rounds=all_rounds)



    @admin_bp.route("/import/leaderboard/report")
    @login_required
    def admin_import_leaderboard_report():
        admin_required()
        from flask import session as flask_session, send_file

        skipped_report_path = flask_session.get('skipped_report_path')
        skipped_report_filename = flask_session.get('skipped_report_filename', 'preskocene_tipy.xlsx')

        if not skipped_report_path or not os.path.exists(skipped_report_path):
            flash("Report přeskočených tipů není k dispozici.", "error")
            return redirect(url_for("admin_import"))

        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 16px 0;">⚠️ Přeskočené tipy při importu</h2>

  <div class="card" style="background:rgba(255,199,79,0.1); border:2px solid rgba(255,199,79,0.6); padding:20px; margin-bottom:20px;">
    <h3 style="margin:0 0 12px 0; color:#FFC74F;">Některé tipy nebyly importovány</h3>
    <div style="font-size:14px; line-height:1.6;">
      Při importu žebříčku byly některé tipy přeskočeny, protože se nepodařilo je správně parsovat.
      <br><br>
      <strong>Obvyklé důvody:</strong><br>
      • Neplatný formát (očekáván "2:1" nebo "2-1")<br>
      • Prázdná buňka nebo neplatná hodnota<br>
      • Text místo čísla<br><br>

      <strong>Stáhni Excel soubor</strong> s detailním přehledem všech přeskočených tipů.
      V souboru najdeš: tipéra, zápas, hodnotu v buňce a důvod přeskočení.
    </div>
  </div>

  <div style="display:flex; gap:12px; margin-bottom:20px;">
    <a href="{{ url_for('admin_import_leaderboard_download') }}"
       class="btn btn-primary"
       style="padding:14px 24px; font-size:16px; font-weight:900; background:#FFC74F; color:#000;">
      📥 Stáhnout Excel report
    </a>
    <a href="{{ url_for('leaderboard') }}" class="btn" style="padding:14px 24px; font-size:16px;">
      ➡️ Přejít do žebříčku
    </a>
  </div>

  <div class="muted" style="font-size:12px;">
    💡 Po opravě tipů v Excelu můžeš znovu importovat žebříček s opravenými daty.
  </div>
</div>
""")



    @admin_bp.route("/import/leaderboard/download")
    @login_required
    def admin_import_leaderboard_download():
        admin_required()
        from flask import session as flask_session, send_file

        skipped_report_path = flask_session.get('skipped_report_path')
        skipped_report_filename = flask_session.get('skipped_report_filename', 'preskocene_tipy.xlsx')

        if not skipped_report_path or not os.path.exists(skipped_report_path):
            flash("Report přeskočených tipů není k dispozici.", "error")
            return redirect(url_for("admin_import"))

        # Smazat ze session po stažení
        flask_session.pop('skipped_report_path', None)
        flask_session.pop('skipped_report_filename', None)

        return send_file(
            skipped_report_path,
            as_attachment=True,
            download_name=skipped_report_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # --- ADMIN AUDIT ---


    @admin_bp.route("/audit")
    @login_required
    def admin_audit():
        admin_required()
        logs = AuditLog.query.order_by(AuditLog.id.desc()).limit(200).all()
        return render_page(r"""
<div class="card">
  <h2 style="margin:0 0 8px 0;">Historie změn (posledních 200)</h2>
  <hr class="sep">
  {% for l in logs %}
    <div>
      <b>{{ l.action }}</b> <span class="muted">{{ l.entity }}{% if l.entity_id %}#{{ l.entity_id }}{% endif %}</span>
      <div class="muted">{{ l.at.strftime("%Y-%m-%d %H:%M:%S") }} | {% if l.actor %}{{ l.actor.username }}{% else %}—{% endif %}</div>
      {% if l.details %}<div class="muted" style="white-space:pre-wrap; margin-top:6px;">{{ l.details }}</div>{% endif %}
    </div>
    {% if not loop.last %}<hr class="sep">{% endif %}
  {% endfor %}
</div>
""", logs=logs)

    # --- PWA (Progressive Web App) ---


    @admin_bp.route("/api-sources")
    @login_required
    def admin_api_sources():
        """Správa API zdrojů"""
        admin_required()

        sources = APISource.query.all()

        return render_page(r"""
<style>
  .api-source-card {
    background: rgba(255,255,255,.03);
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 16px;
    margin-bottom: 16px;
  }

  .api-source-card.active {
    border-color: rgba(51,209,122,.5);
    background: rgba(51,209,122,.08);
  }

  .api-badge {
    display: inline-block;
    padding: 4px 12px;
    border-radius: 12px;
    font-size: 12px;
    font-weight: 700;
  }

  .api-badge.nhl { background: rgba(110,168,254,.2); color: #6ea8fe; }
  .api-badge.api-football { background: rgba(51,209,122,.2); color: #33d17a; }
</style>

<div class="card">
  <div class="row" style="justify-content: space-between; align-items: center;">
    <div>
      <h2 style="margin: 0 0 8px 0;">🔌 API Zdroje</h2>
      <div class="muted">Automatický import zápasů a výsledků</div>
    </div>
    <a href="{{ url_for('admin_api_source_new') }}" class="btn btn-primary">+ Nový zdroj</a>
  </div>
</div>

{% if sources|length == 0 %}
<div class="card">
  <div style="text-align: center; padding: 40px;">
    <div style="font-size: 48px; margin-bottom: 12px;">🔌</div>
    <h3 style="margin: 0 0 8px 0;">Žádné API zdroje</h3>
    <div class="muted">Přidej první API zdroj pro automatický import</div>
    <a href="{{ url_for('admin_api_source_new') }}" class="btn btn-primary" style="margin-top: 16px;">+ Přidat zdroj</a>
  </div>
</div>
{% else %}
  {% for source in sources %}
  <div class="api-source-card {% if source.is_active %}active{% endif %}">
    <div class="row" style="justify-content: space-between; align-items: flex-start; margin-bottom: 12px;">
      <div>
        <div class="row" style="gap: 8px; align-items: center; margin-bottom: 8px;">
          <span class="api-badge {{ source.api_type }}">{{ source.api_type.upper() }}</span>
          {% if source.is_active %}
            <span class="tag pill-ok">✅ Aktivní</span>
          {% else %}
            <span class="tag pill-bad">⏸️ Neaktivní</span>
          {% endif %}
        </div>
        <h3 style="margin: 0 0 4px 0;">{% if source.round %}{{ source.round.name }}{% else %}[Soutěž smazána]{% endif %}</h3>
        <div class="muted" style="font-size: 13px;">
          {% if source.league_id %}Liga/Sezóna: <b>{{ source.league_id }}</b> • {% endif %}
          Vytvořeno: {{ source.created_at.strftime('%d.%m.%Y') }}
        </div>
      </div>
      <div class="row" style="gap: 8px;">
        <a href="{{ url_for('admin_api_import_preview', source_id=source.id, import_type='matches') }}"
           class="btn btn-sm">📥 Import zápasů</a>
        <a href="{{ url_for('admin_api_import_preview', source_id=source.id, import_type='results') }}"
           class="btn btn-sm">📊 Import výsledků</a>
        <a href="{{ url_for('admin_api_source_edit', source_id=source.id) }}"
           class="btn btn-sm">✏️ Upravit</a>
        <form method="post" action="{{ url_for('admin_api_source_delete', source_id=source.id) }}"
              style="display:inline;"
              onsubmit="return confirm('Opravdu smazat API zdroj {% if source.round %}\'{{ source.round.name }}\'{% else %}[Smazaná soutěž]{% endif %}?')">
          <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
          <button type="submit" class="btn btn-sm btn-danger"
                  style="background:rgba(255,77,109,0.2); color:#ff4d6d; border:none; cursor:pointer;">
            🗑️ Smazat
          </button>
        </form>
      </div>
    </div>

    <div class="row" style="gap: 16px; font-size: 13px;">
      <div>
        <div class="muted">Auto import zápasů:</div>
        <div><strong>{% if source.auto_import_matches %}Ano{% else %}Ne{% endif %}</strong></div>
      </div>
      <div>
        <div class="muted">Auto import výsledků:</div>
        <div><strong>{% if source.auto_import_results %}Ano{% else %}Ne{% endif %}</strong></div>
      </div>
      <div>
        <div class="muted">Vyžadovat potvrzení:</div>
        <div><strong>{% if source.require_admin_approval %}Ano{% else %}Ne{% endif %}</strong></div>
      </div>
      <div>
        <div class="muted">Ignorovat OT/SO:</div>
        <div><strong>{% if source.exclude_overtime %}Ano{% else %}Ne{% endif %}</strong></div>
      </div>
    </div>

    {% if source.last_import_at %}
    <div class="muted" style="margin-top: 12px; font-size: 12px;">
      Poslední import: {{ source.last_import_at.strftime('%d.%m.%Y %H:%M') }}
    </div>
    {% endif %}
  </div>
  {% endfor %}
{% endif %}

<div class="card" style="background: rgba(110,168,254,.08); border-color: rgba(110,168,254,.3);">
  <h3 style="margin: 0 0 12px 0;">ℹ️ Podporované API</h3>
  <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 12px;">
    <div>
      <div style="font-weight: 700; margin-bottom: 4px;">🏒 NHL API</div>
      <div class="muted" style="font-size: 13px;">Oficiální NHL API - zdarma, bez registrace</div>
    </div>
    <div>
      <div style="font-weight: 700; margin-bottom: 4px;">⚽ API-Football</div>
      <div class="muted" style="font-size: 13px;">Fotbalové ligy - vyžaduje API klíč (api-football.com)</div>
    </div>
    <div>
      <div style="font-weight: 700; margin-bottom: 4px;">⚽ TheSportsDB</div>
      <div class="muted" style="font-size: 13px;">Fotbalové ligy - zdarma (pozor na dostupnost aktuální sezóny)</div>
    </div>
    <div>
      <div style="font-weight: 700; margin-bottom: 4px;">🏆 UEFA UCL (All fixtures)</div>
      <div class="muted" style="font-size: 13px;">Liga mistrů - zdarma z oficiální UEFA stránky (scrape)</div>
    </div>
  </div>
</div>
""", sources=sources)



    @admin_bp.route("/api-source/new", methods=["GET", "POST"])
    @login_required
    def admin_api_source_new():
        """Vytvoření nového API zdroje"""
        admin_required()

        if request.method == "POST":
            round_id = int(request.form["round_id"])
            api_type = request.form["api_type"]
            league_id = request.form.get("league_id", "").strip() or None
            api_key = request.form.get("api_key", "").strip() or None

            auto_import_matches = request.form.get("auto_import_matches") == "on"
            auto_import_results = request.form.get("auto_import_results") == "on"
            require_admin_approval = request.form.get("require_admin_approval") == "on"
            exclude_overtime = request.form.get("exclude_overtime") == "on"

            source = APISource(
                round_id=round_id,
                api_type=api_type,
                league_id=league_id,
                api_key=api_key,
                auto_import_matches=auto_import_matches,
                auto_import_results=auto_import_results,
                require_admin_approval=require_admin_approval,
                exclude_overtime=exclude_overtime,
                is_active=True,
                created_by_id=current_user.id
            )

            db.session.add(source)
            db.session.commit()

            audit("api_source.create", "APISource", source.id)
            flash("✅ API zdroj vytvořen!", "ok")
            return redirect(url_for("admin_api_sources"))

        rounds = Round.query.order_by(Round.id.desc()).all()

        return render_page(r"""
<div class="card">
  <h2 style="margin: 0 0 8px 0;">+ Nový API zdroj</h2>
  <div class="muted">Nastav automatický import zápasů a výsledků</div>
  <hr class="sep">

  <form method="post">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div style="display: grid; gap: 16px;">

      <div>
        <label class="muted" style="margin-bottom: 6px; display: block;">Soutěž *</label>
        <select name="round_id" required>
          <option value="">-- Vyber soutěž --</option>
          {% for r in rounds %}
            <option value="{{ r.id }}">{{ r.name }}</option>
          {% endfor %}
        </select>
      </div>

      <div>
        <label class="muted" style="margin-bottom: 6px; display: block;">Typ API *</label>
        <select name="api_type" id="api_type" required onchange="toggleAPIFields()">
          <option value="">-- Vyber API --</option>
          <option value="nhl">🏒 NHL API (hokej)</option>
          <option value="api-football">⚽ API-Football (fotbal - placené)</option>
          <option value="thesportsdb">⚽ TheSportsDB (fotbal - ZDARMA!)</option>
          <option value="uefa-ucl">🏆 UEFA UCL (All fixtures - ZDARMA!)</option>
        </select>
      </div>

      <div id="nhl_fields" style="display: none;">
        <label class="muted" style="margin-bottom: 6px; display: block;">Sezóna</label>
        <input id="nhl_league_id" name="league_id" placeholder="20252026" value="20252026" disabled>
        <div class="muted" style="font-size: 12px; margin-top: 4px;">
          Formát: YYYYYYYY (např. 20252026 pro sezónu 2025/26)
        </div>
      </div>

      <div id="football_fields" style="display: none;">
        <div style="margin-bottom: 12px;">
          <label class="muted" style="margin-bottom: 6px; display: block;">ID Ligy *</label>
          <input id="football_league_id" name="league_id" placeholder="39" disabled>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            Např: 39 = Premier League, 140 = La Liga, 78 = Bundesliga, 345 = Chance Liga (CZ)<br>
            <a href="https://www.api-football.com/documentation-v3#tag/Leagues" target="_blank">Najdi ID ligy →</a>
          </div>
        </div>

        <div>
          <label class="muted" style="margin-bottom: 6px; display: block;">API Klíč *</label>
          <input id="football_api_key" name="api_key" type="password" placeholder="tvůj-api-klíč" disabled>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            Registrace: <a href="https://www.api-football.com/" target="_blank">api-football.com</a>
          </div>
        </div>
      </div>

      <div id="thesportsdb_fields" style="display: none;">
        <div style="margin-bottom: 12px;">
          <label class="muted" style="margin-bottom: 6px; display: block;">ID Ligy *</label>
          <input id="thesportsdb_league_id" name="league_id" placeholder="4631" disabled>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            <strong>🇨🇿 Česko:</strong> 4631 = Chance Liga (1. liga)<br>
            <strong>🏆 Evropa:</strong> 4480 = Liga mistrů (Champions League)<br>
            <strong>🌍 TOP Ligy:</strong> 4328 = Premier League, 4335 = La Liga, 4332 = Serie A, 4331 = Bundesliga<br>
            <a href="https://www.thesportsdb.com/sport/Soccer" target="_blank">Najdi ID ligy →</a>
          </div>
        </div>

        <div>
          <label class="muted" style="margin-bottom: 6px; display: block;">Sezóna (volitelné)</label>
          <input id="thesportsdb_season" name="api_key" placeholder="2024-2025" disabled>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            Formát: YYYY-YYYY (např. 2024-2025)<br>
            Ponech prázdné pro posledních 15 zápasů
          </div>
        </div>

        <div class="muted" style="font-size: 12px; margin-top: 12px; padding: 12px; background: rgba(46,213,115,.08); border-radius: 8px; border: 1px solid rgba(46,213,115,.2);">
          ✅ <strong>ZDARMA!</strong> Žádný API klíč není potřeba. Všechny sezóny dostupné. Bez rate limitů.
        </div>
      </div>

      <hr class="sep">

      <div>
        <h3 style="margin: 0 0 12px 0;">⚙️ Nastavení</h3>

        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="require_admin_approval" checked>
          <span>Vyžadovat potvrzení adminem před importem</span>
        </label>

        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="exclude_overtime" checked>
          <span>Ignorovat prodloužení/nájezdy (jen základní hrací doba)</span>
        </label>

        <div class="muted" style="font-size: 12px; margin-top: 8px; padding: 12px; background: rgba(110,168,254,.08); border-radius: 8px;">
          💡 <strong>Tip:</strong> Automatický import můžeš nastavit později po otestování manuálního importu.
        </div>
      </div>

      <div class="row" style="gap: 8px;">
        <button type="submit" class="btn btn-primary">✅ Vytvořit zdroj</button>
        <a href="{{ url_for('admin_api_sources') }}" class="btn">Zrušit</a>
      </div>
    </div>
  </form>
</div>

<script>
function toggleAPIFields() {
  const apiType = document.getElementById('api_type').value;
  const nhlFields = document.getElementById('nhl_fields');
  const footballFields = document.getElementById('football_fields');
  const thesportsdbFields = document.getElementById('thesportsdb_fields');
  const nhlInput = document.getElementById('nhl_league_id');
  const footballLeagueInput = document.getElementById('football_league_id');
  const footballKeyInput = document.getElementById('football_api_key');
  const thesportsdbLeagueInput = document.getElementById('thesportsdb_league_id');
  const thesportsdbSeasonInput = document.getElementById('thesportsdb_season');
  const uefaFields = document.getElementById('uefa_fields');
  const uefaUrlInput = document.getElementById('uefa_url');

  if (apiType === 'nhl') {
    // Show NHL, hide others
    nhlFields.style.display = 'block';
    footballFields.style.display = 'none';
    thesportsdbFields.style.display = 'none';
    uefaFields.style.display = 'none';
    uefaFields.style.display = 'none';
    uefaFields.style.display = 'none';
    // Enable NHL input, disable others
    nhlInput.disabled = false;
    footballLeagueInput.disabled = true;
    footballKeyInput.disabled = true;
    thesportsdbLeagueInput.disabled = true;
    thesportsdbSeasonInput.disabled = true;
    uefaUrlInput.disabled = true;
    uefaUrlInput.disabled = true;
    uefaUrlInput.disabled = true;
  } else if (apiType === 'api-football') {
    // Show API-Football, hide others
    nhlFields.style.display = 'none';
    footballFields.style.display = 'block';
    thesportsdbFields.style.display = 'none';
    // Enable Football inputs, disable others
    nhlInput.disabled = true;
    footballLeagueInput.disabled = false;
    footballKeyInput.disabled = false;
    thesportsdbLeagueInput.disabled = true;
    thesportsdbSeasonInput.disabled = true;
  } else if (apiType === 'thesportsdb') {
    // Show TheSportsDB, hide others
    nhlFields.style.display = 'none';
    footballFields.style.display = 'none';
    thesportsdbFields.style.display = 'block';
    uefaFields.style.display = 'none';
    // Enable TheSportsDB inputs, disable others
    nhlInput.disabled = true;
    footballLeagueInput.disabled = true;
    footballKeyInput.disabled = true;
    thesportsdbLeagueInput.disabled = false;
    thesportsdbSeasonInput.disabled = false;
    uefaUrlInput.disabled = true;
  } else if (apiType === 'uefa-ucl') {
    // Show UEFA UCL, hide others
    nhlFields.style.display = 'none';
    footballFields.style.display = 'none';
    thesportsdbFields.style.display = 'none';
    uefaFields.style.display = 'block';
    // Disable other inputs, enable UEFA URL
    nhlInput.disabled = true;
    footballLeagueInput.disabled = true;
    footballKeyInput.disabled = true;
    thesportsdbLeagueInput.disabled = true;
    thesportsdbSeasonInput.disabled = true;
    uefaUrlInput.disabled = false;
  } else {
    // Hide all
    nhlFields.style.display = 'none';
    footballFields.style.display = 'none';
    thesportsdbFields.style.display = 'none';
    nhlInput.disabled = true;
    footballLeagueInput.disabled = true;
    footballKeyInput.disabled = true;
    thesportsdbLeagueInput.disabled = true;
    thesportsdbSeasonInput.disabled = true;
  }
}
</script>
""", rounds=rounds)



    @admin_bp.route("/api-source/<int:source_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_api_source_edit(source_id: int):
        """Úprava API zdroje"""
        admin_required()

        api_source = db.session.get(APISource, source_id)
        if not api_source:
            flash("❌ API zdroj nenalezen.", "error")
            return redirect(url_for("admin_api_sources"))

        if request.method == "POST":
            try:
                api_source.round_id = int(request.form["round_id"])
                api_source.api_type = request.form["api_type"]
                api_source.league_id = request.form.get("league_id", "").strip() or None

                # API klíč - pokud je prázdný (placeholder), ponechat původní
                new_api_key = request.form.get("api_key", "").strip()
                if new_api_key and new_api_key != "":
                    api_source.api_key = new_api_key

                api_source.auto_import_matches = request.form.get("auto_import_matches") == "on"
                api_source.auto_import_results = request.form.get("auto_import_results") == "on"
                api_source.require_admin_approval = request.form.get("require_admin_approval") == "on"
                api_source.exclude_overtime = request.form.get("exclude_overtime") == "on"
                api_source.is_active = request.form.get("is_active") == "on"

                db.session.commit()

                audit("api_source.edit", "APISource", api_source.id)
                flash("✅ API zdroj aktualizován!", "ok")
                return redirect(url_for("admin_api_sources"))
            except Exception as e:
                db.session.rollback()
                flash(f"❌ Chyba při ukládání: {str(e)}", "error")
                return redirect(url_for("admin_api_source_edit", source_id=source_id))

        try:
            rounds = Round.query.order_by(Round.id.desc()).all()

            # Fallback pokud není žádná soutěž
            if not rounds:
                flash("⚠️ Nejsou k dispozici žádné soutěže. Vytvoř nejprve soutěž.", "error")
                return redirect(url_for("admin_api_sources"))

            return render_page(r"""
<div class="card">
  <h2 style="margin: 0 0 8px 0;">✏️ Upravit API zdroj</h2>
  <div class="muted">{% if api_source.round %}{{ api_source.round.name }}{% else %}Soutěž smazána{% endif %} - {{ api_source.api_type.upper() }}</div>
  <hr class="sep">

  <form method="post">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div style="display: grid; gap: 16px;">

      <div>
        <label class="muted" style="margin-bottom: 6px; display: block;">Soutěž *</label>
        <select name="round_id" required>
          {% for r in rounds %}
            <option value="{{ r.id }}" {% if r.id == api_source.round_id %}selected{% endif %}>{{ r.name }}</option>
          {% endfor %}
        </select>
      </div>

      <div>
        <label class="muted" style="margin-bottom: 6px; display: block;">Typ API *</label>
        <select name="api_type" id="api_type" required onchange="toggleAPIFields()">
          <option value="nhl" {% if api_source.api_type == 'nhl' %}selected{% endif %}>🏒 NHL API (hokej)</option>
          <option value="api-football" {% if api_source.api_type == 'api-football' %}selected{% endif %}>⚽ API-Football (fotbal)</option>
        </select>
      </div>

      <div id="nhl_fields" style="{% if api_source.api_type != 'nhl' %}display: none;{% endif %}">
        <label class="muted" style="margin-bottom: 6px; display: block;">Sezóna</label>
        <input id="nhl_league_id" name="league_id" placeholder="20252026" value="{{ api_source.league_id or '20252026' }}" {% if api_source.api_type != 'nhl' %}disabled{% endif %}>
        <div class="muted" style="font-size: 12px; margin-top: 4px;">
          Formát: YYYYYYYY (např. 20252026 pro sezónu 2025/26)
        </div>
      </div>

      <div id="football_fields" style="{% if api_source.api_type != 'api-football' %}display: none;{% endif %}">
        <div style="margin-bottom: 12px;">
          <label class="muted" style="margin-bottom: 6px; display: block;">ID Ligy *</label>
          <input id="football_league_id" name="league_id" placeholder="39" value="{{ api_source.league_id or '' }}" {% if api_source.api_type != 'api-football' %}disabled{% endif %}>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            Např: 39 = Premier League, 140 = La Liga, 78 = Bundesliga, 345 = Chance Liga (CZ)<br>
            <a href="https://www.api-football.com/documentation-v3#tag/Leagues" target="_blank">Najdi ID ligy →</a>
          </div>
        </div>

        <div>
          <label class="muted" style="margin-bottom: 6px; display: block;">API Klíč *</label>
          <input id="football_api_key" name="api_key" type="password" placeholder="{% if api_source.api_key %}••••••••{% else %}tvůj-api-klíč{% endif %}" value="{{ api_source.api_key or '' }}" {% if api_source.api_type != 'api-football' %}disabled{% endif %}>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            Registrace: <a href="https://www.api-football.com/" target="_blank">api-football.com</a>
          </div>
        </div>
      </div>

      <hr class="sep">

      <div>
        <h3 style="margin: 0 0 12px 0;">⚙️ Nastavení</h3>

        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="is_active" {% if api_source.is_active %}checked{% endif %}>
          <span>✅ Aktivní (zapnuto)</span>
        </label>


      <!-- UEFA UCL (All fixtures) -->
      <div id="uefa_fields" style="display: none;">
        <div style="margin-bottom: 12px;">
          <label class="muted" style="margin-bottom: 6px; display: block;">UEFA "All fixtures" URL (volitelné)</label>
          <input id="uefa_url" name="league_id" placeholder="(nech prázdné pro default UEFA link)" {% if api_source is defined %}value="{{ api_source.league_id or '' }}"{% endif %} disabled>
          <div class="muted" style="font-size: 12px; margin-top: 4px;">
            Zdroj je oficiální UEFA článek "All the fixtures and results".<br>
            Pokud necháš prázdné, použije se výchozí URL pro sezonu 2025/26.
          </div>
        </div>
        <div class="muted" style="font-size: 12px; margin-top: 12px; padding: 12px; background: rgba(46,213,115,.08); border-radius: 8px; border: 1px solid rgba(46,213,115,.2);">
          ✅ <strong>ZDARMA!</strong> Nevyžaduje API klíč. Získáš: domácí/hosté/čas + po zápase skóre.
        </div>
      </div>

<label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="auto_import_matches" {% if api_source.auto_import_matches %}checked{% endif %}>
          <span>Automatický import zápasů</span>
        </label>

        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="auto_import_results" {% if api_source.auto_import_results %}checked{% endif %}>
          <span>Automatický import výsledků</span>
        </label>

        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="require_admin_approval" {% if api_source.require_admin_approval %}checked{% endif %}>
          <span>Vyžadovat potvrzení adminem</span>
        </label>

        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 8px;">
          <input type="checkbox" name="exclude_overtime" {% if api_source.exclude_overtime %}checked{% endif %}>
          <span>Ignorovat prodloužení/nájezdy</span>
        </label>
      </div>

      <div class="row" style="gap: 12px; margin-top: 16px;">
        <button class="btn btn-primary" type="submit">💾 Uložit změny</button>
        <a class="btn" href="{{ url_for('admin_api_sources') }}">Zrušit</a>
      </div>

    </div>
  </form>
</div>

<script>
function toggleAPIFields() {
  const apiType = document.getElementById('api_type').value;
  const nhlFields = document.getElementById('nhl_fields');
  const footballFields = document.getElementById('football_fields');
  const nhlInput = document.getElementById('nhl_league_id');
  const footballLeagueInput = document.getElementById('football_league_id');
  const footballKeyInput = document.getElementById('football_api_key');

  if (apiType === 'nhl') {
    // Show NHL, hide Football
    nhlFields.style.display = 'block';
    footballFields.style.display = 'none';
    // Enable NHL input, disable Football inputs
    nhlInput.disabled = false;
    footballLeagueInput.disabled = true;
    footballKeyInput.disabled = true;
  } else if (apiType === 'api-football') {
    // Show Football, hide NHL
    nhlFields.style.display = 'none';
    footballFields.style.display = 'block';
    // Enable Football inputs, disable NHL input
    nhlInput.disabled = true;
    footballLeagueInput.disabled = false;
    footballKeyInput.disabled = false;
  } else {
    // Hide both
    nhlFields.style.display = 'none';
    footballFields.style.display = 'none';
    nhlInput.disabled = true;
    footballLeagueInput.disabled = true;
    footballKeyInput.disabled = true;
  }
}
</script>
""", api_source=api_source, rounds=rounds)
        except Exception as e:
            flash(f"❌ Chyba při zobrazení formuláře: {str(e)}", "error")
            return redirect(url_for("admin_api_sources"))



    @admin_bp.route("/api-source/<int:source_id>/delete", methods=["POST"])
    @login_required
    def admin_api_source_delete(source_id: int):
        """Smazání API zdroje"""
        admin_required()

        api_source = db.session.get(APISource, source_id)
        if not api_source:
            flash("❌ API zdroj nenalezen.", "error")
            return redirect(url_for("admin_api_sources"))

        # Název pro audit log
        if api_source.round:
            source_name = f"{api_source.round.name} - {api_source.api_type.upper()}"
        else:
            source_name = f"[Smazaná soutěž] - {api_source.api_type.upper()}"

        # Smazat všechny import logy tohoto zdroje
        APIImportLog.query.filter_by(source_id=api_source.id).delete()

        # Smazat všechny mapování
        MatchAPIMapping.query.filter_by(source_id=api_source.id).delete()

        # Smazat zdroj
        db.session.delete(api_source)
        db.session.commit()

        audit("api_source.delete", "APISource", source_id, name=source_name)
        flash(f"🗑️ API zdroj '{source_name}' byl smazán.", "ok")
        return redirect(url_for("admin_api_sources"))



    @admin_bp.route("/api-import/preview/<int:source_id>/<import_type>")
    @login_required
    def admin_api_import_preview(source_id: int, import_type: str):
        """Preview importu před potvrzením"""
        admin_required()

        source = db.session.get(APISource, source_id)
        if not source:
            abort(404)

        if import_type not in ['matches', 'results']:
            abort(400)

        # Stáhnout data z API
        try:
            games = fetch_api_games(source, import_type=import_type)

            if not games:
                # Konkrétnější chybová zpráva podle typu API
                if source.api_type == 'nhl':
                    flash("❌ Nepodařilo se stáhnout zápasy z NHL API. Zkontroluj Koyeb logy pro detaily.", "error")
                elif source.api_type == 'api-football':
                    flash("❌ Nepodařilo se stáhnout zápasy z API-Football. Zkontroluj API klíč a limit requestů (100/den).", "error")
                elif source.api_type == 'thesportsdb':
                    flash("❌ TheSportsDB nic nevrátil. Buď je špatné ID ligy, nebo free endpoint nemá data pro danou sezónu.", "error")
                elif source.api_type == 'uefa-ucl':
                    flash("❌ Nepodařilo se načíst data z UEFA (All fixtures). Zkontroluj, že stránka je dostupná a že se nezměnila struktura.", "error")
                else:
                    flash("❌ Nepodařilo se stáhnout data z API. Zkontroluj nastavení zdroje.", "error")
                return redirect(url_for("admin_api_sources"))

            # Vytvoř preview
            preview = {
                'source_id': source.id,
                'import_type': import_type,
                'total_games': len(games),
                'games': games[:50]  # Max 50 pro preview
            }

            # Dry run - zjisti co by se importovalo
            if import_type == 'matches':
                imported, skipped, errors = import_matches_from_api(source, games, commit=False)
            else:  # results
                imported, skipped, errors = import_results_from_api(source, games, commit=False)

            # Ulož preview do session nebo DB
            preview_json = json.dumps(preview)

            # Vytvoř import log s preview
            import_log = APIImportLog(
                source_id=source.id,
                import_type=import_type,
                status='pending',
                imported_count=imported,
                skipped_count=skipped,
                error_count=len(errors),
                preview_data=preview_json,
                error_details=json.dumps(errors) if errors else None
            )
            db.session.add(import_log)
            db.session.commit()

            return redirect(url_for('admin_api_import_confirm', log_id=import_log.id))

        except Exception as e:
            flash(f"❌ Chyba: {str(e)}", "error")
            return redirect(url_for("admin_api_sources"))



    @admin_bp.route("/api-import/confirm/<int:log_id>", methods=["GET", "POST"])
    @login_required
    def admin_api_import_confirm(log_id: int):
        """Potvrzení nebo zamítnutí importu"""
        admin_required()

        import_log = db.session.get(APIImportLog, log_id)
        if not import_log:
            abort(404)

        if import_log.status != 'pending':
            flash("Import už byl zpracován", "error")
            return redirect(url_for("admin_api_sources"))

        if request.method == "POST":
            action = request.form.get("action")

            if action == "approve":
                # Proveď import
                try:
                    preview = json.loads(import_log.preview_data)
                    source = db.session.get(APISource, import_log.source_id)

                    # Získat indexy vybraných zápasů
                    selected_indices = request.form.getlist("selected_games")

                    if not selected_indices:
                        flash("⚠️ Nevybrali jste žádné zápasy k importu.", "error")
                        return redirect(url_for("admin_api_import_confirm", log_id=log_id))

                    selected_indices = [int(idx) for idx in selected_indices]

                    # Použij data z preview (zachová pořadí i filtrování), ne refetch
                    preview_games = preview.get('games', [])
                    selected_games = [preview_games[i] for i in selected_indices if 0 <= i < len(preview_games)]

                    if not selected_games:
                        flash("❌ Vybrané zápasy nebyly nalezeny.", "error")
                        return redirect(url_for("admin_api_sources"))

                    if import_log.import_type == 'matches':
                        imported, skipped, errors = import_matches_from_api(source, selected_games, commit=True)
                    else:  # results
                        imported, skipped, errors = import_results_from_api(source, selected_games, commit=True)

                    # Update log
                    import_log.status = 'completed'
                    import_log.imported_count = imported
                    import_log.skipped_count = skipped
                    import_log.error_count = len(errors)
                    import_log.error_details = json.dumps(errors) if errors else None
                    import_log.approved_by_id = current_user.id
                    import_log.approved_at = datetime.utcnow()
                    import_log.completed_at = datetime.utcnow()

                    # Update source last import
                    source.last_import_at = datetime.utcnow()

                    db.session.commit()

                    audit("api_import.approved", "APIImportLog", import_log.id,
                          imported=imported, skipped=skipped, errors=len(errors),
                          selected=len(selected_games))

                    flash(f"✅ Import dokončen! Importováno: {imported}, Přeskočeno: {skipped}, Vybraných: {len(selected_games)}", "ok")

                except Exception as e:
                    import_log.status = 'failed'
                    import_log.error_details = str(e)
                    db.session.commit()

                    flash(f"❌ Chyba při importu: {str(e)}", "error")

            elif action == "reject":
                import_log.status = 'rejected'
                import_log.approved_by_id = current_user.id
                import_log.approved_at = datetime.utcnow()
                db.session.commit()

                audit("api_import.rejected", "APIImportLog", import_log.id)
                flash("Import zamítnut", "ok")

            return redirect(url_for("admin_api_sources"))

        # Parse preview data
        preview = json.loads(import_log.preview_data) if import_log.preview_data else {}
        games = preview.get('games', [])

        errors = []
        if import_log.error_details:
            try:
                errors = json.loads(import_log.error_details)
            except:
                errors = [import_log.error_details]

        return render_page(r"""
<style>
  .preview-table {
    width: 100%;
    border-collapse: collapse;
  }

  .preview-table th,
  .preview-table td {
    padding: 10px 8px;
    text-align: left;
    border-bottom: 1px solid var(--line);
  }

  .preview-table th {
    background: rgba(255,255,255,.03);
    font-weight: 900;
    position: sticky;
    top: 0;
  }

  .new-import { background: rgba(51,209,122,.08); }
  .skip-import { background: rgba(167,178,214,.08); }
  .error-import { background: rgba(255,77,109,.08); }
</style>

<div class="card">
  <h2 style="margin: 0 0 8px 0;">
    {% if import_log.import_type == 'matches' %}📥 Preview importu zápasů{% else %}📊 Preview importu výsledků{% endif %}
  </h2>
  <div class="muted">{{ import_log.source.round.name }} • {{ import_log.source.api_type.upper() }}</div>
</div>

<div class="card">
  <div class="row" style="justify-content: space-between; flex-wrap: wrap; gap: 16px;">
    <div>
      <div class="muted">Celkem záznamů</div>
      <div style="font-size: 24px; font-weight: 900;">{{ import_log.imported_count + import_log.skipped_count }}</div>
    </div>
    <div>
      <div class="muted">Nové k importu</div>
      <div style="font-size: 24px; font-weight: 900; color: var(--ok);">{{ import_log.imported_count }}</div>
    </div>
    <div>
      <div class="muted">Přeskočit (existující)</div>
      <div style="font-size: 24px; font-weight: 900; color: var(--muted);">{{ import_log.skipped_count }}</div>
    </div>
    {% if import_log.error_count > 0 %}
    <div>
      <div class="muted">Chyby</div>
      <div style="font-size: 24px; font-weight: 900; color: var(--danger);">{{ import_log.error_count }}</div>
    </div>
    {% endif %}
  </div>
</div>

{% if errors|length > 0 %}
<div class="card" style="background: rgba(255,77,109,.08); border-color: rgba(255,77,109,.3);">
  <h3 style="margin: 0 0 12px 0;">⚠️ Chyby ({{ errors|length }})</h3>
  <div style="max-height: 200px; overflow-y: auto;">
    {% for error in errors %}
      <div class="muted" style="font-size: 13px; margin-bottom: 4px;">• {{ error }}</div>
    {% endfor %}
  </div>
</div>
{% endif %}

<form method="post">
  <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
<div class="card">
  <h3 style="margin: 0 0 12px 0;">📋 Data k importu (zobrazeno max {{ games|length }})</h3>

  <div style="margin-bottom: 12px;">
    <button type="button" class="btn btn-sm" onclick="selectAll()">✅ Vybrat vše</button>
    <button type="button" class="btn btn-sm" onclick="deselectAll()">☐ Zrušit výběr</button>
    <span class="muted" style="margin-left: 12px;" id="selected-count">Vybrané: 0</span>
  </div>

  <div style="overflow-x: auto;">
    <table class="preview-table">
      <thead>
        <tr>
          <th style="width: 40px;">
            <input type="checkbox" id="select-all-header" onchange="toggleAll(this)">
          </th>
          <th style="width: 40px;">#</th>
          <th>Domácí</th>
          {% if import_log.import_type == 'results' %}
          <th style="width: 80px; text-align: center;">Výsledek</th>
          {% endif %}
          <th>Hosté</th>
          <th style="width: 140px;">Datum/Čas</th>
          {% if import_log.import_type == 'results' %}
          <th style="width: 80px;">OT/SO</th>
          {% endif %}
        </tr>
      </thead>
      <tbody>
        {% for game in games %}
        <tr>
          <td>
            <input type="checkbox" name="selected_games" value="{{ loop.index0 }}"
                   class="game-checkbox" onchange="updateCount()">
          </td>
          <td>{{ loop.index }}</td>
          <td><strong>{{ game.home_team }}</strong></td>
          {% if import_log.import_type == 'results' %}
          <td style="text-align: center;">
            {% if game.home_score is not none %}
              <strong>{{ game.home_score }}:{{ game.away_score }}</strong>
            {% else %}
              <span class="muted">—</span>
            {% endif %}
          </td>
          {% endif %}
          <td><strong>{{ game.away_team }}</strong></td>
          <td class="muted">
            {% if game.start_time %}
              {{ game.start_time[:16].replace('T', ' ') }}
            {% else %}
              —
            {% endif %}
          </td>
          {% if import_log.import_type == 'results' %}
          <td>
            {% if game.overtime %}
              <span class="tag pill-warn">OT</span>
            {% elif game.shootout %}
              <span class="tag pill-warn">SO</span>
            {% else %}
              <span class="muted">—</span>
            {% endif %}
          </td>
          {% endif %}
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>

<script>
function toggleAll(checkbox) {
  const checkboxes = document.querySelectorAll('.game-checkbox');
  checkboxes.forEach(cb => cb.checked = checkbox.checked);
  updateCount();
}

function selectAll() {
  const checkboxes = document.querySelectorAll('.game-checkbox');
  checkboxes.forEach(cb => cb.checked = true);
  document.getElementById('select-all-header').checked = true;
  updateCount();
}

function deselectAll() {
  const checkboxes = document.querySelectorAll('.game-checkbox');
  checkboxes.forEach(cb => cb.checked = false);
  document.getElementById('select-all-header').checked = false;
  updateCount();
}

function updateCount() {
  const checked = document.querySelectorAll('.game-checkbox:checked').length;
  document.getElementById('selected-count').textContent = `Vybrané: ${checked}`;
}

// Auto-select all on load
window.addEventListener('DOMContentLoaded', function() {
  selectAll();
});
</script>

<div class="card" style="background: rgba(110,168,254,.08); border-color: rgba(110,168,254,.3);">
  <h3 style="margin: 0 0 12px 0;">ℹ️ Co se stane po potvrzení?</h3>

  {% if import_log.import_type == 'matches' %}
  <ul style="margin: 0; padding-left: 20px;">
    <li>Vytvoří se <strong>{{ import_log.imported_count }} nových zápasů</strong></li>
    <li>Týmy budou vytvořeny automaticky (pokud neexistují)</li>
    <li>{{ import_log.skipped_count }} existujících zápasů bude přeskočeno</li>
    <li>Vytvoří se mapování mezi API ID a našimi záznamy</li>
  </ul>
  {% else %}
  <ul style="margin: 0; padding-left: 20px;">
    <li>Aktualizuje se <strong>{{ import_log.imported_count }} výsledků zápasů</strong></li>
    {% if import_log.source.exclude_overtime %}
    <li><strong>Zápasy s OT/SO budou přeskočeny</strong> (kontroluj manuálně)</li>
    {% endif %}
    <li>{{ import_log.skipped_count }} zápasů už má výsledek (přeskočeno)</li>
  </ul>
  {% endif %}
</div>

  <div class="card">
    <div class="row" style="gap: 12px;">
      <button type="submit" name="action" value="approve" class="btn btn-primary"
              style="flex: 1;">
        ✅ Potvrdit a importovat
      </button>
      <button type="submit" name="action" value="reject" class="btn"
              style="flex: 1; background: rgba(255,77,109,.15); color: var(--danger);">
        ❌ Zamítnout
      </button>
    </div>
  </div>
</form>
""", import_log=import_log, games=games, errors=errors)



    @admin_bp.route("/api-import/history")
    @login_required
    def admin_api_import_history():
        """Historie importů"""
        admin_required()

        logs = APIImportLog.query.order_by(APIImportLog.id.desc()).limit(100).all()

        return render_page(r"""
<div class="card">
  <h2 style="margin: 0 0 8px 0;">📋 Historie importů</h2>
  <div class="muted">Posledních 100 importů</div>
</div>

{% if logs|length == 0 %}
<div class="card">
  <div style="text-align: center; padding: 40px;">
    <div style="font-size: 48px; margin-bottom: 12px;">📋</div>
    <h3 style="margin: 0;">Žádná historie</h3>
  </div>
</div>
{% else %}
  {% for log in logs %}
  <div class="card" style="margin-bottom: 12px;">
    <div class="row" style="justify-content: space-between; align-items: center;">
      <div>
        <div class="row" style="gap: 8px; margin-bottom: 4px;">
          {% if log.import_type == 'matches' %}
            <span class="tag">📥 Zápasy</span>
          {% else %}
            <span class="tag">📊 Výsledky</span>
          {% endif %}

          {% if log.status == 'completed' %}
            <span class="tag pill-ok">✅ Hotovo</span>
          {% elif log.status == 'pending' %}
            <span class="tag pill-warn">⏳ Čeká</span>
          {% elif log.status == 'rejected' %}
            <span class="tag pill-bad">❌ Zamítnuto</span>
          {% elif log.status == 'failed' %}
            <span class="tag pill-bad">⚠️ Chyba</span>
          {% endif %}
        </div>

        <div style="font-weight: 700; margin-bottom: 4px;">
          {{ log.source.round.name }} • {{ log.source.api_type.upper() }}
        </div>

        <div class="muted" style="font-size: 13px;">
          {{ log.created_at.strftime('%d.%m.%Y %H:%M') }}
          {% if log.approved_by %}
            • {{ log.approved_by.username }}
          {% endif %}
        </div>
      </div>

      <div style="text-align: right;">
        <div style="font-size: 20px; font-weight: 900; color: var(--ok);">
          +{{ log.imported_count }}
        </div>
        <div class="muted" style="font-size: 12px;">
          Přeskočeno: {{ log.skipped_count }}
          {% if log.error_count > 0 %} • Chyby: {{ log.error_count }}{% endif %}
        </div>
      </div>
    </div>
  </div>
  {% endfor %}
{% endif %}
""", logs=logs)

    # === EXPORT HUB ===


    @admin_bp.route("/export-hub", methods=["GET", "POST"])
    @login_required
    def admin_export_hub():
        """Centrální export hub s filtry"""
        admin_required()

        if request.method == "POST":
            # Co exportovat
            export_teams = request.form.get("export_teams") == "1"
            export_matches = request.form.get("export_matches") == "1"
            export_tips = request.form.get("export_tips") == "1"
            export_extras = request.form.get("export_extras") == "1"
            export_leaderboard = request.form.get("export_leaderboard") == "1"

            # Filtry
            round_id_str = request.form.get("round_id", "")
            round_id = int(round_id_str) if round_id_str and round_id_str != "0" else None
            user_id_str = request.form.get("user_id", "")
            user_id = int(user_id_str) if user_id_str and user_id_str != "0" else None
            only_finished = request.form.get("only_finished") == "1"
            include_deleted = request.form.get("include_deleted") == "1"
            format_type = request.form.get("format", "xlsx")

            if not any([export_teams, export_matches, export_tips, export_extras, export_leaderboard]):
                flash("Vyber alespoň jednu kategorii dat.", "error")
                return redirect(url_for("admin_export_hub"))

            # Vytvoř export
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from io import BytesIO
            import csv

            if format_type == "xlsx":
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

                # TÝMY
                if export_teams:
                    ws = wb.create_sheet("Týmy")
                    headers = ["ID", "Soutěž", "Název", "Skupina", "Země"]
                    header_fill = PatternFill("solid", fgColor="4472C4")
                    header_font = Font(color="FFFFFF", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = Team.query
                    if round_id:
                        query = query.filter_by(round_id=round_id)
                    if not include_deleted:
                        query = query.filter_by(is_deleted=False)

                    teams = query.all()
                    for row_i, team in enumerate(teams, 2):
                        ws.cell(row=row_i, column=1, value=team.id)
                        ws.cell(row=row_i, column=2, value=team.round.name if team.round else "")
                        ws.cell(row=row_i, column=3, value=team.name)
                        ws.cell(row=row_i, column=4, value=team.group or "")
                        ws.cell(row=row_i, column=5, value=team.country_code or "")

                    for col in range(1, 6):
                        ws.column_dimensions[chr(64 + col)].width = 20

                # ZÁPASY
                if export_matches:
                    ws = wb.create_sheet("Zápasy")
                    headers = ["ID", "Soutěž", "Datum", "Čas", "Domácí", "Hosté", "Skóre D", "Skóre H", "Stav"]
                    header_fill = PatternFill("solid", fgColor="217346")
                    header_font = Font(color="FFFFFF", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = Match.query
                    if round_id:
                        query = query.filter_by(round_id=round_id)
                    if not include_deleted:
                        query = query.filter_by(is_deleted=False)
                    if only_finished:
                        query = query.filter(Match.home_score != None, Match.away_score != None)

                    matches = query.order_by(Match.start_time.asc()).all()
                    for row_i, match in enumerate(matches, 2):
                        ws.cell(row=row_i, column=1, value=match.id)
                        ws.cell(row=row_i, column=2, value=match.round.name if match.round else "")
                        ws.cell(row=row_i, column=3, value=match.start_time.strftime("%Y-%m-%d") if match.start_time else "")
                        ws.cell(row=row_i, column=4, value=match.start_time.strftime("%H:%M") if match.start_time else "")
                        ws.cell(row=row_i, column=5, value=match.home_team.name if match.home_team else "")
                        ws.cell(row=row_i, column=6, value=match.away_team.name if match.away_team else "")
                        ws.cell(row=row_i, column=7, value=match.home_score if match.home_score is not None else "")
                        ws.cell(row=row_i, column=8, value=match.away_score if match.away_score is not None else "")

                        status = "Ukončen" if match.home_score is not None else ("Smazán" if match.is_deleted else "Naplánován")
                        ws.cell(row=row_i, column=9, value=status)

                    for col in range(1, 10):
                        ws.column_dimensions[chr(64 + col)].width = 15

                # TIPY
                if export_tips:
                    ws = wb.create_sheet("Tipy")
                    headers = ["ID", "Soutěž", "Uživatel", "Zápas", "Tip D", "Tip H", "Datum"]
                    header_fill = PatternFill("solid", fgColor="FFC000")
                    header_font = Font(color="000000", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = Tip.query.join(Match)
                    if round_id:
                        query = query.filter(Match.round_id == round_id)
                    if user_id:
                        query = query.filter(Tip.user_id == user_id)
                    if not include_deleted:
                        query = query.filter(Match.is_deleted == False)
                    if only_finished:
                        query = query.filter(Match.home_score != None)

                    tips = query.order_by(Tip.created_at.desc()).limit(10000).all()
                    for row_i, tip in enumerate(tips, 2):
                        ws.cell(row=row_i, column=1, value=tip.id)
                        ws.cell(row=row_i, column=2, value=tip.match.round.name if tip.match and tip.match.round else "")
                        ws.cell(row=row_i, column=3, value=tip.user.username if tip.user else "")
                        match_str = f"{tip.match.home_team.name} vs {tip.match.away_team.name}" if tip.match else ""
                        ws.cell(row=row_i, column=4, value=match_str)
                        ws.cell(row=row_i, column=5, value=tip.tip_home)
                        ws.cell(row=row_i, column=6, value=tip.tip_away)
                        ws.cell(row=row_i, column=7, value=tip.created_at.strftime("%Y-%m-%d %H:%M") if tip.created_at else "")

                    for col in range(1, 8):
                        ws.column_dimensions[chr(64 + col)].width = 20

                # EXTRA OTÁZKY
                if export_extras:
                    ws = wb.create_sheet("Extra")
                    headers = ["ID", "Soutěž", "Otázka", "Uzávěrka", "Odpovědí"]
                    header_fill = PatternFill("solid", fgColor="C65911")
                    header_font = Font(color="FFFFFF", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = ExtraQuestion.query
                    if round_id:
                        query = query.filter_by(round_id=round_id)
                    if not include_deleted:
                        query = query.filter_by(is_deleted=False)

                    questions = query.all()
                    for row_i, q in enumerate(questions, 2):
                        ws.cell(row=row_i, column=1, value=q.id)
                        ws.cell(row=row_i, column=2, value=q.round.name if q.round else "")
                        ws.cell(row=row_i, column=3, value=q.question)
                        ws.cell(row=row_i, column=4, value=q.deadline.strftime("%Y-%m-%d %H:%M") if q.deadline else "")
                        answer_count = ExtraAnswer.query.filter_by(question_id=q.id).count()
                        ws.cell(row=row_i, column=5, value=answer_count)

                    for col in range(1, 6):
                        ws.column_dimensions[chr(64 + col)].width = 25

                # ŽEBŘÍČEK
                if export_leaderboard and round_id:
                    r = db.session.get(Round, round_id)
                    if r:
                        ws = wb.create_sheet("Žebříček")
                        headers = ["Pořadí", "Uživatel", "Body", "Přesné", "Rozdíl", "Tendence", "Chybné"]
                        header_fill = PatternFill("solid", fgColor="70AD47")
                        header_font = Font(color="FFFFFF", bold=True)

                        for col, h in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col, value=h)
                            cell.fill = header_fill
                            cell.font = header_font

                        leaderboard = compute_leaderboard(r.id)
                        for row_i, entry in enumerate(leaderboard, 2):
                            ws.cell(row=row_i, column=1, value=entry.get("rank", ""))
                            ws.cell(row=row_i, column=2, value=entry.get("username", ""))
                            ws.cell(row=row_i, column=3, value=entry.get("total_points", 0))
                            ws.cell(row=row_i, column=4, value=entry.get("exact", 0))
                            ws.cell(row=row_i, column=5, value=entry.get("diff", 0))
                            ws.cell(row=row_i, column=6, value=entry.get("tend", 0))
                            ws.cell(row=row_i, column=7, value=entry.get("wrong", 0))

                        for col in range(1, 8):
                            ws.column_dimensions[chr(64 + col)].width = 15

                out = BytesIO()
                wb.save(out)
                out.seek(0)
                filename = f"export_{round_id or 'all'}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
                return send_file(out,
                               mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               download_name=filename,
                               as_attachment=True)

            else:  # CSV
                flash("CSV export coming soon!", "info")
                return redirect(url_for("admin_export_hub"))

        # GET: Formulář
        rounds = Round.query.order_by(Round.id.desc()).all()
        users = User.query.order_by(User.username.asc()).all()

        return render_page(r"""
<style>
.export-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
@media (max-width: 768px) { .export-grid { grid-template-columns: 1fr; } }
.export-category {
  padding: 16px; background: rgba(255,255,255,.03); border: 1px solid var(--line); border-radius: 10px;
}
.export-category:has(input:checked) { background: rgba(110,168,254,.1); border-color: rgba(110,168,254,.5); }
</style>

<div class="card">
  <h2 style="margin:0 0 4px 0;">📤 Export dat</h2>
  <div class="muted">Stáhni data z tipovačky ve strukturovaném formátu</div>
  <hr class="sep">

  <form method="post">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <h3 style="margin:20px 0 12px 0;">Co exportovat</h3>
    <div class="export-grid">
      <label class="export-category">
        <input type="checkbox" name="export_teams" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">🏟️ Týmy</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Názvy týmů, skupiny, kódy zemí</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_matches" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">⚽ Zápasy</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Rozpis zápasů, data, časy, výsledky</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_tips" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">🎯 Tipy</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Všechny tipy uživatelů</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_extras" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">❓ Extra otázky</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Otázky a odpovědi</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_leaderboard" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">🏆 Žebříček</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Aktuální pořadí a statistiky</div>
      </label>
    </div>

    <h3 style="margin:24px 0 12px 0;">Filtry</h3>
    <div class="row" style="gap:16px; flex-wrap:wrap;">
      <div class="form-group" style="flex:1; min-width:200px;">
        <label>Soutěž</label>
        <select name="round_id">
          <option value="">Všechny soutěže</option>
          {% for round in rounds %}
            <option value="{{ round.id }}">{{ round.name }}</option>
          {% endfor %}
        </select>
      </div>

      <div class="form-group" style="flex:1; min-width:200px;">
        <label>Uživatel (jen pro tipy)</label>
        <select name="user_id">
          <option value="">Všichni uživatelé</option>
          {% for user in users %}
            <option value="{{ user.id }}">{{ user.username }}</option>
          {% endfor %}
        </select>
      </div>
    </div>

    <div style="margin:12px 0; display:flex; flex-direction:column; gap:8px;">
      <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
        <input type="checkbox" name="only_finished" value="1" style="width:18px; height:18px;">
        <span>Jen dokončené zápasy</span>
      </label>

      <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
        <input type="checkbox" name="include_deleted" value="1" style="width:18px; height:18px;">
        <span>Včetně smazaných</span>
      </label>
    </div>

    <h3 style="margin:24px 0 12px 0;">Formát</h3>
    <div class="row" style="gap:16px;">
      <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
        <input type="radio" name="format" value="xlsx" checked style="width:18px; height:18px;">
        <span>📊 Excel (.xlsx)</span>
      </label>
    </div>

    <div class="row" style="gap:12px; margin-top:24px;">
      <button type="submit" class="btn btn-primary">📥 Stáhnout export</button>
      <a href="{{ url_for('admin_dashboard') }}" class="btn">← Zpět</a>
    </div>
  </form>
</div>
""", rounds=rounds, users=users)





"""
routes/admin_import.py
"""

import datetime
from io import BytesIO
import os
import csv
import pickle
import tempfile

from werkzeug.security import generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from reportlab.lib.pagesizes import A4
from flask import request, flash, redirect, url_for, abort, send_file, session
from flask_login import login_required

from models import ExtraAnswer, ExtraQuestion, Match, Round, Team, Tip, User
from app_utils import admin_required, audit, ensure_selected_round, parse_naive_datetime, render_page, set_selected_round_id
from extensions import db

def register_admin_import(app):
    @app.route("/admin/import")
    @login_required
    def admin_import():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Import (CSV)</h2>
      <div class="muted">Vybraná soutěž: <b>{{ r.name if r else "—" }}</b>. Pokud CSV obsahuje sloupec <b>round_id</b>, použije se.</div>
      <hr class="sep">
      <div class="row">
    <a class="btn btn-primary" href="{{ url_for('admin_import_teams') }}">Import týmů</a>
    <a class="btn btn-primary" href="{{ url_for('admin_import_matches') }}">Import zápasů</a>
    <a class="btn btn-primary" href="{{ url_for('admin_import_extras') }}">Import extra</a>
      </div>
      <hr class="sep">
      <div style="margin-top:20px;">
    <h3 style="margin:0 0 8px 0;">Import dat z jiné tipovačky</h3>
    <div class="muted" style="margin-bottom:10px;">Importuj kompletní data (uživatele, zápasy a tipy) z Excel žebříčku</div>
    <div class="row" style="gap: 12px;">
      <a class="btn btn-primary" href="{{ url_for('admin_import_leaderboard_smart') }}" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
        ✨ Smart Import (s kontrolou duplicit)
      </a>
      <a class="btn" href="{{ url_for('admin_import_leaderboard') }}">
        📥 Klasický Import (bez kontroly)
      </a>
    </div>
    <div class="muted" style="margin-top: 8px; font-size: 12px;">
      💡 <strong>Smart Import</strong> ti ukáže preview a označí možné duplicity před importem!
    </div>
      </div>
    </div>
    """, r=r)

    # --- ADMIN SMART IMPORT (s kontrolou duplicit) ---

    @app.route("/admin/import/leaderboard-smart/template")
    @login_required
    def admin_import_leaderboard_smart_template():
        """Stáhne Excel šablonu pro smart import žebříčku"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Smart Import"

            # Hlavička - Jméno | Zápasy
            headers = ['Jméno', 'Sparta-Slavia', 'Plzeň-Brno', 'Baník-Bohemians']
            ws.append(headers)

            # Stylování
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Příklady
            ws.append(['Jan Novák', '2:1', '1:1', '3:0'])
            ws.append(['Petr Svoboda', '1:2', '2:0', '1:1'])

            # Šířka
            ws.column_dimensions['A'].width = 20
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 15

            # Poznámky
            notes = wb.create_sheet("Poznámky")
            notes['A1'] = "SMART IMPORT - S KONTROLOU DUPLICIT"
            notes['A1'].font = Font(bold=True, size=14)
            notes['A3'] = "Formát stejný jako běžný import žebříčku."
            notes['A4'] = "Výhoda: Automaticky detekuje duplicitní zápasy!"
            notes['A6'] = "HLAVIČKA:"
            notes['A7'] = "   Jméno | Domácí-Hosté | Domácí-Hosté ..."
            notes['A9'] = "TIPY:"
            notes['A10'] = "   Jan | 2:1 | 1:0 | 3:2"
            notes.column_dimensions['A'].width = 50

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name='smart_import_sablona.xlsx')
        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_import"))

    @app.route("/admin/import/leaderboard-smart", methods=["GET", "POST"])
    @login_required
    def admin_import_leaderboard_smart():
        admin_required()

        if request.method == "POST":
            file = request.files.get('excel_file')
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_leaderboard_smart"))

            # Určit cílovou soutěž
            import_target = request.form.get('import_target', 'existing')

            if import_target == 'new':
                new_round_name = request.form.get('new_round_name', '').strip()
                if not new_round_name:
                    flash("Zadej název nové soutěže.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                sport = Sport.query.filter_by(name="Fotbal").first()
                if not sport:
                    sport = Sport(name="Fotbal")
                    db.session.add(sport)
                    db.session.flush()

                r = Round(name=new_round_name, sport_id=sport.id, is_active=False)
                db.session.add(r)
                db.session.flush()
                round_id = r.id
                audit("round.create.import", "Round", r.id, name=r.name)
            else:
                round_id = int(request.form.get('round_id', 0))
                r = db.session.get(Round, round_id)
                if not r:
                    flash("Vybraná soutěž neexistuje.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

            try:
                import openpyxl
                from io import BytesIO

                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active

                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    flash("Soubor musí mít alespoň hlavičku a jeden řádek.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                # Parsovat zápasy z hlavičky
                matches_data = []
                for i, match_name in enumerate(header[1:], start=1):
                    if not match_name or str(match_name).strip() == "":
                        continue

                    match_str = str(match_name).strip()
                    for sep in ['-', ' vs ', ' x ', ':', ' – ']:
                        if sep in match_str:
                            parts = match_str.split(sep, 1)
                            if len(parts) == 2:
                                home = parts[0].strip()
                                away = parts[1].strip()

                                # Detekovat možnou duplicitu
                                home_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == home.lower()
                                ).first()

                                away_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == away.lower()
                                ).first()

                                is_duplicate = False
                                existing_match_id = None
                                status = 'new'  # NOVÉ!

                                if home_team_check and away_team_check:
                                    existing_match = Match.query.filter_by(
                                        round_id=round_id,
                                        home_team_id=home_team_check.id,
                                        away_team_id=away_team_check.id,
                                        is_deleted=False
                                    ).first()

                                    if existing_match:
                                        is_duplicate = True
                                        existing_match_id = existing_match.id
                                        status = 'overwrite'  # Může být přepsán!

                                matches_data.append({
                                    'col': i,
                                    'home': home,
                                    'away': away,
                                    'is_duplicate': is_duplicate,  # Zpětná kompatibilita
                                    'status': status,  # NOVÉ!
                                    'existing_match_id': existing_match_id,
                                    'match_str': match_str
                                })
                                break

                if not matches_data:
                    flash("Nenalezeny žádné platné zápasy v hlavičce.", "error")
                    return redirect(url_for("admin_import_leaderboard_smart"))

                # Uložit preview data do TEMP FILE (ne session - cookie overflow!)
                import tempfile
                import pickle

                preview_data = {
                    'round_id': round_id,
                    'round_name': r.name,
                    'matches': matches_data,  # Velké data!
                    'total_matches': len(matches_data),
                    'overwrite_count': sum(1 for m in matches_data if m['status'] == 'overwrite'),
                    'new_count': sum(1 for m in matches_data if m['status'] == 'new')
                }

                # Uložit preview data do pickle file
                preview_fd, preview_path = tempfile.mkstemp(suffix='.pkl', prefix='smart_preview_')
                with os.fdopen(preview_fd, 'wb') as f:
                    pickle.dump(preview_data, f)

                # Do session jen cesta + základní info (malé!)
                session['smart_import_preview'] = {
                    'preview_file': preview_path,
                    'round_id': round_id,
                    'round_name': r.name,
                    'total_matches': len(matches_data),
                    'overwrite_count': sum(1 for m in matches_data if m['status'] == 'overwrite'),
                    'new_count': sum(1 for m in matches_data if m['status'] == 'new')
                }

                # Uložit Excel file do TEMP FILE
                file.seek(0)  # Reset file pointer
                excel_fd, excel_path = tempfile.mkstemp(suffix='.xlsx', prefix='smart_excel_')
                with os.fdopen(excel_fd, 'wb') as f:
                    f.write(file.read())
                session['smart_import_file'] = excel_path

                return redirect(url_for("admin_import_leaderboard_smart_preview"))

            except Exception as e:
                flash(f"Chyba při načítání souboru: {str(e)}", "error")
                db.session.rollback()
                return redirect(url_for("admin_import_leaderboard_smart"))

        # GET
        all_rounds = Round.query.order_by(Round.id.desc()).all()

        return render_page(r"""
    <div class="card">
      <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <div>
      <h2 style="margin: 0 0 8px 0;">✨ Smart Import Žebříčku</h2>
      <div class="muted">Import s kontrolou duplicit a možností výběru</div>
    </div>
    <a href="{{ url_for('admin_import_leaderboard_smart_template') }}" class="btn" style="background:#667eea; color:white;">
      📥 Stáhnout šablonu Excel
    </a>
      </div>
      <hr class="sep">

      <div class="card" style="background: rgba(110,168,254,.08); border-color: rgba(110,168,254,.3); margin-bottom: 16px;">
    <h3 style="margin: 0 0 12px 0;">💡 Jak to funguje?</h3>
    <ul style="margin: 0; padding-left: 20px;">
      <li><strong>Preview před importem</strong> - Vidíš CO se chystá importovat</li>
      <li><strong>Detekce existujících</strong> - Modře označí zápasy které již existují (mohou být použity)</li>
      <li><strong>Checkboxy pro výběr</strong> - Ručně vyber co chceš/nechceš importovat</li>
      <li><strong>Quick select</strong> - Tlačítka "Jen nové" / "Přepsat existující"</li>
    </ul>
      </div>

      <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; align-items:stretch; gap:16px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">1️⃣ Kam importovat data?</h3>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="existing" {% if all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_existing">
        <span>Do existující soutěže</span>
      </label>

      <div id="existing_round_select" style="margin-left:28px; margin-bottom:16px;">
        <select name="round_id" style="width:100%; max-width:400px;">
          {% for rnd in all_rounds %}
            <option value="{{ rnd.id }}">
              {% if rnd.is_active %}★ {% endif %}{{ rnd.name }}
            </option>
          {% endfor %}
          {% if not all_rounds %}
            <option value="">-- Žádné soutěže --</option>
          {% endif %}
        </select>
      </div>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="new" {% if not all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_new">
        <span>Vytvořit novou soutěž</span>
      </label>

      <div id="new_round_input" style="margin-left:28px; display:none;">
        <input type="text" name="new_round_name" placeholder="Název nové soutěže"
               style="width:100%; max-width:400px;">
      </div>
    </div>

    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">2️⃣ Vyber Excel soubor</h3>
      <input type="file" name="excel_file" accept=".xlsx,.xls" required>
    </div>

    <button class="btn btn-primary" type="submit" style="padding:14px; font-size:16px; font-weight:900;">
      📥 Načíst preview
    </button>
    <a class="btn" href="{{ url_for('admin_import') }}">Zpět</a>
      </form>
    </div>

    <script>
    function toggleImportTarget() {
      const existingChecked = document.getElementById('radio_existing').checked;
      const existingDiv = document.getElementById('existing_round_select');
      const newDiv = document.getElementById('new_round_input');

      if (existingChecked) {
    existingDiv.style.display = 'block';
    newDiv.style.display = 'none';
      } else {
    existingDiv.style.display = 'none';
    newDiv.style.display = 'block';
      }
    }

    toggleImportTarget();
    </script>
    """, all_rounds=all_rounds)

    @app.route("/admin/import/leaderboard-smart/preview")
    @login_required
    def admin_import_leaderboard_smart_preview():
        admin_required()

        preview_meta = session.get('smart_import_preview')
        if not preview_meta:
            flash("Nejprve nahraj soubor.", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

        # Načti matches data z pickle file
        preview_file = preview_meta.get('preview_file')
        if not preview_file or not os.path.exists(preview_file):
            flash("Preview data expirovala, nahraj soubor znovu.", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

        import pickle
        with open(preview_file, 'rb') as f:
            preview = pickle.load(f)

        return render_page(r"""
    <style>
      .match-preview {
    display: flex;
    align-items: center;
    padding: 12px;
    border: 1px solid var(--line);
    border-radius: 8px;
    margin-bottom: 8px;
    transition: all 0.2s ease;
      }

      .match-preview:hover {
    background: rgba(255,255,255,.03);
      }

      .match-preview.overwrite {
    background: rgba(99,179,237,.08);
    border-color: rgba(99,179,237,.3);
      }

      .match-preview.new {
    background: rgba(51,209,122,.05);
    border-color: rgba(51,209,122,.2);
      }

      .match-checkbox {
    margin-right: 12px;
    width: 20px;
    height: 20px;
    cursor: pointer;
      }
    </style>

    <div class="card">
      <h2 style="margin: 0 0 8px 0;">✨ Preview importu</h2>
      <div class="muted">Soutěž: <b>{{ preview.round_name }}</b></div>
      <hr class="sep">

      <div class="row" style="justify-content: space-between; margin-bottom: 16px;">
    <div>
      <div class="tag pill-ok" style="font-size: 14px;">✅ {{ preview.new_count }} nových</div>
      <div class="tag pill-info" style="font-size: 14px; background: rgba(99,179,237,.15); color: #63b3ed;">🔵 {{ preview.overwrite_count }} přepsat</div>
    </div>

    <div class="row" style="gap: 8px;">
      <button type="button" onclick="selectAll()" class="btn btn-sm">☑️ Vybrat vše</button>
      <button type="button" onclick="deselectAll()" class="btn btn-sm">☐ Zrušit výběr</button>
      <button type="button" onclick="selectNew()" class="btn btn-sm">🟢 Jen nové</button>
      <button type="button" onclick="selectOverwrite()" class="btn btn-sm btn-primary">🔵 Přepsat existující</button>
    </div>
      </div>
    </div>

    <form method="post" action="{{ url_for('admin_import_leaderboard_smart_confirm') }}">
      <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
      <div class="card">
    <h3 style="margin: 0 0 16px 0;">📋 Zápasy k importu ({{ preview.total_matches }})</h3>

    {% for match in preview.matches %}
      <div class="match-preview {{ match.status }}" data-status="{{ match.status }}">
        <input type="checkbox" 
               name="selected_matches" 
               value="{{ loop.index0 }}"
               class="match-checkbox"
               {% if match.status == 'new' %}checked{% endif %}
               id="match_{{ loop.index0 }}">

        <label for="match_{{ loop.index0 }}" style="flex: 1; cursor: pointer; margin: 0;">
          <div class="row" style="justify-content: space-between; align-items: center;">
            <div>
              <strong>{{ match.home }}</strong> 
              <span class="muted">vs</span> 
              <strong>{{ match.away }}</strong>
            </div>

            <div>
              {% if match.status == 'overwrite' %}
                <span class="tag pill-info" style="background: rgba(99,179,237,.15); color: #63b3ed;">🔵 PŘEPSAT</span>
              {% else %}
                <span class="tag pill-ok">✅ NOVÝ</span>
              {% endif %}
            </div>
          </div>
        </label>
      </div>
    {% endfor %}
      </div>

      <div class="card">
    <div class="row" style="justify-content: space-between; align-items: center;">
      <div>
        <div style="font-size: 14px; margin-bottom: 4px;">
          Importuje se: <strong><span id="selected-count">{{ preview.new_count }}</span></strong> zápasů
        </div>
        <div class="muted" style="font-size: 12px;">
          Duplicity budou přeskočeny
        </div>
      </div>

      <div class="row" style="gap: 12px;">
        <a href="{{ url_for('admin_import_leaderboard_smart') }}" class="btn">❌ Zrušit</a>
        <button type="submit" class="btn btn-primary" style="font-size: 16px; padding: 12px 24px;">
          ✅ Potvrdit a importovat
        </button>
      </div>
    </div>
      </div>
    </form>

    <script>
    function updateCount() {
      const checkboxes = document.querySelectorAll('.match-checkbox:checked');
      document.getElementById('selected-count').textContent = checkboxes.length;
    }

    function selectAll() {
      document.querySelectorAll('.match-checkbox').forEach(cb => {
    cb.checked = true;
      });
      updateCount();
    }

    function deselectAll() {
      document.querySelectorAll('.match-checkbox').forEach(cb => {
    cb.checked = false;
      });
      updateCount();
    }

    function selectNew() {
      deselectAll();
      document.querySelectorAll('.match-preview').forEach(preview => {
    const status = preview.getAttribute('data-status');
    const checkbox = preview.querySelector('.match-checkbox');
    if (status === 'new' && checkbox) {
      checkbox.checked = true;
    }
      });
      updateCount();
    }

    function selectOverwrite() {
      deselectAll();
      document.querySelectorAll('.match-preview').forEach(preview => {
    const status = preview.getAttribute('data-status');
    const checkbox = preview.querySelector('.match-checkbox');
    if (status === 'overwrite' && checkbox) {
      checkbox.checked = true;
    }
      });
      updateCount();
    }

    // Update count on checkbox change
    document.querySelectorAll('.match-checkbox').forEach(cb => {
      cb.addEventListener('change', updateCount);
    });
    </script>
    """, preview=preview)

    def _resolve_round_id(row: dict[str, str], fallback_round_id: int) -> int:
        val = (row.get("round_id") or row.get("round") or "").strip()
        if val:
            try:
                return int(val)
            except Exception:
                pass
        return fallback_round_id

    @app.route("/admin/import/leaderboard-smart/confirm", methods=["POST"])
    @login_required
    def admin_import_leaderboard_smart_confirm():
        admin_required()

        preview_meta = session.get('smart_import_preview')
        temp_file = session.get('smart_import_file')

        if not preview_meta or not temp_file:
            flash("Session vypršela, nahraj soubor znovu.", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

        # Zkontroluj jestli temp files existují
        preview_file = preview_meta.get('preview_file')
        if not os.path.exists(temp_file) or not os.path.exists(preview_file):
            flash("Soubor expiroval, nahraj znovu.", "error")
            session.pop('smart_import_file', None)
            session.pop('smart_import_preview', None)
            return redirect(url_for("admin_import_leaderboard_smart"))

        try:
            # Načti preview data z pickle file
            import pickle
            with open(preview_file, 'rb') as f:
                preview = pickle.load(f)

            # Načti vybrané indexy
            selected_indices = request.form.getlist('selected_matches')
            selected_indices = [int(i) for i in selected_indices]

            if not selected_indices:
                flash("Nevybral jsi žádné zápasy k importu.", "error")
                return redirect(url_for("admin_import_leaderboard_smart_preview"))

            # Načti Excel file z TEMP FILE
            import openpyxl

            wb = openpyxl.load_workbook(temp_file)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]

            round_id = preview['round_id']
            r = db.session.get(Round, round_id)

            if not r:
                flash("Soutěž nenalezena.", "error")
                return redirect(url_for("admin_import_leaderboard_smart"))

            # Importuj pouze vybrané zápasy (vytvoř nové nebo použij existující)
            created_matches = 0
            skipped_matches = 0
            match_map = {}

            # KROK 1: Vytvoř/použij vybrané zápasy
            for idx in selected_indices:
                if idx >= len(preview['matches']):
                    continue

                match_data = preview['matches'][idx]

                home_team = _get_or_create_team(r.id, match_data['home'])
                away_team = _get_or_create_team(r.id, match_data['away'])

                # Zkusit najít existující zápas
                m = Match.query.filter_by(
                    round_id=r.id,
                    home_team_id=home_team.id,
                    away_team_id=away_team.id,
                    is_deleted=False
                ).first()

                if not m:
                    m = Match(round_id=r.id, home_team_id=home_team.id, away_team_id=away_team.id)
                    db.session.add(m)
                    db.session.flush()
                    created_matches += 1
                else:
                    skipped_matches += 1

                match_map[match_data['col']] = m

            # KROK 2: Pro NEVybrané zápasy - najdi existující (pro import tipů)
            for idx, match_data in enumerate(preview['matches']):
                if idx in selected_indices:
                    continue  # Už je v match_map

                # Zkus najít existující zápas (bez vytváření nového)
                # CASE-INSENSITIVE vyhledávání!
                home_team = Team.query.filter(
                    Team.round_id == r.id,
                    db.func.lower(Team.name) == match_data['home'].lower(),
                    Team.is_deleted == False
                ).first()

                away_team = Team.query.filter(
                    Team.round_id == r.id,
                    db.func.lower(Team.name) == match_data['away'].lower(),
                    Team.is_deleted == False
                ).first()

                if home_team and away_team:
                    existing_match = Match.query.filter_by(
                        round_id=r.id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if existing_match:
                        # Přidej do match_map pro import tipů
                        match_map[match_data['col']] = existing_match

            # Import tipů (stejně jako v klasickém importu, ale jen pro vybrané zápasy)
            results_imported = 0
            data_rows_start = 1
            data_rows_end = len(rows)

            # Kontrola výsledků v prvním řádku
            if len(rows) > 1:
                first_data_row = rows[1]
                first_cell = str(first_data_row[0]).strip() if first_data_row[0] else ""

                if first_cell.lower() in ['', 'výsledek', 'result', 'skóre', 'score', 'vysledek']:
                    for col_idx, match in match_map.items():
                        if col_idx >= len(first_data_row):
                            continue

                        result_value = first_data_row[col_idx]
                        if not result_value:
                            continue

                        home_score = None
                        away_score = None

                        if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                            home_score = result_value.hour if result_value.hour < 20 else None
                            away_score = result_value.minute if result_value.minute < 60 else None
                        elif isinstance(result_value, str) and ':' in result_value:
                            parts = result_value.split(':')
                            if len(parts) == 2:
                                try:
                                    home_score = int(parts[0].strip())
                                    away_score = int(parts[1].strip())
                                except:
                                    pass

                        if home_score is not None and away_score is not None:
                            match.home_score = home_score
                            match.away_score = away_score
                            results_imported += 1

                    data_rows_start = 2

            # Kontrola výsledků v posledním řádku
            if len(rows) > data_rows_start:
                last_row = rows[-1]
                last_cell = str(last_row[0]).strip() if last_row[0] else ""

                if last_cell.lower() in ['výsledek', 'result', 'skóre', 'score', 'vysledek']:
                    for col_idx, match in match_map.items():
                        if col_idx >= len(last_row):
                            continue

                        result_value = last_row[col_idx]
                        if not result_value:
                            continue

                        home_score = None
                        away_score = None

                        if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                            home_score = result_value.hour if result_value.hour < 20 else None
                            away_score = result_value.minute if result_value.minute < 60 else None
                        elif isinstance(result_value, str) and ':' in result_value:
                            parts = result_value.split(':')
                            if len(parts) == 2:
                                try:
                                    home_score = int(parts[0].strip())
                                    away_score = int(parts[1].strip())
                                except:
                                    pass

                        if home_score is not None and away_score is not None:
                            match.home_score = home_score
                            match.away_score = away_score
                            results_imported += 1

                    data_rows_end = len(rows) - 1

            # Import tipů
            users_created = 0
            tips_imported = 0
            tips_overwritten = 0

            for row_idx in range(data_rows_start, data_rows_end):
                if row_idx >= len(rows):
                    break

                row = rows[row_idx]
                if not row or len(row) < 2:
                    continue

                username = str(row[0]).strip() if row[0] else ""
                if not username or username.lower() in ['výsledek', 'result', 'skóre', 'score', 'vysledek']:
                    continue

                # Normalize Unicode
                import unicodedata
                username = unicodedata.normalize('NFC', username)

                # Flush aby jsme viděli nově vytvořené uživatele
                db.session.flush()

                # PYTHON-based case-insensitive search (SPOLEHLIVĚJŠÍ!)
                all_users = User.query.all()
                u = None
                username_lower = username.lower()
                for user in all_users:
                    user_norm = unicodedata.normalize('NFC', user.username)
                    if user_norm.lower() == username_lower:
                        u = user
                        break

                if not u:
                    u = User(
                        username=username,
                        email=f"{username.lower()}@imported.local",
                        password_hash=generate_password_hash("changeme123"),
                        is_admin=False
                    )
                    db.session.add(u)
                    db.session.flush()
                    users_created += 1

                # Import tipů pro vybrané zápasy
                for col_idx, match in match_map.items():
                    if col_idx >= len(row):
                        continue

                    tip_value = row[col_idx]
                    if not tip_value:
                        continue

                    home_guess = None
                    away_guess = None

                    if hasattr(tip_value, 'hour') and hasattr(tip_value, 'minute'):
                        home_guess = tip_value.hour if tip_value.hour < 20 else None
                        away_guess = tip_value.minute if tip_value.minute < 60 else None
                    elif isinstance(tip_value, str) and ':' in tip_value:
                        parts = tip_value.split(':')
                        if len(parts) == 2:
                            try:
                                home_guess = int(parts[0].strip())
                                away_guess = int(parts[1].strip())
                            except:
                                pass

                    if home_guess is not None and away_guess is not None:
                        existing_tip = Tip.query.filter_by(user_id=u.id, match_id=match.id).first()
                        if existing_tip:
                            # PŘEPSAT existující tip novými daty
                            existing_tip.tip_home = home_guess
                            existing_tip.tip_away = away_guess
                            tips_overwritten += 1
                        else:
                            # Vytvořit nový tip
                            tip = Tip(
                                user_id=u.id,
                                match_id=match.id,
                                tip_home=home_guess,
                                tip_away=away_guess
                            )
                            db.session.add(tip)
                            tips_imported += 1

            db.session.commit()

            # Smaž temp files
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass

            if preview_file and os.path.exists(preview_file):
                try:
                    os.unlink(preview_file)
                except:
                    pass

            # Vyčisti session
            session.pop('smart_import_preview', None)
            session.pop('smart_import_file', None)

            audit("leaderboard.smart_import", "Round", r.id,
                  created_matches=created_matches,
                  skipped_matches=skipped_matches,
                  users_created=users_created,
                  tips_imported=tips_imported,
                  tips_overwritten=tips_overwritten,
                  results_imported=results_imported)

            tips_message = f"Importováno {tips_imported} nových tipů"
            if tips_overwritten > 0:
                tips_message += f", přepsáno {tips_overwritten} existujících"

            flash(f"""✅ Smart Import dokončen!
                  Vytvořeno {created_matches} nových zápasů
                  Použito {skipped_matches} existujících zápasů
                  Vytvořeno {users_created} nových uživatelů
                  {tips_message}
                  Importováno {results_imported} výsledků""", "ok")

            return redirect(url_for("admin_import"))

        except Exception as e:
            db.session.rollback()

            # Smaž temp files i při chybě
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass

            if preview_file and os.path.exists(preview_file):
                try:
                    os.unlink(preview_file)
                except:
                    pass

            flash(f"Chyba při importu: {str(e)}", "error")
            return redirect(url_for("admin_import_leaderboard_smart"))

    def _get_or_create_team(round_id: int, name: str) -> Team:
        import unicodedata

        name = (name or "").strip()
        if not name:
            raise ValueError("empty team")

        # Normalize Unicode (pro české znaky Š, Č, Ř, ...)
        name = unicodedata.normalize('NFC', name)

        # Flush změny aby jsme viděli všechny týmy včetně nově vytvořených
        db.session.flush()

        # PYTHON-based case-insensitive search (NEJSPOLEHLIVĚJŠÍ!)
        # Načti všechny týmy pro tento round
        all_teams = Team.query.filter_by(
            round_id=round_id,
            is_deleted=False
        ).all()

        # Najdi case-insensitive match s Unicode normalizací
        name_lower = name.lower()
        for t in all_teams:
            t_normalized = unicodedata.normalize('NFC', t.name)
            if t_normalized.lower() == name_lower:
                return t  # Našli jsme!

        # Tým neexistuje - vytvoř nový
        t = Team(round_id=round_id, name=name)
        db.session.add(t)
        db.session.flush()  # Flush aby byl dostupný pro další volání

        audit("team.create.auto", "Team", t.id, round_id=round_id, name=name)
        return t

    @app.route("/admin/import/teams/template")
    @login_required
    def admin_import_teams_template():
        """Stáhni Excel šablonu pro import týmů"""
        admin_required()
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Týmy"

        # Header
        headers = ["name", "group", "country_code"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Příklady
        examples = [
            ("Sparta Praha", "A", "CZ"),
            ("Slavia Praha", "A", "CZ"),
            ("Barcelona", "B", "ES"),
            ("Real Madrid", "B", "ES"),
        ]

        gray_fill = PatternFill("solid", fgColor="F2F2F2")
        for row_i, (name, grp, code) in enumerate(examples, 2):
            fill = gray_fill if row_i % 2 == 0 else PatternFill()
            for col, val in enumerate([name, grp, code], 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.fill = fill

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15

        # Popis
        ws2 = wb.create_sheet("Popis")
        ws2["A1"], ws2["B1"] = "Sloupec", "Popis"
        ws2["A1"].font = ws2["B1"].font = Font(bold=True)

        popis = [
            ("name", "POVINNÉ - název týmu"),
            ("group", "VOLITELNÉ - skupina (A, B, C...)"),
            ("country_code", "VOLITELNÉ - kód země (CZ, SK, DE...)"),
        ]
        for row_i, (col, desc) in enumerate(popis, 2):
            ws2[f"A{row_i}"] = col
            ws2[f"B{row_i}"] = desc
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 40

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        from flask import send_file
        return send_file(out,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="sablona_tymy.xlsx",
                         as_attachment=True)

    @app.route("/admin/import/teams", methods=["GET", "POST"])
    @login_required
    def admin_import_teams():
        admin_required()
        import os, pickle
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            abort(400)

        if request.method == "POST":
            action = request.form.get("action", "upload")

            # ── KROK 2: POTVRZENÍ ──────────────────────────────────────────
            if action == "confirm":
                preview_file = session.get("teams_import_preview_file")
                round_id = session.get("teams_import_round_id")
                overwrite = session.get("teams_import_overwrite", False)

                if not preview_file or not os.path.exists(preview_file):
                    flash("Session vypršela, nahraj soubor znovu.", "error")
                    return redirect(url_for("admin_import_teams"))

                r2 = db.session.get(Round, round_id) if round_id else r

                with open(preview_file, "rb") as f:
                    rows_data = pickle.load(f)
                os.unlink(preview_file)
                session.pop("teams_import_preview_file", None)
                session.pop("teams_import_round_id", None)
                session.pop("teams_import_overwrite", None)

                created = skipped = updated = 0
                for row in rows_data:
                    if row["status"] == "skip":
                        skipped += 1
                        continue

                    name = row["name"]
                    existing = Team.query.filter_by(
                        round_id=r2.id, is_deleted=False
                    ).filter(db.func.lower(Team.name) == name.lower()).first()

                    if existing:
                        if overwrite:
                            # Aktualizuj group/country_code pokud je v importu
                            if row.get("group"):
                                existing.group = row["group"]
                            if row.get("country_code"):
                                existing.country_code = row["country_code"]
                            db.session.flush()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        team = Team(
                            round_id=r2.id,
                            name=name,
                            group=row.get("group"),
                            country_code=row.get("country_code")
                        )
                        db.session.add(team)
                        db.session.flush()
                        created += 1

                db.session.commit()
                audit("import.teams.excel", "Round", r2.id, created=created, updated=updated, skipped=skipped)
                flash(f"Import týmů hotov: vytvořeno {created}, aktualizováno {updated}, přeskočeno {skipped}.", "ok")
                return redirect(url_for("teams"))

            # ── KROK 1: UPLOAD + PREVIEW ───────────────────────────────────
            import openpyxl, tempfile
            from io import BytesIO

            file = request.files.get("excel_file")
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_teams"))

            overwrite = request.form.get("overwrite") == "1"

            try:
                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))

                if len(rows) < 2:
                    flash("Soubor musí mít hlavičku a alespoň jeden řádek.", "error")
                    return redirect(url_for("admin_import_teams"))

                header = [str(h).strip().lower() if h else "" for h in rows[0]]

                # Najdi sloupce
                def col(name):
                    return header.index(name) if name in header else None

                name_col = col("name")
                if name_col is None:
                    flash("Chybí sloupec 'name'. Zkontroluj šablonu.", "error")
                    return redirect(url_for("admin_import_teams"))

                grp_col = col("group")
                code_col = col("country_code")

                # Existující týmy
                existing_teams = Team.query.filter_by(
                    round_id=r.id, is_deleted=False
                ).all()
                existing_lower = {t.name.lower(): t for t in existing_teams}

                preview_rows = []
                for data_row in rows[1:]:
                    name = str(data_row[name_col]).strip() if data_row[name_col] else ""
                    if not name or name.lower() == "none":
                        continue

                    group = str(data_row[grp_col]).strip() if grp_col is not None and data_row[grp_col] else None
                    country_code = str(data_row[code_col]).strip() if code_col is not None and data_row[code_col] else None

                    # Status
                    if name.lower() in existing_lower:
                        status = "overwrite" if overwrite else "skip"
                    else:
                        status = "new"

                    preview_rows.append({
                        "name": name,
                        "group": group,
                        "country_code": country_code,
                        "status": status,
                    })

                if not preview_rows:
                    flash("Soubor neobsahuje žádné platné týmy.", "error")
                    return redirect(url_for("admin_import_teams"))

                # Ulož do temp file
                fd, preview_path = tempfile.mkstemp(suffix=".pkl", prefix="teams_import_")
                with os.fdopen(fd, "wb") as f:
                    pickle.dump(preview_rows, f)

                session["teams_import_preview_file"] = preview_path
                session["teams_import_round_id"] = r.id
                session["teams_import_overwrite"] = overwrite

                new_count = sum(1 for x in preview_rows if x["status"] == "new")
                overwrite_count = sum(1 for x in preview_rows if x["status"] == "overwrite")
                skip_count = sum(1 for x in preview_rows if x["status"] == "skip")

                return render_page(r"""
    <style>
    .preview-table { width: 100%; border-collapse: collapse; font-size: 13px; }
    .preview-table th { background: rgba(255,255,255,.07); padding: 10px 12px; text-align: left; }
    .preview-table td { padding: 10px 12px; border-bottom: 1px solid var(--line); }
    .badge-new    { background: rgba(51,209,122,.2);  color: #33d17a; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }
    .badge-skip   { background: rgba(167,178,214,.15); color: #a7b2d6; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }
    .badge-over   { background: rgba(110,168,254,.2);  color: #6ea8fe; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }
    </style>

    <div class="card">
      <h2 style="margin:0 0 4px 0;">📋 Preview – Týmy</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>

      <div class="row" style="gap:16px; margin: 16px 0; flex-wrap:wrap;">
    <div style="padding:12px 20px; background:rgba(51,209,122,.1); border:1px solid rgba(51,209,122,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#33d17a;">{{ new_count }}</div>
      <div class="muted" style="font-size:12px;">Nových</div>
    </div>
    <div style="padding:12px 20px; background:rgba(110,168,254,.1); border:1px solid rgba(110,168,254,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#6ea8fe;">{{ overwrite_count }}</div>
      <div class="muted" style="font-size:12px;">Aktualizovaných</div>
    </div>
    <div style="padding:12px 20px; background:rgba(167,178,214,.1); border:1px solid rgba(167,178,214,.2); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#a7b2d6;">{{ skip_count }}</div>
      <div class="muted" style="font-size:12px;">Přeskočených</div>
    </div>
      </div>

      <div style="overflow-x:auto; margin-bottom:20px;">
    <table class="preview-table">
      <thead>
        <tr>
          <th>#</th>
          <th>Název týmu</th>
          <th>Skupina</th>
          <th>Země</th>
          <th>Status</th>
        </tr>
      </thead>
      <tbody>
        {% for row in preview_rows %}
        <tr>
          <td class="muted">{{ loop.index }}</td>
          <td>{{ row.name }}</td>
          <td class="muted">{{ row.group or '—' }}</td>
          <td class="muted">{{ row.country_code or '—' }}</td>
          <td>
            {% if row.status == 'new' %}
              <span class="badge-new">✨ Nový</span>
            {% elif row.status == 'overwrite' %}
              <span class="badge-over">✏️ Aktualizovat</span>
            {% else %}
              <span class="badge-skip">⏭ Přeskočit</span>
            {% endif %}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
      </div>

      <form method="post" style="display:flex; gap:12px; flex-wrap:wrap;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="action" value="confirm">
    <button type="submit" class="btn btn-primary">✅ Potvrdit import</button>
    <a href="{{ url_for('admin_import_teams') }}" class="btn">✕ Zrušit</a>
      </form>
    </div>
    """, r=r, preview_rows=preview_rows,
     new_count=new_count, overwrite_count=overwrite_count, skip_count=skip_count)

            except Exception as e:
                flash(f"Chyba při čtení souboru: {e}", "error")
                return redirect(url_for("admin_import_teams"))

        # === GET: Zobraz formulář ===
        return render_page(r"""
    <style>
    .upload-zone {
      border: 2px dashed rgba(110,168,254,0.4);
      border-radius: 12px;
      padding: 40px;
      text-align: center;
      background: rgba(110,168,254,0.05);
      cursor: pointer;
      transition: all 0.2s;
    }
    .upload-zone:hover {
      border-color: rgba(110,168,254,0.7);
      background: rgba(110,168,254,0.1);
    }
    </style>

    <div class="card">
      <div class="row" style="justify-content:space-between; margin-bottom:20px;">
    <div>
      <h2 style="margin:0 0 4px 0;">📥 Import týmů (Excel)</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <a href="{{ url_for('admin_import_teams_template') }}" class="btn"
       style="background:rgba(51,209,122,.15); color:#33d17a; border:1px solid rgba(51,209,122,.3);">
      📄 Stáhnout šablonu
    </a>
      </div>

      <div class="muted" style="margin-bottom:20px; padding:12px 16px; background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); border-radius:8px;">
    💡 <strong>Formát Excel (sloupce):</strong><br>
    <code>name</code> – povinný, název týmu<br>
    <code>group</code> – volitelný, skupina (A, B, C...)<br>
    <code>country_code</code> – volitelný, kód země (CZ, SK, DE...)
      </div>

      <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="upload-zone" onclick="document.getElementById('xl').click()">
      <div style="font-size:48px; margin-bottom:12px;">📊</div>
      <div style="font-weight:700; margin-bottom:8px;">Klikni nebo přetáhni Excel soubor</div>
      <div class="muted">.xlsx • max 10 MB</div>
      <input type="file" id="xl" name="excel_file" accept=".xlsx" style="display:none"
             onchange="document.getElementById('fname').textContent = this.files[0]?.name || ''">
    </div>
    <div id="fname" class="muted" style="margin-bottom:16px; text-align:center;"></div>

    <div style="padding:14px; background:rgba(110,168,254,.08); border:1px solid rgba(110,168,254,.2); border-radius:8px; margin-bottom:20px;">
      <label style="display:flex; align-items:center; gap:10px; cursor:pointer;">
        <input type="checkbox" name="overwrite" value="1" style="width:18px; height:18px;">
        <div>
          <strong>Přepsat existující týmy</strong>
          <div class="muted" style="font-size:12px;">Aktualizuje skupinu a kód země. Bez zaškrtnutí se přeskočí.</div>
        </div>
      </label>
    </div>

    <div class="row" style="gap:12px;">
      <button type="submit" class="btn btn-primary">📋 Zobrazit preview</button>
      <a href="{{ url_for('admin_import') }}" class="btn">← Zpět</a>
    </div>
      </form>
    </div>
    """, r=r)

    @app.route("/admin/import/matches/template")
    @login_required
    def admin_import_matches_template():
        """Stáhne Excel šablonu pro import zápasů"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Zápasy"

            # Hlavička
            headers = ['Datum', 'Domácí', 'Hosté', 'Čas']
            ws.append(headers)

            # Stylování
            for col in range(1, 5):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="6EA8FE", end_color="6EA8FE", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Příklad
            ws.append(['2024-03-15', 'Sparta Praha', 'Slavia Praha', '18:00'])
            ws.append(['2024-03-16', 'Plzeň', 'Brno', '16:30'])

            # Šířka sloupců
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10

            # Poznámky
            notes = wb.create_sheet("Poznámky")
            notes['A1'] = "FORMÁT ZÁPASŮ"
            notes['A1'].font = Font(bold=True, size=14)
            notes['A3'] = "Datum: YYYY-MM-DD (např. 2024-03-15)"
            notes['A4'] = "Domácí: Název týmu"
            notes['A5'] = "Hosté: Název týmu"
            notes['A6'] = "Čas: HH:MM (např. 18:00)"
            notes.column_dimensions['A'].width = 50

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name='import_zapasu_sablona.xlsx')
        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_import"))

    @app.route("/admin/import/matches", methods=["GET", "POST"])
    @login_required
    def admin_import_matches():
        admin_required()
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            abort(400)

        if request.method == "POST":
            # Kontrola, jestli je CSV nebo Excel
            if 'excel_file' in request.files and request.files['excel_file'].filename:
                # Import z Excelu
                file = request.files['excel_file']
                try:
                    import openpyxl
                    from io import BytesIO

                    wb = openpyxl.load_workbook(BytesIO(file.read()))
                    ws = wb.active

                    created = 0
                    skipped = 0

                    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                        if not row or not any(row):
                            continue

                        # Formát: Datum, Domácí, Hosté, Čas, Liga
                        date_val = row[0] if len(row) > 0 else None
                        home = (str(row[1]) if len(row) > 1 and row[1] else "").strip()
                        away = (str(row[2]) if len(row) > 2 and row[2] else "").strip()
                        time_val = row[3] if len(row) > 3 else None

                        if not home or not away or home == away:
                            skipped += 1
                            continue

                        # Parsování datumu a času
                        start = None
                        if date_val and time_val:
                            try:
                                from datetime import datetime as dt_mod
                                if isinstance(date_val, str):
                                    date_str = date_val
                                else:
                                    date_str = date_val.strftime("%Y-%m-%d")

                                if isinstance(time_val, str):
                                    time_str = time_val
                                else:
                                    time_str = time_val.strftime("%H:%M")

                                start = dt_mod.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                            except:
                                pass

                        ht = _get_or_create_team(r.id, home)
                        at = _get_or_create_team(r.id, away)

                        existing = Match.query.filter_by(
                            round_id=r.id,
                            home_team_id=ht.id,
                            away_team_id=at.id,
                            is_deleted=False
                        ).first()

                        if existing:
                            skipped += 1
                            continue

                        m = Match(round_id=r.id, home_team_id=ht.id, away_team_id=at.id, start_time=start)
                        db.session.add(m)
                        created += 1

                    db.session.commit()
                    audit("import.matches.excel", "Round", r.id, created=created, skipped=skipped)
                    flash(f"Import z Excelu hotov: vytvořeno {created}, přeskočeno {skipped}.", "ok")
                    return redirect(url_for("matches"))

                except Exception as e:
                    flash(f"Chyba při importu z Excelu: {str(e)}", "error")
                    return redirect(url_for("admin_import_matches"))
            else:
                # CSV import (původní kód)
                csv_text = (request.form.get("csv_text") or "").strip()
                if not csv_text:
                    flash("Vlož CSV nebo nahraj Excel soubor.", "error")
                    return redirect(url_for("admin_import_matches"))
                reader = csv.DictReader(io.StringIO(csv_text))
                created = 0
                skipped = 0
                for row in reader:
                    rrid = _resolve_round_id(row, r.id)
                    home = (row.get("home_team") or row.get("home") or "").strip()
                    away = (row.get("away_team") or row.get("away") or "").strip()
                    start = parse_naive_datetime((row.get("start_time") or row.get("start") or "").strip())
                    if not home or not away or home == away:
                        skipped += 1
                        continue
                    ht = _get_or_create_team(rrid, home)
                    at = _get_or_create_team(rrid, away)
                    existing = Match.query.filter_by(round_id=rrid, home_team_id=ht.id, away_team_id=at.id, start_time=start, is_deleted=False).first()
                    if existing:
                        skipped += 1
                        continue
                    m = Match(round_id=rrid, home_team_id=ht.id, away_team_id=at.id, start_time=start)
                    hs = (row.get("home_score") or "").strip()
                    aas = (row.get("away_score") or "").strip()
                    if hs != "":
                        m.home_score = int(hs)
                    if aas != "":
                        m.away_score = int(aas)
                    db.session.add(m)
                    created += 1
                db.session.commit()
                audit("import.matches", "Round", r.id, created=created, skipped=skipped)
                flash(f"Import zápasů hotov: vytvořeno {created}, přeskočeno {skipped}.", "ok")
                return redirect(url_for("matches"))

        return render_page(r"""
    <div class="card">
      <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <div>
      <h2 style="margin:0 0 8px 0;">Import zápasů</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <a href="{{ url_for('admin_import_matches_template') }}" class="btn" style="background:#6ea8fe; color:white;">
      📥 Stáhnout šablonu Excel
    </a>
      </div>
      <hr class="sep">

      <h3 style="margin:20px 0 10px 0;">Import z Excel souboru</h3>
      <div class="muted" style="margin-bottom:10px;">
    Formát: 1. sloupec = Datum, 2. = Domácí, 3. = Hosté, 4. = Čas, 5. = Liga (volitelné)<br>
    První řádek je hlavička (přeskočí se)
      </div>
      <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="file" name="excel_file" accept=".xlsx,.xls" required>
    <button class="btn btn-primary" type="submit">Importovat z Excelu</button>
      </form>

      <hr class="sep" style="margin:20px 0;">

      <h3 style="margin:20px 0 10px 0;">Import z CSV (klasický)</h3>
      <div class="muted">Sloupce: <b>round_id</b> (volitelné), <b>home_team</b>, <b>away_team</b>, <b>start_time</b> (YYYY-MM-DD HH:MM), volitelně <b>home_score</b>, <b>away_score</b></div>
      <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <textarea name="csv_text" placeholder="round_id,home_team,away_team,start_time,home_score,away_score&#10;{{ r.id }},Sparta,Slavia,2026-02-10 18:00,,"></textarea>
    <button class="btn btn-primary" type="submit">Importovat CSV</button>
      </form>

      <hr class="sep">
      <a class="btn" href="{{ url_for('admin_import') }}">Zpět</a>
    </div>
    """, r=r)

    @app.route("/admin/import/extras/template")
    @login_required
    def admin_import_extras_template():
        """Stáhni Excel šablonu pro import extra otázek"""
        admin_required()
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extra otázky"

        # Header - user + příklady otázek
        headers = ["user", "Vítězný tým", "Kanadaské bodování", "Střelec turnaje"]
        header_fill_user = PatternFill("solid", fgColor="2F4F8F")
        header_fill_q    = PatternFill("solid", fgColor="4472C4")
        header_font      = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill_user if col == 1 else header_fill_q
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Příklady
        examples = [
            ("Václav",   "Kanada",  "MacKinnon", "McDavid"),
            ("Mejla",    "Kanada",  "McDavid",   "McDavid"),
            ("Ondejsek", "Švédsko", "McDavid",   "MacKinnon"),
        ]
        gray_fill = PatternFill("solid", fgColor="F2F2F2")
        for row_i, row_data in enumerate(examples, 2):
            fill = gray_fill if row_i % 2 == 0 else PatternFill()
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.fill = fill

        ws.column_dimensions["A"].width = 18
        for col_letter in ["B", "C", "D"]:
            ws.column_dimensions[col_letter].width = 28

        # Sheet 2 - popis
        ws2 = wb.create_sheet("Popis")
        ws2["A1"], ws2["B1"] = "Sloupec", "Popis"
        ws2["A1"].font = ws2["B1"].font = Font(bold=True)
        ws2["A2"] = "user"
        ws2["B2"] = "POVINNÉ - přesný username uživatele (case-insensitive)"
        ws2["A3"] = "Název otázky (2. sloupec a dál)"
        ws2["B3"] = "Každý sloupec = jedna extra otázka. Buňka = odpověď uživatele."
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 55

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        from flask import send_file
        return send_file(out,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="sablona_extra_otazky.xlsx",
                         as_attachment=True)

    @app.route("/admin/import/extras", methods=["GET", "POST"])
    @login_required
    def admin_import_extras():
        admin_required()
        import os, pickle
        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None
        if not r:
            abort(400)

        if request.method == "POST":
            action = request.form.get("action", "upload")

            # ── KROK 2: POTVRZENÍ ──────────────────────────────────────────
            if action == "confirm":

                preview_file = session.get("extra_import_preview_file")
                round_id     = session.get("extra_import_round_id")
                overwrite    = session.get("extra_import_overwrite", False)

                if not preview_file or not os.path.exists(preview_file):
                    flash("Session vypršela, nahraj soubor znovu.", "error")
                    return redirect(url_for("admin_import_extras"))

                r2 = db.session.get(Round, round_id) if round_id else r

                with open(preview_file, "rb") as f:
                    preview_data = pickle.load(f)
                os.unlink(preview_file)
                session.pop("extra_import_preview_file", None)
                session.pop("extra_import_round_id", None)
                session.pop("extra_import_overwrite", None)

                questions_data = preview_data["questions"]   # [{text, q_obj_or_None}]
                rows_data      = preview_data["rows"]        # [{user, answers:[{q_idx, answer, status}]}]

                # 1) Vytvoř/najdi otázky
                q_objects = []
                for qd in questions_data:
                    q_text = qd["text"]
                    existing_q = ExtraQuestion.query.filter_by(
                        round_id=r2.id, is_deleted=False
                    ).filter(db.func.lower(ExtraQuestion.question) == q_text.lower()).first()
                    if not existing_q:
                        existing_q = ExtraQuestion(round_id=r2.id, question=q_text)
                        db.session.add(existing_q)
                        db.session.flush()
                    q_objects.append(existing_q)

                # 2) Zpracuj odpovědi
                created = skipped = updated = 0
                for row in rows_data:
                    user = row["user"]
                    if user is None:
                        skipped += len(row["answers"])
                        continue
                    for ans_data in row["answers"]:
                        if ans_data["status"] == "skip":
                            skipped += 1
                            continue
                        q_obj   = q_objects[ans_data["q_idx"]]
                        ans_txt = ans_data["answer"]
                        existing_ans = ExtraAnswer.query.filter_by(
                            question_id=q_obj.id, user_id=user.id
                        ).first()
                        if existing_ans:
                            if overwrite:
                                existing_ans.answer_text = ans_txt
                                updated += 1
                            else:
                                skipped += 1
                        else:
                            db.session.add(ExtraAnswer(
                                question_id=q_obj.id,
                                user_id=user.id,
                                answer_text=ans_txt
                            ))
                            created += 1

                db.session.commit()
                audit("import.extras.excel", "Round", r2.id,
                      created=created, updated=updated, skipped=skipped)
                flash(f"Import extra otázek hotov: vytvořeno {created}, "
                      f"aktualizováno {updated}, přeskočeno {skipped}.", "ok")
                return redirect(url_for("extras"))

            # ── KROK 1: UPLOAD + PREVIEW ───────────────────────────────────
            import openpyxl, tempfile
            from io import BytesIO

            file = request.files.get("excel_file")
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_extras"))

            overwrite = request.form.get("overwrite") == "1"

            try:
                wb   = openpyxl.load_workbook(BytesIO(file.read()))
                ws   = wb.active
                rows = list(ws.iter_rows(values_only=True))

                if len(rows) < 2:
                    flash("Soubor musí mít hlavičku a alespoň jeden řádek.", "error")
                    return redirect(url_for("admin_import_extras"))

                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce (user + otázka).", "error")
                    return redirect(url_for("admin_import_extras"))

                # Otázky = sloupce 2..N
                question_texts = []
                for h in header[1:]:
                    if h and str(h).strip():
                        question_texts.append(str(h).strip())

                if not question_texts:
                    flash("Nenašel jsem žádné otázky v hlavičce.", "error")
                    return redirect(url_for("admin_import_extras"))

                # Existující otázky a odpovědi
                existing_qs = {
                    eq.question.lower(): eq
                    for eq in ExtraQuestion.query.filter_by(round_id=r.id, is_deleted=False).all()
                }
                existing_answers = {}  # (question_id, user_id) → ExtraAnswer
                for ans in ExtraAnswer.query.join(ExtraQuestion).filter(
                    ExtraQuestion.round_id == r.id,
                    ExtraQuestion.is_deleted == False
                ).all():
                    existing_answers[(ans.question_id, ans.user_id)] = ans

                # Všichni uživatelé (case-insensitive lookup)
                import unicodedata
                all_users = User.query.all()
                user_map  = {}
                for u in all_users:
                    norm = unicodedata.normalize("NFC", u.username).lower()
                    user_map[norm] = u

                # Info o otázkách pro preview
                questions_preview = []
                for qt in question_texts:
                    is_new = qt.lower() not in existing_qs
                    questions_preview.append({"text": qt, "is_new": is_new})

                # Zpracuj řádky
                preview_rows = []
                for data_row in rows[1:]:
                    username_raw = str(data_row[0]).strip() if data_row[0] else ""
                    if not username_raw or username_raw.lower() == "none":
                        continue

                    norm_name = unicodedata.normalize("NFC", username_raw).lower()
                    user_obj  = user_map.get(norm_name)

                    answers = []
                    for q_idx, qt in enumerate(question_texts):
                        col_idx = q_idx + 1
                        ans_val = str(data_row[col_idx]).strip() if (col_idx < len(data_row) and data_row[col_idx]) else ""
                        if not ans_val or ans_val.lower() == "none":
                            status = "skip"
                        elif user_obj is None:
                            status = "unknown_user"
                        else:
                            existing_q = existing_qs.get(qt.lower())
                            if existing_q and (existing_q.id, user_obj.id) in existing_answers:
                                status = "overwrite" if overwrite else "skip"
                            else:
                                status = "new"

                        answers.append({
                            "q_idx":  q_idx,
                            "answer": ans_val,
                            "status": status,
                        })

                    preview_rows.append({
                        "username_raw": username_raw,
                        "user":         user_obj,
                        "answers":      answers,
                    })

                if not preview_rows:
                    flash("Soubor neobsahuje žádné platné řádky.", "error")
                    return redirect(url_for("admin_import_extras"))

                # Ulož do temp souboru
                fd, preview_path = tempfile.mkstemp(suffix=".pkl", prefix="extra_import_")
                with os.fdopen(fd, "wb") as f:
                    pickle.dump({
                        "questions": [{"text": qt} for qt in question_texts],
                        "rows":      preview_rows,
                    }, f)

                session["extra_import_preview_file"] = preview_path
                session["extra_import_round_id"]     = r.id
                session["extra_import_overwrite"]    = overwrite

                # Počty pro preview
                new_cnt  = sum(1 for row in preview_rows for a in row["answers"] if a["status"] == "new")
                over_cnt = sum(1 for row in preview_rows for a in row["answers"] if a["status"] == "overwrite")
                skip_cnt = sum(1 for row in preview_rows for a in row["answers"] if a["status"] in ("skip", "unknown_user"))
                unk_users = [row["username_raw"] for row in preview_rows if row["user"] is None]

                return render_page(r"""
    <style>
    .prev-table { width:100%; border-collapse:collapse; font-size:13px; }
    .prev-table th { background:rgba(255,255,255,.07); padding:10px 12px; text-align:left; white-space:nowrap; }
    .prev-table td { padding:10px 12px; border-bottom:1px solid var(--line); vertical-align:middle; }
    .badge-new  { background:rgba(51,209,122,.2);  color:#33d17a; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
    .badge-skip { background:rgba(167,178,214,.15);color:#a7b2d6; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
    .badge-over { background:rgba(110,168,254,.2); color:#6ea8fe; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
    .badge-unk  { background:rgba(255,77,109,.2);  color:#ff4d6d; padding:3px 7px; border-radius:4px; font-size:11px; font-weight:700; }
    .q-header   { writing-mode:vertical-rl; transform:rotate(180deg); white-space:nowrap; padding:8px 4px; max-height:120px; font-size:12px; }
    </style>

    <div class="card">
      <h2 style="margin:0 0 4px 0;">📋 Preview – Extra otázky</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>

      {% if unk_users %}
      <div style="margin:16px 0; padding:12px 16px; background:rgba(255,77,109,.1); border:1px solid rgba(255,77,109,.3); border-radius:8px;">
    ⚠️ <strong>Nenalezení uživatelé</strong> (budou přeskočeni):
    {% for u in unk_users %}<code style="margin-left:8px;">{{ u }}</code>{% endfor %}
      </div>
      {% endif %}

      <div class="row" style="gap:16px; margin:16px 0; flex-wrap:wrap;">
    <div style="padding:12px 20px; background:rgba(51,209,122,.1); border:1px solid rgba(51,209,122,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#33d17a;">{{ new_cnt }}</div>
      <div class="muted" style="font-size:12px;">Nových odpovědí</div>
    </div>
    <div style="padding:12px 20px; background:rgba(110,168,254,.1); border:1px solid rgba(110,168,254,.3); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#6ea8fe;">{{ over_cnt }}</div>
      <div class="muted" style="font-size:12px;">Aktualizovaných</div>
    </div>
    <div style="padding:12px 20px; background:rgba(167,178,214,.1); border:1px solid rgba(167,178,214,.2); border-radius:8px; text-align:center;">
      <div style="font-size:28px; font-weight:900; color:#a7b2d6;">{{ skip_cnt }}</div>
      <div class="muted" style="font-size:12px;">Přeskočených</div>
    </div>
      </div>

      <div style="overflow-x:auto; margin-bottom:20px;">
    <table class="prev-table">
      <thead>
        <tr>
          <th>User</th>
          {% for q in questions_preview %}
            <th style="text-align:center;">
              <div class="q-header">
                {{ q.text }}
                {% if q.is_new %}<span style="color:#33d17a; font-size:10px;"> ✨</span>{% endif %}
              </div>
            </th>
          {% endfor %}
        </tr>
      </thead>
      <tbody>
        {% for row in preview_rows %}
        <tr>
          <td>
            {% if row.user %}
              <strong>{{ row.user.display_name }}</strong>
            {% else %}
              <span class="badge-unk">❓ {{ row.username_raw }}</span>
            {% endif %}
          </td>
          {% for ans in row.answers %}
          <td style="text-align:center;">
            {% if ans.status == 'new' %}
              <div style="font-size:12px; margin-bottom:3px;">{{ ans.answer }}</div>
              <span class="badge-new">✨</span>
            {% elif ans.status == 'overwrite' %}
              <div style="font-size:12px; margin-bottom:3px;">{{ ans.answer }}</div>
              <span class="badge-over">✏️</span>
            {% elif ans.status == 'unknown_user' %}
              <span class="badge-unk">—</span>
            {% else %}
              <span class="badge-skip">—</span>
            {% endif %}
          </td>
          {% endfor %}
        </tr>
        {% endfor %}
      </tbody>
    </table>
      </div>

      <form method="post" style="display:flex; gap:12px; flex-wrap:wrap;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="action" value="confirm">
    <button type="submit" class="btn btn-primary">✅ Potvrdit import</button>
    <a href="{{ url_for('admin_import_extras') }}" class="btn">✕ Zrušit</a>
      </form>
    </div>
    """, r=r,
     preview_rows=preview_rows,
     questions_preview=questions_preview,
     new_cnt=new_cnt, over_cnt=over_cnt, skip_cnt=skip_cnt,
     unk_users=unk_users)

            except Exception as e:
                import traceback
                flash(f"Chyba při čtení souboru: {e}", "error")
                return redirect(url_for("admin_import_extras"))

        # ── GET: formulář ──────────────────────────────────────────────────
        return render_page(r"""
    <style>
    .upload-zone {
      border:2px dashed rgba(110,168,254,0.4); border-radius:12px; padding:40px;
      text-align:center; background:rgba(110,168,254,0.05); cursor:pointer; transition:all .2s;
    }
    .upload-zone:hover { border-color:rgba(110,168,254,0.7); background:rgba(110,168,254,0.1); }
    </style>

    <div class="card">
      <div class="row" style="justify-content:space-between; margin-bottom:20px;">
    <div>
      <h2 style="margin:0 0 4px 0;">📥 Import extra otázek (Excel)</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <a href="{{ url_for('admin_import_extras_template') }}" class="btn"
       style="background:rgba(51,209,122,.15); color:#33d17a; border:1px solid rgba(51,209,122,.3);">
      📄 Stáhnout šablonu
    </a>
      </div>

      <div style="margin-bottom:20px; padding:14px 16px; background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); border-radius:8px; font-size:13px;">
    💡 <strong>Formát Excel:</strong><br><br>
    <table style="border-collapse:collapse; font-size:12px;">
      <thead>
        <tr style="background:rgba(255,255,255,.07);">
          <th style="padding:6px 12px; border:1px solid var(--line);">user</th>
          <th style="padding:6px 12px; border:1px solid var(--line);">Vítězný tým</th>
          <th style="padding:6px 12px; border:1px solid var(--line);">Kanadaské bodování</th>
          <th style="padding:6px 12px; border:1px solid var(--line);">Střelec</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:6px 12px; border:1px solid var(--line);">Václav</td><td style="padding:6px 12px; border:1px solid var(--line);">Kanada</td><td style="padding:6px 12px; border:1px solid var(--line);">MacKinnon</td><td style="padding:6px 12px; border:1px solid var(--line);">McDavid</td></tr>
        <tr><td style="padding:6px 12px; border:1px solid var(--line);">Mejla</td><td style="padding:6px 12px; border:1px solid var(--line);">Kanada</td><td style="padding:6px 12px; border:1px solid var(--line);">McDavid</td><td style="padding:6px 12px; border:1px solid var(--line);">McDavid</td></tr>
      </tbody>
    </table>
    <div class="muted" style="margin-top:10px;">
      • 1. sloupec = <strong>user</strong> (username uživatele)<br>
      • Další sloupce = názvy extra otázek<br>
      • Buňky = odpovědi uživatelů<br>
      • Nové otázky se vytvoří automaticky ✨
    </div>
      </div>

      <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="upload-zone" onclick="document.getElementById('xl').click()">
      <div style="font-size:48px; margin-bottom:12px;">📊</div>
      <div style="font-weight:700; margin-bottom:8px;">Klikni nebo přetáhni Excel soubor</div>
      <div class="muted">.xlsx • max 10 MB</div>
      <input type="file" id="xl" name="excel_file" accept=".xlsx" style="display:none"
             onchange="document.getElementById('fname').textContent = this.files[0]?.name || ''">
    </div>
    <div id="fname" class="muted" style="margin:8px 0 16px 0; text-align:center;"></div>

    <div style="padding:14px; background:rgba(110,168,254,.08); border:1px solid rgba(110,168,254,.2); border-radius:8px; margin-bottom:20px;">
      <label style="display:flex; align-items:center; gap:10px; cursor:pointer;">
        <input type="checkbox" name="overwrite" value="1" style="width:18px; height:18px;">
        <div>
          <strong>Přepsat existující odpovědi</strong>
          <div class="muted" style="font-size:12px;">Bez zaškrtnutí se existující odpovědi přeskočí.</div>
        </div>
      </label>
    </div>

    <div class="row" style="gap:12px;">
      <button type="submit" class="btn btn-primary">📋 Zobrazit preview</button>
      <a href="{{ url_for('admin_import') }}" class="btn">← Zpět</a>
    </div>
      </form>
    </div>
    """, r=r)

    @app.route("/admin/import/leaderboard/template")
    @login_required
    def admin_import_leaderboard_template():
        """Stáhne Excel šablonu pro import žebříčku"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Žebříček"

            # Hlavička - Jméno | Zápasy (formát: "Domácí-Hosté")
            headers = ['Jméno', 'Sparta-Slavia', 'Plzeň-Brno', 'Baník-Bohemians']
            ws.append(headers)

            # Stylování
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="6EA8FE", end_color="6EA8FE", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Příklady
            ws.append(['Jan Novák', '2:1', '1:1', '3:0'])
            ws.append(['Petr Svoboda', '1:2', '2:0', '1:1'])

            # Šířka sloupců
            ws.column_dimensions['A'].width = 20
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 15

            # Poznámky
            notes = wb.create_sheet("Poznámky")
            notes['A1'] = "FORMÁT ŽEBŘÍČKU"
            notes['A1'].font = Font(bold=True, size=14)
            notes['A3'] = "1. řádek = Hlavička"
            notes['A4'] = "   - První sloupec: 'Jméno'"
            notes['A5'] = "   - Další sloupce: názvy zápasů (formát: 'Domácí-Hosté')"
            notes['A7'] = "2+ řádky = Tipy uživatelů"
            notes['A8'] = "   - První sloupec: Jméno uživatele"
            notes['A9'] = "   - Další sloupce: tipy (formát: '2:1' nebo '1:0')"
            notes.column_dimensions['A'].width = 50

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name='import_zebricku_sablona.xlsx')
        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_import"))

    @app.route("/admin/import/leaderboard", methods=["GET", "POST"])
    @login_required
    def admin_import_leaderboard():
        admin_required()

        if request.method == "POST":
            file = request.files.get('excel_file')
            if not file or not file.filename:
                flash("Nahraj Excel soubor.", "error")
                return redirect(url_for("admin_import_leaderboard"))

            # Určit cílovou soutěž
            import_target = request.form.get('import_target', 'existing')

            if import_target == 'new':
                new_round_name = request.form.get('new_round_name', '').strip()
                if not new_round_name:
                    flash("Zadej název nové soutěže.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                sport = Sport.query.filter_by(name="Fotbal").first()
                if not sport:
                    sport = Sport(name="Fotbal")
                    db.session.add(sport)
                    db.session.flush()

                r = Round(name=new_round_name, sport_id=sport.id, is_active=False)
                db.session.add(r)
                db.session.flush()
                round_id = r.id
            else:
                round_id = int(request.form.get('round_id', 0))
                r = db.session.get(Round, round_id)
                if not r:
                    flash("Vybraná soutěž neexistuje.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

            try:
                import openpyxl
                from io import BytesIO
                from datetime import datetime as dt_parse

                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active

                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    flash("Soubor musí mít alespoň hlavičku a jeden řádek.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Parsovat zápasy z hlavičky
                matches_data = []
                for i, match_name in enumerate(header[1:], start=1):
                    if not match_name or str(match_name).strip() == "":
                        continue

                    match_str = str(match_name).strip()
                    for sep in ['-', ' vs ', ' x ', ':', ' – ']:
                        if sep in match_str:
                            parts = match_str.split(sep, 1)
                            if len(parts) == 2:
                                home = parts[0].strip()
                                away = parts[1].strip()

                                # Detekovat možnou duplicitu
                                home_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == home.lower()
                                ).first()

                                away_team_check = Team.query.filter(
                                    Team.round_id == round_id,
                                    Team.is_deleted == False,
                                    db.func.lower(Team.name) == away.lower()
                                ).first()

                                is_duplicate = False
                                if home_team_check and away_team_check:
                                    existing_match = Match.query.filter_by(
                                        round_id=round_id,
                                        home_team_id=home_team_check.id,
                                        away_team_id=away_team_check.id,
                                        is_deleted=False
                                    ).first()

                                    if existing_match:
                                        is_duplicate = True

                                matches_data.append({
                                    'col': i,
                                    'home': home,
                                    'away': away,
                                    'is_duplicate': is_duplicate,
                                    'match_str': match_str
                                })
                                break

                if not matches_data:
                    flash("Nenalezeny žádné platné zápasy v hlavičce.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Uložit preview data do session
                session['import_preview'] = {
                    'round_id': round_id,
                    'round_name': r.name,
                    'matches': matches_data,
                    'total_matches': len(matches_data),
                    'duplicates_count': sum(1 for m in matches_data if m['is_duplicate'])
                }

                return redirect(url_for("admin_import_leaderboard_preview"))

            except Exception as e:
                flash(f"Chyba při načítání souboru: {str(e)}", "error")
                db.session.rollback()
                return redirect(url_for("admin_import_leaderboard"))

            try:
                import openpyxl
                from io import BytesIO

                wb = openpyxl.load_workbook(BytesIO(file.read()))
                ws = wb.active

                # Načíst všechny řádky
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    flash("Soubor musí mít alespoň hlavičku a jeden řádek.", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # První řádek = hlavička
                header = rows[0]
                if not header or len(header) < 2:
                    flash("Hlavička musí mít alespoň 2 sloupce (Tipér + zápasy).", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Parsovat hlavičku zápasů (sloupce od 2. dál)
                matches_data = []
                for i, match_name in enumerate(header[1:], start=1):
                    if not match_name or str(match_name).strip() == "":
                        continue

                    match_str = str(match_name).strip()
                    # Očekáváme formát jako "Slavia-Sparta" nebo "Slavia vs Sparta"
                    for sep in ['-', ' vs ', ' x ', ':']:
                        if sep in match_str:
                            parts = match_str.split(sep, 1)
                            if len(parts) == 2:
                                home = parts[0].strip()
                                away = parts[1].strip()
                                matches_data.append({'col': i, 'home': home, 'away': away})
                                break

                if not matches_data:
                    flash("Nenalezeny žádné platné zápasy v hlavičce (očekáván formát 'Domácí-Hosté').", "error")
                    return redirect(url_for("admin_import_leaderboard"))

                # Vytvořit/najít zápasy
                created_matches = 0
                match_map = {}  # col_index -> Match
                for md in matches_data:
                    home_team = _get_or_create_team(r.id, md['home'])
                    away_team = _get_or_create_team(r.id, md['away'])

                    # Zkusit najít existující zápas
                    m = Match.query.filter_by(
                        round_id=r.id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if not m:
                        m = Match(round_id=r.id, home_team_id=home_team.id, away_team_id=away_team.id)
                        db.session.add(m)
                        db.session.flush()
                        created_matches += 1

                    match_map[md['col']] = m

                # Zkontrolovat první řádek po hlavičce - může obsahovat výsledky
                # NEBO poslední řádek - výsledky mohou být i dole
                results_imported = 0
                data_rows_start = 1  # Od kterého řádku začínají tipy
                data_rows_end = len(rows)  # Kde končí tipy (může být zkráceno o poslední řádek)

                # KONTROLA 1: První řádek po hlavičce (rows[1])
                if len(rows) > 1:
                    first_data_row = rows[1]
                    first_cell = str(first_data_row[0]).strip() if first_data_row[0] else ""

                    # Pokud první buňka je prázdná, "Výsledek", "Result" nebo "Skóre", jedná se o řádek s výsledky
                    if first_cell.lower() in ['', 'výsledek', 'result', 'skóre', 'score', 'vysledek']:
                        # Importovat výsledky z druhého řádku
                        for col_idx, match in match_map.items():
                            if col_idx >= len(first_data_row):
                                continue

                            result_value = first_data_row[col_idx]
                            if not result_value:
                                continue

                            # Parsovat výsledek
                            home_score = None
                            away_score = None

                            # Excel time format
                            if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                                home_score = result_value.hour
                                away_score = result_value.minute
                            else:
                                # String format
                                result_str = str(result_value).strip()
                                if result_str and result_str not in ['-', '—', '']:
                                    for sep in [':', '-']:
                                        if sep in result_str:
                                            parts = result_str.split(sep, 1)
                                            if len(parts) == 2:
                                                try:
                                                    home_score = int(parts[0].strip())
                                                    away_score = int(parts[1].strip())
                                                    break
                                                except:
                                                    pass

                            if home_score is not None and away_score is not None:
                                match.home_score = home_score
                                match.away_score = away_score
                                results_imported += 1

                        # Přeskočit tento řádek při zpracování tipů
                        data_rows_start = 2
                        db.session.flush()

                # KONTROLA 2: Poslední řádek (rows[-1]) - NOVĚ!
                # Pouze pokud jsme NENAŠLI výsledky v prvním řádku
                if results_imported == 0 and len(rows) > 2:
                    last_row = rows[-1]
                    last_cell = str(last_row[0]).strip() if last_row[0] else ""

                    # Pokud poslední řádek má "Výsledek" nebo je prázdný
                    if last_cell.lower() in ['', 'výsledek', 'result', 'skóre', 'score', 'vysledek']:
                        # Importovat výsledky z posledního řádku
                        for col_idx, match in match_map.items():
                            if col_idx >= len(last_row):
                                continue

                            result_value = last_row[col_idx]
                            if not result_value:
                                continue

                            # Parsovat výsledek
                            home_score = None
                            away_score = None

                            # Excel time format
                            if hasattr(result_value, 'hour') and hasattr(result_value, 'minute'):
                                home_score = result_value.hour
                                away_score = result_value.minute
                            else:
                                # String format
                                result_str = str(result_value).strip()
                                if result_str and result_str not in ['-', '—', '']:
                                    for sep in [':', '-']:
                                        if sep in result_str:
                                            parts = result_str.split(sep, 1)
                                            if len(parts) == 2:
                                                try:
                                                    home_score = int(parts[0].strip())
                                                    away_score = int(parts[1].strip())
                                                    break
                                                except:
                                                    pass

                            if home_score is not None and away_score is not None:
                                match.home_score = home_score
                                match.away_score = away_score
                                results_imported += 1

                        # Vynechat poslední řádek při zpracování tipů
                        data_rows_end = len(rows) - 1
                        db.session.flush()

                # Procházet řádky s tipéry
                created_users = 0
                created_tips = 0
                skipped_tips = 0
                skipped_tips_details = []  # Seznam přeskočených tipů pro Excel report

                for row in rows[data_rows_start:data_rows_end]:  # Skip header, případně výsledky nahoře a případně výsledky dole
                    if not row or not any(row):
                        continue

                    username = str(row[0]).strip() if row[0] else ""
                    if not username:
                        continue

                    # Najít nebo vytvořit uživatele
                    email = f"{username.lower().replace(' ', '_')}@import.tipovacka"
                    user = User.query.filter_by(username=username).first()

                    if not user:
                        user = User(
                            email=email,
                            username=username,
                            role='user'
                        )
                        user.set_password('tipovacka123')  # Výchozí heslo
                        db.session.add(user)
                        db.session.flush()
                        created_users += 1

                    # Importovat tipy
                    for col_idx, match in match_map.items():
                        if col_idx >= len(row):
                            continue

                        cell_value = row[col_idx]
                        if not cell_value:
                            continue

                        # Parsovat tip
                        tip_home = None
                        tip_away = None

                        # OPRAVA: Excel často převádí "2:1" na časový formát "2:01:00"
                        # Zkontrolovat jestli je to datetime.time objekt
                        if hasattr(cell_value, 'hour') and hasattr(cell_value, 'minute'):
                            # Je to time objekt - použít hour jako domácí, minute jako hosté
                            tip_home = cell_value.hour
                            tip_away = cell_value.minute
                        else:
                            # Je to string - parsovat běžně
                            tip_str = str(cell_value).strip()
                            if not tip_str or tip_str in ['-', '—', '']:
                                continue

                            # Parsovat formát "2:1" nebo "2-1"
                            for sep in [':', '-']:
                                if sep in tip_str:
                                    parts = tip_str.split(sep, 1)
                                    if len(parts) == 2:
                                        try:
                                            tip_home = int(parts[0].strip())
                                            tip_away = int(parts[1].strip())
                                            break
                                        except:
                                            pass

                        if tip_home is None or tip_away is None:
                            skipped_tips += 1
                            # Zaznamenat detail přeskočeného tipu
                            match_name = f"{match.home_team.name if match.home_team else '?'}-{match.away_team.name if match.away_team else '?'}"
                            skipped_tips_details.append({
                                'user': username,
                                'match': match_name,
                                'value': str(cell_value) if cell_value else '(prázdné)',
                                'reason': 'Nepodařilo se parsovat tip (očekáván formát "2:1" nebo "2-1")'
                            })
                            continue

                        # Vytvořit nebo aktualizovat tip
                        existing_tip = Tip.query.filter_by(
                            user_id=user.id,
                            match_id=match.id
                        ).first()

                        if existing_tip:
                            existing_tip.tip_home = tip_home
                            existing_tip.tip_away = tip_away
                        else:
                            db.session.add(Tip(
                                user_id=user.id,
                                match_id=match.id,
                                tip_home=tip_home,
                                tip_away=tip_away
                            ))
                            created_tips += 1

                db.session.commit()
                audit("import.leaderboard", "Round", r.id,
                      users=created_users, matches=created_matches, tips=created_tips, skipped=skipped_tips, results=results_imported)

                # Vytvořit Excel report pro přeskočené tipy
                skipped_report_path = None
                if skipped_tips > 0 and skipped_tips_details:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment

                    wb_report = openpyxl.Workbook()
                    ws_report = wb_report.active
                    ws_report.title = "Přeskočené tipy"

                    # Hlavička
                    headers = ['Tipér', 'Zápas', 'Hodnota v buňce', 'Důvod přeskočení']
                    ws_report.append(headers)

                    # Styling hlavičky
                    for cell in ws_report[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="FF4D6D", end_color="FF4D6D", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Data
                    for detail in skipped_tips_details:
                        ws_report.append([
                            detail['user'],
                            detail['match'],
                            detail['value'],
                            detail['reason']
                        ])

                    # Automatická šířka sloupců
                    for column in ws_report.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_report.column_dimensions[column_letter].width = adjusted_width

                    # Uložit report do outputs
                    import os
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_dir = "/mnt/user-data/outputs"
                    os.makedirs(output_dir, exist_ok=True)
                    skipped_report_filename = f"preskocene_tipy_{timestamp}.xlsx"
                    skipped_report_path = os.path.join(output_dir, skipped_report_filename)
                    wb_report.save(skipped_report_path)

                msg = f"✅ Import dokončen do soutěže '{r.name}'! Vytvořeno: {created_users} uživatelů, {created_matches} zápasů, {created_tips} tipů."
                if results_imported > 0:
                    msg += f" Importováno {results_imported} výsledků."
                if skipped_tips > 0:
                    msg += f" Přeskočeno: {skipped_tips} tipů."
                flash(msg, "ok")

                # Přepnout na nově importovanou soutěž
                set_selected_round_id(r.id)

                # Pokud byly přeskočené tipy, nabídnout stažení reportu
                if skipped_report_path:
                    # Uložit cestu do session pro download
                    from flask import session as flask_session
                    flask_session['skipped_report_path'] = skipped_report_path
                    flask_session['skipped_report_filename'] = skipped_report_filename
                    return redirect(url_for("admin_import_leaderboard_report"))

                return redirect(url_for("leaderboard"))

            except Exception as e:
                db.session.rollback()
                flash(f"Chyba při importu: {str(e)}", "error")
                return redirect(url_for("admin_import_leaderboard"))

        # GET request - zobrazit formulář
        all_rounds = Round.query.order_by(Round.is_active.desc(), Round.id.desc()).all()

        return render_page(r"""
    <div class="card">
      <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <h2 style="margin:0 0 8px 0;">📊 Import žebříčku z jiné tipovačky</h2>
    <a href="{{ url_for('admin_import_leaderboard_template') }}" class="btn" style="background:#6ea8fe; color:white;">
      📥 Stáhnout šablonu Excel
    </a>
      </div>

      <div class="muted" style="margin-bottom:16px;">
    Importuj kompletní data (uživatele, zápasy a tipy) z Excel souboru s žebříčkem.
      </div>

      <div class="card" style="background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); margin-bottom:16px;">
    <h3 style="margin:0 0 8px 0;">📋 Formát souboru</h3>
    <div class="muted" style="font-size:13px; line-height:1.6;">
      <strong>První řádek (hlavička):</strong><br>
      • Sloupec A: "Tipér" nebo "Jméno" (ignoruje se)<br>
      • Sloupec B+: Názvy zápasů ve formátu <strong>"Domácí-Hosté"</strong><br>
      &nbsp;&nbsp;Příklad: "Slavia-Sparta", "Plzeň-Liberec"<br><br>

      <strong>Druhý řádek NEBO poslední řádek (VOLITELNĚ - výsledky):</strong><br>
      • Sloupec A: prázdné NEBO "Výsledek" NEBO "Result"<br>
      • Sloupec B+: Výsledky zápasů <strong>"2:1"</strong><br>
      &nbsp;&nbsp;💡 Výsledky můžou být na <strong>druhém řádku</strong> (hned po hlavičce)<br>
      &nbsp;&nbsp;💡 NEBO na <strong>posledním řádku</strong> (dole pod tipéry) - NOVĚ!<br>
      &nbsp;&nbsp;💡 Pokud první sloupec je prázdný nebo obsahuje "Výsledek", importují se výsledky<br>
      &nbsp;&nbsp;💡 Pokud první sloupec obsahuje jméno, považuje se za tipéra<br><br>

      <strong>Další řádky (tipéři):</strong><br>
      • Sloupec A: Jméno tipéra<br>
      • Sloupec B+: Tipy ve formátu <strong>"2:1"</strong> nebo <strong>"2-1"</strong><br>
      &nbsp;&nbsp;💡 Funguje i časový formát "2:01" (Excel často automaticky převádí)<br><br>

      <strong>Co se stane:</strong><br>
      ✅ Vytvoří se uživatelé (heslo: "tipovacka123")<br>
      ✅ Vytvoří se zápasy a týmy<br>
      ✅ Naimportují se tipy<br>
      ✅ Naimportují se výsledky (pokud jsou)<br>
      ℹ️ Email uživatelů: <code>jmeno@import.tipovacka</code>
    </div>
      </div>

      <div class="card" style="background:rgba(255,199,79,0.08); border:1px solid rgba(255,199,79,0.2); margin-bottom:16px;">
    <strong>⚠️ Příklad Excel souboru S VÝSLEDKY:</strong>
    <table style="margin-top:8px; border-collapse:collapse; font-size:12px;">
      <tr style="background:rgba(255,255,255,0.05);">
        <th style="border:1px solid var(--line); padding:4px;">Tipér</th>
        <th style="border:1px solid var(--line); padding:4px;">Slavia-Sparta</th>
        <th style="border:1px solid var(--line); padding:4px;">Plzeň-Liberec</th>
        <th style="border:1px solid var(--line); padding:4px;">Baník-Sigma</th>
      </tr>
      <tr style="background:rgba(76,175,80,0.1);">
        <td style="border:1px solid var(--line); padding:4px; font-style:italic; color:#4CAF50;">Výsledek</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">3:1</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">2:2</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">1:0</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Petr</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">1:0</td>
        <td style="border:1px solid var(--line); padding:4px;">3:3</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Jana</td>
        <td style="border:1px solid var(--line); padding:4px;">1:1</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">0:2</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Martin</td>
        <td style="border:1px solid var(--line); padding:4px;">2:01</td>
        <td style="border:1px solid var(--line); padding:4px; font-size:10px;" class="muted">časový formát OK ✓</td>
        <td style="border:1px solid var(--line); padding:4px;">1:3</td>
      </tr>
    </table>
    <div class="muted" style="font-size:11px; margin-top:6px;">
      💡 Řádek "Výsledek" je volitelný. Pokud první sloupec je prázdný nebo obsahuje "Výsledek"/"Result"/"Skóre", importují se výsledky zápasů.
    </div>
      </div>

      <div class="card" style="background:rgba(255,199,79,0.08); border:1px solid rgba(255,199,79,0.2); margin-bottom:16px;">
    <strong>⚠️ Příklad Excel souboru s výsledky DOLE (NOVĚ):</strong>
    <table style="margin-top:8px; border-collapse:collapse; font-size:12px;">
      <tr style="background:rgba(255,255,255,0.05);">
        <th style="border:1px solid var(--line); padding:4px;">Tipér</th>
        <th style="border:1px solid var(--line); padding:4px;">Slavia-Sparta</th>
        <th style="border:1px solid var(--line); padding:4px;">Plzeň-Liberec</th>
        <th style="border:1px solid var(--line); padding:4px;">Baník-Sigma</th>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Petr</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">1:0</td>
        <td style="border:1px solid var(--line); padding:4px;">3:3</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Jana</td>
        <td style="border:1px solid var(--line); padding:4px;">1:1</td>
        <td style="border:1px solid var(--line); padding:4px;">2:1</td>
        <td style="border:1px solid var(--line); padding:4px;">0:2</td>
      </tr>
      <tr>
        <td style="border:1px solid var(--line); padding:4px;">Martin</td>
        <td style="border:1px solid var(--line); padding:4px;">2:01</td>
        <td style="border:1px solid var(--line); padding:4px;">1:1</td>
        <td style="border:1px solid var(--line); padding:4px;">1:3</td>
      </tr>
      <tr style="background:rgba(76,175,80,0.1);">
        <td style="border:1px solid var(--line); padding:4px; font-style:italic; color:#4CAF50;">Výsledek</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">3:1</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">2:2</td>
        <td style="border:1px solid var(--line); padding:4px; font-weight:900; color:#4CAF50;">1:0</td>
      </tr>
    </table>
    <div class="muted" style="font-size:11px; margin-top:6px;">
      💡 <strong>NOVĚ:</strong> Výsledky můžou být i v posledním řádku! Import automaticky detekuje řádek "Výsledek" ať je nahoře nebo dole.
    </div>
      </div>

      <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; align-items:stretch; gap:16px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    {# Krok 1: Kam importovat #}
    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">1️⃣ Kam importovat data?</h3>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="existing" {% if all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_existing">
        <span>Do existující soutěže</span>
      </label>

      <div id="existing_round_select" style="margin-left:28px; margin-bottom:16px;">
        <select name="round_id" style="width:100%; max-width:400px;">
          {% for rnd in all_rounds %}
            <option value="{{ rnd.id }}">
              {% if rnd.is_active %}★ {% endif %}{{ rnd.name }}
            </option>
          {% endfor %}
          {% if not all_rounds %}
            <option value="">-- Žádné soutěže --</option>
          {% endif %}
        </select>
      </div>

      <label style="display:flex; align-items:center; gap:8px; margin-bottom:10px; cursor:pointer;">
        <input type="radio" name="import_target" value="new" {% if not all_rounds %}checked{% endif %}
               onchange="toggleImportTarget()" id="radio_new">
        <span>Vytvořit novou soutěž</span>
      </label>

      <div id="new_round_input" style="margin-left:28px; display:none;">
        <input type="text" name="new_round_name" placeholder="Název nové soutěže (např. CZ Liga fotbal 2026)"
               style="width:100%; max-width:400px;">
        <div class="muted" style="margin-top:6px; font-size:12px;">
          Vytvoří se nová soutěž se sportem "Fotbal"
        </div>
      </div>
    </div>

    {# Krok 2: Soubor #}
    <div class="card" style="background:rgba(255,255,255,0.03); padding:16px;">
      <h3 style="margin:0 0 12px 0;">2️⃣ Vyber Excel soubor</h3>
      <input type="file" name="excel_file" accept=".xlsx,.xls" required>
    </div>

    <button class="btn btn-primary" type="submit" style="padding:14px; font-size:16px; font-weight:900;">
      📥 Importovat žebříček
    </button>
    <a class="btn" href="{{ url_for('admin_import') }}">Zpět</a>
      </form>
    </div>

    <script>
    function toggleImportTarget() {
      const existingChecked = document.getElementById('radio_existing').checked;
      const existingDiv = document.getElementById('existing_round_select');
      const newDiv = document.getElementById('new_round_input');

      if (existingChecked) {
    existingDiv.style.display = 'block';
    newDiv.style.display = 'none';
      } else {
    existingDiv.style.display = 'none';
    newDiv.style.display = 'block';
      }
    }

    // Inicializace při načtení
    toggleImportTarget();
    </script>
    """, all_rounds=all_rounds)

    @app.route("/admin/import/leaderboard/report")
    @login_required
    def admin_import_leaderboard_report():
        admin_required()
        from flask import session as flask_session, send_file

        skipped_report_path = flask_session.get('skipped_report_path')
        skipped_report_filename = flask_session.get('skipped_report_filename', 'preskocene_tipy.xlsx')

        if not skipped_report_path or not os.path.exists(skipped_report_path):
            flash("Report přeskočených tipů není k dispozici.", "error")
            return redirect(url_for("admin_import"))

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 16px 0;">⚠️ Přeskočené tipy při importu</h2>

      <div class="card" style="background:rgba(255,199,79,0.1); border:2px solid rgba(255,199,79,0.6); padding:20px; margin-bottom:20px;">
    <h3 style="margin:0 0 12px 0; color:#FFC74F;">Některé tipy nebyly importovány</h3>
    <div style="font-size:14px; line-height:1.6;">
      Při importu žebříčku byly některé tipy přeskočeny, protože se nepodařilo je správně parsovat.
      <br><br>
      <strong>Obvyklé důvody:</strong><br>
      • Neplatný formát (očekáván "2:1" nebo "2-1")<br>
      • Prázdná buňka nebo neplatná hodnota<br>
      • Text místo čísla<br><br>

      <strong>Stáhni Excel soubor</strong> s detailním přehledem všech přeskočených tipů.
      V souboru najdeš: tipéra, zápas, hodnotu v buňce a důvod přeskočení.
    </div>
      </div>

      <div style="display:flex; gap:12px; margin-bottom:20px;">
    <a href="{{ url_for('admin_import_leaderboard_download') }}"
       class="btn btn-primary"
       style="padding:14px 24px; font-size:16px; font-weight:900; background:#FFC74F; color:#000;">
      📥 Stáhnout Excel report
    </a>
    <a href="{{ url_for('leaderboard') }}" class="btn" style="padding:14px 24px; font-size:16px;">
      ➡️ Přejít do žebříčku
    </a>
      </div>

      <div class="muted" style="font-size:12px;">
    💡 Po opravě tipů v Excelu můžeš znovu importovat žebříček s opravenými daty.
      </div>
    </div>
    """)

    @app.route("/admin/import/leaderboard/download")
    @login_required
    def admin_import_leaderboard_download():
        admin_required()
        from flask import session as flask_session, send_file

        skipped_report_path = flask_session.get('skipped_report_path')
        skipped_report_filename = flask_session.get('skipped_report_filename', 'preskocene_tipy.xlsx')

        if not skipped_report_path or not os.path.exists(skipped_report_path):
            flash("Report přeskočených tipů není k dispozici.", "error")
            return redirect(url_for("admin_import"))

        # Smazat ze session po stažení
        flask_session.pop('skipped_report_path', None)
        flask_session.pop('skipped_report_filename', None)

        return send_file(
            skipped_report_path,
            as_attachment=True,
            download_name=skipped_report_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # --- ADMIN AUDIT ---


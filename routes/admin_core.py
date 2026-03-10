"""
routes/admin_core.py
"""

import datetime
import re
import io
import os
import csv
import json
import tempfile
import zipfile
from io import BytesIO
from typing import Dict, List, Optional, Tuple


try:
    import pytesseract
    from PIL import Image as PILImage
    TESSERACT_AVAILABLE = True
except ImportError:
    pytesseract = None
    PILImage = None
    TESSERACT_AVAILABLE = False


def extract_text_from_screenshot(image_data: bytes) -> Optional[str]:
    """OCR z clipboard screenshotu."""
    try:
        import pytesseract as _tess
        from PIL import Image as _Img, ImageEnhance
        image = _Img.open(io.BytesIO(image_data))
        try:
            image = ImageEnhance.Contrast(image).enhance(2.0)
        except Exception:
            pass
        text = _tess.image_to_string(image, lang='ces+eng')
        return text.strip() if text.strip() else None
    except ImportError:
        return None
    except Exception as e:
        print(f"❌ OCR error: {e}")
        return None


def _split_joined_lines(text: str) -> str:
    """
    CRITICAL FIX: Table copy/paste joins all lines

    New strategy: Extract complete match patterns
    Pattern: DD. MM. YYYY + text + time/score (N:N or NN:NN)
    """

    # Skip if already properly formatted
    lines = text.split('\n')
    if len(lines) > 3:
        return text

    # Rejoin text
    text_joined = ' '.join(lines)

    # Pattern for complete match:
    # Date + any text + final number:number
    # Use non-greedy match and lookahead
    pattern = r'(\d{1,2}\.\s*\d{1,2}\.\s*\d{4}.*?\d{1,2}:\d{1,2})(?=\d{1,2}\.\s*\d{1,2}\.\s*\d{4}|$)'

    matches = re.findall(pattern, text_joined)

    if len(matches) >= 2:
        # Add any header before first match
        parts = []
        first_match_pos = text_joined.find(matches[0])
        if first_match_pos > 0:
            header = text_joined[:first_match_pos].strip()
            if header:
                parts.append(header)

        # Add all matches
        for match in matches:
            parts.append(match.strip())

        if len(parts) > 1:
            result = '\n'.join(parts)
            print(f"🔧 Extracted {len(parts)} lines from joined input")
            return result

    return text




def _split_joined_lines(text: str) -> str:
    """
    🔧 FIX: Multiple matches joined on one line (from table copy/paste)

    Example:
    Input:  "27. 2. 2026DuklaSlavia0:228. 2. 2026Liberec..."
    Output: Each match on separate line

    Finds all dates and splits before each (except first)
    """

    result_lines = []

    for line in text.strip().split('\n'):
        line = line.strip()
        if not line:
            continue

        # Find all date patterns in this line
        date_pattern = r'\d{1,2}\.\s*\d{1,2}\.\s*\d{4}'
        matches = list(re.finditer(date_pattern, line))

        if len(matches) <= 1:
            # 0 or 1 date - keep as is
            result_lines.append(line)
            continue

        # Multiple dates found - split before each (except first)
        print(f"🔧 Found {len(matches)} dates in one line - splitting...")

        parts = []
        last_end = 0

        for i, match in enumerate(matches):
            if i == 0:
                # First date - take from start until next date
                if len(matches) > 1:
                    next_start = matches[1].start()
                    parts.append(line[last_end:next_start])
                    last_end = next_start
                else:
                    parts.append(line[last_end:])
            else:
                # Subsequent dates
                if i < len(matches) - 1:
                    next_start = matches[i+1].start()
                    parts.append(line[last_end:next_start])
                    last_end = next_start
                else:
                    # Last date - take rest of line
                    parts.append(line[last_end:])

        # Add all parts
        for part in parts:
            part = part.strip()
            if part:
                result_lines.append(part)

    return '\n'.join(result_lines)


def _clean_ocr_artifacts(text: str) -> str:
    """
    Clean up OCR artifacts from screenshot text

    OCR often adds extra characters like:
    - "vice >" or "více>" (from web UI buttons)
    - "|" pipes
    - Extra whitespace
    - Weird unicode characters

    Returns cleaned text ready for parsing
    """
    lines = text.strip().split('\n')
    cleaned_lines = []

    for line in lines:
        # Remove common OCR artifacts
        # Remove "vice >", "více>", "vic>", "víc>" and similar
        line = re.sub(r'\s*v[ií]ce?\s*>?\s*', ' ', line, flags=re.IGNORECASE)

        # Remove pipe symbols (often from table borders)
        line = re.sub(r'\s*\|\s*', ' ', line)

        # Remove extra whitespace between words (but keep spaces)
        line = re.sub(r'\s+', ' ', line)

        # Remove leading/trailing whitespace
        # Fix OCR misreading 0 as 'o'/'O' at end of line (trailing score)
        line = re.sub(r' [oO]$', ' 0', line)

        line = line.strip()

        if line:
            cleaned_lines.append(line)

    result = '\n'.join(cleaned_lines)
    print(f"🧹 OCR cleanup: {len(lines)} lines → {len(cleaned_lines)} cleaned lines")
    return result

def _parse_multiline_app_format(text: str) -> List[Dict]:
    """
    Parse multi-line format from app/website:

    07.03. 15:00
    Bohemians
    Slovan Liberec
    --
    """
    matches = []
    blocks = text.split('--')
    current_year = 2026

    for block in blocks:
        lines = [line.strip() for line in block.strip().split('\n') if line.strip()]

        if len(lines) < 3:
            continue

        # Skip "25. kolo" headers
        if lines[0].lower().endswith('kolo'):
            lines = lines[1:]

        if len(lines) < 3:
            continue

        datetime_line = lines[0]
        home_team = lines[1]
        away_team = lines[2]

        # Parse "07.03. 15:00"
        match = re.match(r'(\d{2})\.(\d{2})\.\s+(\d{1,2}):(\d{2})', datetime_line)
        if not match:
            continue

        day = int(match.group(1))
        month = int(match.group(2))
        hour = int(match.group(3))
        minute = int(match.group(4))

        try:
            dt = datetime(current_year, month, day, hour, minute)
            matches.append({
                'home_team': home_team.strip(),
                'away_team': away_team.strip(),
                'start_time': dt,
            })
        except ValueError:
            continue

    return matches

def _parse_multiline_ocr_format(text: str) -> List[Dict]:
    """
    Parser pro OCR screenshot z mobilní appky Fortuny/iSport.
    OCR vrací řádky s prefixovými ikonami, vzor je:
      (© Bohemians :      <- tým1
      úz 07.03. 15:00     <- datum+čas
      @ Slovan Liberec -  <- tým2
      (© Slovácko -       <- tým1
      k 07.03.15:00       <- datum+čas
      ...
    """
    import re
    from datetime import datetime

    current_year = datetime.now().year

    # Regex na datum+čas: "07.03. 15:00" nebo "07.03.15:00" nebo "7.3. 15:00"
    dt_re = re.compile(r'(\d{1,2})\.(\d{1,2})\.?\s*(\d{1,2}):(\d{2})')
    # Regex na samotné datum bez času: "25.KoLo", "25. kolo" -> skip
    kolo_re = re.compile(r'^\d+\.\s*kolo', re.I)
    # Junk header patterns
    skip_re = re.compile(
        r'^(bude se hrát|česko|chance liga|jme\s|\d+\.\s*kolo)',
        re.I
    )

    def strip_prefix(line):
        """Odstraní OCR prefix ikony na začátku řádku, zachová jméno týmu začínající velkým písmenem."""
        # Odstraní vše před prvním velkým písmenem (začátek jména týmu)
        m = re.search(r'[A-ZÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ]', line)
        if m:
            return line[m.start():].strip()
        return line.strip()

    def clean_team(name):
        """Odstraní trailing ' -', ' :', ' =' a podobné artefakty."""
        return re.sub(r'\s*[-:=]\s*$', '', name).strip()

    def parse_dt(line):
        """Vrátí datetime pokud řádek obsahuje datum+čas, jinak None."""
        m = dt_re.search(line)
        if m:
            try:
                return datetime(current_year, int(m.group(2)), int(m.group(1)),
                                int(m.group(3)), int(m.group(4)))
            except Exception:
                return None
        return None

    lines = text.strip().split('\n')

    # Vyčisti a oklasifikuj každý řádek
    # Každý item: ('dt', datetime) | ('team', str, score_or_None)
    classified = []
    score_re = re.compile(r'\s+(\d{1,2})\s*$')  # trailing score digit(s)
    for raw in lines:
        raw = raw.strip()
        if not raw or len(raw) < 3:
            continue
        if skip_re.match(raw.lower()):
            continue
        if kolo_re.match(raw):
            continue

        dt = parse_dt(raw)
        if dt:
            classified.append(('dt', dt, None))
        else:
            # Check for trailing score: '(© Sigma Olomouc 1' -> team='Sigma Olomouc', score=1
            score_match = score_re.search(raw)
            score = int(score_match.group(1)) if score_match else None
            line_no_score = raw[:score_match.start()] if score_match else raw
            clean = clean_team(strip_prefix(line_no_score))
            if clean and len(clean) >= 3:
                classified.append(('team', clean, score))

    # Nyní hledáme vzory: team, dt, team  NEBO  dt, team, team
    results = []
    i = 0
    while i < len(classified):
        # Vzor 1: team1(+score), dt, team2(+score)  (Fortuna/iSport appka s výsledky)
        if (i + 2 < len(classified)
                and classified[i][0] == 'team'
                and classified[i+1][0] == 'dt'
                and classified[i+2][0] == 'team'):
            home = classified[i][1]
            home_score = classified[i][2]
            dt = classified[i+1][1]
            away = classified[i+2][1]
            away_score = classified[i+2][2]
            entry = {
                'home_team': home,
                'away_team': away,
                'start_time': dt.isoformat(),
            }
            if home_score is not None and away_score is not None:
                entry['home_score'] = home_score
                entry['away_score'] = away_score
            results.append(entry)
            i += 3
            continue
        # Vzor 2: dt, team1(+score), team2(+score)  (čistý screenshot)
        if (i + 2 < len(classified)
                and classified[i][0] == 'dt'
                and classified[i+1][0] == 'team'
                and classified[i+2][0] == 'team'):
            dt = classified[i][1]
            home = classified[i+1][1]
            home_score = classified[i+1][2]
            away = classified[i+2][1]
            away_score = classified[i+2][2]
            entry = {
                'home_team': home,
                'away_team': away,
                'start_time': dt.isoformat(),
            }
            if home_score is not None and away_score is not None:
                entry['home_score'] = home_score
                entry['away_score'] = away_score
            results.append(entry)
            i += 3
            continue
        i += 1

    return results


def smart_parse_matches(text: str, round_id: int = None) -> List[Dict]:
    """
    🤖 ULTRA SMART PARSER V2 - Better whitespace handling + OCR cleanup

    Podporované formáty:
    - Table format: "27. 2. 2026DuklaSlavia18:00" 
    - With spaces: "27. 2. 2026 Sparta - Slavia 18:00"
    - Date + time first: "27. 2. 20:00 Sparta vs Slavia"
    - CSV, Fortuna, UEFA, etc.
    - OCR from screenshots (with cleanup)

    Improvements:
    - Skip non-match lines (headers like "24. kolo")
    - Auto year-fill for partial dates
    - Better team splitting
    - OCR artifact cleanup
    """

    # TRY 1: Multi-line app format (with -- separators)
    if '--' in text:
        print("🎯 Detected multi-line app format")
        matches = _parse_multiline_app_format(text)
        if matches:
            print(f"✅ Multi-line parser: Found {len(matches)} matches")
            # Convert datetime to ISO
            for match in matches:
                if 'start_time' in match and isinstance(match['start_time'], datetime):
                    match['start_time'] = match['start_time'].isoformat()
            return matches

    # TRY: Fragmented column OCR (mobile app) - BEFORE cleanup (needs raw OCR)
    fragmented = _parse_fragmented_column_ocr(text)
    if fragmented:
        return fragmented

    # Step 1: Clean OCR artifacts (vice>, |, extra whitespace)
    text = _clean_ocr_artifacts(text)

    # TRY 0: Screenshot multi-line OCR format (datetime / team1 / team2)
    ocr_multiline = _parse_multiline_ocr_format(text)
    if ocr_multiline:
        print(f"✅ OCR multiline parser: Nalezeno {len(ocr_multiline)} zápasů")
        return ocr_multiline

    # Step 2: Handle table copy/paste (all lines joined)
    text = _split_joined_lines(text)

    matches = []
    lines = text.strip().split('\n')

    print(f"🤖 Smart Parser V2: Zpracovávám {len(lines)} řádků")

    for line_num, line in enumerate(lines, 1):
        # Don't normalize whitespace yet - some formats need it
        line_raw = line.strip()

        if not line_raw or len(line_raw) < 5:
            continue

        # Skip lines that are obviously not matches
        if line_raw.lower() in ['kolo', 'round', 'zápasy', 'matches']:
            print(f"⏭️ Přeskakuji header: {line_raw}")
            continue
        if re.match(r'^\d+\.\s*kolo', line_raw, re.I):
            print(f"⏭️ Přeskakuji kolo header: {line_raw}")
            continue
        if re.match(r'česko|chance liga|jme\s|bude se hrát', line_raw.lower()):
            print(f"⏭️ Přeskakuji soutěž header: {line_raw}")
            continue

        match_data = _parse_single_line(line_raw, line_num)
        if match_data:
            # Convert datetime to ISO string for JSON
            if 'start_time' in match_data and match_data['start_time']:
                if isinstance(match_data['start_time'], datetime):
                    match_data['start_time'] = match_data['start_time'].isoformat()

            matches.append(match_data)

    print(f"✅ Smart Parser V2: Nalezeno {len(matches)} zápasů")
    return matches


def _is_date(text: str) -> bool:
    """Check if text looks like a date"""
    date_patterns = [
        r'^\d{1,2}\.\s*\d{1,2}\.\s*\d{4}$',  # 27. 2. 2026
        r'^\d{1,2}/\d{1,2}/\d{4}$',          # 27/2/2026
        r'^\d{4}-\d{1,2}-\d{1,2}$',          # 2026-2-27
        r'^\d{1,2}\.\s*\d{1,2}\.\s*$',       # 27. 2.
        r'^\d{1,2}\.\s*\d{1,2}\.$',          # 27.2.
    ]
    text = text.strip()
    for pattern in date_patterns:
        if re.match(pattern, text):
            return True
    return False



def _parse_compact_date_format(line: str, parts: List[str]) -> Optional[Dict]:
    """
    Parse compact date format (common from OCR):
    "7.3.2026 Teplice Dukla 15:00"
    Format: DATE TEAM1 TEAM2 TIME
    """
    date_str = parts[0]
    rest_parts = parts[1:]

    dt = _parse_datetime(date_str)
    if not dt:
        return None

    # Check if last part is time
    time_str = None
    if rest_parts and re.match(r'^\d{1,2}:\d{2}$', rest_parts[-1]):
        time_str = rest_parts[-1]
        team_parts = rest_parts[:-1]
    else:
        team_parts = rest_parts

    if len(team_parts) < 2:
        return None

    # Split teams
    if len(team_parts) == 2:
        home_team, away_team = team_parts[0], team_parts[1]
    else:
        teams_text = ' '.join(team_parts)
        home_team, away_team = _smart_split_teams(teams_text)
        if not home_team or not away_team:
            mid = len(team_parts) // 2
            home_team = ' '.join(team_parts[:mid])
            away_team = ' '.join(team_parts[mid:])

    # Add time to datetime
    if time_str and dt:
        try:
            time_obj = datetime.strptime(time_str, "%H:%M").time()
            dt = dt.replace(hour=time_obj.hour, minute=time_obj.minute)
        except:
            pass

    return {
        'home_team': home_team.strip(),
        'away_team': away_team.strip(),
        'start_time': dt,
    }

def _parse_date_first_format(line: str, parts: List[str]) -> Optional[Dict]:
    """Parse: 27. 2. 2026 Dukla Slavia 18:00 OR 27. 2. 2026 Dukla Slavia 0:2"""

    date_str = ' '.join(parts[:3])  # "27. 2. 2026"
    rest = ' '.join(parts[3:])       # "Dukla Slavia 18:00" or "Dukla Slavia 0:2"

    dt = _parse_datetime(date_str)

    # Check for time/score at end
    time_or_score_match = re.search(r'(\d{1,2}):(\d{1,2})$', rest)

    is_score = False
    home_score = None
    away_score = None

    if time_or_score_match:
        first_num = int(time_or_score_match.group(1))
        second_num = int(time_or_score_match.group(2))
        full_match = time_or_score_match.group(0)
        teams_part = rest[:rest.rfind(full_match)].strip()

        # Determine if it's time or score
        if first_num >= 10 and first_num <= 23:
            is_score = False  # Likely time
        elif second_num in [0, 15, 30, 45] and first_num <= 23:
            is_score = False  # Common time minutes
        elif first_num < 10 and second_num < 10:
            is_score = True  # Both small - likely score
        else:
            is_score = first_num > 23 or second_num > 59

        if is_score:
            home_score = first_num
            away_score = second_num
        else:
            # It's a time
            if dt:
                try:
                    time_obj = datetime.strptime(full_match, "%H:%M").time()
                    dt = dt.replace(hour=time_obj.hour, minute=time_obj.minute)
                except:
                    pass
    else:
        teams_part = rest

    # Parse teams
    if ' - ' in teams_part:
        teams = teams_part.split(' - ', 1)
    elif ' vs ' in teams_part.lower():
        teams = re.split(r'\s+vs\s+', teams_part, flags=re.I, maxsplit=1)
    else:
        # Try to split by space
        teams = teams_part.split()
        if len(teams) < 2:
            return {
                'home_team': teams_part.strip(),
                'away_team': '',
                'start_time': dt,
            }

    if len(teams) >= 2:
        result = {
            'home_team': teams[0].strip(),
            'away_team': teams[1].strip() if len(teams) > 1 else teams[-1].strip(),
            'start_time': dt,
        }

        if home_score is not None:
            result['home_score'] = home_score
            result['away_score'] = away_score

        return result

    return None

def _parse_date_time_first(line: str, parts: List[str]) -> Optional[Dict]:
    """Parse: 27. 2. 20:00 Sparta vs Slavia"""

    date_str = ' '.join(parts[:2])  # "27. 2."
    time_str = parts[2]              # "20:00"
    rest = ' '.join(parts[3:])       # "Sparta vs Slavia"

    dt = _parse_datetime(date_str + ' ' + time_str)

    # Parse teams
    if ' - ' in rest:
        teams = rest.split(' - ', 1)
    elif ' vs ' in rest.lower():
        teams = re.split(r'\s+vs\s+', rest, flags=re.I, maxsplit=1)
    else:
        return None

    if len(teams) >= 2:
        return {
            'home_team': teams[0].strip(),
            'away_team': teams[1].strip(),
            'start_time': dt,
        }

    return None



def _parse_table_format(line: str) -> Optional[Dict]:
    """
    IMPROVED: Parse table copy/paste - handles NO SPACES between parts

    Examples:
    - 27. 2. 2026DuklaSlavia18:00      (time)
    - 27. 2. 2026DuklaSlavia0:2        (score)
    - 28. 2. 2026Ml. Boleslav__Jablonec__15:00
    """

    # Pattern: DATE (required) + TEAMS (any) + TIME/SCORE (required)

    # Match date at start
    date_match = re.match(r'^(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})', line)
    if not date_match:
        return None

    date_str = date_match.group(1)
    rest = line[len(date_str):].strip()

    # Match time OR score at end
    # Time: HH:MM (hour 0-23, minute 00-59)
    # Score: N:N (any digits)
    time_score_match = re.search(r'(\d{1,2}:\d{1,2})$', rest)
    if not time_score_match:
        return None

    time_or_score = time_score_match.group(1)
    teams_part = rest[:rest.rfind(time_or_score)].strip()

    if not teams_part:
        return None

    # Determine if it's time or score
    parts = time_or_score.split(':')
    hour_or_home = int(parts[0])
    min_or_away = int(parts[1])

    # Heuristic: If first number > 23 OR second number > 59, it's probably score
    # OR if both numbers are single digit and < 10, could be score
    is_score = False
    is_time = False

    # Definite time: HH:MM format where HH in 0-23 and MM in 00-59
    if 0 <= hour_or_home <= 23 and 0 <= min_or_away <= 59:
        # Could be time... but also could be low score like 2:1
        # Check: if MM is 00, 15, 30, 45 → likely time
        if min_or_away in [0, 15, 30, 45]:
            is_time = True
        # If HH >= 10, likely time (games don't usually have 10+ goals)
        elif hour_or_home >= 10:
            is_time = True
        # If both small (< 10), likely score
        else:
            is_score = True
    else:
        # Outside valid time range → score
        is_score = True

    # Parse teams
    if '__' in teams_part:
        parts = [p.strip() for p in teams_part.split('__') if p.strip()]
        if len(parts) >= 2:
            home_team = parts[0]
            away_team = parts[1]
        else:
            return None
    else:
        home_team, away_team = _smart_split_teams(teams_part)
        if not home_team or not away_team:
            return None

    # Build result
    result = {
        'home_team': home_team,
        'away_team': away_team,
    }

    if is_time:
        # Parse as time
        try:
            dt = datetime.strptime(f"{date_str} {time_or_score}", "%d. %m. %Y %H:%M")
            result['start_time'] = dt
        except:
            result['start_time'] = None
    else:
        # Parse as score
        result['home_score'] = hour_or_home
        result['away_score'] = min_or_away
        # Parse just the date (no time)
        try:
            dt = datetime.strptime(date_str, "%d. %m. %Y")
            result['start_time'] = dt
        except:
            result['start_time'] = None

    return result


def _smart_split_teams(text: str) -> Tuple[Optional[str], Optional[str]]:
    """
    IMPROVED: Smart split teams from text without separator

    Examples:
    - "DuklaSlavia" → "Dukla", "Slavia"
    - "SpartaOstrava" → "Sparta", "Ostrava"
    - "Ml. BoleslavJablonec" → "Ml. Boleslav", "Jablonec"
    """

    known_teams = [
        # Longer names first for greedy matching
        'Ml. Boleslav', 'Mladá Boleslav',
        'Hradec Kr.', 'Hradec Králové', 'Hradec',
        'FK Dukla Praha', 'SK Slavia Praha', 'AC Sparta Praha',
        'FC Viktoria Plzeň', 'FC Zlín', 'FC Slovan Liberec',
        'FC Baník Ostrava', 'FK Teplice', 'FK Jablonec',
        'MFK Karviná', '1.FC Slovácko', '1. FC Slovácko', 'FK Pardubice',
        'SK Sigma Olomouc', 'Bohemians Praha 1905',
        'FK Mladá Boleslav', 'FC Hradec Králové',
        # Short names
        'Dukla', 'Slavia', 'Sparta', 'Plzeň', 'Zlín',
        'Liberec', 'Ostrava', 'Baník', 'Teplice', 'Jablonec',
        'Karviná', 'Slovácko', 'Pardubice', 'Olomouc', 'Sigma',
        'Bohemians',
    ]

    # Try to match known team from start
    for team in sorted(known_teams, key=len, reverse=True):
        if text.startswith(team):
            home = team
            away = text[len(team):].strip()

            # Check if away is also a known team
            for away_team in known_teams:
                if away == away_team or away.startswith(away_team):
                    return home, away_team

            # Try away as-is
            if away:
                return home, away

    # Fallback: split at second capital letter
    capitals = [i for i, c in enumerate(text) if c.isupper()]
    if len(capitals) >= 2:
        split_pos = capitals[1]
        return text[:split_pos], text[split_pos:]

    return None, None


def _parse_single_line(line: str, line_num: int = 0) -> Optional[Dict]:
    """
    IMPROVED: Better handling of different formats

    Priority:
    1. Table format (no spaces): 27. 2. 2026DuklaSlavia18:00
    2. Date with spaces: 27. 2. 2026 Dukla - Slavia 18:00
    3. Other formats
    """

    # TRY 1: Table format FIRST (before whitespace normalization!)
    result = _parse_table_format(line)
    if result:
        return result

    # TRY 2: Now normalize whitespace for other formats
    line_normalized = re.sub(r'\s+', ' ', line.replace('\t', ' ')).strip()
    parts = line_normalized.split()  # initialize here to avoid UnboundLocalError

    # TRY 3: Compact date format (OCR): "7.3.2026 Team1 Team2 15:00"
    if len(parts) >= 4 and _is_date(parts[0]):
        result = _parse_compact_date_format(line_normalized, parts)
        if result:
            return result

    # TRY 4: Date at start with spaces
    # parts already assigned above

    # "27. 2. 2026 Dukla - Slavia 18:00"
    if len(parts) >= 3 and _is_date(' '.join(parts[:3])):
        return _parse_date_first_format(line_normalized, parts)

    # "27. 2. 20:00 Sparta vs Slavia"
    if len(parts) >= 4:
        potential_date = ' '.join(parts[:2])
        potential_time = parts[2]

        if _is_date(potential_date) and re.match(r'^\d{1,2}:\d{2}$', potential_time):
            return _parse_date_time_first(line_normalized, parts)

    # TRY 5: Other parsers
    parsers = [
        _parse_fortuna_style,
        _parse_uefa_style,
        _parse_csv_style,
        _parse_pipe_style,
        _parse_score_style,
        _parse_vs_style,
        _parse_dash_style,
    ]

    for parser in parsers:
        try:
            result = parser(line_normalized)
            if result:
                return result
        except Exception:
            continue

    print(f"⚠️ Nepodařilo se parsovat řádek {line_num}: {line[:50]}")
    return None


def _parse_fortuna_style(line: str) -> Optional[Dict]:
    """Fortuna Liga: #22 14/02/26Sat 15:00 DUK 0:0 FCZ"""
    pattern = r'#(\d+)\s+(\d{2})/(\d{2})/(\d{2})(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\d{1,2}):(\d{2})\s+([A-Z]{3,4})\s+(?:(\d+):(\d+)|[-–])\s+([A-Z]{3,4})'
    m = re.search(pattern, line)
    if not m:
        return None

    day, month, year_short = int(m.group(2)), int(m.group(3)), int(m.group(4))
    hour, minute = int(m.group(5)), int(m.group(6))
    home_code, away_code = m.group(7), m.group(10)
    home_score = int(m.group(8)) if m.group(8) else None
    away_score = int(m.group(9)) if m.group(9) else None

    year = 2000 + year_short if year_short < 50 else 1900 + year_short

    # Team code to name
    team_map = {
        'ACS': 'AC Sparta Praha', 'SKS': 'SK Slavia Praha',
        'PLZ': 'FC Viktoria Plzeň', 'LIB': 'FC Slovan Liberec',
        'FCB': 'FC Baník Ostrava', 'MBL': 'FK Mladá Boleslav',
        'FKJ': 'FK Jablonec', 'BOH': 'Bohemians Praha 1905',
        'FCS': '1.FC Slovácko', 'SIG': 'SK Sigma Olomouc',
        'TEP': 'FK Teplice', 'HKR': 'FC Hradec Králové',
        'PCE': 'FK Pardubice', 'KAR': 'MFK Karviná',
        'FCZ': 'FC Zlín', 'DUK': 'FK Dukla Praha',
    }

    return {
        'home_team': team_map.get(home_code, home_code),
        'away_team': team_map.get(away_code, away_code),
        'start_time': datetime(year, month, day, hour, minute),
        'home_score': home_score,
        'away_score': away_score,
    }


def _parse_uefa_style(line: str) -> Optional[Dict]:
    """UEFA: Juventus 2-1 Galatasaray (18:45 CET)"""
    # S časem
    pattern = r'(.+?)\s+(\d+)[-–:](\d+)\s+(.+?)\s*\((\d{1,2}):(\d{2})'
    m = re.search(pattern, line)
    if m:
        home = m.group(1).strip()
        away = m.group(4).strip()
        hour, minute = int(m.group(5)), int(m.group(6))
        return {
            'home_team': home,
            'away_team': away,
            'home_score': int(m.group(2)),
            'away_score': int(m.group(3)),
            'start_time': datetime.now().replace(hour=hour, minute=minute),
        }

    # Bez času
    pattern = r'(.+?)\s+(\d+)[-–](\d+)\s+(.+?)$'
    m = re.search(pattern, line)
    if m:
        return {
            'home_team': m.group(1).strip(),
            'away_team': m.group(4).strip(),
            'home_score': int(m.group(2)),
            'away_score': int(m.group(3)),
        }

    return None


def _parse_csv_style(line: str) -> Optional[Dict]:
    """CSV: Sparta,Slavia,2,1,14.2.2026 20:00

    Robustnější:
    - umí ignorovat úvodní kódový sloupec (např. "Ml."/"ACS"/"MBL")
    - umí ignorovat úvodní datumový sloupec (např. "27. 2. 2026,Sparta,Slavia,18:00")
    - umí sloučit rozpadlý název týmu typu "Ml." + "Boleslav"
    """
    if ',' not in line:
        return None

    parts = [p.strip() for p in line.split(',') if p.strip()]
    if len(parts) < 2:
        return None

    def _looks_like_code(tok: str) -> bool:
        t = (tok or '').strip()
        if not t or ' ' in t or len(t) > 5:
            return False
        if re.match(r'^[A-Za-z]{1,4}\.?$', t):
            return True
        if t.isupper() and re.match(r'^[A-Z0-9]{2,5}$', t):
            return True
        return False

    leading_dt = None
    if len(parts) >= 3 and _is_date(parts[0]):
        leading_dt = parts[0]
        parts = parts[1:]

    if len(parts) >= 3 and _looks_like_code(parts[0]) and (len(parts[1]) >= 5 or ' ' in parts[1]):
        parts = parts[1:]

    if len(parts) >= 3 and parts[0].endswith('.') and len(parts[0]) <= 4 and ' ' not in parts[1] and parts[1].isalpha():
        merged = f"{parts[0]} {parts[1]}"
        parts = [merged] + parts[2:]

    if len(parts) < 2:
        return None

    result: Dict[str, Any] = {'home_team': parts[0], 'away_team': parts[1]}
    tail = parts[2:]

    def _is_probable_time(tok: str) -> bool:
        # HH:MM typical kick-off time (minutes usually 00/15/30/45)
        m = re.match(r'^(\d{1,2}):(\d{2})$', (tok or '').strip())
        if not m:
            return False
        hh = int(m.group(1))
        mm = int(m.group(2))
        return 0 <= hh <= 23 and 0 <= mm <= 59 and mm in (0, 15, 30, 45) and hh >= 8

    def _is_score_colon(tok: str) -> bool:
        return bool(re.match(r'^(\d{1,2}):(\d{1,2})$', (tok or '').strip()))

    def _apply_dt_from_tail(tokens: list[str]) -> None:
        nonlocal result, leading_dt
        if not tokens:
            return
        last = tokens[-1].strip()

        # if we have leading date from first column, allow time-only last column
        if leading_dt and _is_probable_time(last):
            dt = _parse_datetime(f"{leading_dt} {last}")
            if dt:
                result['start_time'] = dt
            return

        # full datetime in last token
        dt = _parse_datetime(last)
        if dt:
            result['start_time'] = dt

    # --- score/time parsing in tail ---
    if tail:
        # Handle scores like "0:2"
        if _is_score_colon(tail[0]):
            try:
                hs, aw = tail[0].split(':', 1)
                result['home_score'] = int(hs)
                result['away_score'] = int(aw)
            except Exception:
                pass
            _apply_dt_from_tail(tail[1:])
            # If we have only date (no time), set a sensible default so the match is tipovatelný.
            if leading_dt and 'start_time' not in result:
                dt = _parse_datetime(f"{leading_dt} 18:00")
                if dt:
                    result['start_time'] = dt
            return result

        # Handle numeric scores "2,1"
        if len(tail) >= 2 and re.match(r'^\d+$', tail[0]) and re.match(r'^\d+$', tail[1]):
            result['home_score'] = int(tail[0])
            result['away_score'] = int(tail[1])
            _apply_dt_from_tail(tail[2:])
            return result

    _apply_dt_from_tail(tail)
    if leading_dt and 'start_time' not in result:
        # date-only without time; keep empty for preview unless you want default time
        dt = _parse_datetime(f"{leading_dt} 18:00")
        if dt:
            result['start_time'] = dt

    return result



def _parse_pipe_style(line: str) -> Optional[Dict]:
    """Pipe separated: Sparta|Slavia|2|1|14.2.2026 20:00

    Robustnější:
    - umí ignorovat úvodní kódový sloupec (např. "Ml."/"ACS"/"MBL")
    - umí ignorovat úvodní datumový sloupec (např. "27. 2. 2026 | Sparta | Slavia | 18:00")
    - umí sloučit rozpadlý název týmu typu "Ml." + "Boleslav"
    """
    if '|' not in line:
        return None

    parts = [p.strip() for p in line.split('|') if p.strip()]
    if len(parts) < 2:
        return None

    def _looks_like_code(tok: str) -> bool:
        t = (tok or '').strip()
        if not t or ' ' in t or len(t) > 5:
            return False
        if re.match(r'^[A-Za-z]{1,4}\.?$', t):
            return True
        if t.isupper() and re.match(r'^[A-Z0-9]{2,5}$', t):
            return True
        return False

    leading_date = None
    if len(parts) >= 3 and _is_date(parts[0]):
        leading_date = parts[0]
        parts = parts[1:]

    if len(parts) >= 3 and _looks_like_code(parts[0]) and (len(parts[1]) >= 5 or ' ' in parts[1]):
        parts = parts[1:]

    if len(parts) >= 3 and parts[0].endswith('.') and len(parts[0]) <= 4 and ' ' not in parts[1] and parts[1].isalpha():
        merged = f"{parts[0]} {parts[1]}"
        parts = [merged] + parts[2:]

    if len(parts) < 2:
        return None

    result: Dict[str, Any] = {'home_team': parts[0], 'away_team': parts[1]}
    tail = parts[2:]

    def _is_probable_time(tok: str) -> bool:
        m = re.match(r'^(\d{1,2}):(\d{2})$', (tok or '').strip())
        if not m:
            return False
        hh = int(m.group(1))
        mm = int(m.group(2))
        return 0 <= hh <= 23 and 0 <= mm <= 59 and mm in (0, 15, 30, 45) and hh >= 8

    def _is_score_colon(tok: str) -> bool:
        return bool(re.match(r'^(\d{1,2}):(\d{1,2})$', (tok or '').strip()))

    # If we have leading date and the next token is a time, treat it as kick-off time
    if leading_date and tail and _is_probable_time(tail[0]):
        dt = _parse_datetime(f"{leading_date} {tail[0]}")
        if dt:
            result['start_time'] = dt
        tail = tail[1:]

    # Handle score in format "0:2" (very common in pasted tables)
    if tail and _is_score_colon(tail[0]) and not _is_probable_time(tail[0]):
        try:
            hs, aw = tail[0].split(':', 1)
            result['home_score'] = int(hs)
            result['away_score'] = int(aw)
        except Exception:
            pass
        tail = tail[1:]
        if leading_date and 'start_time' not in result:
            dt = _parse_datetime(f"{leading_date} 18:00")
            if dt:
                result['start_time'] = dt

    # Handle numeric scores "2|1"
    if len(tail) >= 2 and re.match(r'^\d+$', str(tail[0])) and re.match(r'^\d+$', str(tail[1])):
        result['home_score'] = int(tail[0])
        result['away_score'] = int(tail[1])
        tail = tail[2:]

    # Remaining tail may contain full datetime
    if tail:
        dt = _parse_datetime(tail[0])
        if dt:
            result['start_time'] = dt
        elif leading_date and _is_probable_time(tail[0]):
            dt = _parse_datetime(f"{leading_date} {tail[0]}")
            if dt:
                result['start_time'] = dt

    if leading_date and 'start_time' not in result:
        dt = _parse_datetime(f"{leading_date} 18:00")
        if dt:
            result['start_time'] = dt

    return result



def _parse_datetime_first(line: str) -> Optional[Dict]:
    """Datum první: 14.2. 20:00 Sparta vs Slavia"""
    pattern = r'(\d{1,2})\.(\d{1,2})\.?\s+(\d{1,2}):(\d{2})\s+(.+?)\s+(?:vs|versus|-|–)\s+(.+?)$'
    m = re.search(pattern, line)
    if not m:
        return None

    day, month = int(m.group(1)), int(m.group(2))
    hour, minute = int(m.group(3)), int(m.group(4))
    year = datetime.now().year

    # Pokud je měsíc v minulosti, použij příští rok
    if month < datetime.now().month:
        year += 1

    return {
        'home_team': m.group(5).strip(),
        'away_team': m.group(6).strip(),
        'start_time': datetime(year, month, day, hour, minute),
    }


def _parse_score_style(line: str) -> Optional[Dict]:
    """Se skóre: Sparta - Slavia 2:1"""
    pattern = r'(.+?)\s+[-–]\s+(.+?)\s+(\d+)[:\-](\d+)'
    m = re.search(pattern, line)
    if not m:
        return None

    return {
        'home_team': m.group(1).strip(),
        'away_team': m.group(2).strip(),
        'home_score': int(m.group(3)),
        'away_score': int(m.group(4)),
    }


def _parse_vs_style(line: str) -> Optional[Dict]:
    """Vs style: Sparta vs Slavia"""
    pattern = r'(.+?)\s+(?:vs|versus)\s+(.+?)$'
    m = re.search(pattern, line, re.I)
    if not m:
        return None

    return {
        'home_team': m.group(1).strip(),
        'away_team': m.group(2).strip(),
    }


def _parse_dash_style(line: str) -> Optional[Dict]:
    """Dash style: Sparta - Slavia"""
    pattern = r'^(.+?)\s+[-–]\s+(.+?)$'
    m = re.search(pattern, line)
    if not m:
        return None

    return {
        'home_team': m.group(1).strip(),
        'away_team': m.group(2).strip(),
    }


def _parse_datetime(s: str) -> Optional[datetime]:
    """
    Parse datetime from string
    IMPROVED: Auto-fill missing year
    """
    s = s.strip()

    formats = [
        '%d. %m. %Y %H:%M',
        '%d. %m. %Y',
        '%d.%m.%Y %H:%M',
        '%d.%m.%Y',
        '%d/%m/%Y %H:%M',
        '%d/%m/%Y',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%d. %m. %H:%M',
        '%d.%m. %H:%M',
        '%d. %m.',
        '%d.%m.',
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            # IMPROVEMENT: If year is missing (1900), add current/next year
            if dt.year == 1900:
                year = datetime.now().year
                if dt.month < datetime.now().month:
                    year += 1
                dt = dt.replace(year=year)
            return dt
        except:
            continue

    return None




from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from flask import current_app, request, flash, redirect, url_for, send_file, session, Response, render_template_string
from flask_login import current_user, login_required

from models import AuditLog, ExtraAnswer, ExtraQuestion, ImportSession, Match, Round, Team, Tip, UndoStack, User, TeamAlias
from app_utils import admin_required, audit, compute_leaderboard, create_undo_point, ensure_selected_round, perform_undo, render_page, send_email_with_attachment, send_results_notification
from extensions import db


def _parse_fragmented_column_ocr(text: str) -> List[Dict]:
    """OCR z mobilni app - sloupce cteny oddelene: data, casy, tymy stridade."""
    dates, times, teams = [], [], []
    for raw in text.splitlines():
        l = raw.strip()
        if not l:
            continue
        if re.search(r'kolo|liga|česko|chance|uefa|premier|bundesliga', l, re.IGNORECASE):
            continue
        l = re.sub(r'^[yYfF]{1,2}\s+', '', l)
        l = re.sub(r'^[J][}]\s*|^[$@©%?()\[\]{}|]+\s*', '', l)
        l = re.sub(r'[\(\)©@$%|]+$', '', l).strip()
        if not l:
            continue
        if re.match(r'^\d{1,2}\.\d{1,2}\.?(\s*\d{4})?$', l):
            dates.append(l.rstrip('.'))
            continue
        if re.match(r'^\d{1,2}:\d{2}$', l):
            times.append(l)
            continue
        if re.match(r'^\d{1,4}$', l):
            continue
        if re.match(r'^[OoNn0\s.\-,«»]+$', l):
            continue
        if len(l) >= 3 and re.search(r'[A-ZÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ]', l):
            teams.append(l)

    print(f"🔍 OCR debug: dates={len(dates)} times={len(times)} teams={len(teams)}")
    print(f"   dates={dates}")
    print(f"   times={times}")
    print(f"   teams={teams}")

    if len(teams) < 2 or not dates:
        return []

    n = len(dates)
    pairs_interleaved = []
    for i in range(min(n, len(teams) // 2)):
        if i*2+1 < len(teams):
            pairs_interleaved.append((teams[i*2], teams[i*2+1]))

    half = len(teams) // 2
    pairs_block = []
    for i in range(min(n, half)):
        if i + half < len(teams):
            pairs_block.append((teams[i], teams[i + half]))

    pairs = pairs_interleaved if len(pairs_interleaved) >= len(pairs_block) else pairs_block
    if not pairs:
        return []

    matches = []
    for i, (home, away) in enumerate(pairs):
        date_str = dates[i] if i < len(dates) else None
        time_str = times[i] if i < len(times) else None
        dt = None
        if date_str:
            if re.match(r'^\d{1,2}\.\d{1,2}$', date_str):
                date_str += f".{datetime.datetime.now().year}"
            try:
                full = (date_str + " " + time_str) if time_str else date_str
                dt = _parse_datetime(full.strip())
            except Exception:
                pass
        match: Dict = {'home_team': home, 'away_team': away}
        if dt:
            match['start_time'] = dt.isoformat()
        matches.append(match)

    if matches:
        print(f"✅ Fragmented column OCR parser: Nalezeno {len(matches)} zapasu")
    return matches

def normalize_team_name(name: str, round_id: int = None) -> str:
    """
    🧠 Inteligentní normalizace jména týmu

    - Opraví běžné překlepy
    - Doplní plný název (Slavia → SK Slavia Praha)
    - Najde podobné týmy v databázi
    """

    if name is None:
        return ""
    if isinstance(name, bool):
        return ""
    if not isinstance(name, str):
        name = str(name)
    name = name.strip()
    name = name.replace("\u00a0", " ").replace("\t", " ").strip()
    name = re.sub(r"\s+", " ", name)
    # Normalize common Czech football prefix patterns
    name = re.sub(r"\b1\.?\s*FC\b", "1. FC", name, flags=re.IGNORECASE)
    name = re.sub(r"\bFK\b", "FK", name)

    # Aliasy z DB (spravované v Admin UI)
    if round_id:
        try:
            al = TeamAlias.query.filter_by(round_id=round_id, alias=name).first()
            if not al:
                # case-insensitive match
                al = TeamAlias.query.filter(TeamAlias.round_id == round_id, db.func.lower(TeamAlias.alias) == name.lower()).first()
            if al and al.canonical_name:
                return al.canonical_name
        except Exception as e:
            print(f"⚠️ Chyba v TeamAlias lookup: {e}")

    # Zkratky → plné názvy
    short_to_full = {
        # Zkratky / běžné názvy → názvy, které typicky chceš mít v DB
        'Dukla': 'FK Dukla Praha',
        'Slavia': 'SK Slavia Praha',
        'Sparta': 'AC Sparta Praha',
        'Ostrava': 'FC Baník Ostrava',
        'Baník': 'FC Baník Ostrava',

        'Liberec': 'FC Slovan Liberec',
        'Hradec Kr.': 'FC Hradec Králové',
        'Hradec': 'FC Hradec Králové',

        'Ml. Boleslav': 'FK Mladá Boleslav',
        'Mladá Boleslav': 'FK Mladá Boleslav',
        'Jablonec': 'FK Jablonec',

        'Pardubice': 'FK Pardubice',
        'Teplice': 'FK Teplice',

        'Karviná': 'MFK Karviná',
        'Slovácko': '1.FC Slovácko',

        'Zlín': 'FC Zlín',
        'Plzeň': 'FC Viktoria Plzeň',

        'Olomouc': 'SK Sigma Olomouc',
        'Sigma': 'SK Sigma Olomouc',

        'Bohemians': 'Bohemians Praha 1905',
    }


    # Přesná shoda
    if name in short_to_full:
        return short_to_full[name]

    # Case-insensitive hledání
    for short, full in short_to_full.items():
        if name.lower() == short.lower():
            return full
        if name.lower() == full.lower():
            return full

    # Částečná shoda
    for short, full in short_to_full.items():
        if short.lower() in name.lower() or name.lower() in short.lower():
            return full
    # Pokud je round_id, zkus najít v DB (fuzzy match na existující Team.name)
    if round_id:
        try:
            if not isinstance(name, str):
                name = str(name)

            # Načti názvy týmů pro tuto soutěž
            team_names = [
                (t.name or "").strip()
                for t in Team.query.filter_by(round_id=round_id, is_deleted=False).all()
                if t and t.name
            ]

            needle = name.lower()
            for tname in team_names:
                t_low = tname.lower()
                if needle in t_low or t_low in needle:
                    return tname
        except Exception as e:
            print(f"⚠️ Chyba v normalize_team_name: {e}")
            pass
    # Žádná shoda - vrať original
    return name


def register_admin_core(app):
    @app.route("/admin/dashboard")
    @login_required
    def admin_dashboard():
        admin_required()

        # Stats
        total_users = User.query.count()
        total_rounds = Round.query.count()
        active_rounds = Round.query.filter_by(is_active=True).count()
        total_matches = Match.query.filter_by(is_deleted=False).count()
        total_tips = Tip.query.count()

        # Recent users
        recent_users = User.query.order_by(User.id.desc()).limit(5).all()

        # Active round stats
        active_round = Round.query.filter_by(is_active=True).first()
        active_round_stats = None
        if active_round:
            matches_total = Match.query.filter_by(round_id=active_round.id, is_deleted=False).count()
            matches_with_results = Match.query.filter(
                Match.round_id == active_round.id,
                Match.is_deleted == False,
                Match.home_score != None,
                Match.away_score != None
            ).count()
            tips_total = Tip.query.join(Match).filter(Match.round_id == active_round.id).count()

            active_round_stats = {
                'round': active_round,
                'matches_total': matches_total,
                'matches_with_results': matches_with_results,
                'tips_total': tips_total,
                'completion': int((matches_with_results / matches_total * 100) if matches_total > 0 else 0)
            }

        # Pending tasks
        matches_no_results = Match.query.filter(
            Match.is_deleted == False,
            db.or_(Match.home_score == None, Match.away_score == None)
        ).count()

        return render_page(r"""
    <style>
      .admin-dashboard {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
      }

      .stat-card {
    background: rgba(255,255,255,.03);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 20px;
    transition: all 0.3s ease;
      }

      .stat-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(0,0,0,.2);
      }

      .stat-value {
    font-size: 36px;
    font-weight: 900;
    margin-bottom: 8px;
    line-height: 1;
      }

      .progress-bar-container {
    background: rgba(255,255,255,.05);
    border-radius: 8px;
    height: 8px;
    margin-top: 12px;
    overflow: hidden;
      }

      .progress-bar-fill {
    height: 100%;
    background: linear-gradient(90deg, var(--accent), var(--ok));
    border-radius: 8px;
    transition: width 0.5s ease;
      }

      /* Admin cards */
      .admin-cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 12px;
      }

      .admin-card {
    background: rgba(110,168,254,0.08);
    border: 1px solid rgba(110,168,254,0.2);
    border-radius: 16px;
    padding: 16px;
    cursor: pointer;
    transition: all 0.2s;
    text-decoration: none;
    color: #e9eefc;
    display: flex;
    flex-direction: column;
    gap: 8px;
      }

      .admin-card:hover {
    background: rgba(110,168,254,0.15);
    border-color: rgba(110,168,254,0.3);
    transform: translateY(-2px);
    box-shadow: 0 4px 16px rgba(0,0,0,0.2);
      }

      .admin-card-header {
    display: flex;
    align-items: center;
    gap: 12px;
      }

      .admin-card-icon {
    font-size: 24px;
      }

      .admin-card-title {
    font-size: 15px;
    font-weight: 700;
      }

      .admin-card-desc {
    font-size: 12px;
    color: #94a3b8;
    line-height: 1.4;
      }
    </style>

    <div class="card">
      <h1 style="margin: 0 0 8px 0;">👨‍💼 Admin Dashboard</h1>
      <div class="muted">Přehled a rychlé akce</div>
    </div>

    <div class="admin-dashboard">
      <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">👥 UŽIVATELÉ</h3>
    <div class="stat-value" style="color: var(--accent);">{{ total_users }}</div>
    <div class="muted" style="font-size: 13px;">Celkem registrovaných</div>
      </div>

      <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">🏆 SOUTĚŽE</h3>
    <div class="stat-value" style="color: var(--ok);">{{ total_rounds }}</div>
    <div class="muted" style="font-size: 13px;">{{ active_rounds }} aktivních</div>
      </div>

      <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">⚽ ZÁPASY</h3>
    <div class="stat-value" style="color: var(--warn);">{{ total_matches }}</div>
    <div class="muted" style="font-size: 13px;">V databázi</div>
      </div>

      <div class="stat-card">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: var(--muted);">🎯 TIPY</h3>
    <div class="stat-value" style="color: var(--danger);">{{ total_tips }}</div>
    <div class="muted" style="font-size: 13px;">Celkem odeslaných</div>
      </div>
    </div>

    {% if active_round_stats %}
    <div class="card">
      <h3 style="margin: 0 0 16px 0;">⭐ Aktivní soutěž: {{ active_round_stats.round.name }}</h3>

      <div class="row" style="justify-content: space-between; margin-bottom: 16px;">
    <div>
      <div class="muted" style="font-size: 13px;">Zápasy s výsledky</div>
      <div style="font-size: 24px; font-weight: 900;">
        {{ active_round_stats.matches_with_results }} / {{ active_round_stats.matches_total }}
      </div>
    </div>

    <div>
      <div class="muted" style="font-size: 13px;">Odesláno tipů</div>
      <div style="font-size: 24px; font-weight: 900;">
        {{ active_round_stats.tips_total }}
      </div>
    </div>

    <div>
      <div class="muted" style="font-size: 13px;">Dokončení</div>
      <div style="font-size: 24px; font-weight: 900; color: var(--ok);">
        {{ active_round_stats.completion }}%
      </div>
    </div>
      </div>

      <div class="progress-bar-container">
    <div class="progress-bar-fill" style="width: {{ active_round_stats.completion }}%;"></div>
      </div>
    </div>
    {% endif %}

    {% if matches_no_results > 0 %}
    <div class="card" style="background: rgba(249,199,79,.08); border-color: rgba(249,199,79,.3);">
      <h3 style="margin: 0 0 12px 0; color: #f9c74f;">⚠️ Čeká na vyřízení</h3>
      <div>
    <strong>{{ matches_no_results }}</strong> zápasů bez výsledků
    <a href="{{ url_for('admin_bulk_edit') }}" class="btn btn-sm" style="margin-left: 12px;">Zadat výsledky</a>
      </div>
    </div>
    {% endif %}

    <div class="card">
      <h3 style="margin: 0 0 16px 0;">👥 Noví uživatelé</h3>
      {% for user in recent_users %}
    <div style="padding: 8px 0; border-bottom: 1px solid var(--line);">
      <strong>{{ user.display_name }}</strong>
      <span class="muted" style="font-size: 12px; margin-left: 8px;">@{{ user.username }}</span>
      {% if user.is_admin %}<span class="tag pill-ok" style="margin-left: 8px;">Admin</span>{% endif %}
    </div>
      {% endfor %}
    </div>

    <div class="card">
      <h3 style="margin: 0 0 16px 0;">🔧 Admin nástroje</h3>
      <div class="admin-cards">

    <a href="{{ url_for('admin_bulk_edit') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">✏️</div>
        <div class="admin-card-title">Bulk Edit</div>
      </div>
      <div class="admin-card-desc">Zadávání výsledků</div>
    </a>

    <a href="{{ url_for('admin_import') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">📥</div>
        <div class="admin-card-title">Import</div>
      </div>
      <div class="admin-card-desc">Importovat data</div>
    </a>

    <a href="{{ url_for('admin_export_hub') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">📤</div>
        <div class="admin-card-title">Export</div>
      </div>
      <div class="admin-card-desc">Exportovat data</div>
    </a>

    <a href="{{ url_for('admin_undo') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">⏪</div>
        <div class="admin-card-title">Undo</div>
      </div>
      <div class="admin-card-desc">Vrátit změny</div>
    </a>

    <a href="{{ url_for('admin_rounds') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">🎮</div>
        <div class="admin-card-title">Soutěže</div>
      </div>
      <div class="admin-card-desc">Správa soutěží</div>
    </a>

    <a href="{{ url_for('admin_users') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">👥</div>
        <div class="admin-card-title">Uživatelé</div>
      </div>
      <div class="admin-card-desc">Správa userů</div>
    </a>

    <a href="{{ url_for('admin_api_sources') if 'admin_api_sources' in current_app.view_functions else '#' }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">🔌</div>
        <div class="admin-card-title">API Zdroje</div>
      </div>
      <div class="admin-card-desc">Fotbal & hokej API</div>
    </a>

    <a href="{{ url_for('admin_smart_import') }}" class="admin-card">
      <div class="admin-card-header">
        <div class="admin-card-icon">🤖</div>
        <div class="admin-card-title">Smart Import</div>
      </div>
      <div class="admin-card-desc">AI parsování zápasů</div>
    </a>

      </div>
    </div>

    """, total_users=total_users, total_rounds=total_rounds, active_rounds=active_rounds,
     total_matches=total_matches, total_tips=total_tips, recent_users=recent_users,
     active_round_stats=active_round_stats, matches_no_results=matches_no_results)

    # --- ADMIN BULK EDIT ---

    @app.route("/admin/bulk-edit")
    @login_required
    def admin_bulk_edit():
        admin_required()

        rid = ensure_selected_round()
        r = db.session.get(Round, rid) if rid else None

        if not r:
            flash("Vyber soutěž pro bulk editaci.", "error")
            return redirect(url_for("admin_rounds"))

        # Načti všechny zápasy
        matches = Match.query.filter_by(
            round_id=r.id,
            is_deleted=False
        ).order_by(Match.start_time.asc(), Match.id.asc()).all()

        # Stats
        total = len(matches)
        with_results = sum(1 for m in matches if m.home_score is not None and m.away_score is not None)
        without_results = total - with_results

        return render_page(r"""
    <style>
      .bulk-table {
    width: 100%;
    border-collapse: collapse;
      }

      .bulk-table th,
      .bulk-table td {
    padding: 12px 8px;
    text-align: left;
    border-bottom: 1px solid var(--line);
      }

      .bulk-table th {
    background: rgba(255,255,255,.03);
    font-weight: 900;
    position: sticky;
    top: 0;
      }

      .bulk-table input[type="number"] {
    width: 60px;
    text-align: center;
      }

      .match-row:hover {
    background: rgba(255,255,255,.03);
      }

      .match-row.has-result {
    background: rgba(51,209,122,.05);
      }
    </style>

    <div class="card">
      <div class="row" style="justify-content: space-between; align-items: center;">
    <div>
      <h2 style="margin: 0 0 8px 0;">✏️ Bulk Edit - Hromadné úpravy</h2>
      <div class="muted">Soutěž: <b>{{ r.name }}</b></div>
    </div>
    <div class="row" style="gap: 8px;">
      <div class="tag pill-ok">✅ {{ with_results }}</div>
      <div class="tag pill-warn">⏳ {{ without_results }}</div>
    </div>
      </div>
    </div>

    <form method="post" action="{{ url_for('admin_bulk_edit_save') }}">
      <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
      <input type="hidden" name="round_id" value="{{ r.id }}">

      <div class="card">
    <div class="row" style="justify-content: space-between; align-items: center; margin-bottom: 16px;">
      <h3 style="margin: 0;">📋 Zápasy ({{ total }})</h3>
      <button type="submit" class="btn btn-primary">💾 Uložit všechny změny</button>
    </div>

    <div class="bulk-table-wrapper" style="overflow-x: auto;">
      <table class="bulk-table">
        <thead>
          <tr>
            <th style="width: 40px;">#</th>
            <th>Domácí</th>
            <th style="width: 60px; text-align: center;">Skóre</th>
            <th>Hosté</th>
            <th style="width: 60px; text-align: center;">Skóre</th>
          </tr>
        </thead>
        <tbody>
          {% for m in matches %}
            <tr class="match-row {% if m.home_score is not none and m.away_score is not none %}has-result{% endif %}">
              <td>{{ loop.index }}</td>
              <td><strong>{{ m.home_team.name if m.home_team else '?' }}</strong></td>
              <td style="text-align: center;">
                <input type="number" 
                       name="match_{{ m.id }}_home" 
                       value="{{ m.home_score if m.home_score is not none else '' }}"
                       min="0" max="20">
              </td>
              <td><strong>{{ m.away_team.name if m.away_team else '?' }}</strong></td>
              <td style="text-align: center;">
                <input type="number" 
                       name="match_{{ m.id }}_away" 
                       value="{{ m.away_score if m.away_score is not none else '' }}"
                       min="0" max="20">
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <div style="margin-top: 16px; text-align: right;">
      <button type="submit" class="btn btn-primary">💾 Uložit všechny změny</button>
    </div>
      </div>
    </form>

    """, r=r, matches=matches, total=total, with_results=with_results, without_results=without_results)

    @app.route("/admin/bulk-edit/save", methods=["POST"])
    @login_required
    def admin_bulk_edit_save():
        admin_required()

        round_id = int(request.form.get("round_id"))
        r = db.session.get(Round, round_id)

        if not r:
            flash("Soutěž nenalezena.", "error")
            return redirect(url_for("admin_bulk_edit"))

        matches = Match.query.filter_by(round_id=r.id, is_deleted=False).all()

        updated_count = 0
        for match in matches:
            match_id = match.id

            home_score_str = request.form.get(f"match_{match_id}_home")
            away_score_str = request.form.get(f"match_{match_id}_away")

            old_home = match.home_score
            old_away = match.away_score

            if home_score_str and away_score_str:
                new_home = int(home_score_str)
                new_away = int(away_score_str)

                if old_home != new_home or old_away != new_away:
                    # Vytvoř undo point PŘED změnou
                    home_team_name = match.home_team.name if match.home_team else "?"
                    away_team_name = match.away_team.name if match.away_team else "?"
                    old_score_str = f"{old_home}:{old_away}" if old_home is not None and old_away is not None else "—"
                    new_score_str = f"{new_home}:{new_away}"

                    create_undo_point(
                        action_type='update_score',
                        entity_type='Match',
                        entity_id=match.id,
                        before_state={
                            'home_score': old_home,
                            'away_score': old_away
                        },
                        description=f"{home_team_name} {old_score_str} → {new_score_str} {away_team_name}"
                    )

                    # Teď proveď změnu
                    match.home_score = new_home
                    match.away_score = new_away
                    updated_count += 1
            elif home_score_str == '' and away_score_str == '':
                if match.home_score is not None or match.away_score is not None:
                    # Vytvoř undo point PŘED smazáním
                    home_team_name = match.home_team.name if match.home_team else "?"
                    away_team_name = match.away_team.name if match.away_team else "?"
                    old_score_str = f"{old_home}:{old_away}" if old_home is not None and old_away is not None else "—"

                    create_undo_point(
                        action_type='clear_score',
                        entity_type='Match',
                        entity_id=match.id,
                        before_state={
                            'home_score': old_home,
                            'away_score': old_away
                        },
                        description=f"{home_team_name} {old_score_str} → smazáno {away_team_name}"
                    )

                    match.home_score = None
                    match.away_score = None
                    updated_count += 1

        db.session.commit()
        audit("bulk_edit.save", "Match", None, details=f"Updated {updated_count} matches")

        # Pošli push notifikace o zadaných výsledcích
        if updated_count > 0:
            try:
                send_results_notification(round_id)
            except Exception as e:
                print(f"Error sending push notifications: {e}")

        flash(f"✅ Aktualizováno {updated_count} zápasů!", "ok")
        return redirect(url_for("admin_bulk_edit"))

    # --- ADMIN BULK IMPORT (CSV) WITH PREVIEW ---

    @app.route("/admin/bulk-import/template")
    @login_required
    def admin_bulk_import_template():
        """Stáhne CSV šablonu pro bulk import"""
        admin_required()

        try:
            from io import StringIO

            # Create CSV template
            output = StringIO()
            output.write("Domácí,Hosté,Datum,Čas\n")
            output.write("Sparta Praha,Slavia Praha,2024-03-15,18:00\n")
            output.write("Plzeň,Brno,2024-03-16,16:30\n")
            output.write("Baník,Bohemians,2024-03-17,15:00\n")

            # Create response
            from flask import Response
            response = Response(output.getvalue(), mimetype='text/csv')
            response.headers['Content-Disposition'] = 'attachment; filename=bulk_import_sablona.csv'
            return response

        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_bulk_import"))

    @app.route("/admin/bulk-import", methods=["GET", "POST"])
    @login_required
    def admin_bulk_import():
        """Hromadný import týmů a zápasů z CSV - STEP 1: Upload"""
        admin_required()

        if request.method == "POST":
            import_type = request.form.get("import_type")
            round_id = request.form.get("round_id")

            if not round_id:
                flash("Vyber soutěž!", "error")
                return redirect(url_for("admin_bulk_import"))

            round_id = int(round_id)
            r = db.session.get(Round, round_id)

            if not r:
                flash("Soutěž nenalezena!", "error")
                return redirect(url_for("admin_bulk_import"))

            if 'csv_file' not in request.files:
                flash("Žádný soubor nevybrán!", "error")
                return redirect(url_for("admin_bulk_import"))

            file = request.files['csv_file']

            if file.filename == '':
                flash("Žádný soubor nevybrán!", "error")
                return redirect(url_for("admin_bulk_import"))

            if not file.filename.endswith('.csv'):
                flash("Musí být CSV soubor!", "error")
                return redirect(url_for("admin_bulk_import"))

            try:
                # Read CSV and store in temporary file (session cookies are limited to 4KB)
                csv_content = file.stream.read().decode("utf-8")

                # Create temporary file
                temp_fd, temp_path = tempfile.mkstemp(suffix='.csv', prefix='bulk_import_')
                with os.fdopen(temp_fd, 'w', encoding='utf-8') as f:
                    f.write(csv_content)

                # Store only the file path in session (small!)
                session['bulk_import_file'] = temp_path
                session['bulk_import_type'] = import_type
                session['bulk_import_round_id'] = round_id

                # Redirect to preview
                return redirect(url_for("admin_bulk_import_preview"))

            except Exception as e:
                flash(f"❌ Chyba při čtení CSV: {str(e)}", "error")
                return redirect(url_for("admin_bulk_import"))

        # GET request - show form
        rounds = Round.query.order_by(Round.id.desc()).all()

        return render_page(r"""
    <div class="card">
      <div class="row" style="justify-content:space-between; align-items:flex-start;">
    <div>
      <h2>📥 Hromadný import z CSV</h2>
      <div class="muted">Import týmů a zápasů z CSV souborů s preview</div>
    </div>
    <a href="{{ url_for('admin_bulk_import_template') }}" class="btn" style="background:#6ea8fe; color:white;">
      📥 Stáhnout šablonu CSV
    </a>
      </div>
    </div>

    <div class="card">
      <h3>📋 Import týmů</h3>
      <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="import_type" value="teams">

    <div style="margin-bottom: 16px;">
      <label class="muted">Soutěž *</label>
      <select name="round_id" required>
        <option value="">-- Vyber soutěž --</option>
        {% for r in rounds %}
          <option value="{{ r.id }}">{{ r.name }}</option>
        {% endfor %}
      </select>
    </div>

    <div style="margin-bottom: 16px;">
      <label class="muted">CSV soubor *</label>
      <input type="file" name="csv_file" accept=".csv" required>
      <div class="muted" style="font-size: 13px; margin-top: 4px;">
        Formát: <code>name</code> (1 sloupec)
      </div>
    </div>

    <button type="submit" class="btn btn-primary">👁️ Zobrazit preview</button>
      </form>
    </div>

    <div class="card">
      <h3>⚽ Import zápasů</h3>
      <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="import_type" value="matches">

    <div style="margin-bottom: 16px;">
      <label class="muted">Soutěž *</label>
      <select name="round_id" required>
        <option value="">-- Vyber soutěž --</option>
        {% for r in rounds %}
          <option value="{{ r.id }}">{{ r.name }}</option>
        {% endfor %}
      </select>
    </div>

    <div style="margin-bottom: 16px;">
      <label class="muted">CSV soubor *</label>
      <input type="file" name="csv_file" accept=".csv" required>
      <div class="muted" style="font-size: 13px; margin-top: 4px;">
        Formát: <code>home_team,away_team,start_time,home_score,away_score</code><br>
        Start time: <code>YYYY-MM-DD HH:MM</code> nebo <code>YYYY-MM-DD</code><br>
        Scores: Nechej prázdné pokud zápas ještě nebyl
      </div>
    </div>

    <button type="submit" class="btn btn-primary">👁️ Zobrazit preview</button>
      </form>
    </div>

    <div class="card">
      <h3>👤 Import tipů</h3>
      <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <input type="hidden" name="import_type" value="tips">

    <div style="margin-bottom: 16px;">
      <label class="muted">Soutěž *</label>
      <select name="round_id" required>
        <option value="">-- Vyber soutěž --</option>
        {% for r in rounds %}
          <option value="{{ r.id }}">{{ r.name }}</option>
        {% endfor %}
      </select>
    </div>

    <div style="margin-bottom: 16px;">
      <label class="muted">CSV soubor *</label>
      <input type="file" name="csv_file" accept=".csv" required>
      <div class="muted" style="font-size: 13px; margin-top: 4px;">
        Formát: <code>user_email,home_team,away_team,home_score,away_score</code><br>
        Příklad: <code>mejlacz@gmail.com,Česko,Kanada,3,2</code><br>
        <strong>Možnost přepsání:</strong> Existující tipy budou označeny jako "Update"
      </div>
    </div>

    <button type="submit" class="btn btn-primary">👁️ Zobrazit preview</button>
      </form>
    </div>

    <div class="card" style="background: rgba(255,255,255,.02);">
      <h3>💡 Jak na to</h3>
      <ol style="margin: 0; padding-left: 20px;">
    <li>Vyber CSV soubor → zobrazí se <strong>preview</strong></li>
    <li>Zkontroluj data v tabulce</li>
    <li>Klikni "Potvrdit" nebo "Zrušit"</li>
    <li>Pořadí: <strong>týmy</strong> → <strong>zápasy</strong> → <strong>tipy</strong></li>
    <li><strong>Update tipů:</strong> Existující tipy lze přepsat (status "Update")</li>
      </ol>
    </div>
    """, rounds=rounds)

    @app.route("/admin/bulk-import/preview")
    @login_required
    def admin_bulk_import_preview():
        """STEP 2: Preview dat před importem"""
        admin_required()

        # Get data from session
        temp_file = session.get('bulk_import_file')
        import_type = session.get('bulk_import_type')
        round_id = session.get('bulk_import_round_id')

        if not temp_file or not import_type or not round_id:
            flash("Session expirovala, nahraj CSV znovu", "error")
            return redirect(url_for("admin_bulk_import"))

        # Check if temp file still exists
        if not os.path.exists(temp_file):
            flash("CSV soubor expiroval, nahraj znovu", "error")
            session.pop('bulk_import_file', None)
            return redirect(url_for("admin_bulk_import"))

        r = db.session.get(Round, round_id)
        if not r:
            flash("Soutěž nenalezena", "error")
            return redirect(url_for("admin_bulk_import"))

        try:
            # Read CSV from temp file
            with open(temp_file, 'r', encoding='utf-8') as f:
                csv_content = f.read()

            stream = io.StringIO(csv_content, newline=None)
            csv_reader = csv.DictReader(stream)
            rows = list(csv_reader)

            preview_data = []
            new_count = 0
            overwrite_count = 0
            error_count = 0
            errors = []

            if import_type == 'teams':
                for idx, row in enumerate(rows):
                    team_name = row.get('name', '').strip()

                    if not team_name:
                        error_count += 1
                        errors.append("Prázdný název týmu")
                        continue

                    existing = Team.query.filter_by(round_id=round_id, name=team_name).first()

                    status = "overwrite" if existing else "new"
                    if status == "new":
                        new_count += 1
                    else:
                        overwrite_count += 1

                    preview_data.append({
                        'index': idx,
                        'name': team_name,
                        'status': status,
                        'selected': status == 'new',  # Auto-select only new
                        'existing_id': existing.id if existing else None
                    })

            elif import_type == 'matches':
                for idx, row in enumerate(rows):
                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    start_time_str = row.get('start_time', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    if not home_name or not away_name:
                        error_count += 1
                        errors.append(f"Chybí název týmu")
                        preview_data.append({
                            'index': idx,
                            'home_team': home_name or '?',
                            'away_team': away_name or '?',
                            'start_time': start_time_str,
                            'home_score': home_score_str,
                            'away_score': away_score_str,
                            'status': 'error',
                            'error': 'Chybí název týmu',
                            'selected': False
                        })
                        continue

                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()

                    if not home_team or not away_team:
                        error_count += 1
                        missing = []
                        if not home_team:
                            missing.append(home_name)
                        if not away_team:
                            missing.append(away_name)
                        error_msg = f"Tým nenalezen: {', '.join(missing)}"
                        errors.append(error_msg)
                        preview_data.append({
                            'index': idx,
                            'home_team': home_name,
                            'away_team': away_name,
                            'start_time': start_time_str,
                            'home_score': home_score_str,
                            'away_score': away_score_str,
                            'status': 'error',
                            'error': error_msg,
                            'selected': False
                        })
                        continue

                    # Check existing
                    existing = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    status = "new"
                    if existing:
                        # Zápas již existuje - můžeš ho přepsat
                        status = "overwrite"
                        overwrite_count += 1  # Count as overwrite
                    else:
                        new_count += 1

                    preview_data.append({
                        'index': idx,
                        'home_team': home_name,
                        'away_team': away_name,
                        'start_time': start_time_str,
                        'home_score': home_score_str if home_score_str else '—',
                        'away_score': away_score_str if away_score_str else '—',
                        'status': status,
                        'selected': status == 'new',  # Auto-select only NEW (not overwrite by default)
                        'existing_id': existing.id if existing else None  # Store ID for update
                    })

            elif import_type == 'tips':
                for idx, row in enumerate(rows):
                    user_email = row.get('user_email', '').strip()
                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    # Validate basic data
                    if not user_email or not home_name or not away_name or not home_score_str or not away_score_str:
                        error_count += 1
                        errors.append(f"Neúplná data na řádku {idx+1}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email or '?',
                            'home_team': home_name or '?',
                            'away_team': away_name or '?',
                            'tip': f"{home_score_str or '?'}:{away_score_str or '?'}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Neúplná data',
                            'selected': False
                        })
                        continue

                    # Find user
                    user = User.query.filter_by(email=user_email).first()
                    if not user:
                        error_count += 1
                        errors.append(f"Uživatel nenalezen: {user_email}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Uživatel nenalezen',
                            'selected': False
                        })
                        continue

                    # Find teams
                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()

                    if not home_team or not away_team:
                        error_count += 1
                        missing = []
                        if not home_team:
                            missing.append(home_name)
                        if not away_team:
                            missing.append(away_name)
                        error_msg = f"Tým nenalezen: {', '.join(missing)}"
                        errors.append(error_msg)
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': error_msg,
                            'selected': False
                        })
                        continue

                    # Find match
                    match = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if not match:
                        error_count += 1
                        errors.append(f"Zápas nenalezen: {home_name} vs {away_name}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Zápas nenalezen',
                            'selected': False
                        })
                        continue

                    # Parse tip scores
                    try:
                        new_home = int(home_score_str)
                        new_away = int(away_score_str)
                    except:
                        error_count += 1
                        errors.append(f"Neplatné skóre: {home_score_str}:{away_score_str}")
                        preview_data.append({
                            'index': idx,
                            'user_email': user_email,
                            'home_team': home_name,
                            'away_team': away_name,
                            'tip': f"{home_score_str}:{away_score_str}",
                            'current_tip': '—',
                            'status': 'error',
                            'error': 'Neplatné skóre',
                            'selected': False
                        })
                        continue

                    # Find existing tip
                    existing_tip = Tip.query.filter_by(
                        user_id=user.id,
                        match_id=match.id
                    ).first()

                    # Determine status
                    status = "new"
                    current_tip = "—"

                    if existing_tip:
                        current_tip = f"{existing_tip.home_score}:{existing_tip.away_score}"
                        if existing_tip.home_score != new_home or existing_tip.away_score != new_away:
                            status = "update"
                        else:
                            status = "skip"

                    if status == "new":
                        new_count += 1
                    elif status == "skip":
                        skip_count += 1
                    elif status == "update":
                        update_count += 1

                    preview_data.append({
                        'index': idx,
                        'user_email': user_email,
                        'user_name': user.username,
                        'home_team': home_name,
                        'away_team': away_name,
                        'tip': f"{new_home}:{new_away}",
                        'current_tip': current_tip,
                        'status': status,
                        'selected': status in ['new', 'update']  # Auto-select new and updates
                    })

            # Render preview
            if import_type == 'teams':
                table_html = """
    <table style="width: 100%; border-collapse: collapse;">
      <tr style="border-bottom: 2px solid var(--line);">
    <th style="padding: 12px 8px; text-align: center; width: 50px;">
      <input type="checkbox" id="select-all" checked>
    </th>
    <th style="padding: 12px 8px; text-align: left;">Název týmu</th>
    <th style="padding: 12px 8px; text-align: center; width: 120px;">Status</th>
      </tr>
      {% for item in preview_data %}
      <tr style="border-bottom: 1px solid var(--line);">
    <td style="padding: 10px 8px; text-align: center;">
      <input type="checkbox" name="selected_rows" value="{{ item.index }}" {% if item.selected %}checked{% endif %} class="row-checkbox">
    </td>
    <td style="padding: 10px 8px;">{{ item.name }}</td>
    <td style="padding: 10px 8px; text-align: center;">
      {% if item.status == 'new' %}
        <span class="tag" style="background: rgba(51,209,122,.15); color: #33d17a;">Nový</span>
      {% elif item.status == 'overwrite' %}
        <span class="tag" style="background: rgba(99,179,237,.15); color: #63b3ed;">Přepsat</span>
      {% endif %}
    </td>
      </tr>
      {% endfor %}
    </table>
    """
            elif import_type == 'matches':
                table_html = """
    <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
      <tr style="border-bottom: 2px solid var(--line);">
    <th style="padding: 10px 8px; text-align: center; width: 50px;">
      <input type="checkbox" id="select-all" checked>
    </th>
    <th style="padding: 10px 8px; text-align: left;">Domácí</th>
    <th style="padding: 10px 8px; text-align: left;">Hosté</th>
    <th style="padding: 10px 8px; text-align: center; width: 100px;">Skóre</th>
    <th style="padding: 10px 8px; text-align: center; width: 120px;">Status</th>
      </tr>
      {% for item in preview_data %}
      <tr style="border-bottom: 1px solid var(--line);">
    <td style="padding: 8px; text-align: center;">
      <input type="checkbox" name="selected_rows" value="{{ item.index }}" {% if item.selected %}checked{% endif %} class="row-checkbox">
    </td>
    <td style="padding: 8px;">{{ item.home_team }}</td>
    <td style="padding: 8px;">{{ item.away_team }}</td>
    <td style="padding: 8px; text-align: center;">
      {% if item.home_score != '—' and item.away_score != '—' %}
        {{ item.home_score }}:{{ item.away_score }}
      {% else %}
        <span class="muted">—</span>
      {% endif %}
    </td>
    <td style="padding: 8px; text-align: center;">
      {% if item.status == 'new' %}
        <span class="tag" style="background: rgba(51,209,122,.15); color: #33d17a;">Nový</span>
      {% elif item.status == 'skip' %}
        <span class="tag" style="background: rgba(255,255,255,.05); color: var(--muted);">Přeskočit</span>
      {% elif item.status == 'overwrite' %}
        <span class="tag" style="background: rgba(99,179,237,.15); color: #63b3ed;">Přepsat</span>
      {% elif item.status == 'update' %}
        <span class="tag" style="background: rgba(249,199,79,.15); color: #f9c74f;">Update</span>
      {% elif item.status == 'error' %}
        <span class="tag" style="background: rgba(255,77,109,.15); color: #ff4d6d;">Chyba</span>
      {% endif %}
    </td>
      </tr>
      {% endfor %}
    </table>
    """
            else:  # tips
                table_html = """
    <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
      <tr style="border-bottom: 2px solid var(--line);">
    <th style="padding: 10px 8px; text-align: center; width: 50px;">
      <input type="checkbox" id="select-all" checked>
    </th>
    <th style="padding: 10px 8px; text-align: left;">Uživatel</th>
    <th style="padding: 10px 8px; text-align: left;">Domácí</th>
    <th style="padding: 10px 8px; text-align: left;">Hosté</th>
    <th style="padding: 10px 8px; text-align: center; width: 80px;">Nový tip</th>
    <th style="padding: 10px 8px; text-align: center; width: 80px;">Současný</th>
    <th style="padding: 10px 8px; text-align: center; width: 100px;">Status</th>
      </tr>
      {% for item in preview_data %}
      <tr style="border-bottom: 1px solid var(--line);">
    <td style="padding: 8px; text-align: center;">
      <input type="checkbox" name="selected_rows" value="{{ item.index }}" {% if item.selected %}checked{% endif %} class="row-checkbox">
    </td>
    <td style="padding: 8px;">
      {% if item.user_name %}
        <strong>{{ item.user_name }}</strong><br>
        <span class="muted" style="font-size: 11px;">{{ item.user_email }}</span>
      {% else %}
        {{ item.user_email }}
      {% endif %}
    </td>
    <td style="padding: 8px;">{{ item.home_team }}</td>
    <td style="padding: 8px;">{{ item.away_team }}</td>
    <td style="padding: 8px; text-align: center;">
      <strong>{{ item.tip }}</strong>
    </td>
    <td style="padding: 8px; text-align: center;">
      {% if item.current_tip != '—' %}
        {{ item.current_tip }}
      {% else %}
        <span class="muted">—</span>
      {% endif %}
    </td>
    <td style="padding: 8px; text-align: center;">
      {% if item.status == 'new' %}
        <span class="tag" style="background: rgba(51,209,122,.15); color: #33d17a;">Nový</span>
      {% elif item.status == 'skip' %}
        <span class="tag" style="background: rgba(255,255,255,.05); color: var(--muted);">Stejný</span>
      {% elif item.status == 'update' %}
        <span class="tag" style="background: rgba(249,199,79,.15); color: #f9c74f;">Přepsat</span>
      {% elif item.status == 'error' %}
        <span class="tag" style="background: rgba(255,77,109,.15); color: #ff4d6d;">Chyba</span>
      {% endif %}
    </td>
      </tr>
      {% endfor %}
    </table>
    """

            return render_page(r"""
    <div class="card">
      <h2>👁️ Preview importu</h2>
      <div class="muted">Zkontroluj data před potvrzením</div>
    </div>

    <div class="card">
      <div class="row" style="gap: 32px;">
    <div>
      <div class="muted">Celkem</div>
      <div style="font-size: 28px; font-weight: 900;">{{ total }}</div>
    </div>
    <div>
      <div class="muted">Nové</div>
      <div style="font-size: 28px; font-weight: 900; color: #33d17a;">{{ new_count }}</div>
    </div>
    {% if overwrite_count > 0 %}
    <div>
      <div class="muted">Přepsat</div>
      <div style="font-size: 28px; font-weight: 900; color: #63b3ed;">{{ overwrite_count }}</div>
    </div>
    {% endif %}
    {% if error_count > 0 %}
    <div>
      <div class="muted">Chyby</div>
      <div style="font-size: 28px; font-weight: 900; color: #ff4d6d;">{{ error_count }}</div>
    </div>
    {% endif %}
      </div>
    </div>

    {% if errors|length > 0 %}
    <div class="card" style="background: rgba(255,77,109,.08);">
      <h3>⚠️ Chyby ({{ errors|length }})</h3>
      {% for error in errors[:10] %}
    <div class="muted" style="font-size: 13px; margin-bottom: 4px;">• {{ error }}</div>
      {% endfor %}
      {% if errors|length > 10 %}
    <div class="muted" style="font-size: 13px; margin-top: 8px;">... a {{ errors|length - 10 }} dalších</div>
      {% endif %}
    </div>
    {% endif %}

    <div class="card">
      <h3>📊 Data ({{ preview_data|length }})</h3>

      <div style="margin-bottom: 12px; display: flex; gap: 12px; flex-wrap: wrap;">
    <button type="button" onclick="selectAll()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      ✅ Vybrat vše
    </button>
    <button type="button" onclick="deselectAll()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      ❌ Zrušit vše
    </button>
    <button type="button" onclick="selectNew()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      🟢 Jen nové
    </button>
    <button type="button" onclick="selectOverwrite()" class="btn" style="font-size: 13px; padding: 6px 12px;">
      🔵 Přepsat existující
    </button>
    <div style="flex: 1; text-align: right; line-height: 32px;">
      <span class="muted" style="font-size: 13px;">
        Vybráno: <strong id="selected-count">{{ new_count }}</strong> / {{ preview_data|length }}
      </span>
    </div>
      </div>

      <div style="overflow-x: auto; margin-top: 12px;">
    """ + table_html + """
      </div>
    </div>

    <script>
    function updateCount() {
      const checked = document.querySelectorAll('.row-checkbox:checked').length;
      document.getElementById('selected-count').textContent = checked;
      document.getElementById('confirm-count').textContent = checked;
    }

    function selectAll() {
      document.querySelectorAll('.row-checkbox').forEach(cb => cb.checked = true);
      document.getElementById('select-all').checked = true;
      updateCount();
    }

    function deselectAll() {
      document.querySelectorAll('.row-checkbox').forEach(cb => cb.checked = false);
      document.getElementById('select-all').checked = false;
      updateCount();
    }

    function selectNew() {
      deselectAll();
      document.querySelectorAll('tr').forEach(row => {
    const tag = row.querySelector('.tag');
    const checkbox = row.querySelector('.row-checkbox');
    if (tag && tag.textContent.trim() === 'Nový' && checkbox) {
      checkbox.checked = true;
    }
      });
      updateCount();
    }

    function selectOverwrite() {
      deselectAll();
      document.querySelectorAll('tr').forEach(row => {
    const tag = row.querySelector('.tag');
    const checkbox = row.querySelector('.row-checkbox');
    if (tag && tag.textContent.trim() === 'Přepsat' && checkbox) {
      checkbox.checked = true;
    }
      });
      updateCount();
    }

    function selectUpdates() {
      deselectAll();
      document.querySelectorAll('tr').forEach(row => {
    const tag = row.querySelector('.tag');
    const checkbox = row.querySelector('.row-checkbox');
    if (tag && tag.textContent.trim() === 'Update' && checkbox) {
      checkbox.checked = true;
    }
      });
      updateCount();
    }

    // Select all checkbox handler
    document.getElementById('select-all').addEventListener('change', function() {
      if (this.checked) {
    selectAll();
      } else {
    deselectAll();
      }
    });

    // Individual checkbox handler
    document.querySelectorAll('.row-checkbox').forEach(cb => {
      cb.addEventListener('change', updateCount);
    });

    // Initial count
    updateCount();
    </script>

    <form method="post" action="{{ url_for('admin_bulk_import_confirm') }}">
      <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
      <div class="card">
    <div class="row" style="gap: 12px;">
      <button type="submit" name="action" value="confirm" class="btn btn-primary" style="flex: 1;">
        ✅ Potvrdit import (<span id="confirm-count">{{ new_count + update_count }}</span>)
      </button>
      <a href="{{ url_for('admin_bulk_import') }}" class="btn" style="flex: 1; text-align: center; line-height: 40px;">
        ❌ Zrušit
      </a>
    </div>
      </div>
    </form>
    """, 
            preview_data=preview_data,
            total=len(preview_data),
            new_count=new_count,
            overwrite_count=overwrite_count,
            error_count=error_count,
            errors=errors,
            import_type=import_type,
            round_name=r.name
        )

        except Exception as e:
            flash(f"❌ Chyba při preview: {str(e)}", "error")
            import traceback
            traceback.print_exc()
            return redirect(url_for("admin_bulk_import"))

    @app.route("/admin/bulk-import/confirm", methods=["POST"])
    @login_required
    def admin_bulk_import_confirm():
        """STEP 3: Potvrzení a provedení importu"""
        admin_required()

        action = request.form.get("action")

        # Get temp file path
        temp_file = session.get('bulk_import_file')

        if action != "confirm":
            flash("Import zrušen", "ok")
            # Cleanup temp file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
            session.pop('bulk_import_file', None)
            session.pop('bulk_import_type', None)
            session.pop('bulk_import_round_id', None)
            return redirect(url_for("admin_bulk_import"))

        # Get data from session
        import_type = session.get('bulk_import_type')
        round_id = session.get('bulk_import_round_id')

        if not temp_file or not import_type or not round_id:
            flash("Session expirovala, nahraj CSV znovu", "error")
            return redirect(url_for("admin_bulk_import"))

        # Check if temp file still exists
        if not os.path.exists(temp_file):
            flash("CSV soubor expiroval, nahraj znovu", "error")
            session.pop('bulk_import_file', None)
            return redirect(url_for("admin_bulk_import"))

        r = db.session.get(Round, round_id)
        if not r:
            flash("Soutěž nenalezena", "error")
            return redirect(url_for("admin_bulk_import"))

        # Get selected rows from form
        selected_rows = request.form.getlist('selected_rows')
        if not selected_rows:
            flash("Nevybrané žádné záznamy k importu", "error")
            # Cleanup temp file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
            return redirect(url_for("admin_bulk_import_preview"))

        # Convert to set of integers for fast lookup
        selected_indices = set(int(idx) for idx in selected_rows)

        try:
            # Read CSV from temp file
            with open(temp_file, 'r', encoding='utf-8') as f:
                csv_content = f.read()

            stream = io.StringIO(csv_content, newline=None)
            csv_reader = csv.DictReader(stream)

            if import_type == 'teams':
                imported = 0
                skipped = 0

                for idx, row in enumerate(csv_reader):
                    # Skip if not selected
                    if idx not in selected_indices:
                        continue

                    team_name = row.get('name', '').strip()

                    if not team_name:
                        continue

                    existing = Team.query.filter_by(round_id=round_id, name=team_name).first()

                    if existing:
                        skipped += 1
                    else:
                        team = Team(round_id=round_id, name=team_name)
                        db.session.add(team)
                        imported += 1

                db.session.commit()
                audit("bulk_import.teams", "Team", None, details=f"Imported {imported}, skipped {skipped}")
                flash(f"✅ Importováno {imported} týmů, přeskočeno {skipped}", "ok")

            elif import_type == 'matches':
                imported = 0
                skipped = 0
                updated = 0
                errors = []

                for idx, row in enumerate(csv_reader):
                    # Skip if not selected
                    if idx not in selected_indices:
                        continue

                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    start_time_str = row.get('start_time', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    if not home_name or not away_name:
                        continue

                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()

                    if not home_team or not away_team:
                        continue

                    # Parse start time
                    start_time = None
                    if start_time_str:
                        try:
                            start_time = datetime.strptime(start_time_str, '%Y-%m-%d %H:%M')
                        except:
                            try:
                                start_time = datetime.strptime(start_time_str, '%Y-%m-%d')
                            except:
                                continue

                    # Parse scores
                    home_score = None
                    away_score = None

                    if home_score_str and away_score_str:
                        try:
                            home_score = int(home_score_str)
                            away_score = int(away_score_str)
                        except:
                            pass

                    # Check if match exists
                    existing = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()

                    if existing:
                        # OVERWRITE - přepsat existující zápas VŠEMI daty z CSV
                        if start_time is not None:
                            existing.start_time = start_time
                        existing.home_score = home_score
                        existing.away_score = away_score
                        updated += 1
                    else:
                        # Create new match
                        match = Match(
                            round_id=round_id,
                            home_team_id=home_team.id,
                            away_team_id=away_team.id,
                            start_time=start_time,
                            home_score=home_score,
                            away_score=away_score
                        )
                        db.session.add(match)
                        imported += 1

                db.session.commit()
                audit("bulk_import.matches", "Match", None, details=f"Imported {imported}, overwritten {updated}")

                if updated > 0:
                    flash(f"✅ Importováno {imported} nových zápasů, přepsáno {updated} existujících", "ok")
                else:
                    flash(f"✅ Importováno {imported} nových zápasů", "ok")

            elif import_type == 'tips':
                imported = 0
                updated = 0
                skipped = 0
                errors = []

                for idx, row in enumerate(csv_reader):
                    # Skip if not selected
                    if idx not in selected_indices:
                        continue

                    user_email = row.get('user_email', '').strip()
                    home_name = row.get('home_team', '').strip()
                    away_name = row.get('away_team', '').strip()
                    home_score_str = row.get('home_score', '').strip()
                    away_score_str = row.get('away_score', '').strip()

                    # Validate
                    if not user_email or not home_name or not away_name or not home_score_str or not away_score_str:
                        continue

                    # Find user
                    user = User.query.filter_by(email=user_email).first()
                    if not user:
                        continue

                    # Find teams
                    home_team = Team.query.filter_by(round_id=round_id, name=home_name).first()
                    away_team = Team.query.filter_by(round_id=round_id, name=away_name).first()
                    if not home_team or not away_team:
                        continue

                    # Find match
                    match = Match.query.filter_by(
                        round_id=round_id,
                        home_team_id=home_team.id,
                        away_team_id=away_team.id,
                        is_deleted=False
                    ).first()
                    if not match:
                        continue

                    # Parse scores
                    try:
                        new_home = int(home_score_str)
                        new_away = int(away_score_str)
                    except:
                        continue

                    # Find existing tip
                    existing_tip = Tip.query.filter_by(
                        user_id=user.id,
                        match_id=match.id
                    ).first()

                    if existing_tip:
                        # Update if different
                        if existing_tip.home_score != new_home or existing_tip.away_score != new_away:
                            existing_tip.home_score = new_home
                            existing_tip.away_score = new_away
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        # Create new tip
                        tip = Tip(
                            user_id=user.id,
                            match_id=match.id,
                            home_score=new_home,
                            away_score=new_away
                        )
                        db.session.add(tip)
                        imported += 1

                db.session.commit()
                audit("bulk_import.tips", "Tip", None, details=f"Imported {imported}, updated {updated}, skipped {skipped}")

                if updated > 0:
                    flash(f"✅ Importováno {imported} tipů, přepsáno {updated}, přeskočeno {skipped}", "ok")
                else:
                    flash(f"✅ Importováno {imported} tipů, přeskočeno {skipped}", "ok")

            # Clear session and cleanup temp file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
            session.pop('bulk_import_file', None)
            session.pop('bulk_import_type', None)
            session.pop('bulk_import_round_id', None)

        except Exception as e:
            flash(f"❌ Chyba při importu: {str(e)}", "error")
            import traceback
            traceback.print_exc()
            # Cleanup temp file on error too
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass

        return redirect(url_for("admin_bulk_import"))

    # --- ADMIN EXPORT ---

    @app.route("/admin/export/<what>")
    @login_required
    def admin_export(what):
        admin_required()

        if what == "users":
            users = User.query.all()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['ID', 'Username', 'Email', 'Jméno', 'Příjmení', 'Admin'])

            for u in users:
                writer.writerow([
                    u.id,
                    u.username,
                    u.email,
                    u.first_name or '',
                    u.last_name or '',
                    'Ano' if u.is_admin else 'Ne'
                ])

            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment; filename=users_export.csv'}
            )

        elif what == "matches":
            rid = ensure_selected_round()
            r = db.session.get(Round, rid) if rid else None

            if not r:
                flash("Vyber soutěž.", "error")
                return redirect(url_for("admin_dashboard"))

            matches = Match.query.filter_by(round_id=r.id, is_deleted=False).order_by(Match.start_time).all()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['ID', 'Domácí', 'Hosté', 'Skóre domácí', 'Skóre hosté'])

            for m in matches:
                writer.writerow([
                    m.id,
                    m.home_team.name if m.home_team else '',
                    m.away_team.name if m.away_team else '',
                    m.home_score if m.home_score is not None else '',
                    m.away_score if m.away_score is not None else ''
                ])

            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=matches_{r.name.replace(" ", "_")}.csv'}
            )

        else:
            flash("Neznámý export typ.", "error")
            return redirect(url_for("admin_dashboard"))

    # --- ADMIN UNDO ---

    @app.route("/admin/undo")
    @login_required
    def admin_undo():
        admin_required()

        # Načti poslední undo pointy (maximálně 20)
        recent_undos = UndoStack.query.filter_by(
            user_id=current_user.id,
            is_undone=False
        ).order_by(UndoStack.created_at.desc()).limit(20).all()

        # Stats
        total_undos = UndoStack.query.filter_by(user_id=current_user.id).count()
        available_undos = UndoStack.query.filter_by(user_id=current_user.id, is_undone=False).count()

        return render_page(r"""
    <div class="card">
      <h2 style="margin: 0 0 8px 0;">🔄 Undo - Vrátit změny</h2>
      <div class="muted">Možnost vrátit poslední akce zpět</div>
      <hr class="sep">

      <div class="row" style="gap: 12px; margin-bottom: 16px;">
    <div class="tag pill-ok">✅ {{ available_undos }} dostupných</div>
    <div class="tag">📊 {{ total_undos }} celkem</div>
      </div>

      {% if not recent_undos %}
    <div class="card" style="background: rgba(255,255,255,.03); text-align: center; padding: 32px;">
      <div style="font-size: 48px; margin-bottom: 16px;">📝</div>
      <div class="muted">Žádné akce k vrácení.</div>
      <div class="muted" style="font-size: 13px; margin-top: 8px;">
        Změny se zaznamenávají automaticky při Bulk Edit.
      </div>
    </div>
      {% else %}
    <table class="datatable">
      <thead>
        <tr>
          <th style="width: 140px;">Čas</th>
          <th style="width: 120px;">Typ</th>
          <th>Popis</th>
          <th style="width: 120px; text-align: center;">Akce</th>
        </tr>
      </thead>
      <tbody>
        {% for undo in recent_undos %}
          <tr>
            <td>{{ undo.created_at.strftime("%d.%m. %H:%M") }}</td>
            <td>
              <span class="tag pill-ok">{{ undo.entity_type }}</span>
            </td>
            <td>{{ undo.description or '-' }}</td>
            <td style="text-align: center;">
              <form method="post" action="{{ url_for('admin_undo_perform', undo_id=undo.id) }}" style="margin: 0;">
                <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                <button type="submit" class="btn btn-sm btn-primary" 
                        onclick="return confirm('Opravdu vrátit tuto změnu?')">
                  ↶ Vrátit
                </button>
              </form>
            </td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
      {% endif %}
    </div>

    <div class="card" style="background: rgba(110,168,254,.08); border-color: rgba(110,168,254,.3);">
      <h3 style="margin: 0 0 12px 0;">💡 Jak to funguje?</h3>
      <ul style="margin: 0; padding-left: 20px;">
    <li>Při každé důležité změně (Bulk Edit) se vytvoří undo point</li>
    <li>Můžeš kdykoli vrátit změnu pomocí tlačítka "Vrátit"</li>
    <li>Každý admin vidí pouze své vlastní undo pointy</li>
    <li>Akce lze vrátit pouze jednou</li>
    <li>Zobrazuje se max 20 posledních změn</li>
      </ul>
    </div>

    """, recent_undos=recent_undos, total_undos=total_undos, available_undos=available_undos)

    @app.route("/admin/undo/<int:undo_id>/perform", methods=["POST"])
    @login_required
    def admin_undo_perform(undo_id):
        admin_required()

        result = perform_undo(undo_id)

        if result['success']:
            flash(result['message'], "ok")
        else:
            flash(f"❌ {result['message']}", "error")

        return redirect(url_for("admin_undo"))

    # === BACKUP DATABÁZE ===

    @app.route("/admin/backup")
    @login_required
    def admin_backup():
        """Stránka pro správu záloh databáze"""
        admin_required()

        # Zjisti velikost aktuální databáze
        db_path = os.path.join(app.instance_path, 'tipovacka.db')
        try:
            db_size = os.path.getsize(db_path)
            db_size_mb = db_size / (1024 * 1024)
        except:
            db_size_mb = 0

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">💾 Záloha databáze</h2>
      <div class="muted">Pravidelně zálohuj data tipovačky</div>
      <hr class="sep">

      <div style="background:rgba(110,168,254,0.08); padding:16px; border-radius:8px; margin-bottom:16px;">
    <div class="row" style="justify-content:space-between;">
      <div>
        <strong>Aktuální databáze:</strong>
      </div>
      <div>
        <span class="tag">{{ "%.2f"|format(db_size_mb) }} MB</span>
      </div>
    </div>
      </div>

      <h3 style="margin:16px 0 8px 0;">Manual záloha</h3>
      <div class="muted" style="margin-bottom:12px;">
    Vytvoř zálohu kdykoliv kliknutím na tlačítko
      </div>

      <div class="row" style="gap:10px; margin-bottom:24px;">
    <a href="{{ url_for('admin_backup_download') }}" class="btn btn-primary">
      📥 Stáhnout backup
    </a>
    <a href="{{ url_for('admin_backup_email') }}" class="btn" style="background:#667eea; color:white;">
      📧 Poslat na email
    </a>
      </div>

      <div style="background:rgba(255,193,7,0.1); padding:12px; border-left:4px solid #ffc107; font-size:13px; margin-bottom:24px;">
    <strong>📧 Email backup:</strong> Odešle se na <strong>{{ current_user.email }}</strong>
      </div>

      <hr class="sep">

      <h3 style="margin:16px 0 8px 0;">⏰ Automatický backup</h3>
      <div class="muted" style="margin-bottom:12px;">
    Záloha se pošle automaticky každý den v 3:00
      </div>

      <div style="background:rgba(110,168,254,0.1); padding:16px; border-radius:8px; margin-bottom:16px;">
    <div class="row" style="gap:16px; align-items:flex-start;">
      <div style="flex:1;">
        <strong style="display:block; margin-bottom:8px;">FREE účet (PythonAnywhere)</strong>
        <div class="muted" style="font-size:13px; line-height:1.6;">
          ❌ Scheduled tasks nejsou dostupné<br>
          ✅ Použij manual backup tlačítka výše
        </div>
      </div>
      <div style="flex:1;">
        <strong style="display:block; margin-bottom:8px;">PAID účet / Wedos</strong>
        <div class="muted" style="font-size:13px; line-height:1.6;">
          ✅ Cron job dostupný<br>
          ✅ Denní automatický backup<br>
          ✅ Email všem adminům
        </div>
      </div>
    </div>
      </div>

      <details style="margin-top:16px;">
    <summary style="cursor:pointer; padding:12px; background:rgba(0,0,0,0.05); border-radius:4px; font-weight:600;">
      📖 Návod: Nastavení automatického backupu (pro budoucnost)
    </summary>
    <div style="padding:16px; background:rgba(0,0,0,0.02); border-radius:4px; margin-top:8px; font-size:13px;">
      <p><strong>Soubor: backup_daily.py</strong></p>
      <pre style="background:#0b1020; color:#6ea8fe; padding:12px; border-radius:4px; overflow-x:auto; font-size:11px; line-height:1.4;">#!/usr/bin/env python3
    import os, sys
    sys.path.insert(0, '/path/to/tipovacka')

    os.environ['SECRET_KEY'] = 'tvuj-secret-key'
    os.environ['SEND_REAL_EMAILS'] = 'true'
    os.environ['FROM_EMAIL'] = 'noreply@tvoje-domena.cz'
    os.environ['FROM_NAME'] = 'Tipovačka Backup'
    os.environ['SMTP_SERVER'] = 'smtp.wedos.com'
    os.environ['SMTP_PORT'] = '587'
    os.environ['SMTP_USERNAME'] = 'noreply@tvoje-domena.cz'
    os.environ['SMTP_PASSWORD'] = 'heslo'

    from app2 import app, db, User, send_email_with_attachment
    from datetime import datetime
    import zipfile, os
    from io import BytesIO

    with app.app_context():
    admins = User.query.filter(User.role.in_(['admin', 'owner'])).all()
    db_path = os.path.join(app.instance_path, 'tipovacka.db')

    memory_file = BytesIO()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(db_path, f'tipovacka_{timestamp}.db')

    memory_file.seek(0)
    zip_data = memory_file.read()
    size_mb = len(zip_data) / (1024 * 1024)

    for admin in admins:
        html = f"&lt;h1&gt;Backup {datetime.now().strftime('%d.%m.%Y')}&lt;/h1&gt;&lt;p&gt;Velikost: {size_mb:.2f} MB&lt;/p&gt;"
        send_email_with_attachment(admin.email, f"Záloha {datetime.now().strftime('%d.%m.%Y')}", html, "Backup", zip_data, f'backup_{timestamp}.zip')

    print(f"{datetime.now()} - Backup odeslán {len(admins)} adminům")</pre>

      <p style="margin-top:16px;"><strong>Cron job (Wedos/VPS):</strong></p>
      <pre style="background:#0b1020; color:#6ea8fe; padding:12px; border-radius:4px; font-size:12px;">0 3 * * * /usr/bin/python3 /path/to/backup_daily.py >> /var/log/backup.log 2>&1</pre>

      <p class="muted" style="margin-top:12px;">
        = Každý den ve 3:00 se spustí backup a pošle email všem adminům
      </p>
    </div>
      </details>

      <hr class="sep">
      <a href="{{ url_for('admin_users') }}" class="btn">← Zpět do admin</a>
    </div>
    """, db_size_mb=db_size_mb)

    @app.route("/admin/backup/download")
    @login_required
    def admin_backup_download():
        """Stáhne aktuální databázi jako .zip"""
        admin_required()

        try:
            import zipfile
            from io import BytesIO
            from datetime import datetime

            db_path = os.path.join(app.instance_path, 'tipovacka.db')

            if not os.path.exists(db_path):
                flash("Databáze nenalezena!", "error")
                return redirect(url_for("admin_backup"))

            # Vytvoř zip v paměti
            memory_file = BytesIO()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(db_path, f'tipovacka_{timestamp}.db')

            memory_file.seek(0)

            audit("backup.download", "Database", None, timestamp=timestamp)

            return send_file(
                memory_file,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'tipovacka_backup_{timestamp}.zip'
            )

        except Exception as e:
            flash(f"Chyba při vytváření backupu: {str(e)}", "error")
            return redirect(url_for("admin_backup"))

    @app.route("/admin/backup/email")
    @login_required
    def admin_backup_email():
        """Pošle backup databáze na email aktuálního admina"""
        admin_required()

        try:
            import zipfile
            from io import BytesIO
            from datetime import datetime

            db_path = os.path.join(app.instance_path, 'tipovacka.db')

            if not os.path.exists(db_path):
                flash("Databáze nenalezena!", "error")
                return redirect(url_for("admin_backup"))

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
            timestamp_file = datetime.now().strftime('%Y%m%d_%H%M%S')

            # Vytvoř zip
            memory_file = BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(db_path, f'tipovacka_{timestamp_file}.db')

            memory_file.seek(0)
            zip_data = memory_file.read()

            # Velikost backupu
            size_mb = len(zip_data) / (1024 * 1024)

            # Pošli email s přílohou
            html = f"""
            <!DOCTYPE html>
            <html>
            <head><meta charset="utf-8"></head>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <div style="background: #0b1020; color: white; padding: 20px; text-align: center;">
                        <h1>💾 Záloha databáze</h1>
                    </div>
                    <div style="background: #f4f4f4; padding: 30px;">
                        <h2>Záloha tipovačky</h2>
                        <p><strong>Datum:</strong> {timestamp}</p>
                        <p><strong>Velikost:</strong> {size_mb:.2f} MB</p>
                        <p><strong>Soubor:</strong> tipovacka_backup_{timestamp_file}.zip</p>

                        <div style="background: #e3f2fd; padding: 15px; border-left: 4px solid #2196f3; margin: 20px 0;">
                            <strong>💡 Doporučení:</strong><br>
                            • Ulož si backup na bezpečné místo<br>
                            • Nezveřejňuj ho (obsahuje všechna data)<br>
                            • Gmail: 15 GB free = ~5000 backupů
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """

            text = f"""Záloha tipovačky

    Datum: {timestamp}
    Velikost: {size_mb:.2f} MB
    Soubor: tipovacka_backup_{timestamp_file}.zip
            """

            # Odešli s přílohou
            success = send_email_with_attachment(
                to_email=current_user.email,
                subject=f"Záloha tipovačky - {timestamp}",
                html_body=html,
                text_body=text,
                attachment_data=zip_data,
                attachment_name=f'tipovacka_backup_{timestamp_file}.zip'
            )

            if success:
                flash(f"✅ Backup odeslán na: {current_user.email} ({size_mb:.2f} MB)", "ok")
                audit("backup.email", "Database", None, size_mb=f"{size_mb:.2f}")
            else:
                flash(f"⚠️ Chyba při odesílání. Zkus stáhnout manuálně.", "warning")

            return redirect(url_for("admin_backup"))

        except Exception as e:
            flash(f"Chyba: {str(e)}", "error")
            return redirect(url_for("admin_backup"))

    # --- ADMIN IMPORT ---

    @app.route("/admin/audit")
    @login_required
    def admin_audit():
        admin_required()
        logs = AuditLog.query.order_by(AuditLog.id.desc()).limit(200).all()
        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Historie změn (posledních 200)</h2>
      <hr class="sep">
      {% for l in logs %}
    <div>
      <b>{{ l.action }}</b> <span class="muted">{{ l.entity }}{% if l.entity_id %}#{{ l.entity_id }}{% endif %}</span>
      <div class="muted">{{ l.at.strftime("%Y-%m-%d %H:%M:%S") }} | {% if l.actor %}{{ l.actor.username }}{% else %}—{% endif %}</div>
      {% if l.details %}<div class="muted" style="white-space:pre-wrap; margin-top:6px;">{{ l.details }}</div>{% endif %}
    </div>
    {% if not loop.last %}<hr class="sep">{% endif %}
      {% endfor %}
    </div>
    """, logs=logs)

    # --- PWA (Progressive Web App) ---

    @app.route("/admin/export-hub", methods=["GET", "POST"])
    @login_required
    def admin_export_hub():
        """Centrální export hub s filtry"""
        admin_required()

        if request.method == "POST":
            # Co exportovat
            export_teams = request.form.get("export_teams") == "1"
            export_matches = request.form.get("export_matches") == "1"
            export_tips = request.form.get("export_tips") == "1"
            export_extras = request.form.get("export_extras") == "1"
            export_leaderboard = request.form.get("export_leaderboard") == "1"

            # Filtry
            round_id_str = request.form.get("round_id", "")
            round_id = int(round_id_str) if round_id_str and round_id_str != "0" else None
            user_id_str = request.form.get("user_id", "")
            user_id = int(user_id_str) if user_id_str and user_id_str != "0" else None
            only_finished = request.form.get("only_finished") == "1"
            include_deleted = request.form.get("include_deleted") == "1"
            format_type = request.form.get("format", "xlsx")

            if not any([export_teams, export_matches, export_tips, export_extras, export_leaderboard]):
                flash("Vyber alespoň jednu kategorii dat.", "error")
                return redirect(url_for("admin_export_hub"))

            # Vytvoř export
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from io import BytesIO
            import csv

            if format_type == "xlsx":
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

                # TÝMY
                if export_teams:
                    ws = wb.create_sheet("Týmy")
                    headers = ["ID", "Soutěž", "Název", "Skupina", "Země"]
                    header_fill = PatternFill("solid", fgColor="4472C4")
                    header_font = Font(color="FFFFFF", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = Team.query
                    if round_id:
                        query = query.filter_by(round_id=round_id)
                    if not include_deleted:
                        query = query.filter_by(is_deleted=False)

                    teams = query.all()
                    for row_i, team in enumerate(teams, 2):
                        ws.cell(row=row_i, column=1, value=team.id)
                        ws.cell(row=row_i, column=2, value=team.round.name if team.round else "")
                        ws.cell(row=row_i, column=3, value=team.name)
                        ws.cell(row=row_i, column=4, value=team.group or "")
                        ws.cell(row=row_i, column=5, value=team.country_code or "")

                    for col in range(1, 6):
                        ws.column_dimensions[chr(64 + col)].width = 20

                # ZÁPASY
                if export_matches:
                    ws = wb.create_sheet("Zápasy")
                    headers = ["ID", "Soutěž", "Datum", "Čas", "Domácí", "Hosté", "Skóre D", "Skóre H", "Stav"]
                    header_fill = PatternFill("solid", fgColor="217346")
                    header_font = Font(color="FFFFFF", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = Match.query
                    if round_id:
                        query = query.filter_by(round_id=round_id)
                    if not include_deleted:
                        query = query.filter_by(is_deleted=False)
                    if only_finished:
                        query = query.filter(Match.home_score != None, Match.away_score != None)

                    matches = query.order_by(Match.start_time.asc()).all()
                    for row_i, match in enumerate(matches, 2):
                        ws.cell(row=row_i, column=1, value=match.id)
                        ws.cell(row=row_i, column=2, value=match.round.name if match.round else "")
                        ws.cell(row=row_i, column=3, value=match.start_time.strftime("%Y-%m-%d") if match.start_time else "")
                        ws.cell(row=row_i, column=4, value=match.start_time.strftime("%H:%M") if match.start_time else "")
                        ws.cell(row=row_i, column=5, value=match.home_team.name if match.home_team else "")
                        ws.cell(row=row_i, column=6, value=match.away_team.name if match.away_team else "")
                        ws.cell(row=row_i, column=7, value=match.home_score if match.home_score is not None else "")
                        ws.cell(row=row_i, column=8, value=match.away_score if match.away_score is not None else "")

                        status = "Ukončen" if match.home_score is not None else ("Smazán" if match.is_deleted else "Naplánován")
                        ws.cell(row=row_i, column=9, value=status)

                    for col in range(1, 10):
                        ws.column_dimensions[chr(64 + col)].width = 15

                # TIPY
                if export_tips:
                    ws = wb.create_sheet("Tipy")
                    headers = ["ID", "Soutěž", "Uživatel", "Zápas", "Tip D", "Tip H", "Datum"]
                    header_fill = PatternFill("solid", fgColor="FFC000")
                    header_font = Font(color="000000", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = Tip.query.join(Match)
                    if round_id:
                        query = query.filter(Match.round_id == round_id)
                    if user_id:
                        query = query.filter(Tip.user_id == user_id)
                    if not include_deleted:
                        query = query.filter(Match.is_deleted == False)
                    if only_finished:
                        query = query.filter(Match.home_score != None)

                    tips = query.order_by(Tip.created_at.desc()).limit(10000).all()
                    for row_i, tip in enumerate(tips, 2):
                        ws.cell(row=row_i, column=1, value=tip.id)
                        ws.cell(row=row_i, column=2, value=tip.match.round.name if tip.match and tip.match.round else "")
                        ws.cell(row=row_i, column=3, value=tip.user.username if tip.user else "")
                        match_str = f"{tip.match.home_team.name} vs {tip.match.away_team.name}" if tip.match else ""
                        ws.cell(row=row_i, column=4, value=match_str)
                        ws.cell(row=row_i, column=5, value=tip.tip_home)
                        ws.cell(row=row_i, column=6, value=tip.tip_away)
                        ws.cell(row=row_i, column=7, value=tip.created_at.strftime("%Y-%m-%d %H:%M") if tip.created_at else "")

                    for col in range(1, 8):
                        ws.column_dimensions[chr(64 + col)].width = 20

                # EXTRA OTÁZKY
                if export_extras:
                    ws = wb.create_sheet("Extra")
                    headers = ["ID", "Soutěž", "Otázka", "Uzávěrka", "Odpovědí"]
                    header_fill = PatternFill("solid", fgColor="C65911")
                    header_font = Font(color="FFFFFF", bold=True)

                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font

                    query = ExtraQuestion.query
                    if round_id:
                        query = query.filter_by(round_id=round_id)
                    if not include_deleted:
                        query = query.filter_by(is_deleted=False)

                    questions = query.all()
                    for row_i, q in enumerate(questions, 2):
                        ws.cell(row=row_i, column=1, value=q.id)
                        ws.cell(row=row_i, column=2, value=q.round.name if q.round else "")
                        ws.cell(row=row_i, column=3, value=q.question)
                        ws.cell(row=row_i, column=4, value=q.deadline.strftime("%Y-%m-%d %H:%M") if q.deadline else "")
                        answer_count = ExtraAnswer.query.filter_by(question_id=q.id).count()
                        ws.cell(row=row_i, column=5, value=answer_count)

                    for col in range(1, 6):
                        ws.column_dimensions[chr(64 + col)].width = 25

                # ŽEBŘÍČEK
                if export_leaderboard and round_id:
                    r = db.session.get(Round, round_id)
                    if r:
                        ws = wb.create_sheet("Žebříček")
                        headers = ["Pořadí", "Uživatel", "Body", "Přesné", "Rozdíl", "Tendence", "Chybné"]
                        header_fill = PatternFill("solid", fgColor="70AD47")
                        header_font = Font(color="FFFFFF", bold=True)

                        for col, h in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col, value=h)
                            cell.fill = header_fill
                            cell.font = header_font

                        leaderboard = compute_leaderboard(r.id)
                        for row_i, entry in enumerate(leaderboard, 2):
                            ws.cell(row=row_i, column=1, value=entry.get("rank", ""))
                            ws.cell(row=row_i, column=2, value=entry.get("username", ""))
                            ws.cell(row=row_i, column=3, value=entry.get("total_points", 0))
                            ws.cell(row=row_i, column=4, value=entry.get("exact", 0))
                            ws.cell(row=row_i, column=5, value=entry.get("diff", 0))
                            ws.cell(row=row_i, column=6, value=entry.get("tend", 0))
                            ws.cell(row=row_i, column=7, value=entry.get("wrong", 0))

                        for col in range(1, 8):
                            ws.column_dimensions[chr(64 + col)].width = 15

                out = BytesIO()
                wb.save(out)
                out.seek(0)
                filename = f"export_{round_id or 'all'}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
                return send_file(out,
                               mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               download_name=filename,
                               as_attachment=True)

            else:  # CSV
                flash("CSV export coming soon!", "info")
                return redirect(url_for("admin_export_hub"))

        # GET: Formulář
        rounds = Round.query.order_by(Round.id.desc()).all()
        users = User.query.order_by(User.username.asc()).all()

        return render_page(r"""
    <style>
    .export-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
    @media (max-width: 768px) { .export-grid { grid-template-columns: 1fr; } }
    .export-category { 
      padding: 16px; background: rgba(255,255,255,.03); border: 1px solid var(--line); border-radius: 10px;
    }
    .export-category:has(input:checked) { background: rgba(110,168,254,.1); border-color: rgba(110,168,254,.5); }
    </style>

    <div class="card">
      <h2 style="margin:0 0 4px 0;">📤 Export dat</h2>
      <div class="muted">Stáhni data z tipovačky ve strukturovaném formátu</div>
      <hr class="sep">

      <form method="post">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <h3 style="margin:20px 0 12px 0;">Co exportovat</h3>
    <div class="export-grid">
      <label class="export-category">
        <input type="checkbox" name="export_teams" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">🏟️ Týmy</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Názvy týmů, skupiny, kódy zemí</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_matches" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">⚽ Zápasy</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Rozpis zápasů, data, časy, výsledky</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_tips" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">🎯 Tipy</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Všechny tipy uživatelů</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_extras" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">❓ Extra otázky</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Otázky a odpovědi</div>
      </label>

      <label class="export-category">
        <input type="checkbox" name="export_leaderboard" value="1" style="width:18px; height:18px;">
        <strong style="margin-left:8px;">🏆 Žebříček</strong>
        <div class="muted" style="font-size:12px; margin-top:4px;">Aktuální pořadí a statistiky</div>
      </label>
    </div>

    <h3 style="margin:24px 0 12px 0;">Filtry</h3>
    <div class="row" style="gap:16px; flex-wrap:wrap;">
      <div class="form-group" style="flex:1; min-width:200px;">
        <label>Soutěž</label>
        <select name="round_id">
          <option value="">Všechny soutěže</option>
          {% for round in rounds %}
            <option value="{{ round.id }}">{{ round.name }}</option>
          {% endfor %}
        </select>
      </div>

      <div class="form-group" style="flex:1; min-width:200px;">
        <label>Uživatel (jen pro tipy)</label>
        <select name="user_id">
          <option value="">Všichni uživatelé</option>
          {% for user in users %}
            <option value="{{ user.id }}">{{ user.username }}</option>
          {% endfor %}
        </select>
      </div>
    </div>

    <div style="margin:12px 0; display:flex; flex-direction:column; gap:8px;">
      <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
        <input type="checkbox" name="only_finished" value="1" style="width:18px; height:18px;">
        <span>Jen dokončené zápasy</span>
      </label>

      <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
        <input type="checkbox" name="include_deleted" value="1" style="width:18px; height:18px;">
        <span>Včetně smazaných</span>
      </label>
    </div>

    <h3 style="margin:24px 0 12px 0;">Formát</h3>
    <div class="row" style="gap:16px;">
      <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
        <input type="radio" name="format" value="xlsx" checked style="width:18px; height:18px;">
        <span>📊 Excel (.xlsx)</span>
      </label>
    </div>

    <div class="row" style="gap:12px; margin-top:24px;">
      <button type="submit" class="btn btn-primary">📥 Stáhnout export</button>
      <a href="{{ url_for('admin_dashboard') }}" class="btn">← Zpět</a>
    </div>
      </form>
    </div>
    """, rounds=rounds, users=users)

    @app.route("/admin/smart-import", methods=["GET", "POST"])
    @login_required
    def admin_smart_import():
        """
        🤖 ULTRA SMART IMPORT

        Nakopíruj zápasy odkudkoliv → AI je naparsuje → Preview → Import
        """
        admin_required()

        rounds = Round.query.order_by(Round.name).all()

        if request.method == "POST":
            action = request.form.get("action", "parse")

            if action == "parse":
                # KROK 1: Parsování
                round_id = request.form.get("round_id")
                import_mode = request.form.get("import_mode", "fixtures")

                # Check if screenshot pasted from clipboard
                screenshot_data = request.form.get('screenshot_data', '')

                if screenshot_data and screenshot_data.startswith('data:image'):
                    # Clipboard screenshot mode
                    print("📸 Clipboard screenshot detected")

                    try:
                        # Extract base64 data
                        import base64
                        # Format: data:image/png;base64,iVBORw0KG...
                        base64_data = screenshot_data.split(',')[1]
                        image_data = base64.b64decode(base64_data)

                        # Extract text using OCR
                        text = extract_text_from_screenshot(image_data)

                        if not text:
                            flash("❌ Nepodařilo se přečíst screenshot. Zkus paste text ručně.", "error")
                            return redirect(url_for("admin_smart_import"))

                        print(f"✅ OCR extracted text ({len(text)} chars)")

                    except Exception as e:
                        print(f"❌ Screenshot processing error: {e}")
                        flash(f"❌ Chyba při zpracování screenshotu: {e}", "error")
                        return redirect(url_for("admin_smart_import"))
                else:
                    # Text mode (default)
                    text = request.form.get("raw_text", "")

                if not text.strip():
                    flash("❌ Zadej nějaký text k parsování!", "error")
                    return redirect(url_for("admin_smart_import"))

                # Parse!
                matches = smart_parse_matches(text, round_id=int(round_id) if round_id else None)

                if not matches:
                    flash("❌ Nepodařilo se naparsovat žádné zápasy. Zkus jiný formát.", "error")
                    return redirect(url_for("admin_smart_import"))

                # Normalize team names
                if round_id:
                    for m in matches:
                        if 'home' in m and 'home_team' not in m:
                            m['home_team'] = m.pop('home')
                        if 'away' in m and 'away_team' not in m:
                            m['away_team'] = m.pop('away')
                        m['home_team'] = normalize_team_name(m['home_team'], int(round_id))
                        m['away_team'] = normalize_team_name(m['away_team'], int(round_id))

                # Store preview server-side (DB) – v cookie jen ID (spolehlivé i na multi-worker hostingu)
                try:
                    payload = json.dumps({"round_id": round_id, "import_mode": import_mode, "matches": matches}, ensure_ascii=False)
                    imp = ImportSession(user_id=current_user.id, kind="smart_import_matches", payload_json=payload)
                    db.session.add(imp)
                    db.session.commit()
                    session['smart_import_session_id'] = imp.id
                    session['import_round_id'] = round_id
                except Exception as e:
                    db.session.rollback()
                    flash(f"❌ Nepodařilo se uložit preview importu: {e}", "error")
                    return redirect(url_for("admin_smart_import"))

                flash(f"✅ Naparsováno {len(matches)} zápasů! Zkontroluj a uprav pokud potřeba.", "ok")
                return redirect(url_for("admin_smart_import") + "?preview=1")

            elif action == "import":
                # KROK 2: Import do DB
                matches_json = request.form.get("matches_data", "[]")
                round_id = request.form.get("round_id")
                import_mode = request.form.get("import_mode", "fixtures")

                if not round_id:
                    flash("❌ Vyber soutěž/kolo!", "error")
                    return redirect(url_for("admin_smart_import"))

                try:
                    matches = json.loads(matches_json)
                except:
                    flash("❌ Chyba při parsování dat!", "error")
                    return redirect(url_for("admin_smart_import"))

                # Get round first
                r = db.session.get(Round, int(round_id))
                if not r:
                    flash("❌ Kolo nenalezeno!", "error")
                    return redirect(url_for("admin_smart_import"))

                # Store round data before expunging
                round_id_int = r.id
                round_name = r.name

                # CRITICAL: Expunge all objects from session to avoid pollution
                db.session.expunge_all()

                print(f"🔍 Using round_id: {round_id_int}, round_name: {round_name}")

                # Import!
                imported = 0
                errors = []
                skipped = 0

                for match_data in matches:
                    print(f"\n🔍 Processing match_data: {match_data}")
                    print(f"🔍 match_data type: {type(match_data)}")
                    print(f"🔍 match_data keys: {match_data.keys() if isinstance(match_data, dict) else 'NOT A DICT!'}")

                    try:
                        home_team = match_data.get('home_team', '').strip()
                        away_team = match_data.get('away_team', '').strip()

                        print(f"🔍 Extracted teams: home='{home_team}' ({type(home_team)}), away='{away_team}' ({type(away_team)})")

                        if not home_team or not away_team:
                            print(f"⚠️ Přeskakuji zápas - chybí tým: home='{home_team}', away='{away_team}'")
                            skipped += 1
                            continue

                        # Start time
                        start_time = None
                        if match_data.get('start_time'):
                            try:
                                start_time_str = match_data['start_time']
                                if isinstance(start_time_str, str):
                                    start_time = datetime.fromisoformat(start_time_str)
                                elif isinstance(start_time_str, datetime):
                                    start_time = start_time_str
                                else:
                                    print(f"⚠️ Neznámý typ start_time: {type(start_time_str)}")
                                    start_time = None
                            except Exception as e:
                                print(f"⚠️ Chyba parsování času '{match_data.get('start_time')}': {e}")
                                start_time = None

                        # Scores
                        home_score = match_data.get('home_score')
                        away_score = match_data.get('away_score')

                        # Import mode handling
                        if import_mode == 'fixtures':
                            home_score = None
                            away_score = None

                        # Validate scores are integers or None
                        if home_score is not None:
                            try:
                                home_score = int(home_score)
                            except:
                                home_score = None

                        if away_score is not None:
                            try:
                                away_score = int(away_score)
                            except:
                                away_score = None

                        # Debug log before creating Match
                        print(f"🔍 Match data:")
                        print(f"   round_id={round_id_int} (type={type(round_id_int)})")
                        print(f"   home_team={home_team} (type={type(home_team)})")
                        print(f"   away_team={away_team} (type={type(away_team)})")
                        print(f"   start_time={start_time} (type={type(start_time)})")
                        print(f"   home_score={home_score} (type={type(home_score)})")
                        print(f"   away_score={away_score} (type={type(away_score)})")

                        # Ensure all values are correct types
                        if not isinstance(round_id_int, int):
                            print(f"❌ round_id není int: {type(round_id_int)}")
                            continue

                        if not isinstance(home_team, str) or not isinstance(away_team, str):
                            print(f"❌ Týmy nejsou string")
                            continue

                        if start_time is not None and not isinstance(start_time, datetime):
                            print(f"❌ start_time není datetime: {type(start_time)}")
                            start_time = None

                        if home_score is not None and not isinstance(home_score, int):
                            print(f"❌ home_score není int: {type(home_score)}")
                            home_score = None

                        if away_score is not None and not isinstance(away_score, int):
                            print(f"❌ away_score není int: {type(away_score)}")
                            away_score = None

                                                # Results mode: must have scores
                        if import_mode == 'results':
                            if home_score is None or away_score is None:
                                skipped += 1
                                continue

                        # Create match
                        # Resolve Team objects (Match expects FK IDs, not strings)
                        home_team_obj = Team.query.filter_by(
                            round_id=round_id_int,
                            name=home_team,
                            is_deleted=False
                        ).first()
                        if not home_team_obj:
                            home_team_obj = Team(round_id=round_id_int, name=home_team)
                            db.session.add(home_team_obj)
                            db.session.flush()

                        away_team_obj = Team.query.filter_by(
                            round_id=round_id_int,
                            name=away_team,
                            is_deleted=False
                        ).first()
                        if not away_team_obj:
                            away_team_obj = Team(round_id=round_id_int, name=away_team)
                            db.session.add(away_team_obj)
                            db.session.flush()

                        # Results mode: try to find existing match and update scores
                        if import_mode == 'results':
                            q = Match.query.filter_by(round_id=round_id_int, home_team_id=home_team_obj.id, away_team_id=away_team_obj.id)
                            if start_time is not None:
                                exact = q.filter(Match.start_time == start_time).first()
                                if exact:
                                    exact.home_score = home_score
                                    exact.away_score = away_score
                                    imported += 1
                                    print(f"✅ Aktualizován výsledek (exact): {home_team} - {away_team} {home_score}:{away_score}")
                                    continue
                            existing = q.order_by(Match.start_time.asc()).first()
                            if existing:
                                existing.home_score = home_score
                                existing.away_score = away_score
                                imported += 1
                                print(f"✅ Aktualizován výsledek: {home_team} - {away_team} {home_score}:{away_score}")
                                continue
                            print(f"⚠️ Nenalezen existující zápas pro výsledek: {home_team} - {away_team}")
                            skipped += 1
                            continue

                        m = Match(
                            round_id=round_id_int,
                            home_team_id=home_team_obj.id,
                            away_team_id=away_team_obj.id,
                            start_time=start_time,
                            home_score=home_score,
                            away_score=away_score,
                        )

                        print(f"🔍 Match object created: {m}")
                        print(f"🔍 Match type: {type(m)}")
                        print(f"🔍 Match.__dict__: {m.__dict__}")

                        db.session.add(m)

                        # Flush immediately to catch errors
                        try:
                            db.session.flush()
                            imported += 1
                            print(f"✅ Přidán zápas: {home_team} - {away_team}")
                        except Exception as flush_error:
                            db.session.rollback()
                            error_msg = f"{home_team} - {away_team}: FLUSH ERROR: {str(flush_error)}"
                            errors.append(error_msg)
                            print(f"❌ Flush error: {flush_error}")
                            print(f"❌ Match object: {m.__dict__}")
                            continue

                    except Exception as e:
                        error_msg = f"{home_team} - {away_team}: {str(e)}"
                        errors.append(error_msg)
                        print(f"❌ Chyba při importu: {error_msg}")

                try:
                    db.session.commit()
                    print(f"✅ DB commit úspěšný: {imported} zápasů")
                except Exception as e:
                    db.session.rollback()
                    print(f"❌ DB commit selhal: {e}")
                    flash(f"❌ Chyba při ukládání do databáze: {str(e)}", "error")
                    return redirect(url_for("admin_smart_import"))

                # Clear session
                session.pop('parsed_matches', None)
                session.pop('import_round_id', None)

                if errors:
                    flash(f"⚠️ Importováno {imported} zápasů, přeskočeno {skipped}, {len(errors)} chyb: {', '.join(errors[:3])}", "warning")
                elif skipped > 0:
                    flash(f"✅ Importováno {imported} zápasů, přeskočeno {skipped} (chybějící týmy)", "ok")
                else:
                    flash(f"✅ Úspěšně importováno {imported} zápasů!", "ok")

                audit("smart_import", "Match", None, count=imported, round=round_name)

                return redirect(url_for("admin_rounds"))

        # GET request
        preview_mode = request.args.get("preview") == "1"
        parsed_matches: List[Dict[str, Any]] = []
        import_round_id = session.get('import_round_id')
        import_mode = 'fixtures'

        if preview_mode:
            sid = session.get('smart_import_session_id')
            if sid:
                imp = db.session.get(ImportSession, int(sid))
                if imp and imp.user_id == current_user.id and imp.kind == "smart_import_matches":
                    try:
                        payload = json.loads(imp.payload_json or "{}")
                        import_round_id = payload.get("round_id", import_round_id)
                        import_mode = payload.get('import_mode', import_mode) or import_mode
                        parsed_matches = payload.get("matches", []) or []
                    except Exception:
                        parsed_matches = []

        return render_template_string(SMART_IMPORT_TEMPLATE,
                                      rounds=rounds,
                                      preview_mode=preview_mode,
                                      parsed_matches=parsed_matches,
                                      import_round_id=import_round_id,
                                      import_mode=import_mode)


    # Template pro Smart Import
    SMART_IMPORT_TEMPLATE = """<!DOCTYPE html>
    <html>
    <head>
      <meta charset="UTF-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>🤖 Smart Import | Tipovačka</title>
      <style>
    * { margin: 0; padding: 0; box-sizing: border-box; }

    body {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
      min-height: 100vh;
      padding: 20px;
      color: #e0e0e0;
    }

    .container {
      max-width: 1200px;
      margin: 0 auto;
    }

    .card {
      background: rgba(30, 30, 46, 0.95);
      border-radius: 16px;
      padding: 32px;
      box-shadow: 0 20px 60px rgba(0,0,0,0.5);
      margin-bottom: 24px;
      border: 1px solid rgba(255,255,255,0.1);
      backdrop-filter: blur(10px);
    }

    h1 {
      font-size: 32px;
      margin-bottom: 8px;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
    }

    .subtitle {
      color: #a0a0a0;
      margin-bottom: 24px;
      font-size: 16px;
    }

    .form-group {
      margin-bottom: 20px;
    }

    label {
      display: block;
      margin-bottom: 8px;
      font-weight: 600;
      color: #e0e0e0;
    }

    .muted {
      color: #888;
      font-size: 14px;
      margin-top: 4px;
    }

    select, textarea, input[type="text"], input[type="number"], input[type="datetime-local"] {
      width: 100%;
      padding: 12px;
      border: 2px solid rgba(255,255,255,0.1);
      border-radius: 8px;
      font-size: 16px;
      font-family: inherit;
      transition: all 0.2s;
      background: rgba(20, 20, 30, 0.8);
      color: #e0e0e0;
    }

    select:focus, textarea:focus, input:focus {
      outline: none;
      border-color: #667eea;
      background: rgba(20, 20, 30, 0.95);
    }

    textarea {
      min-height: 300px;
      font-family: 'Monaco', 'Courier New', monospace;
      resize: vertical;
      line-height: 1.6;
    }

    .btn {
      padding: 14px 28px;
      border: none;
      border-radius: 8px;
      font-size: 16px;
      font-weight: 600;
      cursor: pointer;
      transition: all 0.2s;
      display: inline-block;
    }

    .btn-primary {
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      color: white;
    }

    .btn-primary:hover {
      transform: translateY(-2px);
      box-shadow: 0 8px 24px rgba(102, 126, 234, 0.4);
    }

    .btn-success {
      background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
      color: white;
    }

    .btn-success:hover {
      transform: translateY(-2px);
      box-shadow: 0 8px 24px rgba(46, 204, 113, 0.4);
    }

    .btn-secondary {
      background: rgba(255,255,255,0.1);
      color: #e0e0e0;
      border: 1px solid rgba(255,255,255,0.2);
    }

    .btn-secondary:hover {
      background: rgba(255,255,255,0.15);
    }

    .examples {
      background: rgba(20, 20, 30, 0.6);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 20px;
      border: 1px solid rgba(255,255,255,0.05);
    }

    .examples h3 {
      margin-bottom: 12px;
      font-size: 18px;
      color: #667eea;
    }

    .example-item {
      background: rgba(10, 10, 15, 0.8);
      padding: 8px 12px;
      border-radius: 4px;
      margin-bottom: 8px;
      font-family: 'Monaco', 'Courier New', monospace;
      font-size: 14px;
      border-left: 4px solid #667eea;
      color: #a0a0a0;
    }

    .preview-table {
      width: 100%;
      border-collapse: collapse;
      margin-top: 20px;
    }

    .preview-table th {
      background: rgba(102, 126, 234, 0.2);
      padding: 12px;
      text-align: left;
      font-weight: 600;
      border-bottom: 2px solid #667eea;
      color: #e0e0e0;
    }

    .preview-table td {
      padding: 10px 12px;
      border-bottom: 1px solid rgba(255,255,255,0.05);
    }

    .preview-table input[type="text"],
    .preview-table input[type="number"],
    .preview-table input[type="datetime-local"],
    .preview-table input[type="checkbox"] {
      padding: 8px;
      font-size: 14px;
      background: rgba(20, 20, 30, 0.8);
      border: 1px solid rgba(255,255,255,0.1);
      color: #e0e0e0;
    }

    .preview-table input:focus {
      border-color: #667eea;
      background: rgba(20, 20, 30, 0.95);
    }

    .radio-group {
      display: flex;
      gap: 20px;
      margin-bottom: 20px;
      flex-wrap: wrap;
    }

    .radio-group label {
      display: flex;
      align-items: center;
      gap: 8px;
      cursor: pointer;
      padding: 10px 16px;
      background: rgba(20, 20, 30, 0.6);
      border-radius: 8px;
      border: 2px solid rgba(255,255,255,0.1);
      transition: all 0.2s;
      margin-bottom: 0;
      font-weight: normal;
    }

    .radio-group label:hover {
      border-color: #667eea;
      background: rgba(20, 20, 30, 0.8);
    }

    .radio-group input[type="radio"] {
      width: auto;
      margin: 0;
    }

    .radio-group input[type="radio"]:checked + span {
      color: #667eea;
      font-weight: 600;
    }

    .flash {
      padding: 12px 20px;
      border-radius: 8px;
      margin-bottom: 20px;
      border-left: 4px solid;
    }

    .flash.ok {
      background: rgba(46, 204, 113, 0.1);
      border-color: #2ecc71;
      color: #2ecc71;
    }

    .flash.error {
      background: rgba(231, 76, 60, 0.1);
      border-color: #e74c3c;
      color: #e74c3c;
    }

    .flash.warning {
      background: rgba(241, 196, 15, 0.1);
      border-color: #f1c40f;
      color: #f1c40f;
    }

    .back-link {
      color: #667eea;
      text-decoration: none;
      display: inline-block;
      margin-bottom: 20px;
      font-weight: 600;
    }

    .back-link:hover {
      text-decoration: underline;
    }

    @media (max-width: 768px) {
      body { padding: 10px; }
      .card { padding: 20px; }
      h1 { font-size: 24px; }
      .radio-group { flex-direction: column; }
    }
      .paste-zone {
      position: relative;
    }

    .paste-instruction {
      text-align: center;
      padding: 20px;
      background: rgba(102, 126, 234, 0.1);
      border-radius: 8px;
      margin-bottom: 16px;
      border: 2px dashed rgba(102, 126, 234, 0.3);
    }

    .paste-zone.has-image .paste-instruction {
      display: none;
    }
      </style>
    </head>
    <body>
      <div class="container">
    <a href="{{ url_for('admin_dashboard') }}" class="back-link">← Zpět na Admin</a>

    {% with messages = get_flashed_messages(with_categories=True) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="flash {{ category }}">{{ message }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    {% if not preview_mode %}
    <div class="card">
      <h1>🤖 Smart Import V2</h1>
      <div class="subtitle">Nakopíruj zápasy odkudkoliv → AI je naparsuje → Preview → Import</div>

      <form method="post" >
        <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
        <input type="hidden" name="action" value="parse"/>

        <div class="form-group">
          <label>Soutěž / Kolo *</label>
          <select name="round_id" required>
            <option value="">-- Vyber soutěž --</option>
            {% for r in rounds %}
            <option value="{{ r.id }}" {% if import_round_id and import_round_id|string == r.id|string %}selected{% endif %}>
              {{ r.name }}
            </option>
            {% endfor %}
          </select>
        </div>

        <div class="form-group">
          <label>Import Mode</label>
          <div class="radio-group">
            <label>
              <input type="radio" name="import_mode" value="fixtures" {% if import_mode != 'results' %}checked{% endif %}>
              <span>📅 Fixtures (rozpisy)</span>
            </label>
            <label>
              <input type="radio" name="import_mode" value="results" {% if import_mode == 'results' %}checked{% endif %}>
              <span>🎯 Results (výsledky)</span>
            </label>
          </div>
          <div class="muted">
            Fixtures = nové zápasy bez skóre | Results = aktualizuje skóre existujících
          </div>
        </div>

        <div class="form-group">
          <label>Zápasy - screenshot NEBO text</label>

          <!-- Screenshot Upload -->


          <div class="examples">
            <h3>💡 Podporované formáty:</h3>
            <div class="example-item">27. 2. 2026DuklaSlavia18:00</div>
            <div class="example-item">27. 2. 2026 Sparta - Slavia 18:00</div>
            <div class="example-item">27. 2. 20:00 Sparta vs Slavia</div>
            <div class="example-item">1. 3. SpartaOstrava20:00 (auto rok)</div>
            <div class="example-item">Sparta - Slavia 2:1</div>
          </div>
          <div class="paste-zone" id="pasteZone">
            <div id="pasteInstruction" class="paste-instruction">
              📸 <strong>Napiš nebo vlož screenshot</strong>
              <div class="muted" style="margin-top: 8px;">
                Win+Shift+S → Ctrl+V zde → Auto OCR ✨
              </div>
            </div>

            <textarea name="raw_text" 
                      id="rawText"
                      placeholder="Paste text NEBO screenshot (Ctrl+V)..."
                      style="min-height: 300px;"></textarea>

            <input type="hidden" name="screenshot_data" id="screenshotData">

            <div id="imagePreview" style="display: none; margin-top: 12px;">
              <div style="color: #2ecc71; margin-bottom: 8px;">
                ✅ Screenshot načten - OCR se spustí při parsování
              </div>
              <img id="previewImg" style="max-width: 100%; max-height: 200px; border-radius: 8px; border: 2px solid #667eea;">
              <button type="button" onclick="clearPastedImage()" class="btn btn-secondary" style="margin-top: 8px;">
                ❌ Smazat screenshot
              </button>
            </div>
          </div>
          <div class="muted">💡 Tip: Headers jako "24. kolo" se automaticky přeskočí</div>
        </div>

        <button type="submit" class="btn btn-primary">🔍 Parsovat & Preview</button>
      </form>
    </div>
    {% else %}
    <div class="card">
      <h1>🔍 Preview - zkontroluj a uprav</h1>
      <div class="subtitle">Naparsováno {{ parsed_matches|length }} zápasů</div>

      <form method="post" id="importForm">
        <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
        <input type="hidden" name="action" value="import"/>
        <input type="hidden" name="round_id" value="{{ import_round_id }}"/>
        <input type="hidden" name="import_mode" value="{{ import_mode }}"/>
        <input type="hidden" name="matches_data" id="matchesData" value=""/>

        <table class="preview-table">
          <thead>
            <tr>
              <th style="width: 40px;">
                <input type="checkbox" id="selectAll" checked>
              </th>
              <th>Domácí</th>
              <th>Hosté</th>
              <th style="width: 80px;">Skóre D</th>
              <th style="width: 80px;">Skóre H</th>
              <th style="width: 200px;">Datum & Čas</th>
            </tr>
          </thead>
          <tbody>
            {% for match in parsed_matches %}
            <tr>
              <td>
                <input type="checkbox" class="match-select" checked>
              </td>
              <td>
                <input type="text" 
                       class="home-team" 
                       value="{{ match.home_team }}"
                       data-original="{{ match.home_team }}">
              </td>
              <td>
                <input type="text" 
                       class="away-team" 
                       value="{{ match.away_team }}"
                       data-original="{{ match.away_team }}">
              </td>
              <td>
                <input type="number" 
                       class="home-score" 
                       value="{{ match.home_score if match.home_score is not none else '' }}"
                       min="0"
                       placeholder="-">
              </td>
              <td>
                <input type="number" 
                       class="away-score" 
                       value="{{ match.away_score if match.away_score is not none else '' }}"
                       min="0"
                       placeholder="-">
              </td>
              <td>
                <input type="datetime-local" 
                       class="start-time" 
                       value="{{ match.start_time[:16] if match.start_time else '' }}">
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>

        <div style="margin-top: 24px; display: flex; gap: 12px;">
          <button type="submit" class="btn btn-success">✅ Importovat vybrané zápasy</button>
          <a href="{{ url_for('admin_smart_import') }}" class="btn btn-secondary">← Zpět</a>
        </div>
      </form>
    </div>
    {% endif %}
      </div>

      <script>
    // Select all checkbox
    document.getElementById('selectAll')?.addEventListener('change', function() {
      document.querySelectorAll('.match-select').forEach(cb => cb.checked = this.checked);
    });

    // Collect data before submit
    document.getElementById('importForm')?.addEventListener('submit', function(e) {
      const rows = document.querySelectorAll('.preview-table tbody tr');
      const matches = [];

      rows.forEach(row => {
        if (row.querySelector('.match-select').checked) {
          const homeTeam = row.querySelector('.home-team').value;
          const awayTeam = row.querySelector('.away-team').value;
          const homeScore = row.querySelector('.home-score').value;
          const awayScore = row.querySelector('.away-score').value;
          const startTime = row.querySelector('.start-time').value;

          matches.push({
            home_team: homeTeam,
            away_team: awayTeam,
            home_score: homeScore !== "" ? parseInt(homeScore, 10) : null,
            away_score: awayScore !== "" ? parseInt(awayScore, 10) : null,
            start_time: startTime || null
          });
        }
      });

      document.getElementById('matchesData').value = JSON.stringify(matches);
    });
      </script>

      <script>
    // Image preview
    document.getElementById('screenshot')?.addEventListener('change', function(e) {
      const file = e.target.files[0];
      if (file) {
        const reader = new FileReader();
        reader.onload = function(e) {
          document.getElementById('previewImg').src = e.target.result;
          document.getElementById('preview').style.display = 'block';
          document.getElementById('uploadZone').querySelector('.upload-label').style.display = 'none';
        };
        reader.readAsDataURL(file);
      }
    });

    function clearImage() {
      document.getElementById('screenshot').value = '';
      document.getElementById('preview').style.display = 'none';
      document.getElementById('uploadZone').querySelector('.upload-label').style.display = 'block';
    }

    // Drag & drop
    const uploadZone = document.getElementById('uploadZone');
    if (uploadZone) {
      ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
        uploadZone.addEventListener(eventName, preventDefaults, false);
      });

      function preventDefaults(e) {
        e.preventDefault();
        e.stopPropagation();
      }

      ['dragenter', 'dragover'].forEach(eventName => {
        uploadZone.addEventListener(eventName, () => {
          uploadZone.classList.add('drag-over');
        });
      });

      ['dragleave', 'drop'].forEach(eventName => {
        uploadZone.addEventListener(eventName, () => {
          uploadZone.classList.remove('drag-over');
        });
      });

      uploadZone.addEventListener('drop', function(e) {
        const dt = e.dataTransfer;
        const files = dt.files;

        if (files.length > 0) {
          document.getElementById('screenshot').files = files;
          const event = new Event('change');
          document.getElementById('screenshot').dispatchEvent(event);
        }
      });
    }
      </script>

      <script>
    let pastedImageData = null;

    // Listen for paste events on textarea
    document.getElementById('rawText')?.addEventListener('paste', function(e) {
      const items = (e.clipboardData || e.originalEvent.clipboardData).items;

      for (let item of items) {
        if (item.type.indexOf('image') !== -1) {
          // Image pasted!
          e.preventDefault();

          const blob = item.getAsFile();
          const reader = new FileReader();

          reader.onload = function(event) {
            const base64 = event.target.result;

            // Store in hidden field
            document.getElementById('screenshotData').value = base64;

            // Show preview
            document.getElementById('previewImg').src = base64;
            document.getElementById('imagePreview').style.display = 'block';
            document.getElementById('pasteZone').classList.add('has-image');

            // Clear textarea (user pasted image, not text)
            document.getElementById('rawText').value = '';

            console.log('✅ Screenshot pasted, will OCR on submit');
          };

          reader.readAsDataURL(blob);
          break;
        }
      }
    });

    function clearPastedImage() {
      document.getElementById('screenshotData').value = '';
      document.getElementById('imagePreview').style.display = 'none';
      document.getElementById('pasteZone').classList.remove('has-image');
      document.getElementById('rawText').focus();
    }
      </script>
    </body>
    </html>
    """

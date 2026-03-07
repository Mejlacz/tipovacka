"""
routes/admin_users.py
"""

import datetime
from io import BytesIO
import os
import csv
import re
import pickle
import tempfile

from typing import Optional, List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from flask import request, flash, redirect, url_for, abort, send_file, session
from flask_login import current_user, login_required

from models import ExtraAnswer, Match, Team, TeamAlias, Tip, User
from app_utils import admin_required, audit, can_see_user_in_admin, render_page, send_welcome_with_reset_link, validate_password
from extensions import db
try:
    import pytesseract
    from PIL import Image
    TESSERACT_AVAILABLE = True
except ImportError:
    pytesseract = None
    Image = None
    TESSERACT_AVAILABLE = False
from config import OWNER_ADMIN_EMAIL

def register_admin_users(app):
    @app.route("/admin/users", methods=["GET", "POST"])
    @login_required
    def admin_users():
        admin_required()

        if request.method == "POST":
            action = request.form.get("bulk_action")
            user_ids = request.form.getlist("user_ids")

            if not user_ids:
                flash("Nevybral jsi žádné uživatele.", "error")
                return redirect(url_for("admin_users"))

            user_ids = [int(uid) for uid in user_ids]
            affected_users = User.query.filter(User.id.in_(user_ids)).all()

            # Ochrana proti smazání/úpravě ownera a sebe sama
            for u in affected_users:
                if u.is_owner:
                    flash(f"Nelze hromadně upravit ownera ({u.username}).", "error")
                    return redirect(url_for("admin_users"))
                if u.id == current_user.id:
                    flash("Nelze hromadně upravit sám sebe.", "error")
                    return redirect(url_for("admin_users"))

            if action == "delete":
                count = len(affected_users)
                for u in affected_users:
                    db.session.delete(u)
                db.session.commit()
                audit("users.bulk_delete", "User", None, count=count)
                flash(f"Smazáno {count} uživatelů.", "ok")

            elif action == "set_role":
                new_role = request.form.get("new_role")
                if new_role not in ["user", "viewer", "moderator", "admin"]:
                    flash("Neplatná role.", "error")
                    return redirect(url_for("admin_users"))

                for u in affected_users:
                    u.role = new_role
                db.session.commit()
                audit("users.bulk_role", "User", None, role=new_role, count=len(affected_users))
                flash(f"Změněna role pro {len(affected_users)} uživatelů na '{new_role}'.", "ok")

            elif action == "reset_password":
                new_password = request.form.get("new_password", "").strip()
                if not new_password:
                    flash("Zadej nové heslo.", "error")
                    return redirect(url_for("admin_users"))

                # Validace síly hesla
                is_valid, error_msg = validate_password(new_password)
                if not is_valid:
                    flash(error_msg, "error")
                    return redirect(url_for("admin_users"))

                for u in affected_users:
                    u.set_password(new_password)
                db.session.commit()
                audit("users.bulk_password", "User", None, count=len(affected_users))
                flash(f"Resetováno heslo pro {len(affected_users)} uživatelů.", "ok")

            elif action == "send_welcome_reset":
                # Pošli welcome email s reset linkem
                base_url = request.url_root.rstrip('/')
                sent_count = 0

                for u in affected_users:
                    try:
                        if send_welcome_with_reset_link(u, base_url):
                            sent_count += 1
                    except Exception as e:
                        continue

                audit("users.bulk_welcome", "User", None, count=sent_count)
                flash(f"📧 Welcome emaily odeslány: {sent_count} uživatelům. Mají 24h na nastavení hesla.", "ok")

            return redirect(url_for("admin_users"))

        users = User.query.order_by(User.username.asc()).all()
        users = [u for u in users if can_see_user_in_admin(u)]
        return render_page(r"""
    <div class="card">
      <div class="row" style="justify-content:space-between;">
    <div>
      <h2 style="margin:0 0 8px 0;">Uživatelé</h2>
      <div class="muted">Owner = <b>{{ owner }}</b>. Tajný user je skrytý pro jiné adminy.</div>
    </div>
    <div class="row">
      {% if current_user.is_owner %}<span class="tag pill-ok">Owner admin</span>{% endif %}
      <a class="btn btn-primary" href="{{ url_for('admin_user_new') }}">➕ Nový uživatel</a>
      <a class="btn" href="{{ url_for('admin_users_import') }}" style="background:#667eea; color:white;">📤 Import uživatelů</a>
    </div>
      </div>

      <hr class="sep">

      {% if current_user.is_owner %}
      {# Hromadná správa #}
      <form method="post" id="bulkForm">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div class="card" style="background:rgba(110,168,254,0.08); border:1px solid rgba(110,168,254,0.2); margin-bottom:16px; padding:16px;">
      <h3 style="margin:0 0 12px 0;">Hromadná správa</h3>

      <div style="display:grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap:10px;">
        {# Změna role #}
        <div>
          <select name="new_role" id="bulk_role" style="width:100%;">
            <option value="">-- Vybrat roli --</option>
            <option value="user">User</option>
            <option value="viewer">Viewer</option>
            <option value="moderator">Moderátor</option>
            <option value="admin">Admin</option>
          </select>
        </div>
        <button type="button" onclick="submitBulkAction('set_role')" class="btn btn-sm">
          👥 Změnit roli
        </button>

        {# Reset hesla #}
        <input type="password" name="new_password" id="bulk_password" placeholder="Nové heslo" style="width:100%;">
        <button type="button" onclick="submitBulkAction('reset_password')" class="btn btn-sm">
          🔑 Resetovat heslo
        </button>

        {# Poslat welcome email #}
        <div></div>
        <button type="button" onclick="submitBulkAction('send_welcome_reset')" class="btn btn-sm" style="background:rgba(110,168,254,0.3); color:#0b1020;">
          📧 Poslat welcome email
        </button>

        {# Smazat #}
        <div></div>
        <button type="button" onclick="submitBulkDelete()" class="btn btn-sm"
                style="background:rgba(255,77,109,0.2); color:#ff4d6d; border:1px solid rgba(255,77,109,0.4);">
          🗑️ Smazat vybrané
        </button>
      </div>

      <input type="hidden" name="bulk_action" id="bulk_action">
      <div class="muted" style="margin-top:10px; font-size:12px;">
        💡 Označ uživatele checkboxem, vyber akci a klikni na tlačítko<br>
        📧 Welcome email = pošle reset link pro nastavení hesla (24h platnost)
      </div>
    </div>

      <table class="datatable">
    <thead>
      <tr>
        <th style="width:40px;">
          <input type="checkbox" id="selectAll" onclick="toggleAll(this)"
                 title="Vybrat všechny">
        </th>
        <th>Username</th>
        <th>Email</th>
        <th style="text-align:center;">Role</th>
        <th style="text-align:center;">Registrace</th>
        <th style="text-align:center; min-width:400px;">Akce</th>
      </tr>
    </thead>
    <tbody>
      {% for u in users %}
      <tr>
        <td>
          {% if u.id != current_user.id and not u.is_owner %}
            <input type="checkbox" name="user_ids" value="{{ u.id }}" class="user-checkbox">
          {% endif %}
        </td>
        <td>
          <strong>{{ u.display_name }}</strong>
          {% if u.is_owner %} <span class="tag pill-ok">Owner</span>{% endif %}
          {% if u.nickname %}<br><span class="muted" style="font-size:11px;">Nick: {{ u.nickname }}</span>{% endif %}
          <div class="muted" style="font-size:11px;">Login: {{ u.username }}</div>
        </td>
        <td class="muted">{{ u.email }}</td>
        <td style="text-align:center;">
          {% if u.effective_role == 'admin' %}
            <span class="tag pill-bad">Admin</span>
          {% elif u.effective_role == 'moderator' %}
            <span class="tag pill-ok">Moderátor</span>
          {% elif u.effective_role == 'viewer' %}
            <span class="tag">Viewer</span>
          {% else %}
            <span class="tag pill-primary">User</span>
          {% endif %}
        </td>
        <td style="text-align:center;" class="muted">{{ u.created_at.strftime('%d.%m.%Y') }}</td>
        <td style="text-align:center;">
          <div style="display:flex; justify-content:center; gap:6px; flex-wrap:nowrap;">
              <a class="btn btn-sm" href="{{ url_for('admin_user_edit', user_id=u.id) }}">Upravit</a>
              <a class="btn btn-sm" href="{{ url_for('admin_user_reset_password', user_id=u.id) }}">Reset hesla</a>
              {% if u.id != current_user.id and not u.is_owner %}
                <a class="btn btn-sm" href="{{ url_for('admin_user_change_role', user_id=u.id) }}">Změnit roli</a>
                <form method="post" action="{{ url_for('admin_user_delete', user_id=u.id) }}" style="display:inline;" 
                      onsubmit="return confirm('Opravdu smazat uživatele {{ u.username }}?')">
                  <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                  <button type="submit" class="btn btn-sm btn-danger" 
                          style="background:rgba(255,77,109,0.2); color:#ff4d6d; border:none; cursor:pointer;">
                    Smazat
                  </button>
                </form>
              {% endif %}
          </div>
        </td>
      </tr>
      {% endfor %}
    </tbody>
      </table>
      </form>
      {% else %}
      {# Non-owner zobrazí jen tabulku bez hromadné správy #}
      <table class="datatable">
    <thead>
      <tr>
        <th>Username</th>
        <th>Email</th>
        <th style="text-align:center;">Role</th>
        <th style="text-align:center;">Registrace</th>
      </tr>
    </thead>
    <tbody>
      {% for u in users %}
      <tr>
        <td>
          <strong>{{ u.display_name }}</strong>
          {% if u.nickname %}<br><span class="muted" style="font-size:11px;">Nick: {{ u.nickname }}</span>{% endif %}
          <div class="muted" style="font-size:11px;">Login: {{ u.username }}</div>
        </td>
        <td class="muted">{{ u.email }}</td>
        <td style="text-align:center;">
          {% if u.effective_role == 'admin' %}
            <span class="tag pill-bad">Admin</span>
          {% elif u.effective_role == 'moderator' %}
            <span class="tag pill-ok">Moderátor</span>
          {% elif u.effective_role == 'viewer' %}
            <span class="tag">Viewer</span>
          {% else %}
            <span class="tag pill-primary">User</span>
          {% endif %}
        </td>
        <td style="text-align:center;" class="muted">{{ u.created_at.strftime('%d.%m.%Y') }}</td>
      </tr>
      {% endfor %}
    </tbody>
      </table>
      {% endif %}
    </div>

    <script>
    function toggleAll(checkbox) {
      const checkboxes = document.querySelectorAll('.user-checkbox');
      checkboxes.forEach(cb => cb.checked = checkbox.checked);
    }

    function getSelectedCount() {
      const checkboxes = document.querySelectorAll('.user-checkbox:checked');
      return checkboxes.length;
    }

    function submitBulkAction(action) {
      const count = getSelectedCount();
      if (count === 0) {
    alert('Nevybral jsi žádné uživatele!');
    return;
      }

      let confirmMsg = '';
      if (action === 'set_role') {
    const role = document.getElementById('bulk_role').value;
    if (!role) {
      alert('Vyber roli!');
      return;
    }
    confirmMsg = `Opravdu změnit roli pro ${count} uživatelů na '${role}'?`;
      } else if (action === 'reset_password') {
    const password = document.getElementById('bulk_password').value;
    if (!password) {
      alert('Zadej nové heslo!');
      return;
    }
    confirmMsg = `Opravdu resetovat heslo pro ${count} uživatelů?`;
      } else if (action === 'send_welcome_reset') {
    confirmMsg = `Poslat welcome email ${count} uživatelům?\n\nKaždý dostane email s odkazem pro nastavení hesla (platnost 24h).`;
      }

      if (confirmMsg && confirm(confirmMsg)) {
    document.getElementById('bulk_action').value = action;
    document.getElementById('bulkForm').submit();
      }
    }

    function submitBulkDelete() {
      const count = getSelectedCount();
      if (count === 0) {
    alert('Nevybral jsi žádné uživatele!');
    return;
      }

      if (confirm(`POZOR: Opravdu smazat ${count} uživatelů?\n\nTato akce je NEVRATNÁ!`)) {
    document.getElementById('bulk_action').value = 'delete';
    document.getElementById('bulkForm').submit();
      }
    }
    </script>

    <style>
      .btn-danger {
    background: rgba(255, 77, 109, 0.15);
    color: #ff4d6d;
      }
      .btn-danger:hover {
    background: rgba(255, 77, 109, 0.25);
      }
    </style>
    """, users=users, owner=OWNER_ADMIN_EMAIL)

    @app.route("/admin/user/<int:user_id>/toggle-admin")
    @login_required
    def admin_toggle_admin(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)
        if u.id == current_user.id:
            flash("Sebe si nepřepínej.", "error")
            return redirect(url_for("admin_users"))
        u.is_admin = not u.is_admin
        db.session.commit()
        audit("user.toggle_admin", "User", u.id, is_admin=u.is_admin)
        return redirect(url_for("admin_users"))

    @app.route("/admin/user/new", methods=["GET", "POST"])
    @login_required
    def admin_user_new():
        admin_required()
        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            email = (request.form.get("email") or "").strip().lower()
            password = (request.form.get("password") or "").strip()
            role = (request.form.get("role") or "user").strip()
            first_name = (request.form.get("first_name") or "").strip()
            last_name = (request.form.get("last_name") or "").strip()
            nickname = (request.form.get("nickname") or "").strip()

            if not username or not email or not password:
                flash("Vyplň všechna povinná pole.", "error")
                return redirect(url_for("admin_user_new"))

            # Validace síly hesla
            is_valid, error_msg = validate_password(password)
            if not is_valid:
                flash(error_msg, "error")
                return redirect(url_for("admin_user_new"))

            # Kontrola, jestli už username nebo email neexistuje
            if User.query.filter_by(username=username).first():
                flash(f"Username '{username}' už existuje.", "error")
                return redirect(url_for("admin_user_new"))

            if User.query.filter_by(email=email).first():
                flash(f"Email '{email}' už existuje.", "error")
                return redirect(url_for("admin_user_new"))

            # Vytvoření nového uživatele
            new_user = User(
                username=username,
                email=email,
                role=role,
                first_name=first_name or None,
                last_name=last_name or None,
                nickname=nickname or None,
                email_verified=True  # Admin vytvořil uživatele, email je ověřený
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            audit("user.create", "User", new_user.id, username=username, email=email, role=role)
            flash(f"Uživatel '{username}' byl vytvořen.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Nový uživatel</h2>
      <hr class="sep">
      <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Username (login) *</label>
      <input name="username" placeholder="username" required>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Email *</label>
      <input name="email" type="email" placeholder="email@example.com" required>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Heslo *</label>
      <input name="password" type="password" placeholder="Min. 8 znaků" required minlength="8">
      <div class="muted" style="font-size:12px; margin-top:4px;">
        Požadavky: min. 8 znaků, velké/malé písmeno, číslo
      </div>
    </div>
    <div class="grid2">
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Jméno</label>
        <input name="first_name" placeholder="Jan">
      </div>
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Příjmení</label>
        <input name="last_name" placeholder="Novák">
      </div>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Přezdívka (Nick)</label>
      <input name="nickname" placeholder="JN23">
      <div class="muted" style="font-size:12px; margin-top:4px;">Zobrazí se v žebříčku</div>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Role</label>
      <select name="role">
        <option value="user">User (může tipovat)</option>
        <option value="viewer">Viewer (jen prohlížet)</option>
        <option value="moderator">Moderátor (může tipovat + částečná správa)</option>
        <option value="admin">Admin (plná správa)</option>
      </select>
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Vytvořit</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
      </form>
    </div>
    """)

    @app.route("/admin/users/import", methods=["GET", "POST"])
    @login_required
    def admin_users_import():
        """Import uživatelů z Excel souboru - krok 1: Upload"""
        admin_required()

        if request.method == "POST":
            if 'file' not in request.files:
                flash("Žádný soubor nebyl vybrán.", "error")
                return redirect(url_for("admin_users_import"))

            file = request.files['file']
            if file.filename == '':
                flash("Žádný soubor nebyl vybrán.", "error")
                return redirect(url_for("admin_users_import"))

            if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
                flash("Podporovány jsou pouze .xlsx, .xls a .csv soubory.", "error")
                return redirect(url_for("admin_users_import"))

            try:
                import openpyxl
                import pandas as pd
                import pickle

                # Načti soubor
                if file.filename.endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                # Očekávané sloupce
                required_cols = ['email', 'username', 'password']
                missing_cols = [col for col in required_cols if col not in df.columns]

                if missing_cols:
                    flash(f"Chybí povinné sloupce: {', '.join(missing_cols)}", "error")
                    return redirect(url_for("admin_users_import"))

                # Parsuj a validuj všechny uživatele
                users_preview = []

                for index, row in df.iterrows():
                    email = str(row['email']).strip().lower() if pd.notna(row['email']) else ''
                    username = str(row['username']).strip() if pd.notna(row['username']) else ''
                    password = str(row['password']).strip() if pd.notna(row['password']) else ''

                    # Validace
                    errors = []
                    status = 'ok'

                    if not email or not username or not password:
                        errors.append("Prázdné povinné pole")
                        status = 'error'

                    # Kontrola duplicit
                    if email and User.query.filter_by(email=email).first():
                        errors.append(f"Email '{email}' už existuje")
                        status = 'duplicate'

                    if username and User.query.filter_by(username=username).first():
                        errors.append(f"Username '{username}' už existuje")
                        status = 'duplicate'

                    # Validace hesla
                    if password and status == 'ok':
                        is_valid, error_msg = validate_password(password)
                        if not is_valid:
                            errors.append(error_msg)
                            status = 'error'

                    # Volitelné sloupce
                    first_name = str(row.get('first_name', '')).strip() if pd.notna(row.get('first_name')) else None
                    last_name = str(row.get('last_name', '')).strip() if pd.notna(row.get('last_name')) else None
                    nickname = str(row.get('nickname', '')).strip() if pd.notna(row.get('nickname')) else None
                    role = str(row.get('role', 'user')).strip().lower() if pd.notna(row.get('role')) else 'user'

                    if role not in ['user', 'viewer', 'moderator', 'admin']:
                        role = 'user'

                    users_preview.append({
                        'row_num': index + 2,  # +2 protože 1=header, Excel je 1-indexed
                        'email': email,
                        'username': username,
                        'password': password,
                        'first_name': first_name,
                        'last_name': last_name,
                        'nickname': nickname,
                        'role': role,
                        'status': status,
                        'errors': errors
                    })

                # Počet podle statusů
                ok_count = sum(1 for u in users_preview if u['status'] == 'ok')
                error_count = sum(1 for u in users_preview if u['status'] in ['error', 'duplicate'])

                # Ulož do temp file
                preview_fd, preview_path = tempfile.mkstemp(suffix='.pkl', prefix='users_import_')
                os.close(preview_fd)

                with open(preview_path, 'wb') as f:
                    pickle.dump(users_preview, f)

                session['users_import_preview_file'] = preview_path

                return redirect(url_for("admin_users_import_preview"))

            except Exception as e:
                flash(f"Chyba při zpracování souboru: {str(e)}", "error")
                return redirect(url_for("admin_users_import"))

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Import uživatelů - Krok 1</h2>
      <div class="muted">Nahraj Excel nebo CSV soubor s uživateli</div>
      <hr class="sep">

      <div style="background:rgba(110,168,254,0.08); padding:16px; border-radius:8px; margin-bottom:16px;">
    <h3 style="margin:0 0 8px 0;">📋 Formát souboru</h3>
    <p><strong>Povinné sloupce:</strong></p>
    <ul>
      <li><code>email</code> - Email uživatele</li>
      <li><code>username</code> - Uživatelské jméno (login)</li>
      <li><code>password</code> - Heslo (min. 8 znaků, velké/malé/číslo)</li>
    </ul>
    <p><strong>Volitelné sloupce:</strong></p>
    <ul>
      <li><code>first_name</code> - Jméno</li>
      <li><code>last_name</code> - Příjmení</li>
      <li><code>nickname</code> - Přezdívka (zobrazí se v žebříčku)</li>
      <li><code>role</code> - Role (user/viewer/moderator/admin, default: user)</li>
    </ul>
      </div>

      <form method="post" enctype="multipart/form-data" class="row" style="flex-direction:column; gap:16px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    <div style="background:rgba(110,168,254,0.1); padding:12px; border-left:4px solid #6ea8fe; margin-bottom:8px;">
      <strong>💡 Nemáš Excel?</strong> Stáhni si šablonu s hlavičkami a příkladem:
      <a href="{{ url_for('admin_users_import_template') }}" class="btn btn-sm" style="margin-left:8px;">
        📥 Stáhnout šablonu Excel
      </a>
    </div>

    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Soubor (.xlsx, .xls, .csv)</label>
      <input type="file" name="file" accept=".xlsx,.xls,.csv" required>
    </div>

    <div class="row" style="gap:10px;">
      <button class="btn btn-primary" type="submit">➡️ Pokračovat na Preview</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
      </form>

      <hr class="sep">

      <div class="muted" style="font-size:12px;">
    <p><strong>💡 Postup:</strong></p>
    <ol>
      <li><strong>Upload</strong> - Nahraj Excel soubor</li>
      <li><strong>Preview</strong> - Zkontroluj kdo se bude importovat a vyber komu poslat email</li>
      <li><strong>Potvrzení</strong> - Importuj vybrané uživatele</li>
    </ol>
      </div>
    </div>
    """)

    @app.route("/admin/users/import/template")
    @login_required
    def admin_users_import_template():
        """Stáhne Excel šablonu pro import uživatelů"""
        admin_required()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            # Vytvoř workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Uživatelé"

            # Hlavička
            headers = ['email', 'username', 'password', 'first_name', 'last_name', 'nickname', 'role']
            ws.append(headers)

            # Stylování hlavičky
            header_fill = PatternFill(start_color="6EA8FE", end_color="6EA8FE", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Příklad řádku (volitelně, pro ukázku)
            example_row = [
                'jan.novak@email.cz',
                'jan123',
                'Test1234',
                'Jan',
                'Novák',
                'JN',
                'user'
            ]
            ws.append(example_row)

            # Nastavení šířky sloupců
            ws.column_dimensions['A'].width = 25  # email
            ws.column_dimensions['B'].width = 15  # username
            ws.column_dimensions['C'].width = 15  # password
            ws.column_dimensions['D'].width = 15  # first_name
            ws.column_dimensions['E'].width = 15  # last_name
            ws.column_dimensions['F'].width = 12  # nickname
            ws.column_dimensions['G'].width = 12  # role

            # Přidej poznámku do druhého sheetu
            notes_ws = wb.create_sheet("Poznámky")
            notes_ws['A1'] = "NÁVOD K POUŽITÍ"
            notes_ws['A1'].font = Font(bold=True, size=14)

            notes = [
                "",
                "POVINNÉ SLOUPCE:",
                "• email - Email uživatele (musí být unikátní)",
                "• username - Uživatelské jméno pro login (musí být unikátní)",
                "• password - Heslo (min. 8 znaků, velké/malé písmeno, číslice)",
                "",
                "VOLITELNÉ SLOUPCE:",
                "• first_name - Křestní jméno",
                "• last_name - Příjmení",
                "• nickname - Přezdívka (zobrazí se v žebříčku)",
                "• role - Role (user/viewer/moderator/admin, výchozí: user)",
                "",
                "POSTUP:",
                "1. Vyplň uživatele do 1. sheetu (Uživatelé)",
                "2. První řádek s ukázkou můžeš smazat nebo upravit",
                "3. Ulož soubor",
                "4. Nahraj na stránce Import uživatelů",
                "",
                "PŘÍKLAD HESEL:",
                "✅ Test1234 - OK (8+ znaků, velké, malé, číslice)",
                "✅ Password123 - OK",
                "❌ test - Špatně (moc krátké, bez velkého, bez čísla)",
                "❌ TestTest - Špatně (bez čísla)",
            ]

            for i, note in enumerate(notes, 2):
                notes_ws[f'A{i}'] = note

            notes_ws.column_dimensions['A'].width = 60

            # Ulož do BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='import_uzivatelu_sablona.xlsx'
            )

        except Exception as e:
            flash(f"Chyba při vytváření šablony: {str(e)}", "error")
            return redirect(url_for("admin_users_import"))

    @app.route("/admin/users/import/preview")
    @login_required
    def admin_users_import_preview():
        """Import uživatelů - krok 2: Preview a výběr komu poslat email"""
        admin_required()

        preview_file = session.get('users_import_preview_file')
        if not preview_file or not os.path.exists(preview_file):
            flash("Preview data nenalezena. Nahraj soubor znovu.", "error")
            return redirect(url_for("admin_users_import"))

        try:
            import pickle
            with open(preview_file, 'rb') as f:
                users_preview = pickle.load(f)

            ok_users = [u for u in users_preview if u['status'] == 'ok']
            error_users = [u for u in users_preview if u['status'] in ['error', 'duplicate']]

            return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Import uživatelů - Preview</h2>
      <div class="muted">Zkontroluj kdo se bude importovat</div>
      <hr class="sep">

      <div class="row" style="gap:12px; margin-bottom:16px;">
    <div class="tag pill-ok">✅ K importu: {{ ok_count }}</div>
    {% if error_count > 0 %}<div class="tag pill-error">❌ Chyby: {{ error_count }}</div>{% endif %}
      </div>

      <form method="post" action="{{ url_for('admin_users_import_confirm') }}">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>

    {% if ok_users %}
    <h3 style="margin:16px 0 8px 0;">✅ Připraveno k importu ({{ ok_count }})</h3>
    <div style="background:#f8f9fa; padding:8px; border-radius:8px; margin-bottom:16px; max-height:400px; overflow-y:auto;">
      <table style="width:100%; font-size:13px;">
        <thead style="position:sticky; top:0; background:#e9ecef;">
          <tr>
            <th style="padding:8px; text-align:left;">Email</th>
            <th style="padding:8px; text-align:left;">Username</th>
            <th style="padding:8px; text-align:left;">Jméno</th>
            <th style="padding:8px; text-align:left;">Role</th>
          </tr>
        </thead>
        <tbody>
          {% for user in ok_users %}
          <tr style="border-bottom:1px solid #dee2e6;">
            <td style="padding:8px;">{{ user.email }}</td>
            <td style="padding:8px;">{{ user.username }}</td>
            <td style="padding:8px;">{{ user.first_name or '' }} {{ user.last_name or '' }}</td>
            <td style="padding:8px;"><span class="tag">{{ user.role }}</span></td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <div style="background:rgba(110,168,254,0.1); padding:12px; border-left:4px solid #6ea8fe; margin-bottom:16px;">
      <strong>💡 Info:</strong> Uživatelé budou importováni s ověřeným emailem a budou se moci přihlásit.
      Welcome emaily můžeš poslat později z Admin → Uživatelé → Hromadná správa.
    </div>
    {% endif %}

    {% if error_users %}
    <h3 style="margin:16px 0 8px 0;">❌ Chyby - nebudou importováni ({{ error_count }})</h3>
    <div style="background:#fff3cd; padding:12px; border-radius:8px; margin-bottom:16px; max-height:300px; overflow-y:auto;">
      {% for user in error_users %}
      <div style="padding:8px; border-bottom:1px solid #ffc107;">
        <strong>Řádek {{ user.row_num }}:</strong> {{ user.email }} / {{ user.username }}
        <div class="muted" style="font-size:12px;">
          {% for error in user.errors %}• {{ error }}<br>{% endfor %}
        </div>
      </div>
      {% endfor %}
    </div>
    {% endif %}

    <hr class="sep">

    <div class="row" style="gap:10px;">
      <button class="btn btn-primary" type="submit" {% if ok_count == 0 %}disabled{% endif %}>
        ✅ Potvrdit import ({{ ok_count }} uživatelů)
      </button>
      <a class="btn" href="{{ url_for('admin_users_import') }}">⬅️ Zpět na upload</a>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
      </form>
    </div>
    """, ok_users=ok_users, error_users=error_users, ok_count=len(ok_users), error_count=len(error_users))

        except Exception as e:
            flash(f"Chyba při načítání preview: {str(e)}", "error")
            return redirect(url_for("admin_users_import"))

    @app.route("/admin/users/import/confirm", methods=["POST"])
    @login_required
    def admin_users_import_confirm():
        """Import uživatelů - krok 3: Potvrzení a provedení importu"""
        admin_required()

        preview_file = session.get('users_import_preview_file')
        if not preview_file or not os.path.exists(preview_file):
            flash("Preview data nenalezena. Nahraj soubor znovu.", "error")
            return redirect(url_for("admin_users_import"))

        try:
            import pickle
            with open(preview_file, 'rb') as f:
                users_preview = pickle.load(f)

            created = 0
            ok_users = [u for u in users_preview if u['status'] == 'ok']

            for user_data in ok_users:
                try:
                    # Vytvoř uživatele
                    new_user = User(
                        email=user_data['email'],
                        username=user_data['username'],
                        first_name=user_data['first_name'],
                        last_name=user_data['last_name'],
                        nickname=user_data['nickname'],
                        role=user_data['role'],
                        email_verified=True  # Importovaní users mají ověřený email
                    )
                    new_user.set_password(user_data['password'])
                    db.session.add(new_user)
                    db.session.flush()

                    audit("user.imported", "User", new_user.id, username=new_user.username, email=new_user.email)
                    created += 1

                except Exception as e:
                    # Rollback tohoto usera, pokračuj dál
                    db.session.rollback()
                    continue

            db.session.commit()

            # Cleanup temp file
            try:
                os.remove(preview_file)
                session.pop('users_import_preview_file', None)
            except:
                pass

            # Flash výsledek
            flash(f"✅ Importováno: {created} uživatelů. Pro poslání welcome emailů jdi do Uživatelé → Hromadná správa.", "ok")
            return redirect(url_for("admin_users"))

        except Exception as e:
            flash(f"Chyba při importu: {str(e)}", "error")
            return redirect(url_for("admin_users_import"))

    @app.route("/admin/user/<int:user_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_user_edit(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if request.method == "POST":
            new_username = (request.form.get("username") or "").strip()
            new_email = (request.form.get("email") or "").strip().lower()
            new_first_name = (request.form.get("first_name") or "").strip()
            new_last_name = (request.form.get("last_name") or "").strip()
            new_nickname = (request.form.get("nickname") or "").strip()

            if not new_username or not new_email:
                flash("Vyplň username a email.", "error")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            # Kontrola duplicit
            existing_user = User.query.filter_by(username=new_username).first()
            if existing_user and existing_user.id != u.id:
                flash(f"Username '{new_username}' už existuje.", "error")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            existing_email = User.query.filter_by(email=new_email).first()
            if existing_email and existing_email.id != u.id:
                flash(f"Email '{new_email}' už existuje.", "error")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            old_username = u.username
            old_email = u.email
            u.username = new_username
            u.email = new_email
            u.first_name = new_first_name or None
            u.last_name = new_last_name or None
            u.nickname = new_nickname or None
            db.session.commit()
            audit("user.update", "User", u.id, old_username=old_username, new_username=new_username, old_email=old_email, new_email=new_email)
            flash(f"Uživatel aktualizován.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Upravit uživatele</h2>
      <hr class="sep">
      <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Username (login) *</label>
      <input name="username" value="{{ u.username }}" required>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Email *</label>
      <input name="email" type="email" value="{{ u.email }}" required>
    </div>
    <div class="grid2">
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Jméno</label>
        <input name="first_name" value="{{ u.first_name or '' }}">
      </div>
      <div>
        <label class="muted" style="margin-bottom:6px; display:block;">Příjmení</label>
        <input name="last_name" value="{{ u.last_name or '' }}">
      </div>
    </div>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Přezdívka (Nick)</label>
      <input name="nickname" value="{{ u.nickname or '' }}">
      <div class="muted" style="font-size:12px; margin-top:4px;">
        Zobrazí se v žebříčku. Současný nick: <strong>{{ u.display_name }}</strong>
      </div>
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Uložit</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
      </form>
    </div>
    """, u=u)

    @app.route("/admin/user/<int:user_id>/reset-password", methods=["GET", "POST"])
    @login_required
    def admin_user_reset_password(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if request.method == "POST":
            new_password = (request.form.get("password") or "").strip()
            if not new_password:
                flash("Vyplň nové heslo.", "error")
                return redirect(url_for("admin_user_reset_password", user_id=user_id))

            # Validace síly hesla
            is_valid, error_msg = validate_password(new_password)
            if not is_valid:
                flash(error_msg, "error")
                return redirect(url_for("admin_user_reset_password", user_id=user_id))

            u.set_password(new_password)
            db.session.commit()
            audit("user.reset_password", "User", u.id)
            flash(f"Heslo pro '{u.username}' bylo resetováno.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Resetovat heslo</h2>
      <div class="muted">Uživatel: <b>{{ u.username }}</b> ({{ u.email }})</div>
      <hr class="sep">
      <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Nové heslo *</label>
      <input name="password" type="password" placeholder="Min. 8 znaků" required minlength="8">
      <div class="muted" style="font-size:12px; margin-top:4px;">
        Požadavky: min. 8 znaků, velké/malé písmeno, číslo
      </div>
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Resetovat heslo</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
      </form>
    </div>
    """, u=u)

    @app.route("/admin/user/<int:user_id>/change-role", methods=["GET", "POST"])
    @login_required
    def admin_user_change_role(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if u.id == current_user.id:
            flash("Nemůžeš měnit roli sám sobě.", "error")
            return redirect(url_for("admin_users"))

        if request.method == "POST":
            new_role = (request.form.get("role") or "user").strip()
            old_role = u.effective_role

            # Nastavení role
            if new_role == "admin":
                u.is_admin = True
                u.role = "admin"
            else:
                u.is_admin = False
                u.role = new_role

            db.session.commit()
            audit("user.change_role", "User", u.id, old_role=old_role, new_role=new_role)
            flash(f"Role uživatele '{u.username}' změněna na '{new_role}'.", "ok")
            return redirect(url_for("admin_users"))

        return render_page(r"""
    <div class="card">
      <h2 style="margin:0 0 8px 0;">Změnit roli</h2>
      <div class="muted">Uživatel: <b>{{ u.username }}</b> ({{ u.email }})</div>
      <div class="muted">Současná role: <b>{{ u.effective_role }}</b></div>
      <hr class="sep">
      <form method="post" class="row" style="flex-direction:column; align-items:stretch; gap:10px;">
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
    <div>
      <label class="muted" style="margin-bottom:6px; display:block;">Nová role</label>
      <select name="role">
        <option value="user" {% if u.effective_role == 'user' %}selected{% endif %}>User (může tipovat)</option>
        <option value="viewer" {% if u.effective_role == 'viewer' %}selected{% endif %}>Viewer (jen prohlížet)</option>
        <option value="moderator" {% if u.effective_role == 'moderator' %}selected{% endif %}>Moderátor (může tipovat + částečná správa)</option>
        <option value="admin" {% if u.effective_role == 'admin' %}selected{% endif %}>Admin (plná správa)</option>
      </select>
    </div>
    <div class="muted" style="margin-top:10px; padding:10px; background:rgba(249,199,79,0.1); border-radius:8px;">
      <b>Vysvětlení rolí:</b><br>
      • <b>User</b> - může tipovat zápasy a odpovídat na extra otázky<br>
      • <b>Viewer</b> - může jen prohlížet žebříček a výsledky, nemůže tipovat<br>
      • <b>Moderátor</b> - může tipovat + má přístup k některým admin funkcím<br>
      • <b>Admin</b> - plný přístup ke všem funkcím
    </div>
    <div class="row" style="gap:10px; margin-top:10px;">
      <button class="btn btn-primary" type="submit">Změnit roli</button>
      <a class="btn" href="{{ url_for('admin_users') }}">Zrušit</a>
    </div>
      </form>
    </div>
    """, u=u)

    @app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
    @login_required
    def admin_user_delete(user_id: int):
        admin_required()
        u = db.session.get(User, user_id)
        if not u:
            abort(404)

        if u.id == current_user.id:
            flash("Nemůžeš smazat sám sebe.", "error")
            return redirect(url_for("admin_users"))

        username = u.username
        # Smazat všechny tipy uživatele
        Tip.query.filter_by(user_id=u.id).delete()
        # Smazat všechny odpovědi na extra otázky
        ExtraAnswer.query.filter_by(user_id=u.id).delete()
        # Smazat uživatele
        db.session.delete(u)
        db.session.commit()
        audit("user.delete", "User", user_id, username=username)
        flash(f"Uživatel '{username}' byl smazán.", "ok")
        return redirect(url_for("admin_users"))

    # --- ADMIN ROUNDS ---
    # --- ADMIN SMART IMPORT ---
    # =========================================================
    # ULTRA SMART IMPORT - AI-POWERED MATCH PARSING
    # =========================================================

    def extract_text_from_screenshot(image_data: bytes) -> Optional[str]:
        """
        Extract text from screenshot using Tesseract OCR (FREE)

        Args:
            image_data: Image file bytes (PNG, JPEG, etc)

        Returns:
            Extracted text or None if failed
        """

        if not TESSERACT_AVAILABLE:
            print("❌ Tesseract not available")
            return None

        try:
            # Open image
            image = Image.open(io.BytesIO(image_data))

            # Preprocessing for better OCR
            # Convert to grayscale
            image = image.convert('L')

            # Optional: increase contrast
            from PIL import ImageEnhance
            enhancer = ImageEnhance.Contrast(image)
            image = enhancer.enhance(2.0)

            # Extract text
            # Use custom config for better table recognition
            custom_config = r'--oem 3 --psm 6'
            text = pytesseract.image_to_string(image, lang='ces+eng', config=custom_config)

            print(f"✅ OCR extracted {len(text)} characters")
            print(f"📝 First 200 chars: {text[:200]}")

            return text.strip()

        except Exception as e:
            print(f"❌ OCR error: {e}")
            return None


    def _split_joined_lines(text: str) -> str:
        """
        CRITICAL FIX: Table copy/paste joins all lines

        New strategy: Extract complete match patterns
        Pattern: DD. MM. YYYY + text + time/score (N:N or NN:NN)
        """

        # Skip if already properly formatted
        lines = text.split('\n')
        if len(lines) > 3:
            return text

        # Rejoin text
        text_joined = ' '.join(lines)

        # Pattern for complete match:
        # Date + any text + final number:number
        # Use non-greedy match and lookahead
        pattern = r'(\d{1,2}\.\s*\d{1,2}\.\s*\d{4}.*?\d{1,2}:\d{1,2})(?=\d{1,2}\.\s*\d{1,2}\.\s*\d{4}|$)'

        matches = re.findall(pattern, text_joined)

        if len(matches) >= 2:
            # Add any header before first match
            parts = []
            first_match_pos = text_joined.find(matches[0])
            if first_match_pos > 0:
                header = text_joined[:first_match_pos].strip()
                if header:
                    parts.append(header)

            # Add all matches
            for match in matches:
                parts.append(match.strip())

            if len(parts) > 1:
                result = '\n'.join(parts)
                print(f"🔧 Extracted {len(parts)} lines from joined input")
                return result

        return text




    def _split_joined_lines(text: str) -> str:
        """
        🔧 FIX: Multiple matches joined on one line (from table copy/paste)

        Example:
        Input:  "27. 2. 2026DuklaSlavia0:228. 2. 2026Liberec..."
        Output: Each match on separate line

        Finds all dates and splits before each (except first)
        """

        result_lines = []

        for line in text.strip().split('\n'):
            line = line.strip()
            if not line:
                continue

            # Find all date patterns in this line
            date_pattern = r'\d{1,2}\.\s*\d{1,2}\.\s*\d{4}'
            matches = list(re.finditer(date_pattern, line))

            if len(matches) <= 1:
                # 0 or 1 date - keep as is
                result_lines.append(line)
                continue

            # Multiple dates found - split before each (except first)
            print(f"🔧 Found {len(matches)} dates in one line - splitting...")

            parts = []
            last_end = 0

            for i, match in enumerate(matches):
                if i == 0:
                    # First date - take from start until next date
                    if len(matches) > 1:
                        next_start = matches[1].start()
                        parts.append(line[last_end:next_start])
                        last_end = next_start
                    else:
                        parts.append(line[last_end:])
                else:
                    # Subsequent dates
                    if i < len(matches) - 1:
                        next_start = matches[i+1].start()
                        parts.append(line[last_end:next_start])
                        last_end = next_start
                    else:
                        # Last date - take rest of line
                        parts.append(line[last_end:])

            # Add all parts
            for part in parts:
                part = part.strip()
                if part:
                    result_lines.append(part)

        return '\n'.join(result_lines)


    def _clean_ocr_artifacts(text: str) -> str:
        """
        Clean up OCR artifacts from screenshot text

        OCR often adds extra characters like:
        - "vice >" or "více>" (from web UI buttons)
        - "|" pipes
        - Extra whitespace
        - Weird unicode characters

        Returns cleaned text ready for parsing
        """
        lines = text.strip().split('\n')
        cleaned_lines = []

        for line in lines:
            # Remove common OCR artifacts
            # Remove "vice >", "více>", "vic>", "víc>" and similar
            line = re.sub(r'\s*v[ií]ce?\s*>?\s*', ' ', line, flags=re.IGNORECASE)

            # Remove pipe symbols (often from table borders)
            line = re.sub(r'\s*\|\s*', ' ', line)

            # Remove extra whitespace between words (but keep spaces)
            line = re.sub(r'\s+', ' ', line)

            # Remove leading/trailing whitespace
            # Fix OCR misreading 0 as 'o'/'O' at end of line (trailing score)
            line = re.sub(r' [oO]$', ' 0', line)

            line = line.strip()

            if line:
                cleaned_lines.append(line)

        result = '\n'.join(cleaned_lines)
        print(f"🧹 OCR cleanup: {len(lines)} lines → {len(cleaned_lines)} cleaned lines")
        return result

    def _parse_multiline_app_format(text: str) -> List[Dict]:
        """
        Parse multi-line format from app/website:

        07.03. 15:00
        Bohemians
        Slovan Liberec
        --
        """
        matches = []
        blocks = text.split('--')
        current_year = 2026

        for block in blocks:
            lines = [line.strip() for line in block.strip().split('\n') if line.strip()]

            if len(lines) < 3:
                continue

            # Skip "25. kolo" headers
            if lines[0].lower().endswith('kolo'):
                lines = lines[1:]

            if len(lines) < 3:
                continue

            datetime_line = lines[0]
            home_team = lines[1]
            away_team = lines[2]

            # Parse "07.03. 15:00"
            match = re.match(r'(\d{2})\.(\d{2})\.\s+(\d{1,2}):(\d{2})', datetime_line)
            if not match:
                continue

            day = int(match.group(1))
            month = int(match.group(2))
            hour = int(match.group(3))
            minute = int(match.group(4))

            try:
                dt = datetime(current_year, month, day, hour, minute)
                matches.append({
                    'home_team': home_team.strip(),
                    'away_team': away_team.strip(),
                    'start_time': dt,
                })
            except ValueError:
                continue

        return matches

    def _parse_multiline_ocr_format(text: str) -> List[Dict]:
        """
        Parser pro OCR screenshot z mobilní appky Fortuny/iSport.
        OCR vrací řádky s prefixovými ikonami, vzor je:
          (© Bohemians :      <- tým1
          úz 07.03. 15:00     <- datum+čas
          @ Slovan Liberec -  <- tým2
          (© Slovácko -       <- tým1
          k 07.03.15:00       <- datum+čas
          ...
        """
        import re
        from datetime import datetime

        current_year = datetime.now().year

        # Regex na datum+čas: "07.03. 15:00" nebo "07.03.15:00" nebo "7.3. 15:00"
        dt_re = re.compile(r'(\d{1,2})\.(\d{1,2})\.?\s*(\d{1,2}):(\d{2})')
        # Regex na samotné datum bez času: "25.KoLo", "25. kolo" -> skip
        kolo_re = re.compile(r'^\d+\.\s*kolo', re.I)
        # Junk header patterns
        skip_re = re.compile(
            r'^(bude se hrát|česko|chance liga|jme\s|\d+\.\s*kolo)',
            re.I
        )

        def strip_prefix(line):
            """Odstraní OCR prefix ikony na začátku řádku, zachová jméno týmu začínající velkým písmenem."""
            # Odstraní vše před prvním velkým písmenem (začátek jména týmu)
            m = re.search(r'[A-ZÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ]', line)
            if m:
                return line[m.start():].strip()
            return line.strip()

        def clean_team(name):
            """Odstraní trailing ' -', ' :', ' =' a podobné artefakty."""
            return re.sub(r'\s*[-:=]\s*$', '', name).strip()

        def parse_dt(line):
            """Vrátí datetime pokud řádek obsahuje datum+čas, jinak None."""
            m = dt_re.search(line)
            if m:
                try:
                    return datetime(current_year, int(m.group(2)), int(m.group(1)),
                                    int(m.group(3)), int(m.group(4)))
                except Exception:
                    return None
            return None

        lines = text.strip().split('\n')

        # Vyčisti a oklasifikuj každý řádek
        # Každý item: ('dt', datetime) | ('team', str, score_or_None)
        classified = []
        score_re = re.compile(r'\s+(\d{1,2})\s*$')  # trailing score digit(s)
        for raw in lines:
            raw = raw.strip()
            if not raw or len(raw) < 3:
                continue
            if skip_re.match(raw.lower()):
                continue
            if kolo_re.match(raw):
                continue

            dt = parse_dt(raw)
            if dt:
                classified.append(('dt', dt, None))
            else:
                # Check for trailing score: '(© Sigma Olomouc 1' -> team='Sigma Olomouc', score=1
                score_match = score_re.search(raw)
                score = int(score_match.group(1)) if score_match else None
                line_no_score = raw[:score_match.start()] if score_match else raw
                clean = clean_team(strip_prefix(line_no_score))
                if clean and len(clean) >= 3:
                    classified.append(('team', clean, score))

        # Nyní hledáme vzory: team, dt, team  NEBO  dt, team, team
        results = []
        i = 0
        while i < len(classified):
            # Vzor 1: team1(+score), dt, team2(+score)  (Fortuna/iSport appka s výsledky)
            if (i + 2 < len(classified)
                    and classified[i][0] == 'team'
                    and classified[i+1][0] == 'dt'
                    and classified[i+2][0] == 'team'):
                home = classified[i][1]
                home_score = classified[i][2]
                dt = classified[i+1][1]
                away = classified[i+2][1]
                away_score = classified[i+2][2]
                entry = {
                    'home_team': home,
                    'away_team': away,
                    'start_time': dt.isoformat(),
                }
                if home_score is not None and away_score is not None:
                    entry['home_score'] = home_score
                    entry['away_score'] = away_score
                results.append(entry)
                i += 3
                continue
            # Vzor 2: dt, team1(+score), team2(+score)  (čistý screenshot)
            if (i + 2 < len(classified)
                    and classified[i][0] == 'dt'
                    and classified[i+1][0] == 'team'
                    and classified[i+2][0] == 'team'):
                dt = classified[i][1]
                home = classified[i+1][1]
                home_score = classified[i+1][2]
                away = classified[i+2][1]
                away_score = classified[i+2][2]
                entry = {
                    'home_team': home,
                    'away_team': away,
                    'start_time': dt.isoformat(),
                }
                if home_score is not None and away_score is not None:
                    entry['home_score'] = home_score
                    entry['away_score'] = away_score
                results.append(entry)
                i += 3
                continue
            i += 1

        return results


    def smart_parse_matches(text: str, round_id: int = None) -> List[Dict]:
        """
        🤖 ULTRA SMART PARSER V2 - Better whitespace handling + OCR cleanup

        Podporované formáty:
        - Table format: "27. 2. 2026DuklaSlavia18:00" 
        - With spaces: "27. 2. 2026 Sparta - Slavia 18:00"
        - Date + time first: "27. 2. 20:00 Sparta vs Slavia"
        - CSV, Fortuna, UEFA, etc.
        - OCR from screenshots (with cleanup)

        Improvements:
        - Skip non-match lines (headers like "24. kolo")
        - Auto year-fill for partial dates
        - Better team splitting
        - OCR artifact cleanup
        """

        # TRY 1: Multi-line app format (with -- separators)
        if '--' in text:
            print("🎯 Detected multi-line app format")
            matches = _parse_multiline_app_format(text)
            if matches:
                print(f"✅ Multi-line parser: Found {len(matches)} matches")
                # Convert datetime to ISO
                for match in matches:
                    if 'start_time' in match and isinstance(match['start_time'], datetime):
                        match['start_time'] = match['start_time'].isoformat()
                return matches

        # Step 1: Clean OCR artifacts (vice>, |, extra whitespace)
        text = _clean_ocr_artifacts(text)

        # TRY 0: Screenshot multi-line OCR format (datetime / team1 / team2)
        ocr_multiline = _parse_multiline_ocr_format(text)
        if ocr_multiline:
            print(f"✅ OCR multiline parser: Nalezeno {len(ocr_multiline)} zápasů")
            return ocr_multiline

        # Step 2: Handle table copy/paste (all lines joined)
        text = _split_joined_lines(text)

        matches = []
        lines = text.strip().split('\n')

        print(f"🤖 Smart Parser V2: Zpracovávám {len(lines)} řádků")

        for line_num, line in enumerate(lines, 1):
            # Don't normalize whitespace yet - some formats need it
            line_raw = line.strip()

            if not line_raw or len(line_raw) < 5:
                continue

            # Skip lines that are obviously not matches
            if line_raw.lower() in ['kolo', 'round', 'zápasy', 'matches']:
                print(f"⏭️ Přeskakuji header: {line_raw}")
                continue
            if re.match(r'^\d+\.\s*kolo', line_raw, re.I):
                print(f"⏭️ Přeskakuji kolo header: {line_raw}")
                continue
            if re.match(r'česko|chance liga|jme\s|bude se hrát', line_raw.lower()):
                print(f"⏭️ Přeskakuji soutěž header: {line_raw}")
                continue

            match_data = _parse_single_line(line_raw, line_num)
            if match_data:
                # Convert datetime to ISO string for JSON
                if 'start_time' in match_data and match_data['start_time']:
                    if isinstance(match_data['start_time'], datetime):
                        match_data['start_time'] = match_data['start_time'].isoformat()

                matches.append(match_data)

        print(f"✅ Smart Parser V2: Nalezeno {len(matches)} zápasů")
        return matches


    def _is_date(text: str) -> bool:
        """Check if text looks like a date"""
        date_patterns = [
            r'^\d{1,2}\.\s*\d{1,2}\.\s*\d{4}$',  # 27. 2. 2026
            r'^\d{1,2}/\d{1,2}/\d{4}$',          # 27/2/2026
            r'^\d{4}-\d{1,2}-\d{1,2}$',          # 2026-2-27
            r'^\d{1,2}\.\s*\d{1,2}\.\s*$',       # 27. 2.
            r'^\d{1,2}\.\s*\d{1,2}\.$',          # 27.2.
        ]
        text = text.strip()
        for pattern in date_patterns:
            if re.match(pattern, text):
                return True
        return False



    def _parse_compact_date_format(line: str, parts: List[str]) -> Optional[Dict]:
        """
        Parse compact date format (common from OCR):
        "7.3.2026 Teplice Dukla 15:00"
        Format: DATE TEAM1 TEAM2 TIME
        """
        date_str = parts[0]
        rest_parts = parts[1:]

        dt = _parse_datetime(date_str)
        if not dt:
            return None

        # Check if last part is time
        time_str = None
        if rest_parts and re.match(r'^\d{1,2}:\d{2}$', rest_parts[-1]):
            time_str = rest_parts[-1]
            team_parts = rest_parts[:-1]
        else:
            team_parts = rest_parts

        if len(team_parts) < 2:
            return None

        # Split teams
        if len(team_parts) == 2:
            home_team, away_team = team_parts[0], team_parts[1]
        else:
            teams_text = ' '.join(team_parts)
            home_team, away_team = _smart_split_teams(teams_text)
            if not home_team or not away_team:
                mid = len(team_parts) // 2
                home_team = ' '.join(team_parts[:mid])
                away_team = ' '.join(team_parts[mid:])

        # Add time to datetime
        if time_str and dt:
            try:
                time_obj = datetime.strptime(time_str, "%H:%M").time()
                dt = dt.replace(hour=time_obj.hour, minute=time_obj.minute)
            except:
                pass

        return {
            'home_team': home_team.strip(),
            'away_team': away_team.strip(),
            'start_time': dt,
        }

    def _parse_date_first_format(line: str, parts: List[str]) -> Optional[Dict]:
        """Parse: 27. 2. 2026 Dukla Slavia 18:00 OR 27. 2. 2026 Dukla Slavia 0:2"""

        date_str = ' '.join(parts[:3])  # "27. 2. 2026"
        rest = ' '.join(parts[3:])       # "Dukla Slavia 18:00" or "Dukla Slavia 0:2"

        dt = _parse_datetime(date_str)

        # Check for time/score at end
        time_or_score_match = re.search(r'(\d{1,2}):(\d{1,2})$', rest)

        is_score = False
        home_score = None
        away_score = None

        if time_or_score_match:
            first_num = int(time_or_score_match.group(1))
            second_num = int(time_or_score_match.group(2))
            full_match = time_or_score_match.group(0)
            teams_part = rest[:rest.rfind(full_match)].strip()

            # Determine if it's time or score
            if first_num >= 10 and first_num <= 23:
                is_score = False  # Likely time
            elif second_num in [0, 15, 30, 45] and first_num <= 23:
                is_score = False  # Common time minutes
            elif first_num < 10 and second_num < 10:
                is_score = True  # Both small - likely score
            else:
                is_score = first_num > 23 or second_num > 59

            if is_score:
                home_score = first_num
                away_score = second_num
            else:
                # It's a time
                if dt:
                    try:
                        time_obj = datetime.strptime(full_match, "%H:%M").time()
                        dt = dt.replace(hour=time_obj.hour, minute=time_obj.minute)
                    except:
                        pass
        else:
            teams_part = rest

        # Parse teams
        if ' - ' in teams_part:
            teams = teams_part.split(' - ', 1)
        elif ' vs ' in teams_part.lower():
            teams = re.split(r'\s+vs\s+', teams_part, flags=re.I, maxsplit=1)
        else:
            # Try to split by space
            teams = teams_part.split()
            if len(teams) < 2:
                return {
                    'home_team': teams_part.strip(),
                    'away_team': '',
                    'start_time': dt,
                }

        if len(teams) >= 2:
            result = {
                'home_team': teams[0].strip(),
                'away_team': teams[1].strip() if len(teams) > 1 else teams[-1].strip(),
                'start_time': dt,
            }

            if home_score is not None:
                result['home_score'] = home_score
                result['away_score'] = away_score

            return result

        return None

    def _parse_date_time_first(line: str, parts: List[str]) -> Optional[Dict]:
        """Parse: 27. 2. 20:00 Sparta vs Slavia"""

        date_str = ' '.join(parts[:2])  # "27. 2."
        time_str = parts[2]              # "20:00"
        rest = ' '.join(parts[3:])       # "Sparta vs Slavia"

        dt = _parse_datetime(date_str + ' ' + time_str)

        # Parse teams
        if ' - ' in rest:
            teams = rest.split(' - ', 1)
        elif ' vs ' in rest.lower():
            teams = re.split(r'\s+vs\s+', rest, flags=re.I, maxsplit=1)
        else:
            return None

        if len(teams) >= 2:
            return {
                'home_team': teams[0].strip(),
                'away_team': teams[1].strip(),
                'start_time': dt,
            }

        return None



    def _parse_table_format(line: str) -> Optional[Dict]:
        """
        IMPROVED: Parse table copy/paste - handles NO SPACES between parts

        Examples:
        - 27. 2. 2026DuklaSlavia18:00      (time)
        - 27. 2. 2026DuklaSlavia0:2        (score)
        - 28. 2. 2026Ml. Boleslav__Jablonec__15:00
        """

        # Pattern: DATE (required) + TEAMS (any) + TIME/SCORE (required)

        # Match date at start
        date_match = re.match(r'^(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})', line)
        if not date_match:
            return None

        date_str = date_match.group(1)
        rest = line[len(date_str):].strip()

        # Match time OR score at end
        # Time: HH:MM (hour 0-23, minute 00-59)
        # Score: N:N (any digits)
        time_score_match = re.search(r'(\d{1,2}:\d{1,2})$', rest)
        if not time_score_match:
            return None

        time_or_score = time_score_match.group(1)
        teams_part = rest[:rest.rfind(time_or_score)].strip()

        if not teams_part:
            return None

        # Determine if it's time or score
        parts = time_or_score.split(':')
        hour_or_home = int(parts[0])
        min_or_away = int(parts[1])

        # Heuristic: If first number > 23 OR second number > 59, it's probably score
        # OR if both numbers are single digit and < 10, could be score
        is_score = False
        is_time = False

        # Definite time: HH:MM format where HH in 0-23 and MM in 00-59
        if 0 <= hour_or_home <= 23 and 0 <= min_or_away <= 59:
            # Could be time... but also could be low score like 2:1
            # Check: if MM is 00, 15, 30, 45 → likely time
            if min_or_away in [0, 15, 30, 45]:
                is_time = True
            # If HH >= 10, likely time (games don't usually have 10+ goals)
            elif hour_or_home >= 10:
                is_time = True
            # If both small (< 10), likely score
            else:
                is_score = True
        else:
            # Outside valid time range → score
            is_score = True

        # Parse teams
        if '__' in teams_part:
            parts = [p.strip() for p in teams_part.split('__') if p.strip()]
            if len(parts) >= 2:
                home_team = parts[0]
                away_team = parts[1]
            else:
                return None
        else:
            home_team, away_team = _smart_split_teams(teams_part)
            if not home_team or not away_team:
                return None

        # Build result
        result = {
            'home_team': home_team,
            'away_team': away_team,
        }

        if is_time:
            # Parse as time
            try:
                dt = datetime.strptime(f"{date_str} {time_or_score}", "%d. %m. %Y %H:%M")
                result['start_time'] = dt
            except:
                result['start_time'] = None
        else:
            # Parse as score
            result['home_score'] = hour_or_home
            result['away_score'] = min_or_away
            # Parse just the date (no time)
            try:
                dt = datetime.strptime(date_str, "%d. %m. %Y")
                result['start_time'] = dt
            except:
                result['start_time'] = None

        return result


    def _smart_split_teams(text: str) -> Tuple[Optional[str], Optional[str]]:
        """
        IMPROVED: Smart split teams from text without separator

        Examples:
        - "DuklaSlavia" → "Dukla", "Slavia"
        - "SpartaOstrava" → "Sparta", "Ostrava"
        - "Ml. BoleslavJablonec" → "Ml. Boleslav", "Jablonec"
        """

        known_teams = [
            # Longer names first for greedy matching
            'Ml. Boleslav', 'Mladá Boleslav',
            'Hradec Kr.', 'Hradec Králové', 'Hradec',
            'FK Dukla Praha', 'SK Slavia Praha', 'AC Sparta Praha',
            'FC Viktoria Plzeň', 'FC Zlín', 'FC Slovan Liberec',
            'FC Baník Ostrava', 'FK Teplice', 'FK Jablonec',
            'MFK Karviná', '1.FC Slovácko', '1. FC Slovácko', 'FK Pardubice',
            'SK Sigma Olomouc', 'Bohemians Praha 1905',
            'FK Mladá Boleslav', 'FC Hradec Králové',
            # Short names
            'Dukla', 'Slavia', 'Sparta', 'Plzeň', 'Zlín',
            'Liberec', 'Ostrava', 'Baník', 'Teplice', 'Jablonec',
            'Karviná', 'Slovácko', 'Pardubice', 'Olomouc', 'Sigma',
            'Bohemians',
        ]

        # Try to match known team from start
        for team in sorted(known_teams, key=len, reverse=True):
            if text.startswith(team):
                home = team
                away = text[len(team):].strip()

                # Check if away is also a known team
                for away_team in known_teams:
                    if away == away_team or away.startswith(away_team):
                        return home, away_team

                # Try away as-is
                if away:
                    return home, away

        # Fallback: split at second capital letter
        capitals = [i for i, c in enumerate(text) if c.isupper()]
        if len(capitals) >= 2:
            split_pos = capitals[1]
            return text[:split_pos], text[split_pos:]

        return None, None


    def _parse_single_line(line: str, line_num: int = 0) -> Optional[Dict]:
        """
        IMPROVED: Better handling of different formats

        Priority:
        1. Table format (no spaces): 27. 2. 2026DuklaSlavia18:00
        2. Date with spaces: 27. 2. 2026 Dukla - Slavia 18:00
        3. Other formats
        """

        # TRY 1: Table format FIRST (before whitespace normalization!)
        result = _parse_table_format(line)
        if result:
            return result

        # TRY 2: Now normalize whitespace for other formats
        line_normalized = re.sub(r'\s+', ' ', line.replace('\t', ' ')).strip()
        parts = line_normalized.split()  # initialize here to avoid UnboundLocalError

        # TRY 3: Compact date format (OCR): "7.3.2026 Team1 Team2 15:00"
        if len(parts) >= 4 and _is_date(parts[0]):
            result = _parse_compact_date_format(line_normalized, parts)
            if result:
                return result

        # TRY 4: Date at start with spaces
        # parts already assigned above

        # "27. 2. 2026 Dukla - Slavia 18:00"
        if len(parts) >= 3 and _is_date(' '.join(parts[:3])):
            return _parse_date_first_format(line_normalized, parts)

        # "27. 2. 20:00 Sparta vs Slavia"
        if len(parts) >= 4:
            potential_date = ' '.join(parts[:2])
            potential_time = parts[2]

            if _is_date(potential_date) and re.match(r'^\d{1,2}:\d{2}$', potential_time):
                return _parse_date_time_first(line_normalized, parts)

        # TRY 5: Other parsers
        parsers = [
            _parse_fortuna_style,
            _parse_uefa_style,
            _parse_csv_style,
            _parse_pipe_style,
            _parse_score_style,
            _parse_vs_style,
            _parse_dash_style,
        ]

        for parser in parsers:
            try:
                result = parser(line_normalized)
                if result:
                    return result
            except Exception:
                continue

        print(f"⚠️ Nepodařilo se parsovat řádek {line_num}: {line[:50]}")
        return None


    def _parse_fortuna_style(line: str) -> Optional[Dict]:
        """Fortuna Liga: #22 14/02/26Sat 15:00 DUK 0:0 FCZ"""
        pattern = r'#(\d+)\s+(\d{2})/(\d{2})/(\d{2})(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\d{1,2}):(\d{2})\s+([A-Z]{3,4})\s+(?:(\d+):(\d+)|[-–])\s+([A-Z]{3,4})'
        m = re.search(pattern, line)
        if not m:
            return None

        day, month, year_short = int(m.group(2)), int(m.group(3)), int(m.group(4))
        hour, minute = int(m.group(5)), int(m.group(6))
        home_code, away_code = m.group(7), m.group(10)
        home_score = int(m.group(8)) if m.group(8) else None
        away_score = int(m.group(9)) if m.group(9) else None

        year = 2000 + year_short if year_short < 50 else 1900 + year_short

        # Team code to name
        team_map = {
            'ACS': 'AC Sparta Praha', 'SKS': 'SK Slavia Praha',
            'PLZ': 'FC Viktoria Plzeň', 'LIB': 'FC Slovan Liberec',
            'FCB': 'FC Baník Ostrava', 'MBL': 'FK Mladá Boleslav',
            'FKJ': 'FK Jablonec', 'BOH': 'Bohemians Praha 1905',
            'FCS': '1.FC Slovácko', 'SIG': 'SK Sigma Olomouc',
            'TEP': 'FK Teplice', 'HKR': 'FC Hradec Králové',
            'PCE': 'FK Pardubice', 'KAR': 'MFK Karviná',
            'FCZ': 'FC Zlín', 'DUK': 'FK Dukla Praha',
        }

        return {
            'home_team': team_map.get(home_code, home_code),
            'away_team': team_map.get(away_code, away_code),
            'start_time': datetime(year, month, day, hour, minute),
            'home_score': home_score,
            'away_score': away_score,
        }


    def _parse_uefa_style(line: str) -> Optional[Dict]:
        """UEFA: Juventus 2-1 Galatasaray (18:45 CET)"""
        # S časem
        pattern = r'(.+?)\s+(\d+)[-–:](\d+)\s+(.+?)\s*\((\d{1,2}):(\d{2})'
        m = re.search(pattern, line)
        if m:
            home = m.group(1).strip()
            away = m.group(4).strip()
            hour, minute = int(m.group(5)), int(m.group(6))
            return {
                'home_team': home,
                'away_team': away,
                'home_score': int(m.group(2)),
                'away_score': int(m.group(3)),
                'start_time': datetime.now().replace(hour=hour, minute=minute),
            }

        # Bez času
        pattern = r'(.+?)\s+(\d+)[-–](\d+)\s+(.+?)$'
        m = re.search(pattern, line)
        if m:
            return {
                'home_team': m.group(1).strip(),
                'away_team': m.group(4).strip(),
                'home_score': int(m.group(2)),
                'away_score': int(m.group(3)),
            }

        return None


    def _parse_csv_style(line: str) -> Optional[Dict]:
        """CSV: Sparta,Slavia,2,1,14.2.2026 20:00

        Robustnější:
        - umí ignorovat úvodní kódový sloupec (např. "Ml."/"ACS"/"MBL")
        - umí ignorovat úvodní datumový sloupec (např. "27. 2. 2026,Sparta,Slavia,18:00")
        - umí sloučit rozpadlý název týmu typu "Ml." + "Boleslav"
        """
        if ',' not in line:
            return None

        parts = [p.strip() for p in line.split(',') if p.strip()]
        if len(parts) < 2:
            return None

        def _looks_like_code(tok: str) -> bool:
            t = (tok or '').strip()
            if not t or ' ' in t or len(t) > 5:
                return False
            if re.match(r'^[A-Za-z]{1,4}\.?$', t):
                return True
            if t.isupper() and re.match(r'^[A-Z0-9]{2,5}$', t):
                return True
            return False

        leading_dt = None
        if len(parts) >= 3 and _is_date(parts[0]):
            leading_dt = parts[0]
            parts = parts[1:]

        if len(parts) >= 3 and _looks_like_code(parts[0]) and (len(parts[1]) >= 5 or ' ' in parts[1]):
            parts = parts[1:]

        if len(parts) >= 3 and parts[0].endswith('.') and len(parts[0]) <= 4 and ' ' not in parts[1] and parts[1].isalpha():
            merged = f"{parts[0]} {parts[1]}"
            parts = [merged] + parts[2:]

        if len(parts) < 2:
            return None

        result: Dict[str, Any] = {'home_team': parts[0], 'away_team': parts[1]}
        tail = parts[2:]

        def _is_probable_time(tok: str) -> bool:
            # HH:MM typical kick-off time (minutes usually 00/15/30/45)
            m = re.match(r'^(\d{1,2}):(\d{2})$', (tok or '').strip())
            if not m:
                return False
            hh = int(m.group(1))
            mm = int(m.group(2))
            return 0 <= hh <= 23 and 0 <= mm <= 59 and mm in (0, 15, 30, 45) and hh >= 8

        def _is_score_colon(tok: str) -> bool:
            return bool(re.match(r'^(\d{1,2}):(\d{1,2})$', (tok or '').strip()))

        def _apply_dt_from_tail(tokens: list[str]) -> None:
            nonlocal result, leading_dt
            if not tokens:
                return
            last = tokens[-1].strip()

            # if we have leading date from first column, allow time-only last column
            if leading_dt and _is_probable_time(last):
                dt = _parse_datetime(f"{leading_dt} {last}")
                if dt:
                    result['start_time'] = dt
                return

            # full datetime in last token
            dt = _parse_datetime(last)
            if dt:
                result['start_time'] = dt

        # --- score/time parsing in tail ---
        if tail:
            # Handle scores like "0:2"
            if _is_score_colon(tail[0]):
                try:
                    hs, aw = tail[0].split(':', 1)
                    result['home_score'] = int(hs)
                    result['away_score'] = int(aw)
                except Exception:
                    pass
                _apply_dt_from_tail(tail[1:])
                # If we have only date (no time), set a sensible default so the match is tipovatelný.
                if leading_dt and 'start_time' not in result:
                    dt = _parse_datetime(f"{leading_dt} 18:00")
                    if dt:
                        result['start_time'] = dt
                return result

            # Handle numeric scores "2,1"
            if len(tail) >= 2 and re.match(r'^\d+$', tail[0]) and re.match(r'^\d+$', tail[1]):
                result['home_score'] = int(tail[0])
                result['away_score'] = int(tail[1])
                _apply_dt_from_tail(tail[2:])
                return result

        _apply_dt_from_tail(tail)
        if leading_dt and 'start_time' not in result:
            # date-only without time; keep empty for preview unless you want default time
            dt = _parse_datetime(f"{leading_dt} 18:00")
            if dt:
                result['start_time'] = dt

        return result



    def _parse_pipe_style(line: str) -> Optional[Dict]:
        """Pipe separated: Sparta|Slavia|2|1|14.2.2026 20:00

        Robustnější:
        - umí ignorovat úvodní kódový sloupec (např. "Ml."/"ACS"/"MBL")
        - umí ignorovat úvodní datumový sloupec (např. "27. 2. 2026 | Sparta | Slavia | 18:00")
        - umí sloučit rozpadlý název týmu typu "Ml." + "Boleslav"
        """
        if '|' not in line:
            return None

        parts = [p.strip() for p in line.split('|') if p.strip()]
        if len(parts) < 2:
            return None

        def _looks_like_code(tok: str) -> bool:
            t = (tok or '').strip()
            if not t or ' ' in t or len(t) > 5:
                return False
            if re.match(r'^[A-Za-z]{1,4}\.?$', t):
                return True
            if t.isupper() and re.match(r'^[A-Z0-9]{2,5}$', t):
                return True
            return False

        leading_date = None
        if len(parts) >= 3 and _is_date(parts[0]):
            leading_date = parts[0]
            parts = parts[1:]

        if len(parts) >= 3 and _looks_like_code(parts[0]) and (len(parts[1]) >= 5 or ' ' in parts[1]):
            parts = parts[1:]

        if len(parts) >= 3 and parts[0].endswith('.') and len(parts[0]) <= 4 and ' ' not in parts[1] and parts[1].isalpha():
            merged = f"{parts[0]} {parts[1]}"
            parts = [merged] + parts[2:]

        if len(parts) < 2:
            return None

        result: Dict[str, Any] = {'home_team': parts[0], 'away_team': parts[1]}
        tail = parts[2:]

        def _is_probable_time(tok: str) -> bool:
            m = re.match(r'^(\d{1,2}):(\d{2})$', (tok or '').strip())
            if not m:
                return False
            hh = int(m.group(1))
            mm = int(m.group(2))
            return 0 <= hh <= 23 and 0 <= mm <= 59 and mm in (0, 15, 30, 45) and hh >= 8

        def _is_score_colon(tok: str) -> bool:
            return bool(re.match(r'^(\d{1,2}):(\d{1,2})$', (tok or '').strip()))

        # If we have leading date and the next token is a time, treat it as kick-off time
        if leading_date and tail and _is_probable_time(tail[0]):
            dt = _parse_datetime(f"{leading_date} {tail[0]}")
            if dt:
                result['start_time'] = dt
            tail = tail[1:]

        # Handle score in format "0:2" (very common in pasted tables)
        if tail and _is_score_colon(tail[0]) and not _is_probable_time(tail[0]):
            try:
                hs, aw = tail[0].split(':', 1)
                result['home_score'] = int(hs)
                result['away_score'] = int(aw)
            except Exception:
                pass
            tail = tail[1:]
            if leading_date and 'start_time' not in result:
                dt = _parse_datetime(f"{leading_date} 18:00")
                if dt:
                    result['start_time'] = dt

        # Handle numeric scores "2|1"
        if len(tail) >= 2 and re.match(r'^\d+$', str(tail[0])) and re.match(r'^\d+$', str(tail[1])):
            result['home_score'] = int(tail[0])
            result['away_score'] = int(tail[1])
            tail = tail[2:]

        # Remaining tail may contain full datetime
        if tail:
            dt = _parse_datetime(tail[0])
            if dt:
                result['start_time'] = dt
            elif leading_date and _is_probable_time(tail[0]):
                dt = _parse_datetime(f"{leading_date} {tail[0]}")
                if dt:
                    result['start_time'] = dt

        if leading_date and 'start_time' not in result:
            dt = _parse_datetime(f"{leading_date} 18:00")
            if dt:
                result['start_time'] = dt

        return result



    def _parse_datetime_first(line: str) -> Optional[Dict]:
        """Datum první: 14.2. 20:00 Sparta vs Slavia"""
        pattern = r'(\d{1,2})\.(\d{1,2})\.?\s+(\d{1,2}):(\d{2})\s+(.+?)\s+(?:vs|versus|-|–)\s+(.+?)$'
        m = re.search(pattern, line)
        if not m:
            return None

        day, month = int(m.group(1)), int(m.group(2))
        hour, minute = int(m.group(3)), int(m.group(4))
        year = datetime.now().year

        # Pokud je měsíc v minulosti, použij příští rok
        if month < datetime.now().month:
            year += 1

        return {
            'home_team': m.group(5).strip(),
            'away_team': m.group(6).strip(),
            'start_time': datetime(year, month, day, hour, minute),
        }


    def _parse_score_style(line: str) -> Optional[Dict]:
        """Se skóre: Sparta - Slavia 2:1"""
        pattern = r'(.+?)\s+[-–]\s+(.+?)\s+(\d+)[:\-](\d+)'
        m = re.search(pattern, line)
        if not m:
            return None

        return {
            'home_team': m.group(1).strip(),
            'away_team': m.group(2).strip(),
            'home_score': int(m.group(3)),
            'away_score': int(m.group(4)),
        }


    def _parse_vs_style(line: str) -> Optional[Dict]:
        """Vs style: Sparta vs Slavia"""
        pattern = r'(.+?)\s+(?:vs|versus)\s+(.+?)$'
        m = re.search(pattern, line, re.I)
        if not m:
            return None

        return {
            'home_team': m.group(1).strip(),
            'away_team': m.group(2).strip(),
        }


    def _parse_dash_style(line: str) -> Optional[Dict]:
        """Dash style: Sparta - Slavia"""
        pattern = r'^(.+?)\s+[-–]\s+(.+?)$'
        m = re.search(pattern, line)
        if not m:
            return None

        return {
            'home_team': m.group(1).strip(),
            'away_team': m.group(2).strip(),
        }


    def _parse_datetime(s: str) -> Optional[datetime]:
        """
        Parse datetime from string
        IMPROVED: Auto-fill missing year
        """
        s = s.strip()

        formats = [
            '%d. %m. %Y %H:%M',
            '%d. %m. %Y',
            '%d.%m.%Y %H:%M',
            '%d.%m.%Y',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%d. %m. %H:%M',
            '%d.%m. %H:%M',
            '%d. %m.',
            '%d.%m.',
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(s, fmt)
                # IMPROVEMENT: If year is missing (1900), add current/next year
                if dt.year == 1900:
                    year = datetime.now().year
                    if dt.month < datetime.now().month:
                        year += 1
                    dt = dt.replace(year=year)
                return dt
            except:
                continue

        return None


    def normalize_team_name(name: str, round_id: int = None) -> str:
        """
        🧠 Inteligentní normalizace jména týmu

        - Opraví běžné překlepy
        - Doplní plný název (Slavia → SK Slavia Praha)
        - Najde podobné týmy v databázi
        """

        if name is None:
            return ""
        if isinstance(name, bool):
            return ""
        if not isinstance(name, str):
            name = str(name)
        name = name.strip()
        name = name.replace("\u00a0", " ").replace("\t", " ").strip()
        name = re.sub(r"\s+", " ", name)
        # Normalize common Czech football prefix patterns
        name = re.sub(r"\b1\.?\s*FC\b", "1. FC", name, flags=re.IGNORECASE)
        name = re.sub(r"\bFK\b", "FK", name)

        # Aliasy z DB (spravované v Admin UI)
        if round_id:
            try:
                al = TeamAlias.query.filter_by(round_id=round_id, alias=name).first()
                if not al:
                    # case-insensitive match
                    al = TeamAlias.query.filter(TeamAlias.round_id == round_id, db.func.lower(TeamAlias.alias) == name.lower()).first()
                if al and al.canonical_name:
                    return al.canonical_name
            except Exception as e:
                print(f"⚠️ Chyba v TeamAlias lookup: {e}")

        # Zkratky → plné názvy
        short_to_full = {
            # Zkratky / běžné názvy → názvy, které typicky chceš mít v DB
            'Dukla': 'FK Dukla Praha',
            'Slavia': 'SK Slavia Praha',
            'Sparta': 'AC Sparta Praha',
            'Ostrava': 'FC Baník Ostrava',
            'Baník': 'FC Baník Ostrava',

            'Liberec': 'FC Slovan Liberec',
            'Hradec Kr.': 'FC Hradec Králové',
            'Hradec': 'FC Hradec Králové',

            'Ml. Boleslav': 'FK Mladá Boleslav',
            'Mladá Boleslav': 'FK Mladá Boleslav',
            'Jablonec': 'FK Jablonec',

            'Pardubice': 'FK Pardubice',
            'Teplice': 'FK Teplice',

            'Karviná': 'MFK Karviná',
            'Slovácko': '1.FC Slovácko',

            'Zlín': 'FC Zlín',
            'Plzeň': 'FC Viktoria Plzeň',

            'Olomouc': 'SK Sigma Olomouc',
            'Sigma': 'SK Sigma Olomouc',

            'Bohemians': 'Bohemians Praha 1905',
        }


        # Přesná shoda
        if name in short_to_full:
            return short_to_full[name]

        # Case-insensitive hledání
        for short, full in short_to_full.items():
            if name.lower() == short.lower():
                return full
            if name.lower() == full.lower():
                return full

        # Částečná shoda
        for short, full in short_to_full.items():
            if short.lower() in name.lower() or name.lower() in short.lower():
                return full
        # Pokud je round_id, zkus najít v DB (fuzzy match na existující Team.name)
        if round_id:
            try:
                if not isinstance(name, str):
                    name = str(name)

                # Načti názvy týmů pro tuto soutěž
                team_names = [
                    (t.name or "").strip()
                    for t in Team.query.filter_by(round_id=round_id, is_deleted=False).all()
                    if t and t.name
                ]

                needle = name.lower()
                for tname in team_names:
                    t_low = tname.lower()
                    if needle in t_low or t_low in needle:
                        return tname
            except Exception as e:
                print(f"⚠️ Chyba v normalize_team_name: {e}")
                pass
        # Žádná shoda - vrať original
        return name

